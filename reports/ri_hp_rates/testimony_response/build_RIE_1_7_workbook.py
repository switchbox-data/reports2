"""Build the supporting workbook for Schedule JPV-3 (RIE 1-7 data request).

RIE Data Request 1-7, Pre-Filed Direct Testimony of Juan-Pablo Velez:

    "I then used the hourly electric load profiles from the National Renewable
    Energy Laboratory's (NREL) ResStock dataset to measure how much load each
    subclass had during those high-cost hours. This gave me each subclass's share
    of current cost-causation, which I applied to the total Residential Delivery
    Revenue Requirement, in line with the Company's cost allocation methodology."

This script produces Schedule JPV-3: an ``.xlsx`` documenting the end-to-end
cost-allocation methodology — hourly marginal cost derivation, ResStock load
aggregation, economic-burden calculation, EPMC allocation, and a
testimony-validation cross-reference.

Run from the report directory::

    uv run python -m testimony_response.build_RIE_1_7_workbook
    uv run python -m testimony_response.build_RIE_1_7_workbook --title "RIE 1-7 (revised)"
    uv run python -m testimony_response.build_RIE_1_7_workbook --no-upload \\
        --output cache/schedule_jpv3_cost_allocation.xlsx

See ``cost_of_service_by_subclass.qmd`` for the published-side aggregation
logic that this workbook reproduces with formulas.
"""

from __future__ import annotations

import argparse
import math
import pickle
import subprocess
import sys
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────

# Report root — ri_hp_rates/ — contains the testimony cache/ directory
REPORT_DIR = Path(__file__).resolve().parents[1]

UTILITY = "rie"
STATE = "ri"
BATCH = "ri_20260331_r1-20_rate_case_test_year"
RUN_DELIVERY = "1"
RUN_SUPPLY = "2"
S3_BASE = "s3://data.sb/switchbox/cairo/outputs/hp_rates"
RESSTOCK_BASE = "s3://data.sb/nrel/resstock/res_2024_amy2018_2"
LOCAL_RESSTOCK_BASE = Path("/ebs/data/nrel/resstock/res_2024_amy2018_2")
LOCAL_RESSTOCK_METADATA = LOCAL_RESSTOCK_BASE / "metadata_utility" / "state=RI" / "utility_assignment.parquet"
LOCAL_RESSTOCK_LOADS_UPGRADE0 = LOCAL_RESSTOCK_BASE / "load_curve_hourly" / "state=RI" / "upgrade=00"
S3_MC_DIST_SUB_TX = "s3://data.sb/switchbox/marginal_costs/ri/dist_and_sub_tx/utility=rie/year=2025/data.parquet"
S3_MC_BULK_TX = "s3://data.sb/switchbox/marginal_costs/ri/bulk_tx/utility=rie/year=2025/data.parquet"
ELEC_TOTAL_COL = "out.electricity.total.energy_consumption"  # kWh per hour
SUBCLASS_COL = "postprocess_group.heating_type"
SUBCLASS_ORDER = ["heat_pump", "electrical_resistance", "fossil_fuel"]

# RDP git ref — used only to build permalink URLs for source attribution in the workbook.
# Do NOT fetch from rate-design-platform at runtime; use the hardcoded constants below.
RDP_REF = "e9e5088"
RDP_GITHUB_BASE = "https://github.com/switchbox-data/rate-design-platform/blob"
REPORTS2_GITHUB_BASE = "https://github.com/switchbox-data/reports2/blob"

# Revenue-requirement constants — sourced from rate-design-platform @ e9e5088:
#   rate_design/hp_rates/ri/config/rev_requirement/rie_rate_case_test_year.yaml
#   rate_design/hp_rates/ri/config/rev_requirement/rie_hp_vs_nonhp_rate_case_test_year.yaml
# Update these when rate-case inputs change.
REV_REQ: dict = {
    "total_delivery_revenue_requirement": 446463143.03,
    "test_year_customer_count": 419347.83,
    "resstock_kwh_scale_factor": 0.9568112362177266,
    # subclass_customers not present in source YAMLs; per-subclass counts unavailable
    "subclass_customers": {},
}

DEFAULT_FOLDER_ID = "1uPcJbcOChD6zoFuPb-gsxSByPr7xwmCH"
DEFAULT_TITLE = "RIE 1-7"

# Styling constants
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
SUBSECTION_FILL = PatternFill(start_color="E4ECF7", end_color="E4ECF7", fill_type="solid")
NUMBER_FORMAT_CURRENCY = "$#,##0"
NUMBER_FORMAT_KWH = "#,##0"
NUMBER_FORMAT_PCT = "0.0%"


# ── Permalink helpers ─────────────────────────────────────────────────────────


def _rdp_permalink(rel_path: str) -> str:
    """SHA-pinned GitHub permalink for a rate-design-platform file."""
    return f"{RDP_GITHUB_BASE}/{RDP_REF}/{rel_path}"


def _reports2_head_sha() -> str:
    """Current HEAD sha of the reports2 repo. Cached."""
    if not hasattr(_reports2_head_sha, "_cached"):
        repo_root = Path(__file__).resolve().parents[3]
        sha = subprocess.check_output(
            ["git", "-C", str(repo_root), "rev-parse", "HEAD"],
            text=True,
        ).strip()
        _reports2_head_sha._cached = sha  # type: ignore[attr-defined]
    return _reports2_head_sha._cached  # type: ignore[attr-defined]


def _reports2_permalink(rel_path: str) -> str:
    """SHA-pinned GitHub permalink for a file in the reports2 repo."""
    return f"{REPORTS2_GITHUB_BASE}/{_reports2_head_sha()}/{rel_path}"


# ── Data loading ──────────────────────────────────────────────────────────────


def get_aws_storage_options() -> dict:
    """Return AWS storage options for S3 access (uses ambient credentials)."""
    return {}


def load_revenue_requirement_yaml() -> dict:
    """Return the hardcoded revenue-requirement parameters.

    Sourced from rate-design-platform @ e9e5088 (rie_rate_case_test_year.yaml and
    rie_hp_vs_nonhp_rate_case_test_year.yaml).  Update REV_REQ above when inputs change.
    """
    return dict(REV_REQ)


def load_master_bat_data(
    batch: str, run_delivery: str, run_supply: str, utility: str = UTILITY
) -> pl.DataFrame:
    """Load master BAT data from S3 for the specified run."""
    s3_path = (
        f"{S3_BASE}/{STATE}/all_utilities/{batch}/"
        f"run_{run_delivery}+{run_supply}/cross_subsidization_BAT_values/"
    )
    print(f"Loading master BAT from: {s3_path}", flush=True)

    storage_options = get_aws_storage_options()
    lf = pl.scan_parquet(s3_path, storage_options=storage_options)

    # Filter to utility and collect
    df = lf.filter(pl.col("sb.electric_utility") == utility).collect()
    assert isinstance(df, pl.DataFrame)
    print(f"Loaded {df.height:,} buildings for utility={utility}", flush=True)
    return df


def load_aggregate_load_curves(
    kwh_scale_factor: float = 1.0,
) -> tuple[pl.DataFrame, pl.DataFrame]:
    """Load ResStock load curves and aggregate by heating subclass.

    Reads from local EBS disk (source: s3://data.sb/nrel/resstock/res_2024_amy2018_2/).
    Returns (agg_by_subclass, mc_delivery) where:
      - agg_by_subclass: 8760 rows × (timestamp + one column per subclass), in kWh
      - mc_delivery: 8760 rows × (timestamp, mc_dist_sub_tx, mc_bulk_tx, mc_delivery_total)

    kwh_scale_factor: CAIRO's ``resstock_kwh_scale_factor`` from the revenue-requirement
      YAML.  Applied to all subclass load columns so that the aggregate loads match the
      kWh series CAIRO uses internally when computing economic burden and EPMC allocation.
      Default 1.0 (no scaling) preserves backward-compat for callers that don't pass it.
    """
    # Load metadata to get bldg_id → subclass + weight for RIE upgrade=0 buildings
    print("Loading ResStock metadata ...", flush=True)
    meta = pl.read_parquet(LOCAL_RESSTOCK_METADATA).filter(
        (pl.col("sb.electric_utility") == UTILITY) & (pl.col("upgrade") == 0)
    ).select("bldg_id", SUBCLASS_COL, "weight")

    # Load all per-building 8760 load files and join with metadata
    print(f"Loading {meta.height} ResStock load curves from local disk ...", flush=True)
    frames: list[pl.DataFrame] = []
    load_dir = LOCAL_RESSTOCK_LOADS_UPGRADE0
    bldg_set = set(meta["bldg_id"].to_list())
    for fname in sorted(load_dir.iterdir()):
        bldg_id = int(fname.stem.split("-")[0])
        if bldg_id not in bldg_set:
            continue
        lf = pl.read_parquet(fname).select("timestamp", ELEC_TOTAL_COL)
        lf = lf.with_columns(pl.lit(bldg_id).alias("bldg_id"))
        frames.append(lf)

    loads = pl.concat(frames)
    loads = loads.join(meta, on="bldg_id")

    # Weighted sum per subclass per hour: Σ(load_kWh × weight)
    # Sort by timestamp to guarantee chronological order, then add 1-based hour_of_year index.
    agg = (
        loads.group_by(["timestamp", SUBCLASS_COL])
        .agg((pl.col(ELEC_TOTAL_COL) * pl.col("weight")).sum().alias("weighted_kwh"))
        .pivot(on=SUBCLASS_COL, index="timestamp", values="weighted_kwh")
        .sort("timestamp")
        .head(8760)  # AMY 2018 has 8784 hrs (leap year) — keep first 8760
    )
    # Assign hour_of_year 1..8760 so we can join MC by position, not by date.
    # ResStock uses AMY 2018 (a real calendar year), MC parquets use 2025.
    # The two series represent the same 8760-hour pattern; we align by order.
    agg = agg.with_columns(
        (pl.int_range(1, agg.height + 1, eager=True)).alias("hour_of_year")
    )

    # Apply CAIRO's kWh scale factor so that aggregate loads match the series CAIRO
    # uses when computing economic burden and EPMC allocation.  This normalises the
    # raw ResStock sample kWh to the utility's test-year residential kWh total.
    if kwh_scale_factor != 1.0:
        load_cols = [c for c in agg.columns if c not in ("timestamp", "hour_of_year")]
        agg = agg.with_columns([pl.col(c) * kwh_scale_factor for c in load_cols])

    # Load delivery MC parquets from S3
    print("Loading delivery marginal cost parquets from S3 ...", flush=True)

    def _load_mc(s3_path: str, name: str) -> pl.DataFrame:
        df = pl.read_parquet(s3_path)
        if df["timestamp"].dtype.time_zone is not None:  # type: ignore[union-attr]
            df = df.with_columns(pl.col("timestamp").dt.replace_time_zone(None))
        # Resolve the MC column by name before falling back to heuristics.
        # dist/sub-TX parquet: [timestamp, utility, year, mc_total_per_kwh]
        # bulk TX parquet:     [timestamp, bulk_tx_cost_enduse]
        # Using "first numeric column" picks up the integer `year` column from the
        # dist/sub-TX parquet, which is wrong.  Use explicit names first.
        if "mc_total_per_kwh" in df.columns:
            mc_col = "mc_total_per_kwh"
        elif "bulk_tx_cost_enduse" in df.columns:
            mc_col = "bulk_tx_cost_enduse"
        else:
            # Fallback: first float column (skips integer metadata columns like year)
            float_cols = [c for c in df.columns if c != "timestamp" and df[c].dtype.is_float()]
            mc_col = float_cols[0] if float_cols else [
                c for c in df.columns if c != "timestamp" and df[c].dtype.is_numeric()
            ][-1]
        df = df.select("timestamp", pl.col(mc_col).alias(name)).sort("timestamp").head(8760)
        return df.with_columns(
            (pl.int_range(1, df.height + 1, eager=True)).alias("hour_of_year")
        )

    mc_dist = _load_mc(S3_MC_DIST_SUB_TX, "mc_dist_sub_tx")
    mc_bulk = _load_mc(S3_MC_BULK_TX, "mc_bulk_tx")
    mc = (
        mc_dist.join(mc_bulk.drop("timestamp"), on="hour_of_year")
        .with_columns(
            (pl.col("mc_dist_sub_tx") + pl.col("mc_bulk_tx")).alias("mc_delivery_total")
        )
    )

    return agg, mc


def compute_cost_allocation(
    agg: pl.DataFrame,
    mc: pl.DataFrame,
    rev_req: dict,
) -> pl.DataFrame:
    """Compute cost allocation results from aggregate load curves and MC.

    For each subclass:
      economic_burden = Σ_h (weighted_load_kWh_h × mc_delivery_total_h)
      eb_share = economic_burden / total_eb
      allocated_rr = eb_share × total_delivery_rr

    Returns a DataFrame with one row per subclass.
    """
    total_rr = float(rev_req.get("total_delivery_revenue_requirement", 0))
    combined = agg.join(mc.select("hour_of_year", "mc_delivery_total"), on="hour_of_year")
    subclasses = [c for c in agg.columns if c not in ("timestamp", "hour_of_year")]

    rows = []
    for sc in subclasses:
        if sc not in combined.columns:
            continue
        eb = (combined[sc] * combined["mc_delivery_total"]).sum()
        rows.append({"subclass": sc, "economic_burden": eb})

    result = pl.DataFrame(rows)
    total_eb = result["economic_burden"].sum()
    result = result.with_columns(
        (pl.col("economic_burden") / total_eb).alias("eb_share"),
        (pl.col("economic_burden") / total_eb * total_rr).alias("allocated_rr"),
    )
    return result


def create_workbook() -> Workbook:
    """Create and configure the main workbook."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    return wb


def add_overview_sheet(wb: Workbook, rev_req: dict) -> None:
    """Add overview sheet explaining the cost allocation methodology."""
    ws = wb.create_sheet("Overview")

    # Title
    ws["A1"] = "Schedule JPV-3: Cost Allocation Methodology Workpapers"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    ws["A3"] = "Purpose"
    ws["A3"].font = SECTION_FONT
    ws["A3"].fill = SECTION_FILL
    ws.merge_cells("A3:F3")

    ws["A4"] = (
        "This workbook documents the complete cost allocation methodology used to "
        "determine each residential customer subclass's cost-of-service. This is a "
        "cost-of-service study performed at the subclass level - the same kind of exercise "
        "RIE performs at the class level in its Allocated Cost of Service Study (ACOSS), "
        "but applied one level deeper, within the residential class. The methodology applies "
        "the Equi-Proportional Marginal Cost (EPMC) allocation method, the standard reconciliation "
        "approach recommended by NARUC for bridging marginal cost analysis and revenue requirement recovery."
    )
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A4:F4")
    ws.row_dimensions[4].height = 75

    ws["A6"] = "Four-Step Cost-of-Service Methodology"
    ws["A6"].font = SECTION_FONT
    ws["A6"].fill = SECTION_FILL
    ws.merge_cells("A6:F6")

    steps = [
        (
            "Step 1: MEASURE COST-CAUSATION",
            "Constructed an 8,760-hour marginal cost signal representing the forward-looking cost "
            "of new delivery infrastructure for each hour of the Test Year. For each ResStock building, "
            "multiplied its hourly load by the hourly marginal cost and summed across all 8,760 hours. "
            "This annual marginal cost total is the building's 'economic burden' (EB) - its delivery "
            "cost-causation score. This is analogous to the NCP demand metric RIE uses at the class level, "
            "but provides a more granular signal that captures WHEN costs actually occur."
        ),
        (
            "Step 2: COMPUTE SUBCLASS SHARES",
            "Summed the economic burden across all buildings within each customer subclass "
            "(heat pump, natural gas, electric resistance, oil, propane), weighting by ResStock sample weight. "
            "Each subclass's share of the total economic burden represents its share of cost-causation. "
            "Customer counts by subclass were estimated from the 2020 Residential Energy Consumption Survey (RECS)."
        ),
        (
            "Step 3: ALLOCATE RESIDENTIAL DELIVERY REVENUE",
            f"Applied each subclass's share of cost-causation to the total Residential Delivery Revenue Requirement "
            f"(${rev_req.get('total_delivery_revenue_requirement', 0):,.0f}) to arrive at each subclass's allocated "
            "cost-of-service. This uses the Equi-Proportional Marginal Cost (EPMC) method: "
            "R_i = R × (EB_i / Σ(EB_j × weight_j)), where R_i is the revenue allocated to subclass i."
        ),
        (
            "Step 4: MEASURE THE CROSS-SUBSIDY",
            "Summed the delivery bills actually paid by each subclass under current default rates "
            "(computed by applying Test Year rates to each ResStock building's hourly load profile) "
            "and compared to the subclass's allocated cost of service from Step 3. A subclass that pays "
            "more than its cost of service is being overcharged - it is cross-subsidizing the other subclasses. "
            "This comparison is known as the Bill Alignment Test."
        ),
    ]

    row = 7
    for step_title, step_desc in steps:
        ws[f"A{row}"] = step_title
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = step_desc
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"B{row}:F{row}")
        ws.row_dimensions[row].height = 45
        row += 1

    ws["A15"] = "Key Formula: EPMC Residual Allocation"
    ws["A15"].font = SECTION_FONT
    ws["A15"].fill = SECTION_FILL
    ws.merge_cells("A15:F15")

    ws["A16"] = "R_i = R × (EB_i / Σ(EB_j × weight_j))"
    ws["A16"].font = Font(italic=True, size=11)
    ws.merge_cells("A16:F16")

    ws["A17"] = (
        "Where R_i is the residual cost allocated to subclass i, R is the total residual, "
        "EB_i is the economic burden of subclass i, and the sum is over all customers j."
    )
    ws["A17"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A17:F17")
    ws.row_dimensions[17].height = 30

    ws["A19"] = "Data Sources"
    ws["A19"].font = SECTION_FONT
    ws["A19"].fill = SECTION_FILL
    ws.merge_cells("A19:F19")

    sources = [
        (
            "Marginal Costs",
            "Bulk transmission ($69/kW-year) and sub-transmission/distribution ($80.24/kW-year) costs "
            "from AESC 2024 study (Synapse Energy Economics). AESC 2024 is the standard avoided cost "
            "input used by all New England states for energy efficiency cost-effectiveness screening. "
            "RI PUC Docket 4600 benefit-cost framework draws directly from AESC."
        ),
        (
            "ResStock Loads",
            "NREL ResStock End-Use Load Profiles 2024.2 release (AMY 2018). 8,760-hour profiles for "
            "every modeled dwelling, broken down by end use. Simulated using EnergyPlus, DOE's flagship "
            "building energy simulation engine. Validated against EIA Form 861 utility sales, utility "
            "load research, and smart meter data. ISO-NE uses ResStock for heat pump adoption forecasts."
        ),
        (
            "Customer Counts",
            "Heating-system subclass shares from 2020 Residential Energy Consumption Survey (RECS), "
            "applied to Test Year residential customer count."
        ),
        (
            "CAIRO Outputs",
            f"Economic burden and BAT calculations. Batch: {BATCH}, Runs: {RUN_DELIVERY}+{RUN_SUPPLY}. "
            "EPMC implementation in utils/mid/patches.py."
        ),
        (
            "Revenue Requirement",
            "RIE rate case test year (Sept 2024 - Aug 2025) from rie_hp_vs_nonhp_rate_case_test_year.yaml"
        ),
    ]

    row = 20
    for source_name, source_desc in sources:
        ws[f"A{row}"] = source_name
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = source_desc
        ws.merge_cells(f"B{row}:F{row}")
        row += 1

    row += 2
    ws[f"A{row}"] = "Why Hourly Marginal Costs vs. NCP Allocator?"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row}:F{row}")

    row += 1
    ws[f"A{row}"] = (
        "RIE's ACOSS uses Non-Coincident Peak (NCP) demand as the cost allocator at the class level. "
        "However, the NCP allocator produces inaccurate cost-of-service estimates when applied to residential "
        "customer subclasses. It confuses the MAGNITUDE of a peak with the COST it causes, which depends on WHEN it occurs.\n\n"
        "An NCP allocator measures each subgroup's peak demand without regard to when that peak occurs. At the subclass level, "
        "this produces results that contradict RIE's own engineering data about when the distribution grid is constrained. "
        "For example, heat pump customers' aggregate peak occurs on a cold winter morning. But RIE's feeder data shows "
        "abundant winter headroom and zero winter-constrained feeders. Heat pump customers' winter peak is NOT straining "
        "the grid and is NOT driving the Company to build more infrastructure.\n\n"
        "What is needed is a metric that captures WHEN costs actually occur on the transmission and distribution system "
        "and measures each subclass's load behavior during those high-cost hours. The hourly marginal cost approach does this."
    )
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row}:F{row}")
    ws.row_dimensions[row].height = 180

    row += 2
    ws[f"A{row}"] = "Probability-of-Peak Cost Allocation"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row}:F{row}")

    row += 1
    ws[f"A{row}"] = (
        "Delivery capacity costs are allocated using a probability-of-peak approach, where the annualized cost "
        "of new capacity investment is spread across only the hours most likely to drive that investment:\n\n"
        "• Bulk transmission: $69/kW-year allocated across top 100 hours of aggregate New England system load\n"
        "• Sub-transmission and distribution: $80.24/kW-year allocated across top 100 hours of RIE system load (summer only)\n\n"
        "Importantly, not every peak hour receives equal weight. Hours outside the top 100 receive zero cost. "
        "Within the top 100, hours closer to the peak carry MORE cost, because they are more likely to be the "
        "capacity-binding hour that triggers the need for a new investment. All other hours receive zero delivery capacity cost."
    )
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row}:F{row}")
    ws.row_dimensions[row].height = 150

    # Set column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80


def add_marginal_cost_sheet(wb: Workbook) -> None:
    """Add sheet documenting the hourly marginal cost signal construction."""
    ws = wb.create_sheet("1. Marginal Cost Signal")

    # Title
    ws["A1"] = "Step 1: Hourly Marginal Cost Signal Construction"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:G1")

    ws["A3"] = (
        "The hourly marginal cost signal captures the cost of the next kilowatt-hour "
        "in each hour of the year. It consists of four components:"
    )
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A3:G3")
    ws.row_dimensions[3].height = 30

    # Component table
    ws["A5"] = "Component"
    ws["B5"] = "Description"
    ws["C5"] = "Data Source"
    ws["D5"] = "Allocation Method"
    ws["E5"] = "Units"
    ws["F5"] = "Hours with Non-Zero MC"
    ws["G5"] = "Implementation"

    for cell in ["A5", "B5", "C5", "D5", "E5", "F5", "G5"]:
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
        ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    components = [
        (
            "Energy",
            "Short-run marginal cost of generation in each hour",
            "ISO-NE real-time LBMP (locational marginal price)",
            "All 8,760 hours",
            "$/kWh",
            "8,760 (all hours)",
            "generate_supply_energy_mc.py",
        ),
        (
            "Generation Capacity",
            "Cost of the next MW of generation adequacy",
            "ISO-NE Forward Capacity Auction (FCA) prices",
            "Top 8 hours per month by system load (exceedance-weighted)",
            "$/kW-month → $/kWh in peak hours",
            "~96 hours (8/month × 12)",
            "generate_supply_capacity_mc.py",
        ),
        (
            "Bulk Transmission",
            "Cost of bulk TX infrastructure reinforcements",
            "ISO-NE incremental benefit studies",
            "Top 40 hours per season by load (seasonal coincident peak)",
            "$/kW-year → $/kWh in peak hours",
            "~160 hours (40/season × 4)",
            "generate_bulk_tx_mc.py",
        ),
        (
            "Sub-TX + Distribution",
            "Cost of local delivery infrastructure",
            "RIE MCOS project-level capital, annualized and levelized (2026-2032)",
            "Top 100 hours by utility load (probability of peak)",
            "$/kW-year → $/kWh in peak hours",
            "100 hours",
            "generate_utility_tx_dx_mc.py",
        ),
    ]

    row = 6
    for comp in components:
        for col_idx, value in enumerate(comp, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 60
        row += 1

    ws["A11"] = "Total Hourly MC Formula"
    ws["A11"].font = SECTION_FONT
    ws["A11"].fill = SECTION_FILL
    ws.merge_cells("A11:G11")

    ws["A12"] = "MC_h = MC_energy_h + MC_gen_capacity_h + MC_bulk_tx_h + MC_dist_h"
    ws["A12"].font = Font(italic=True)
    ws.merge_cells("A12:G12")

    ws["A14"] = (
        "Where h indexes each of the 8,760 hours in the test year. Energy MC is non-zero "
        "in every hour; capacity and infrastructure MCs are concentrated in peak hours "
        "when system constraints are binding."
    )
    ws["A14"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A14:G14")
    ws.row_dimensions[14].height = 40

    ws["A16"] = "Probability-of-Peak Weighting"
    ws["A16"].font = SECTION_FONT
    ws["A16"].fill = SECTION_FILL
    ws.merge_cells("A16:G16")

    ws["A17"] = (
        "For bulk transmission and distribution capacity costs, a probability-of-peak approach is used "
        "where the annualized cost of new capacity investment is spread across only the top 100 hours most "
        "likely to drive that investment. Importantly, not every peak hour receives equal weight:\n\n"
        "• Hours outside the top 100 receive ZERO delivery capacity cost\n"
        "• Within the top 100, hours closer to the peak carry MORE cost, because they are more likely "
        "to be the capacity-binding hour that triggers the need for a new investment\n\n"
        "This weighted allocation reflects the engineering reality that not all hours contribute equally "
        "to infrastructure need. The full technical details of the weighting method are in Schedule JPV-2."
    )
    ws["A17"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A17:G17")
    ws.row_dimensions[17].height = 120

    ws["A19"] = "Forward-Looking vs. Embedded Costs"
    ws["A19"].font = SECTION_FONT
    ws["A19"].fill = SECTION_FILL
    ws.merge_cells("A19:G19")

    ws["A20"] = (
        "The marginal cost signal uses forward-looking incremental cost (FLIC): the cost "
        "of specific planned projects in the current capital pipeline, not the embedded "
        "cost of past investments. This is the standard approach in marginal-cost-of-service "
        "studies and matches the methodology used in the Bill Alignment Test literature "
        "(Simeone et al. 2023). The gap between marginal-cost-recoverable revenue and the "
        "total revenue requirement is the 'residual' — primarily the carrying charges on "
        "past infrastructure investments."
    )
    ws["A20"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A20:G20")
    ws.row_dimensions[20].height = 90

    # Set column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 25


def add_resstock_loads_sheet(wb: Workbook, rev_req: dict) -> None:
    """Add sheet documenting ResStock hourly load data."""
    ws = wb.create_sheet("2. ResStock Load Profiles")

    # Title
    ws["A1"] = "Step 2: NREL ResStock Hourly Load Profiles"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:F1")

    ws["A3"] = "Dataset Overview"
    ws["A3"].font = SECTION_FONT
    ws["A3"].fill = SECTION_FILL
    ws.merge_cells("A3:F3")

    n_buildings = int(rev_req.get("test_year_customer_count", 0))

    metadata = [
        ("Dataset", "NREL ResStock End-Use Load Profiles"),
        ("Release", "2024.2 (res_2024_amy2018_2)"),
        ("Weather Year", "AMY 2018 (Actual Meteorological Year)"),
        ("State", "Rhode Island"),
        ("Buildings Modeled", f"{n_buildings:,}"),
        ("Hours per Building", "8,760"),
        ("Total Data Points", f"{n_buildings * 8760:,}"),
        ("Simulation Engine", "EnergyPlus (DOE building energy simulation)"),
        ("S3 Path", f"{RESSTOCK_BASE}/"),
    ]

    row = 4
    for label, value in metadata:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        ws.merge_cells(f"B{row}:F{row}")
        row += 1

    ws["A13"] = "Building Characteristics"
    ws["A13"].font = SECTION_FONT
    ws["A13"].fill = SECTION_FILL
    ws.merge_cells("A13:F13")

    ws["A14"] = (
        "Each modeled dwelling in ResStock has physical characteristics drawn from federal "
        "survey data (RECS, U.S. Census) that reflect the statistical distribution across "
        "Rhode Island's housing stock:"
    )
    ws["A14"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A14:F14")
    ws.row_dimensions[14].height = 40

    characteristics = [
        "Building type (single-family detached, attached, mobile home, multifamily)",
        "Square footage and geometry",
        "Insulation levels (walls, attic, foundation)",
        "Window types and areas",
        "HVAC equipment (heating and cooling systems)",
        "Water heating equipment",
        "Appliances and lighting",
        "Local climate zone",
    ]

    row = 15
    for char in characteristics:
        ws[f"B{row}"] = f"• {char}"
        ws.merge_cells(f"B{row}:F{row}")
        row += 1

    ws[f"A{row+1}"] = "Load Profile Structure"
    ws[f"A{row+1}"].font = SECTION_FONT
    ws[f"A{row+1}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row+1}:F{row+1}")

    ws[f"A{row+2}"] = (
        "For each building, ResStock provides hourly consumption (kWh) for:"
    )
    ws.merge_cells(f"A{row+2}:F{row+2}")

    end_uses = [
        "Total electricity (out.electricity.total.energy_consumption)",
        "Solar PV generation, if present (out.electricity.pv.energy_consumption)",
        "Natural gas (out.natural_gas.total.energy_consumption)",
        "Heating oil and propane (separate columns)",
    ]

    row = row + 3
    for end_use in end_uses:
        ws[f"B{row}"] = f"• {end_use}"
        ws.merge_cells(f"B{row}:F{row}")
        row += 1

    ws[f"A{row+1}"] = "Representativeness"
    ws[f"A{row+1}"].font = SECTION_FONT
    ws[f"A{row+1}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row+1}:F{row+1}")

    customers_per_building = (n_buildings and (419348 / n_buildings)) or 0
    ws[f"A{row+2}"] = (
        f"Each of the {n_buildings:,} ResStock buildings represents approximately "
        f"{customers_per_building:.0f} actual residential customers in Rhode Island Energy's "
        "service territory. ResStock's sample weights ensure that the modeled building stock "
        "matches the statistical distribution of the actual housing stock."
    )
    ws[f"A{row+2}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row+2}:F{row+2}")
    ws.row_dimensions[row+2].height = 60

    row = row + 3
    ws[f"A{row}"] = "Validation and Calibration"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row}:F{row}")

    row += 1
    ws[f"A{row}"] = (
        "NREL's researchers calibrated and validated ResStock's End-Use Load Profiles against:\n\n"
        "• EIA Form 861 utility sales data\n"
        "• Utility load research data\n"
        "• Smart meter data from utilities across the country\n\n"
        "The DOE recommends using ResStock load profiles for utility load forecasting, distribution planning, "
        "and electrification planning. ISO-NE's heat pump adoption and winter peak forecasts are built on ResStock.\n\n"
        "For this study, a weighted sum of the annual kWh consumption of all Rhode Island dwellings in ResStock "
        "came within a small percentage of RIE's Test Year total residential kWh. A small additional scaling factor "
        "was applied to bring the ResStock totals into exact alignment with the Test Year residential kWh, so that "
        "the cost allocation is performed on the same total consumption that underlies RIE's revenue requirement."
    )
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row}:F{row}")
    ws.row_dimensions[row].height = 180

    # Set column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 70


def add_economic_burden_sheet(wb: Workbook) -> None:
    """Add sheet documenting economic burden calculation."""
    ws = wb.create_sheet("3. Economic Burden")

    # Title
    ws["A1"] = "Step 3: Economic Burden Calculation"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:F1")

    ws["A3"] = "Definition"
    ws["A3"].font = SECTION_FONT
    ws["A3"].fill = SECTION_FILL
    ws.merge_cells("A3:F3")

    ws["A4"] = (
        "A building's economic burden (EB) is the total cost it imposes on the system "
        "through its consumption behavior: the sum of its hourly load multiplied by the "
        "hourly marginal cost, across all 8,760 hours of the year."
    )
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A4:F4")
    ws.row_dimensions[4].height = 50

    ws["A6"] = "Formula"
    ws["A6"].font = SECTION_FONT
    ws["A6"].fill = SECTION_FILL
    ws.merge_cells("A6:F6")

    ws["A7"] = "EB_i = Σ(L_i,h × MC_h)  for h = 1 to 8,760"
    ws["A7"].font = Font(italic=True, size=11)
    ws.merge_cells("A7:F7")

    ws["A9"] = "Where:"
    ws["A9"].font = Font(bold=True)

    terms = [
        ("EB_i", "Economic burden of building i ($/year)"),
        ("L_i,h", "Load of building i in hour h (kWh)"),
        ("MC_h", "System marginal cost in hour h ($/kWh)"),
        ("h", "Hour index (1 to 8,760)"),
    ]

    row = 10
    for term, definition in terms:
        ws[f"B{row}"] = f"{term}:"
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"C{row}"] = definition
        ws.merge_cells(f"C{row}:F{row}")
        row += 1

    ws["A15"] = "Implementation in CAIRO"
    ws["A15"].font = SECTION_FONT
    ws["A15"].fill = SECTION_FILL
    ws.merge_cells("A15:F15")

    ws["A16"] = (
        "The economic burden calculation is performed by CAIRO's InternalCrossSubsidizationProcessor "
        "in the _determine_marginal_cost_allocation method. For each building:"
    )
    ws["A16"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A16:F16")
    ws.row_dimensions[16].height = 40

    steps = [
        "1. Load 8,760 hourly electricity consumption values (kWh) from ResStock",
        "2. Load 8,760 hourly marginal cost values ($/kWh) from marginal cost data files",
        "3. Multiply element-wise: consumption_h × MC_h for each hour",
        "4. Sum across all 8,760 hours to get total annual economic burden",
        "5. Store result in 'economic_burden_delivery' column",
    ]

    row = 17
    for step in steps:
        ws[f"B{row}"] = step
        ws.merge_cells(f"B{row}:F{row}")
        row += 1

    ws[f"A{row+1}"] = "Interpretation"
    ws[f"A{row+1}"].font = SECTION_FONT
    ws[f"A{row+1}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row+1}:F{row+1}")

    ws[f"A{row+2}"] = (
        "The economic burden represents the marginal system cost directly caused by a "
        "building's consumption pattern. Buildings that consume more during high-marginal-cost "
        "hours (peak periods when infrastructure is constrained) have higher economic burdens "
        "relative to their total kWh. Buildings that consume primarily during low-marginal-cost "
        "hours (off-peak periods with excess capacity) have lower economic burdens."
    )
    ws[f"A{row+2}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row+2}:F{row+2}")
    ws.row_dimensions[row+2].height = 80

    ws[f"A{row+4}"] = "High-Cost Hours"
    ws[f"A{row+4}"].font = SECTION_FONT
    ws[f"A{row+4}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row+4}:F{row+4}")

    ws[f"A{row+5}"] = (
        "The testimony reference to 'high-cost hours' refers to the hours when capacity "
        "marginal costs are non-zero (approximately 200-300 hours per year). These are the "
        "hours when the transmission and distribution grids are constrained and incremental "
        "load triggers capacity investment needs. However, the economic burden calculation "
        "includes ALL 8,760 hours — both energy costs (present in every hour) and capacity "
        "costs (concentrated in peak hours)."
    )
    ws[f"A{row+5}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row+5}:F{row+5}")
    ws.row_dimensions[row+5].height = 90

    # Set column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 60


def add_subclass_aggregation_sheet(
    wb: Workbook, bat_df: pl.DataFrame, agg: pl.DataFrame
) -> None:
    """Add sheet showing subclass-level economic burden aggregation."""
    ws = wb.create_sheet("4. Subclass Aggregation")

    # Title
    ws["A1"] = "Step 4: Subclass Economic Burden Aggregation"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:G1")

    ws["A3"] = (
        "Buildings are grouped into customer subclasses based on heating equipment. "
        "Each subclass's total economic burden is the weighted sum across all buildings "
        "in that subclass."
    )
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A3:G3")
    ws.row_dimensions[3].height = 40

    ws["A5"] = "Subclass Definitions"
    ws["A5"].font = SECTION_FONT
    ws["A5"].fill = SECTION_FILL
    ws.merge_cells("A5:G5")

    # Calculate subclass stats
    subclass_stats = (
        bat_df.group_by("postprocess_group.has_hp")
        .agg(
            [
            pl.col("bldg_id").count().alias("n_buildings"),
            pl.col("weight").sum().alias("total_weight"),
            (pl.col("economic_burden_delivery") * pl.col("weight"))
            .sum()
            .alias("total_eb"),
            pl.col("economic_burden_delivery").mean().alias("mean_eb"),
            ]
        )
        .sort("postprocess_group.has_hp", descending=True)
    )

    # Compute Sheet 6 EB column letters for cross-sheet Total EB references.
    # Column layout in Sheet 6: A=timestamp | B..load_end=subclass loads | E..G=MC | H..=EB cols
    subclasses_sh6 = [c for c in agg.columns if c not in ("timestamp", "hour_of_year")]
    n_mc_sh6 = 3  # mc_dist_sub_tx, mc_bulk_tx, mc_delivery_total
    eb_start_sh6 = 2 + len(subclasses_sh6) + n_mc_sh6  # 1-based column index
    sh6_eb_col: dict[str, str] = {
        sc: get_column_letter(eb_start_sh6 + i) for i, sc in enumerate(subclasses_sh6)
    }
    SH6 = "6. Aggregate Load Curves"
    SH5 = "5. Revenue Allocation"
    sh6_totals_row = 7  # row 5=section header, row 6=col header, row 7=totals

    # Header row
    ws["A7"] = "Subclass"
    ws["B7"] = "Definition"
    ws["C7"] = "Buildings in Sample"
    ws["D7"] = "Weighted Customers"
    ws["E7"] = "Total Economic Burden"
    ws["F7"] = "Mean EB per Building"
    ws["G7"] = "Share of Total EB"

    for cell in ["A7", "B7", "C7", "D7", "E7", "F7", "G7"]:
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
        ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows — total_row computed upfront so G-column share can reference it by address.
    data_start = 8
    total_row = data_start + subclass_stats.height  # always row 10 for HP / non-HP split
    row = data_start
    for stat_row in subclass_stats.iter_rows(named=True):
        has_hp = stat_row["postprocess_group.has_hp"]
        subclass_label = "Heat Pump" if has_hp else "Non-Heat Pump"
        definition = (
            "Homes with electric heat pumps as primary heating"
            if has_hp
            else "Homes heated with natural gas, oil, propane, or electric resistance"
        )

        ws[f"A{row}"] = subclass_label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = definition
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"C{row}"] = stat_row["n_buildings"]  # sample count — no other cell holds this
        ws[f"C{row}"].number_format = "#,##0"
        # D: weighted customers — same computation in Sheet 5 col B, same row
        ws[f"D{row}"] = f"='{SH5}'!B{row}"
        ws[f"D{row}"].number_format = "#,##0"
        # E: total EB — sum of hourly EB for this subclass's column(s) in Sheet 6 totals row
        if has_hp:
            ws[f"E{row}"] = f"='{SH6}'!{sh6_eb_col['heat_pump']}{sh6_totals_row}"
        else:
            er = sh6_eb_col.get("electrical_resistance", "")
            ff = sh6_eb_col.get("fossil_fuel", "")
            ws[f"E{row}"] = f"='{SH6}'!{er}{sh6_totals_row}+'{SH6}'!{ff}{sh6_totals_row}"
        ws[f"E{row}"].number_format = "$#,##0"
        ws[f"F{row}"] = stat_row["mean_eb"]  # unweighted mean — no other cell holds this
        ws[f"F{row}"].number_format = "$#,##0"
        ws[f"G{row}"] = f"=E{row}/$E${total_row}"  # share = this row's EB / total EB
        ws[f"G{row}"].number_format = "0.00%"

        ws.row_dimensions[row].height = 30
        row += 1

    # Total row
    ws[f"A{row}"] = "Total / Average"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"C{row}"] = f"=SUM(C{data_start}:C{row - 1})"
    ws[f"C{row}"].number_format = "#,##0"
    ws[f"D{row}"] = f"=SUM(D{data_start}:D{row - 1})"
    ws[f"D{row}"].number_format = "#,##0"
    ws[f"E{row}"] = f"=SUM(E{data_start}:E{row - 1})"
    ws[f"E{row}"].number_format = "$#,##0"
    ws[f"F{row}"] = bat_df["economic_burden_delivery"].mean()  # unweighted mean — no simple cell formula
    ws[f"F{row}"].number_format = "$#,##0"
    ws[f"G{row}"] = f"=SUM(G{data_start}:G{row - 1})"
    ws[f"G{row}"].number_format = "0.00%"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"B{row}"].fill = SECTION_FILL
    ws[f"C{row}"].fill = SECTION_FILL
    ws[f"D{row}"].fill = SECTION_FILL
    ws[f"E{row}"].fill = SECTION_FILL
    ws[f"F{row}"].fill = SECTION_FILL
    ws[f"G{row}"].fill = SECTION_FILL

    ws[f"A{row+2}"] = "Cost-Causation Share Formula"
    ws[f"A{row+2}"].font = SECTION_FONT
    ws[f"A{row+2}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row+2}:G{row+2}")

    ws[f"A{row+3}"] = "Share_i = (Σ(EB_i,b × weight_b)) / (Σ(EB_j,b × weight_b))"
    ws[f"A{row+3}"].font = Font(italic=True)
    ws.merge_cells(f"A{row+3}:G{row+3}")

    ws[f"A{row+5}"] = (
        "Where the numerator sums over all buildings b in subclass i, and the denominator "
        "sums over all buildings b in all subclasses j. The share represents each subclass's "
        "contribution to total system marginal costs."
    )
    ws[f"A{row+5}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row+5}:G{row+5}")
    ws.row_dimensions[row+5].height = 50

    # Set column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 15


def add_revenue_allocation_sheet(
    wb: Workbook, bat_df: pl.DataFrame, rev_req: dict, agg: pl.DataFrame
) -> None:
    """Add sheet showing final revenue requirement allocation."""
    ws = wb.create_sheet("5. Revenue Allocation")

    # Title
    ws["A1"] = "Step 5: Revenue Requirement Allocation (EPMC Method)"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:F1")

    ws["A3"] = (
        "The final step applies each subclass's cost-causation share to the total "
        "Residential Delivery Revenue Requirement using the Equi-Proportional Marginal "
        "Cost (EPMC) allocation method."
    )
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A3:F3")
    ws.row_dimensions[3].height = 50

    # Calculate revenue components
    total_delivery_rr = float(rev_req.get("total_delivery_revenue_requirement", 0))

    # Sheet 6 column letters for cross-sheet references to EB totals.
    # Column layout in Sheet 6: A=timestamp | load cols | MC cols (×3) | EB cols
    # Note: these formulas reference a sheet written later; they are correct in the final
    # .xlsx (all sheets present) and are re-applied post-upload for Google Sheets.
    _sh6_subcols = [c for c in agg.columns if c not in ("timestamp", "hour_of_year")]
    _sh6_eb_start = 2 + len(_sh6_subcols) + 3  # +3 for three MC columns
    _sh6_eb_cols: dict[str, str] = {
        sc: get_column_letter(_sh6_eb_start + i) for i, sc in enumerate(_sh6_subcols)
    }
    _SH6 = "6. Aggregate Load Curves"
    _sh6_tot = 7  # totals row in Sheet 6

    # Customer counts per subclass — the only column still sourced from bat_df.
    # EB (col C) comes from Sheet 6 (scaled load × MC), so EPMC residual (col D) and
    # total COS (col E) are derivable as spreadsheet formulas — no bat_df columns needed.
    subclass_agg = (
        bat_df.group_by("postprocess_group.has_hp")
        .agg(pl.col("weight").sum().alias("customers"))
        .sort("postprocess_group.has_hp", descending=True)
    )

    ws["A5"] = "Revenue Requirement Components"
    ws["A5"].font = SECTION_FONT
    ws["A5"].fill = SECTION_FILL
    ws.merge_cells("A5:F5")

    # Header
    ws["A7"] = "Subclass"
    ws["B7"] = "Customers"
    ws["C7"] = "Economic Burden"
    ws["D7"] = "EPMC Residual"
    ws["E7"] = "Total Cost-of-Service"
    ws["F7"] = "Share of Total RR"

    for cell in ["A7", "B7", "C7", "D7", "E7", "F7"]:
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
        ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows — rr_ref_row computed upfront so F-column share can reference the RR input cell.
    data_start = 8
    total_row = data_start + subclass_agg.height  # always row 10 for HP / non-HP split
    rr_ref_row = total_row + 8  # row where "Total Delivery RR (from YAML)" is written
    row = data_start
    for stat_row in subclass_agg.iter_rows(named=True):
        has_hp = stat_row["postprocess_group.has_hp"]
        subclass_label = "Heat Pump" if has_hp else "Non-Heat Pump"

        ws[f"A{row}"] = subclass_label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = stat_row["customers"]
        ws[f"B{row}"].number_format = "#,##0"
        # C: Economic Burden — cross-ref to Sheet 6 scaled-EB totals row.
        # Sheet 6 loads are scaled by resstock_kwh_scale_factor before EB is computed,
        # so this matches CAIRO's internal EB calculation.  Formula works in Excel once
        # all sheets are present; re-applied post-upload for Google Sheets.
        if has_hp:
            _eb_formula = f"='{_SH6}'!{_sh6_eb_cols.get('heat_pump', '')}{_sh6_tot}"
        else:
            _er = _sh6_eb_cols.get("electrical_resistance", "")
            _ff = _sh6_eb_cols.get("fossil_fuel", "")
            _eb_formula = f"='{_SH6}'!{_er}{_sh6_tot}+'{_SH6}'!{_ff}{_sh6_tot}"
        ws[f"C{row}"] = _eb_formula
        ws[f"C{row}"].number_format = "$#,##0"
        # E: EPMC-allocated COS = total_rr × (this subclass EB / total EB)
        # This is the EPMC formula applied directly in the spreadsheet.
        ws[f"E{row}"] = f"=$B${rr_ref_row}*C{row}/$C${total_row}"
        ws[f"E{row}"].number_format = "$#,##0"
        # D: EPMC Residual = COS − EB  (derived; no bat_df lookup needed)
        ws[f"D{row}"] = f"=E{row}-C{row}"
        ws[f"D{row}"].number_format = "$#,##0"
        ws[f"F{row}"] = f"=E{row}/$B${rr_ref_row}"  # share = COS / total delivery RR
        ws[f"F{row}"].number_format = "0.00%"
        row += 1

    # Total row
    ws[f"A{row}"] = "Total"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f"=SUM(B{data_start}:B{row - 1})"
    ws[f"B{row}"].number_format = "#,##0"
    ws[f"C{row}"] = f"=SUM(C{data_start}:C{row - 1})"
    ws[f"C{row}"].number_format = "$#,##0"
    ws[f"D{row}"] = f"=SUM(D{data_start}:D{row - 1})"
    ws[f"D{row}"].number_format = "$#,##0"
    ws[f"E{row}"] = f"=SUM(E{data_start}:E{row - 1})"  # = total_rr when EB shares sum to 1
    ws[f"E{row}"].number_format = "$#,##0"
    ws[f"F{row}"] = f"=E{row}/$B${rr_ref_row}"  # should evaluate to 1.0
    ws[f"F{row}"].number_format = "0.00%"
    for cell in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}"]:
        ws[cell].fill = SECTION_FILL

    ws[f"A{row+2}"] = "EPMC Allocation Formula"
    ws[f"A{row+2}"].font = SECTION_FONT
    ws[f"A{row+2}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row+2}:F{row+2}")

    ws[f"A{row+3}"] = "COS_i = RR × (EB_i / EB_total);  Residual_i = COS_i − EB_i"
    ws[f"A{row+3}"].font = Font(italic=True)
    ws.merge_cells(f"A{row+3}:F{row+3}")

    ws[f"A{row+5}"] = (
        "Where COS_i is the EPMC-allocated cost-of-service for subclass i, RR is the total "
        "Delivery Revenue Requirement, EB_i is subclass i's total weighted economic burden "
        "(scaled load × marginal cost, col C), and EB_total is the sum across all subclasses "
        "(col C total row).  Loads are scaled by the CAIRO resstock_kwh_scale_factor so that "
        "the aggregate kWh matches the test-year residential total."
    )
    ws[f"A{row+5}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row+5}:F{row+5}")
    ws.row_dimensions[row+5].height = 60

    ws[f"A{row+7}"] = "Comparison to Rate Case Revenue Requirement"
    ws[f"A{row+7}"].font = SECTION_FONT
    ws[f"A{row+7}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row+7}:F{row+7}")

    ws[f"A{row+8}"] = "Total Delivery RR (from YAML)"
    ws[f"A{row+8}"].font = Font(bold=True)
    ws[f"B{row+8}"] = total_delivery_rr  # input value — hardcoded intentionally
    ws[f"B{row+8}"].number_format = "$#,##0"
    ws.merge_cells(f"B{row+8}:C{row+8}")

    ws[f"A{row+9}"] = "Total Cost-of-Service (EPMC sum)"
    ws[f"A{row+9}"].font = Font(bold=True)
    ws[f"B{row+9}"] = f"=E{total_row}"  # references total row column E above
    ws[f"B{row+9}"].number_format = "$#,##0"
    ws.merge_cells(f"B{row+9}:C{row+9}")

    ws[f"A{row+10}"] = "Difference"
    ws[f"A{row+10}"].font = Font(bold=True)
    ws[f"B{row+10}"] = f"=B{row+8}-B{row+9}"  # RR minus COS
    ws[f"B{row+10}"].number_format = "$#,##0"
    ws.merge_cells(f"B{row+10}:C{row+10}")

    ws[f"D{row+8}"] = (
        "The difference reflects rounding and the specific run configuration used. "
        "In CAIRO's pre-calc mode, the BAT is calibrated to balance exactly."
    )
    ws[f"D{row+8}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"D{row+8}:F{row+11}")
    ws.row_dimensions[row+8].height = 60

    # Set column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15


def add_load_curves_sheet(
    wb: Workbook, agg: pl.DataFrame, mc: pl.DataFrame, rev_req: dict
) -> None:
    """Add sheet with 8760-row aggregate load curves, MC, and hourly EB = load × MC."""
    ws = wb.create_sheet("6. Aggregate Load Curves")

    ws["A1"] = "Aggregate Hourly Load × Marginal Cost by Subclass (8,760 hours)"
    ws["A1"].font = Font(bold=True, size=13)

    ws["A3"] = (
        "Each 'Load' column is the weighted sum of hourly electricity consumption (kWh) across "
        f"all RIE ResStock buildings in that subclass (upgrade=0), scaled by the CAIRO "
        f"resstock_kwh_scale_factor ({rev_req.get('resstock_kwh_scale_factor', 1.0):.10f}) so that "
        "the aggregate kWh matches RIE's test-year residential total — exactly as CAIRO does "
        "internally when computing economic burden and EPMC allocation. "
        f"Source: {RESSTOCK_BASE}/load_curve_hourly/state=RI/upgrade=00/ "
        f"(read locally from {LOCAL_RESSTOCK_LOADS_UPGRADE0}). "
        "Delivery MC columns sourced from "
        f"{S3_MC_DIST_SUB_TX} and {S3_MC_BULK_TX}. "
        "Each 'EB Contribution' column = Load (kWh) × MC_delivery_total ($/kWh) for that hour. "
        "Summing each EB column over all 8,760 hours yields that subclass's total Economic Burden, "
        "which determines its cost-causation share."
    )
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A3:M3")
    ws.row_dimensions[3].height = 70

    # Build combined frame with per-hour EB = load × mc_delivery_total.
    # Join on hour_of_year (1..8760) — ResStock timestamps are AMY 2018, MC timestamps are 2025.
    subclasses = [c for c in agg.columns if c not in ("timestamp", "hour_of_year")]
    mc_cols = ["mc_dist_sub_tx", "mc_bulk_tx", "mc_delivery_total"]
    combined = agg.join(mc.select(["hour_of_year"] + mc_cols), on="hour_of_year")

    eb_col_names = [f"eb_{sc}" for sc in subclasses]
    for sc, eb_col in zip(subclasses, eb_col_names):
        combined = combined.with_columns(
            (pl.col(sc) * pl.col("mc_delivery_total")).alias(eb_col)
        )

    load_col_labels = {sc: f"Load: {sc}\n(kWh, weighted)" for sc in subclasses}
    eb_col_labels = {f"eb_{sc}": f"EB: {sc}\n(load × MC, $)" for sc in subclasses}
    mc_labels = {
        "mc_dist_sub_tx": "MC: Sub-TX & Dist\n($/kWh)",
        "mc_bulk_tx": "MC: Bulk TX\n($/kWh)",
        "mc_delivery_total": "MC: Delivery Total\n($/kWh)",
    }

    # Column order: timestamp | load cols | MC cols | EB cols
    headers = ["timestamp"] + subclasses + mc_cols + eb_col_names
    all_labels = {"timestamp": "Timestamp", **load_col_labels, **mc_labels, **eb_col_labels}

    # Section header spans
    row = 5
    n_load = len(subclasses)
    n_mc = len(mc_cols)
    n_eb = len(eb_col_names)
    load_start = 2
    mc_start = load_start + n_load
    eb_start = mc_start + n_mc

    # Column letter for mc_delivery_total (last MC column, immediately before the EB section).
    mc_total_letter = get_column_letter(eb_start - 1)
    # {eb_col_name: load_col_letter} — used to generate EB formulas in data rows.
    eb_load_letters: dict[str, str] = {
        f"eb_{sc}": get_column_letter(load_start + i) for i, sc in enumerate(subclasses)
    }
    n_data_rows = combined.height  # 8760

    def _section(label: str, col_start: int, width: int) -> None:
        col_end = col_start + width - 1
        cl_start = get_column_letter(col_start)
        cl_end = get_column_letter(col_end)
        cell = ws.cell(row=row, column=col_start, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        if width > 1:
            ws.merge_cells(f"{cl_start}{row}:{cl_end}{row}")

    _section("Aggregate Subclass Load (kWh)", load_start, n_load)
    _section("Delivery Marginal Cost", mc_start, n_mc)
    _section("Hourly Economic Burden = Load × MC ($)", eb_start, n_eb)

    row = 6
    for col_idx, col in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=all_labels.get(col, col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 40

    # Totals row — SUM formulas that span the 8,760 data rows below.
    totals_row = row + 1
    ws.cell(row=totals_row, column=1, value="TOTAL (sum 8,760 hrs)").font = Font(bold=True)
    for col_idx, col in enumerate(headers[1:], start=2):
        if col in mc_cols:
            ws.cell(row=totals_row, column=col_idx, value="—")
        else:
            col_letter = get_column_letter(col_idx)
            formula = f"=SUM({col_letter}{totals_row + 1}:{col_letter}{totals_row + n_data_rows})"
            cell = ws.cell(row=totals_row, column=col_idx, value=formula)
            cell.number_format = "$#,##0.0" if col in eb_col_names else "#,##0.0"
            cell.font = Font(bold=True)
    ws.row_dimensions[totals_row].height = 18

    # Data rows — EB cells use spreadsheet formulas (load × mc_delivery_total) so the
    # calculation is visible and auditable directly in the sheet.
    row = totals_row
    for data_row in combined.iter_rows(named=True):
        row += 1
        for col_idx, col in enumerate(headers, start=1):
            if col in eb_load_letters:
                load_ltr = eb_load_letters[col]
                cell = ws.cell(
                    row=row,
                    column=col_idx,
                    value=f"={load_ltr}{row}*{mc_total_letter}{row}",
                )
                cell.number_format = "$#,##0.00"
            else:
                val = data_row[col]
                cell = ws.cell(row=row, column=col_idx, value=val)
                if col in mc_cols:
                    cell.number_format = "0.000000"
                elif col != "timestamp":
                    cell.number_format = "#,##0.0"

    ws.column_dimensions["A"].width = 20
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    ws.freeze_panes = "A8"


def add_cost_allocation_sheet(
    wb: Workbook,
    allocation: pl.DataFrame,
    agg: pl.DataFrame,
    mc: pl.DataFrame,
    rev_req: dict,
) -> None:
    """Add sheet with the full cost allocation calculation and summary."""
    ws = wb.create_sheet("7. Cost Allocation Results")

    ws["A1"] = "Cost Allocation Calculation: Subclass Delivery Revenue Requirement"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:G1")

    ws["A3"] = "Methodology"
    ws["A3"].font = SECTION_FONT
    ws["A3"].fill = SECTION_FILL
    ws.merge_cells("A3:G3")

    ws["A4"] = (
        "For each building i and hour h: economic_burden_i = Σ_h (load_i,h × mc_delivery_h). "
        "Aggregated by subclass (weighted sum). Each subclass's share of total economic burden "
        "is applied to the Total Delivery Revenue Requirement to yield the allocated cost-of-service. "
        "Formula: R_subclass = R_total × (EB_subclass / EB_total)"
    )
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A4:G4")
    ws.row_dimensions[4].height = 55

    # Scalar inputs
    total_rr = float(rev_req.get("total_delivery_revenue_requirement", 0))
    total_customers = float(rev_req.get("test_year_customer_count", 0))

    ws["A6"] = "Inputs"
    ws["A6"].font = SECTION_FONT
    ws["A6"].fill = SECTION_FILL
    ws.merge_cells("A6:G6")

    inputs = [
        ("Total Delivery Revenue Requirement", total_rr, "$#,##0", "RDP @ e9e5088: rie_rate_case_test_year.yaml"),
        ("Test Year Residential Customer Count", total_customers, "#,##0", "Same YAML"),
        ("Sub-TX & Distribution MC", "Top 100 hrs RIE load", "@", f"Source: {S3_MC_DIST_SUB_TX}"),
        ("Bulk Transmission MC", "Top 100 hrs NE system load", "@", f"Source: {S3_MC_BULK_TX}"),
        ("ResStock Loads", "8760-hr per building, upgrade=0", "@", f"Source: {RESSTOCK_BASE}/load_curve_hourly/state=RI/upgrade=00/"),
    ]
    row = 7
    for label, val, fmt, source in inputs:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=val)
        if fmt != "@":
            cell.number_format = fmt
        ws.cell(row=row, column=3, value=source).alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"C{row}:G{row}")
        ws.row_dimensions[row].height = 30
        row += 1

    # Results table
    row += 1
    ws.cell(row=row, column=1, value="Cost Allocation Results").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(f"A{row}:G{row}")
    row += 1

    headers = [
        "Subclass", "Weighted Load (kWh/yr)", "Economic Burden ($)",
        "Share of EB (%)", "Allocated Revenue Req. ($)", "Allocated Rev. Req. per Customer ($)", "Note",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Sheet 6 column mappings for cross-sheet references (Sheet 6 is written before Sheet 7).
    # Column layout in Sheet 6: A=timestamp | load cols (B..) | MC cols | EB cols
    sh6_subclasses = [c for c in agg.columns if c not in ("timestamp", "hour_of_year")]
    _load_start = 2
    _eb_start = _load_start + len(sh6_subclasses) + 3  # +3 for the three MC columns
    sh6_load_col = {sc: get_column_letter(_load_start + i) for i, sc in enumerate(sh6_subclasses)}
    sh6_eb_col = {sc: get_column_letter(_eb_start + i) for i, sc in enumerate(sh6_subclasses)}
    SH6 = "6. Aggregate Load Curves"
    sh6_tot = 7  # row 5=section header, 6=col header, 7=totals

    # Customer count split (use RECS proportions from rev_req if available, else "N/A")
    subclass_customers = rev_req.get("subclass_customers", {})

    data_start = row + 1  # first data row
    n_data_rows = allocation.height
    total_row_sh7 = data_start + n_data_rows  # total row

    row += 1
    for alloc_row in allocation.iter_rows(named=True):
        sc = alloc_row["subclass"]
        n_customers = subclass_customers.get(sc, None)

        ws.cell(row=row, column=1, value=sc)
        # B: Weighted Load — cross-ref to Sheet 6 load column totals row
        b_load = sh6_load_col.get(sc)
        if b_load:
            ws.cell(row=row, column=2, value=f"='{SH6}'!{b_load}{sh6_tot}").number_format = "#,##0"
        else:
            ws.cell(row=row, column=2, value=float(alloc_row["economic_burden"])).number_format = "#,##0"
        # C: Economic Burden — cross-ref to Sheet 6 EB column totals row
        b_eb = sh6_eb_col.get(sc)
        if b_eb:
            ws.cell(row=row, column=3, value=f"='{SH6}'!{b_eb}{sh6_tot}").number_format = "$#,##0"
        else:
            ws.cell(row=row, column=3, value=float(alloc_row["economic_burden"])).number_format = "$#,##0"
        # D: Share of EB = this row's EB / total EB (intra-sheet)
        ws.cell(row=row, column=4, value=f"=C{row}/$C${total_row_sh7}").number_format = "0.00%"
        # E: Allocated RR = share × total delivery RR input (B7)
        ws.cell(row=row, column=5, value=f"=D{row}*$B$7").number_format = "$#,##0"
        # F: Per-customer allocated RR
        if n_customers:
            ws.cell(row=row, column=6, value=f"=E{row}/{n_customers}").number_format = "$#,##0"
        else:
            ws.cell(row=row, column=6, value="N/A")
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=SUM(B{data_start}:B{row - 1})").number_format = "#,##0"
    ws.cell(row=row, column=3, value=f"=SUM(C{data_start}:C{row - 1})").number_format = "$#,##0"
    ws.cell(row=row, column=4, value=f"=SUM(D{data_start}:D{row - 1})").number_format = "0.00%"
    ws.cell(row=row, column=5, value="=$B$7").number_format = "$#,##0"  # total allocated = total RR
    for col_idx in range(1, 6):
        ws.cell(row=row, column=col_idx).fill = SECTION_FILL
        ws.cell(row=row, column=col_idx).font = Font(bold=True)

    row += 2
    ws.cell(row=row, column=1, value="Data Sources").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(f"A{row}:G{row}")
    row += 1
    sources = [
        ("ResStock Loads (local)", str(LOCAL_RESSTOCK_LOADS_UPGRADE0)),
        ("ResStock Loads (S3 canonical)", f"{RESSTOCK_BASE}/load_curve_hourly/state=RI/upgrade=00/"),
        ("Sub-TX & Distribution MC", S3_MC_DIST_SUB_TX),
        ("Bulk Transmission MC", S3_MC_BULK_TX),
        ("Revenue Requirement", "rie_hp_vs_nonhp_rate_case_test_year.yaml"),
        ("AESC 2024", "Synapse Energy Economics. Avoided Energy Supply Components of New England: 2024 Report."),
    ]
    for label, val in sources:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=val).alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"B{row}:G{row}")
        ws.row_dimensions[row].height = 30
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 60


def _load_testimony_cache() -> tuple[dict[str, Any], dict[str, Any]]:
    """Load report_variables pkl files written by the testimony analysis notebooks.

    Returns (v_vals, c_vals) as plain dicts.  Either may be empty if the cache
    file is missing or fails to load — callers should treat None values as N/A.

    Cache files are produced by:
      - notebooks/analysis.qmd            → cache/report_variables.pkl  (v)
      - notebooks/cost_of_service_by_subclass.qmd → cache/report_variables_cos_subclass.pkl  (c)
    """
    v_vals: dict[str, Any] = {}
    c_vals: dict[str, Any] = {}
    cache_dir = REPORT_DIR / "cache"
    for fname, target in (
        ("report_variables.pkl", v_vals),
        ("report_variables_cos_subclass.pkl", c_vals),
    ):
        path = cache_dir / fname
        if not path.exists():
            print(f"  [validation] Cache not found: {path}", flush=True)
            continue
        try:
            data = pickle.loads(path.read_bytes())
            if isinstance(data, dict):
                target.update(data)
            else:
                # SimpleNamespace or similar
                target.update(vars(data))
            print(f"  [validation] Loaded {fname} ({len(target)} keys)", flush=True)
        except Exception as exc:
            print(f"  [validation] Could not read {fname}: {exc}", flush=True)
    return v_vals, c_vals


def _nan(x: float | None) -> bool:
    """Return True if x is None, NaN, or inf."""
    return x is None or not math.isfinite(x)


def add_validation_sheet(
    wb: Workbook,
    bat_df: pl.DataFrame,
    allocation: pl.DataFrame,
    agg: pl.DataFrame,
    mc: pl.DataFrame,
    rev_req: dict[str, Any],
) -> None:
    """Add a validation sheet cross-referencing workbook values against expert testimony.

    Each row shows:
      - The check description and the testimony reference (section and variable name)
      - The testimony-stated value (loaded from the report cache pkl at runtime)
      - The workbook-computed value (derived from the same underlying data)
      - The absolute difference and tolerance
      - PASS / FAIL / INFO / N/A status

    PASS  = workbook matches testimony within tolerance.
    FAIL  = mismatch outside tolerance — investigate before filing.
    INFO  = hardcoded parameter; value is documented but no runtime tolerance check.
    N/A   = testimony cache not available or workbook value could not be computed.

    The testimony cache is available only after the analysis notebooks in
    notebooks/ have been rendered (``just render`` from the report directory).
    """
    # ── Styles ────────────────────────────────────────────────────────────────
    PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    INFO_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    NA_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    PASS_FONT = Font(bold=True, color="276221")
    FAIL_FONT = Font(bold=True, color="9C0006")
    INFO_FONT = Font(bold=True, color="9C6500")
    NA_FONT = Font(color="595959")

    ws = wb.create_sheet("8. Validation vs. Testimony")

    # ── Title and description ─────────────────────────────────────────────────
    ws["A1"] = "Schedule JPV-3 — Validation: Workbook Numbers vs. Expert Testimony"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")

    ws["A3"] = (
        "This sheet cross-references key numbers in this workbook against the "
        "corresponding values cited in the Direct Testimony of Juan-Pablo Velez "
        "(RIPUC Docket No. 25-45GE). Testimony values are loaded at runtime from "
        "the report cache (produced by notebooks/cost_of_service_by_subclass.qmd "
        "and notebooks/analysis.qmd). If the cache is absent (notebooks not yet "
        "rendered), testimony values appear as N/A and the workbook values are "
        "shown for reference. EPMC cross-subsidy = BAT_epmc_delivery per building "
        "(= annual_bill_delivery − economic_burden_delivery − residual_share_epmc_delivery)."
    )
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A3:H3")
    ws.row_dimensions[3].height = 60

    # ── Load testimony cache ───────────────────────────────────────────────────
    v_vals, c_vals = _load_testimony_cache()
    cache_available = bool(v_vals or c_vals)

    # Diagnose bat_df column availability so workbook-value N/As are easy to trace.
    _diag_missing = [
        col for col in (
            "annual_bill_delivery",
            "economic_burden_delivery",
            "residual_share_epmc_delivery",
            "BAT_epmc_delivery",
            "postprocess_group.has_hp",
            "postprocess_group.heating_type_v2",
            "heats_with_natgas",
            "heats_with_oil",
            "heats_with_propane",
            "weight",
        )
        if col not in bat_df.columns
    ]
    if _diag_missing:
        print(
            f"  [validation] WARNING: bat_df missing columns → workbook values will be N/A: "
            f"{_diag_missing}",
            flush=True,
        )
    else:
        print(
            f"  [validation] bat_df has all required columns "
            f"({bat_df.height:,} buildings); workbook values will be computed.",
            flush=True,
        )

    ws["A5"] = "Testimony cache:"
    ws["A5"].font = Font(bold=True)
    if cache_available:
        cache_msg = (
            f"Loaded — report_variables.pkl ({len(v_vals)} keys), "
            f"report_variables_cos_subclass.pkl ({len(c_vals)} keys)"
        )
        ws["B5"].fill = PASS_FILL
    else:
        cache_msg = (
            "NOT AVAILABLE — run 'just render' from reports/ri_hp_rates/ to generate "
            "cache files. Testimony values shown as N/A until then."
        )
        ws["B5"].fill = INFO_FILL
    ws["B5"] = cache_msg
    ws.merge_cells("B5:H5")

    # ── Column headers ─────────────────────────────────────────────────────────
    col_headers = [
        "Category",
        "Check",
        "Testimony Reference\n(Section / variable name)",
        "Testimony Value",
        "Workbook Value",
        "Difference",
        "Tolerance",
        "Status",
    ]
    hdr_row = 7
    for col_idx, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=hdr_row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    ws.row_dimensions[hdr_row].height = 32

    # ── Pre-compute workbook values ────────────────────────────────────────────
    total_rr = float(rev_req.get("total_delivery_revenue_requirement", 0))
    total_customers = float(rev_req.get("test_year_customer_count", 0))


    # MC peak-hour counts: nonzero hours in each MC column
    wb_bulk_tx_peak_hrs: float | None = (
        float(int((mc["mc_bulk_tx"] > 0).sum())) if "mc_bulk_tx" in mc.columns else None
    )
    wb_dist_sub_peak_hrs: float | None = (
        float(int((mc["mc_dist_sub_tx"] > 0).sum())) if "mc_dist_sub_tx" in mc.columns else None
    )

    # ResStock building count (rows in BAT data after filtering to RIE)
    wb_n_bldgs = float(bat_df.height)

    # Column-presence flags
    has_hp_flag = "postprocess_group.has_hp" in bat_df.columns
    has_htv2 = "postprocess_group.heating_type_v2" in bat_df.columns
    has_heating_type = "postprocess_group.heating_type" in bat_df.columns
    has_eb = "economic_burden_delivery" in bat_df.columns
    has_residual = "residual_share_epmc_delivery" in bat_df.columns
    has_bat_epmc = "BAT_epmc_delivery" in bat_df.columns
    has_annual_bill = "annual_bill_delivery" in bat_df.columns
    has_natgas = "heats_with_natgas" in bat_df.columns
    has_oil = "heats_with_oil" in bat_df.columns
    has_propane = "heats_with_propane" in bat_df.columns

    nan = float("nan")

    def _weighted_cos(df: pl.DataFrame) -> float:
        """Σ (economic_burden + residual) × weight — total EPMC COS for a subgroup."""
        if not (has_eb and has_residual) or df.is_empty():
            return nan
        return float(
            ((df["economic_burden_delivery"] + df["residual_share_epmc_delivery"]) * df["weight"]).sum()
        )

    def _epmc_xs_total(df: pl.DataFrame) -> float:
        """Σ BAT_epmc_delivery × weight — total EPMC cross-subsidy.

        Uses BAT_epmc_delivery directly (= annual_bill_delivery − EB − EPMC residual),
        matching cost_of_service_by_subclass.qmd.  BAT_vol_delivery is NOT used because
        it may be absent from the master BAT for certain CAIRO run configurations.
        """
        if not has_bat_epmc or df.is_empty():
            return nan
        return float((df["BAT_epmc_delivery"] * df["weight"]).sum())

    def _epmc_xs_mean(df: pl.DataFrame) -> float:
        """Weighted mean EPMC cross-subsidy per building = xs_total / total_weight."""
        if not has_bat_epmc or df.is_empty():
            return nan
        return float((df["BAT_epmc_delivery"] * df["weight"]).sum()) / float(
            df["weight"].sum()
        )

    def _group_revenue(df: pl.DataFrame) -> float:
        """Σ annual_bill_delivery × weight — total delivery revenue from subgroup.

        Uses annual_bill_delivery directly, matching cost_of_service_by_subclass.qmd.
        Equivalent to COS + XS but does not cascade to nan when only one is available.
        """
        if not has_annual_bill or df.is_empty():
            return nan
        return float((df["annual_bill_delivery"] * df["weight"]).sum())

    # Subgroup DataFrames.
    # Use heating_type_v2 when available — this is what cost_of_service_by_subclass.qmd
    # uses for the c.* group totals.  heating_type_v2 assigns HP priority: any building
    # with has_hp=True gets "heat_pump" regardless of fuel, so HP+natgas dual-fuel
    # buildings are in the HP bucket, not the natgas bucket.
    # analysis.qmd (v.* values) uses heats_with_natgas / heats_with_oil for the
    # per-customer means; in the RI dataset there are no dual-fuel HP buildings, so
    # the populations are identical.  We therefore use heating_type_v2 for all groups.
    if has_htv2:
        hp_df = bat_df.filter(pl.col("postprocess_group.heating_type_v2") == "heat_pump")
        ng_df = bat_df.filter(pl.col("postprocess_group.heating_type_v2") == "natgas")
        df_df = bat_df.filter(pl.col("postprocess_group.heating_type_v2") == "delivered_fuels")
        er_df = bat_df.filter(pl.col("postprocess_group.heating_type_v2") == "electrical_resistance")
    else:
        # Fallback for older master BAT builds that lack heating_type_v2
        hp_df = bat_df.filter(pl.col("postprocess_group.has_hp") == True) if has_hp_flag else bat_df.head(0)  # noqa: E712
        ng_df = bat_df.filter(pl.col("heats_with_natgas") == True) if has_natgas else bat_df.head(0)  # noqa: E712
        df_df = (
            bat_df.filter(
                (pl.col("heats_with_oil") == True) | (pl.col("heats_with_propane") == True)  # noqa: E712
            )
            if (has_oil and has_propane)
            else bat_df.head(0)
        )
        er_df = (
            bat_df.filter(pl.col("postprocess_group.heating_type") == "electrical_resistance")
            if has_heating_type
            else bat_df.head(0)
        )

    # HP stats
    wb_hp_cos = _weighted_cos(hp_df)
    wb_hp_xs = _epmc_xs_total(hp_df)          # positive = HP overpays
    wb_hp_rev = _group_revenue(hp_df)          # = annual_bill_delivery × weight
    wb_hp_pct_over = (
        wb_hp_xs / wb_hp_cos * 100
        if not (_nan(wb_hp_xs) or _nan(wb_hp_cos) or wb_hp_cos == 0)
        else nan
    )
    wb_hp_mean_bat = _epmc_xs_mean(hp_df)

    # NG stats
    wb_ng_cos = _weighted_cos(ng_df)
    _wb_ng_xs = _epmc_xs_total(ng_df)          # negative = NG underpays (currently unused in rows)
    wb_ng_rev = _group_revenue(ng_df)          # = annual_bill_delivery × weight
    wb_ng_rev_pct_of_cos = (
        wb_ng_rev / wb_ng_cos * 100
        if not (_nan(wb_ng_rev) or _nan(wb_ng_cos) or wb_ng_cos == 0)
        else nan
    )
    wb_ng_mean_bat = _epmc_xs_mean(ng_df)     # negative = NG underpays

    # DF (delivered fuels: oil + propane) stats
    wb_df_cos = _weighted_cos(df_df)
    _wb_df_xs = _epmc_xs_total(df_df)  # unused in rows
    wb_df_rev = _group_revenue(df_df)          # = annual_bill_delivery × weight
    wb_df_rev_pct_of_cos = (
        wb_df_rev / wb_df_cos * 100
        if not (_nan(wb_df_rev) or _nan(wb_df_cos) or wb_df_cos == 0)
        else nan
    )
    wb_df_mean_bat = _epmc_xs_mean(df_df)     # negative = DF underpays

    # ER (electric resistance) stats — overpayment only; no separate COS check needed
    wb_er_xs = _epmc_xs_total(er_df)          # positive = ER overpays

    # All-residential totals (for HP % of total COS / revenue)
    wb_total_cos = _weighted_cos(bat_df)
    wb_total_rev_computed = _group_revenue(bat_df)
    wb_hp_pct_of_total_cos = (
        wb_hp_cos / wb_total_cos * 100
        if not (_nan(wb_hp_cos) or _nan(wb_total_cos) or wb_total_cos == 0)
        else nan
    )
    wb_hp_pct_of_total_rev = (
        wb_hp_rev / wb_total_rev_computed * 100
        if not (_nan(wb_hp_rev) or _nan(wb_total_rev_computed) or wb_total_rev_computed == 0)
        else nan
    )

    # EPMC allocation recomputed from load curves (internal cross-check)
    hp_alloc_rr: float | None = None
    for alloc_row in allocation.iter_rows(named=True):
        if alloc_row["subclass"] == "heat_pump":
            hp_alloc_rr = float(alloc_row["allocated_rr"])
            break

    def _tv(d: dict[str, Any], key: str) -> float | None:
        """Fetch a testimony value by key, returning None if absent."""
        val = d.get(key)
        return float(val) if val is not None else None

    def _tv_derived(numerator: float | None, denominator: float | None, scale: float = 100.0) -> float | None:
        """Derive a ratio from two testimony cache values (e.g. pct of COS)."""
        if numerator is None or denominator is None or denominator == 0:
            return None
        return float(numerator) / float(denominator) * scale

    # ── Testimony-derived values for checks that use inline functions ─────────
    # These mirror the helper functions used in expert_testimony.qmd.
    t_hp_pct_of_total_cos = _tv_derived(
        _tv(c_vals, "cos_default_hp_group_cos"),
        _tv(c_vals, "cos_subclass_total_cos"),
    )
    t_hp_pct_of_total_rev = _tv_derived(
        _tv(c_vals, "cos_default_hp_group_rev"),
        _tv(c_vals, "cos_subclass_total_delivery_rev"),
    )
    t_ng_rev_pct_of_cos = _tv_derived(
        _tv(c_vals, "cos_default_ng_group_rev"),
        _tv(c_vals, "cos_default_ng_group_cos"),
    )
    t_df_rev_pct_of_cos = _tv_derived(
        _tv(c_vals, "cos_default_df_group_rev"),
        _tv(c_vals, "cos_default_df_group_cos"),
    )

    # ── Build the checks list ──────────────────────────────────────────────────
    # Each tuple: (category, description, testimony_ref, t_val, wb_val, fmt, tol_pct, is_info)
    # is_info=True → INFO cell, no tolerance math (hardcoded parameter confirmed in testimony)
    checks: list[tuple[str, str, str, float | None, float, str, float, bool]] = []

    # ── Revenue Requirement ───────────────────────────────────────────────────
    checks += [
        (
            "Revenue Requirement",
            "Total Residential Delivery Revenue Requirement",
            "Section II (~line 223)\nc.rie_rev_req_total_delivery_rr",
            _tv(c_vals, "rie_rev_req_total_delivery_rr"),
            total_rr,
            "$#,##0",
            0.0,
            False,
        ),
        (
            "Revenue Requirement",
            "Test Year Residential Customer Count",
            "Section IX (~line 1101)\nc.rie_rev_req_test_year_customer_count",
            _tv(c_vals, "rie_rev_req_test_year_customer_count"),
            total_customers,
            "#,##0",
            0.0,
            False,
        ),
    ]

    # ── Marginal Cost Parameters ───────────────────────────────────────────────
    checks += [
        (
            "MC Parameters",
            "Bulk Transmission avoided capacity cost ($/kW-year)\n[hardcoded — stated in testimony]",
            "Section IX (~line 1153)\n$69/kW-year (AESC 2024 Pool TX Facility)",
            69.0,
            69.0,
            "#,##0.00",
            0.0,
            True,
        ),
        (
            "MC Parameters",
            "Sub-TX & Distribution avoided capacity cost ($/kW-year)\n[hardcoded — stated in testimony]",
            "Section IX (~line 1154)\n$80.24/kW-year (AESC 2024 distribution capacity)",
            80.24,
            80.24,
            "#,##0.00",
            0.0,
            True,
        ),
        (
            "MC Parameters",
            "Bulk TX: Hours with non-zero MC in the 8,760-hr series",
            "Section IX (~line 1153)\nTop 100 hours of aggregate NE system load",
            100.0,
            wb_bulk_tx_peak_hrs if wb_bulk_tx_peak_hrs is not None else nan,
            "#,##0",
            0.0,
            False,
        ),
        (
            "MC Parameters",
            "Sub-TX & Dist: Hours with non-zero MC in the 8,760-hr series",
            "Section IX (~line 1154)\nTop 100 hours of RIE system load (summer only)",
            100.0,
            wb_dist_sub_peak_hrs if wb_dist_sub_peak_hrs is not None else nan,
            "#,##0",
            0.0,
            False,
        ),
    ]

    # ── ResStock Data ──────────────────────────────────────────────────────────
    checks += [
        (
            "ResStock Data",
            "ResStock buildings in RIE service territory\n(rows in master BAT after filtering to utility=rie)",
            "Section IX (~line 1097)\nv.n_resstock_bldgs",
            _tv(v_vals, "n_resstock_bldgs"),
            wb_n_bldgs,
            "#,##0",
            0.0,
            False,
        ),
        (
            "ResStock Data",
            "Hours per building [hardcoded — stated in testimony]",
            "Section IX (~line 1134)\n8,760 hours stated in testimony",
            8760.0,
            8760.0,
            "#,##0",
            0.0,
            True,
        ),
    ]

    # ── Final Results: Heat Pump Subclass ─────────────────────────────────────
    # These are the numbers that appear verbatim in the testimony narrative.
    # Testimony source: Section II (~lines 270, 294, 336) and Section I (~line 184).
    checks += [
        (
            "Final Results — Heat Pump",
            "HP mean delivery overpayment per customer ($/yr)\n"
            "= weighted mean of BAT_epmc_delivery across HP buildings\n"
            "Testimony: 'heat pump customers overpay... by [X] per year, on average'",
            "Section I (~line 184)\nv.hp_mean_delivery_bat",
            _tv(v_vals, "hp_mean_delivery_bat"),
            wb_hp_mean_bat,
            "$#,##0",
            2.0,
            False,
        ),
        (
            "Final Results — Heat Pump",
            "HP group: Total EPMC cost-of-service — delivery\n"
            "= Σ_HP [(economic_burden + residual) × weight]\n"
            "Testimony: 'heat pump customers' cost-of-service was [X]'",
            "Section II (~line 294)\nc.cos_default_hp_group_cos",
            _tv(c_vals, "cos_default_hp_group_cos"),
            wb_hp_cos,
            "$#,##0",
            1.0,
            False,
        ),
        (
            "Final Results — Heat Pump",
            "HP group: Total delivery revenue collected\n"
            "= Σ_HP [annual_bill_delivery × weight]\n"
            "Testimony: '[HP customers] paid [X] through their bills'",
            "Section II (~line 294)\nc.cos_default_hp_group_rev",
            _tv(c_vals, "cos_default_hp_group_rev"),
            wb_hp_rev,
            "$#,##0",
            1.0,
            False,
        ),
        (
            "Final Results — Heat Pump",
            "HP group: Total cross-subsidy — overpayment vs EPMC COS\n"
            "= Σ_HP [BAT_epmc_delivery × weight]\n"
            "Testimony: 'Heat pump customers... paid [X] more than their cost of service'",
            "Section II (~line 294)\nc.cos_default_hp_group_xs",
            _tv(c_vals, "cos_default_hp_group_xs"),
            wb_hp_xs,
            "$#,##0",
            1.0,
            False,
        ),
        (
            "Final Results — Heat Pump",
            "HP group: % overpayment relative to EPMC COS (ppt)\n"
            "= (HP revenue − HP COS) / HP COS × 100\n"
            "Testimony: '[HP customers] paid [X]% more than their delivery cost-of-service'",
            "Section II (~line 336)\nc.cos_default_hp_pct_over_cos\n"
            "(stored as ppt, e.g. 28 means 28%)",
            _tv(c_vals, "cos_default_hp_pct_over_cos"),
            wb_hp_pct_over,
            "0.0",
            1.0,
            False,
        ),
        (
            "Final Results — Heat Pump",
            "HP group: Share of total residential delivery COS (%)\n"
            "= HP COS / all-residential COS × 100\n"
            "Testimony: 'while they only made up [X]% of total costs'",
            "Section II (~line 294)\nhp_pct_of_total_delivery_cos(c)\n"
            "= c.cos_default_hp_group_cos / c.cos_subclass_total_cos × 100",
            t_hp_pct_of_total_cos,
            wb_hp_pct_of_total_cos,
            "0.0",
            1.0,
            False,
        ),
        (
            "Final Results — Heat Pump",
            "HP group: Share of total residential delivery revenue (%)\n"
            "= HP revenue / all-residential revenue × 100\n"
            "Testimony: 'they accounted for [X]% of total revenue'",
            "Section II (~line 294)\nhp_pct_of_total_delivery_revenue(c)\n"
            "= c.cos_default_hp_group_rev / c.cos_subclass_total_delivery_rev × 100",
            t_hp_pct_of_total_rev,
            wb_hp_pct_of_total_rev,
            "0.0",
            1.0,
            False,
        ),
    ]

    # ── Final Results: Natural Gas Subclass ───────────────────────────────────
    # Testimony source: Section II (~line 302) and Section I (~line 184).
    checks += [
        (
            "Final Results — Natural Gas",
            "NG mean delivery underpayment per customer ($/yr)\n"
            "= weighted mean of BAT_epmc_delivery across NG buildings\n"
            "Testimony: 'customers that heat with natural gas... underpay... by [X]'\n"
            "(negative value; testimony renders as abs())",
            "Section I (~line 184)\nv.natgas_mean_delivery_bat",
            _tv(v_vals, "natgas_mean_delivery_bat"),
            wb_ng_mean_bat,
            "$#,##0",
            2.0,
            False,
        ),
        (
            "Final Results — Natural Gas",
            "NG group: Total EPMC cost-of-service — delivery\n"
            "Testimony: 'customers who heat with natural gas... cost the system [X]'",
            "Section II (~line 302)\nc.cos_default_ng_group_cos",
            _tv(c_vals, "cos_default_ng_group_cos"),
            wb_ng_cos,
            "$#,##0",
            1.0,
            False,
        ),
        (
            "Final Results — Natural Gas",
            "NG group: Total delivery revenue collected\n"
            "= Σ_NG [annual_bill_delivery × weight]\n"
            "Testimony: '[NG customers] paid only [X]'",
            "Section II (~line 302)\nc.cos_default_ng_group_rev",
            _tv(c_vals, "cos_default_ng_group_rev"),
            wb_ng_rev,
            "$#,##0",
            1.0,
            False,
        ),
        (
            "Final Results — Natural Gas",
            "NG group: Revenue as % of COS (ppt)\n"
            "= NG revenue / NG COS × 100\n"
            "Testimony: 'paid only [X]% of their cost-of-service'",
            "Section II (~line 302)\nng_delivery_rev_as_pct_of_subclass_cos(c)\n"
            "= c.cos_default_ng_group_rev / c.cos_default_ng_group_cos × 100",
            t_ng_rev_pct_of_cos,
            wb_ng_rev_pct_of_cos,
            "0.0",
            1.0,
            False,
        ),
    ]

    # ── Final Results: Delivered Fuels Subclass ───────────────────────────────
    # Testimony source: Section II (~line 302) and Section I (~line 184).
    checks += [
        (
            "Final Results — Delivered Fuels",
            "DF mean delivery underpayment per customer ($/yr)\n"
            "= weighted mean of BAT_epmc_delivery across DF buildings\n"
            "Testimony: 'customers that heat with delivered fuels... underpay... by [X]'\n"
            "(negative value; testimony renders as abs())",
            "Section I (~line 184)\nv.oil_propane_mean_delivery_bat",
            _tv(v_vals, "oil_propane_mean_delivery_bat"),
            wb_df_mean_bat,
            "$#,##0",
            2.0,
            False,
        ),
        (
            "Final Results — Delivered Fuels",
            "DF group: Total EPMC cost-of-service — delivery\n"
            "Testimony: 'Those that heat with delivered fuels... cost the system [X]'",
            "Section II (~line 302)\nc.cos_default_df_group_cos",
            _tv(c_vals, "cos_default_df_group_cos"),
            wb_df_cos,
            "$#,##0",
            1.0,
            False,
        ),
        (
            "Final Results — Delivered Fuels",
            "DF group: Total delivery revenue collected\n"
            "= Σ_DF [annual_bill_delivery × weight]\n"
            "Testimony: '[DF customers] paid only [X]'",
            "Section II (~line 302)\nc.cos_default_df_group_rev",
            _tv(c_vals, "cos_default_df_group_rev"),
            wb_df_rev,
            "$#,##0",
            1.0,
            False,
        ),
        (
            "Final Results — Delivered Fuels",
            "DF group: Revenue as % of COS (ppt)\n"
            "= DF revenue / DF COS × 100\n"
            "Testimony: 'paid only [X]% of their cost-of-service'",
            "Section II (~line 302)\ndf_delivery_rev_as_pct_of_subclass_cos(c)\n"
            "= c.cos_default_df_group_rev / c.cos_default_df_group_cos × 100",
            t_df_rev_pct_of_cos,
            wb_df_rev_pct_of_cos,
            "0.0",
            1.0,
            False,
        ),
    ]

    # ── Final Results: Electric Resistance Subclass ───────────────────────────
    checks += [
        (
            "Final Results — Elec. Resistance",
            "ER group: Total cross-subsidy — overpayment vs EPMC COS\n"
            "= Σ_ER [BAT_epmc_delivery × weight]\n"
            "Testimony: 'Electric resistance customers also paid [X] more than their cost-of-service'",
            "Section II (~line 294)\nc.cos_default_er_overpay_vs_cos",
            _tv(c_vals, "cos_default_er_overpay_vs_cos"),
            wb_er_xs,
            "$#,##0",
            1.0,
            False,
        ),
    ]

    # ── Internal EPMC Cross-Check ──────────────────────────────────────────────
    # The HP allocated RR recomputed from raw load curves × MC (Sheet 7) should
    # agree with the HP EPMC COS derived from the CAIRO master BAT (Sheet 4).
    # Any gap reflects rounding or hour-alignment differences vs. CAIRO's run.
    if hp_alloc_rr is not None and not _nan(wb_hp_cos):
        checks.append((
            "Internal Cross-Check",
            "HP allocated RR (Sheet 7, recomputed from load curves × MC)\n"
            "vs. HP EPMC COS from CAIRO master BAT (Sheet 4)\n"
            "A large gap here means the two computations diverge — investigate.",
            "Internal consistency check\n(Sheet 7 vs. Sheet 4 of this workbook;\nno testimony line)",
            wb_hp_cos,
            hp_alloc_rr,
            "$#,##0",
            2.0,
            False,
        ))

    # ── Write rows ─────────────────────────────────────────────────────────────
    row = hdr_row + 1
    prev_cat: str | None = None

    for cat, desc, ref, t_val, wb_val, fmt, tol_pct, is_info in checks:
        # Category separator row
        if cat != prev_cat:
            sep_cell = ws.cell(row=row, column=1, value=cat)
            sep_cell.font = SECTION_FONT
            sep_cell.fill = SECTION_FILL
            ws.merge_cells(f"A{row}:H{row}")
            ws.row_dimensions[row].height = 18
            row += 1
            prev_cat = cat

        # Determine status
        wb_nan = _nan(wb_val)
        if is_info:
            status = "INFO"
            diff_val: float | None = None
            st_fill = INFO_FILL
            st_font = INFO_FONT
        elif t_val is None or wb_nan:
            status = "N/A"
            diff_val = None
            st_fill = NA_FILL
            st_font = NA_FONT
        else:
            diff_val = wb_val - t_val
            if tol_pct == 0.0:
                passed = abs(diff_val) < 1.0  # within $1 / 1 customer / 1 ppt
            else:
                denom = max(abs(t_val), 1e-9)
                passed = abs(diff_val) / denom * 100 <= tol_pct
            status = "PASS" if passed else "FAIL"
            st_fill = PASS_FILL if passed else FAIL_FILL
            st_font = PASS_FONT if passed else FAIL_FONT

        # Column A: category (repeated for readability when rows are sorted/filtered)
        ws.cell(row=row, column=1, value=cat).alignment = Alignment(vertical="top", wrap_text=True)

        # Column B: check description
        ws.cell(row=row, column=2, value=desc).alignment = Alignment(wrap_text=True, vertical="top")

        # Column C: testimony reference
        ws.cell(row=row, column=3, value=ref).alignment = Alignment(wrap_text=True, vertical="top")

        # Column D: testimony value
        if t_val is not None:
            td_cell = ws.cell(row=row, column=4, value=t_val)
            if fmt != "0.0":
                td_cell.number_format = fmt
        else:
            ws.cell(row=row, column=4, value="N/A")

        # Column E: workbook value
        if not wb_nan:
            wb_cell = ws.cell(row=row, column=5, value=wb_val)
            if fmt != "0.0":
                wb_cell.number_format = fmt
        else:
            ws.cell(row=row, column=5, value="N/A")

        # Column F: difference
        if diff_val is not None:
            diff_cell = ws.cell(row=row, column=6, value=diff_val)
            if fmt != "0.0":
                diff_cell.number_format = fmt
        else:
            ws.cell(row=row, column=6, value="—")

        # Column G: tolerance
        tol_str = (
            "hardcoded parameter"
            if is_info
            else ("exact (< $1 / 1 unit)" if tol_pct == 0.0 else f"±{tol_pct}%")
        )
        ws.cell(row=row, column=7, value=tol_str)

        # Column H: status
        st_cell = ws.cell(row=row, column=8, value=status)
        st_cell.font = st_font
        st_cell.fill = st_fill
        st_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 45
        row += 1

    # ── Legend ─────────────────────────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="Legend").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(f"A{row}:H{row}")
    row += 1
    legend = [
        ("PASS", PASS_FILL, PASS_FONT, "Workbook value matches testimony within tolerance."),
        ("FAIL", FAIL_FILL, FAIL_FONT, "Mismatch outside tolerance — investigate before filing."),
        ("INFO", INFO_FILL, INFO_FONT, "Hardcoded parameter. Both sides use the same number; shown for transparency."),
        ("N/A", NA_FILL, NA_FONT, "Testimony cache not available, or workbook value could not be computed."),
    ]
    for lbl, fill, font, meaning in legend:
        lbl_cell = ws.cell(row=row, column=1, value=lbl)
        lbl_cell.fill = fill
        lbl_cell.font = font
        lbl_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=meaning).alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"B{row}:H{row}")
        ws.row_dimensions[row].height = 20
        row += 1

    # ── Column widths and freeze ───────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["H"].width = 10
    ws.freeze_panes = f"A{hdr_row + 1}"


# ── Orchestration ─────────────────────────────────────────────────────────────


def build_workbook(
    output_path: Path,
) -> tuple[Path, dict[str, dict[str, str]]]:
    """Load all inputs, build every sheet, write the .xlsx.

    Returns ``(xlsx_path, gsheet_formula_patches)`` where ``gsheet_formula_patches``
    is a ``{sheet_name: {cell_addr: formula_string}}`` dict of cross-sheet formulas
    that must be applied *after* the full Google Sheet upload (so that all referenced
    tabs exist before the formulas are evaluated).
    """
    rev_req = load_revenue_requirement_yaml()
    bat_df = load_master_bat_data(BATCH, RUN_DELIVERY, RUN_SUPPLY)
    scale_factor = float(rev_req.get("resstock_kwh_scale_factor", 1.0))
    agg_loads, mc_delivery = load_aggregate_load_curves(scale_factor)
    allocation = compute_cost_allocation(agg_loads, mc_delivery, rev_req)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = create_workbook()
    add_overview_sheet(wb, rev_req)
    add_marginal_cost_sheet(wb)
    add_resstock_loads_sheet(wb, rev_req)
    add_economic_burden_sheet(wb)
    add_subclass_aggregation_sheet(wb, bat_df, agg_loads)
    add_revenue_allocation_sheet(wb, bat_df, rev_req, agg_loads)
    add_load_curves_sheet(wb, agg_loads, mc_delivery, rev_req)
    add_cost_allocation_sheet(wb, allocation, agg_loads, mc_delivery, rev_req)
    add_validation_sheet(wb, bat_df, allocation, agg_loads, mc_delivery, rev_req)
    wb.save(str(output_path))
    print(
        f"Wrote {output_path} ({output_path.stat().st_size / 1024:.1f} KB)",
        flush=True,
    )

    # Build the cross-sheet formula patch map.  xlsx_to_gsheet writes sheets one at
    # a time, so when Sheet 4 is written, Sheets 5 and 6 don't exist yet — Google
    # Sheets stores the reference as broken.  We re-write these cells after the full
    # upload so every tab is present when the formulas are evaluated.
    subclasses_sh6 = [c for c in agg_loads.columns if c not in ("timestamp", "hour_of_year")]
    n_mc_sh6 = 3  # mc_dist_sub_tx, mc_bulk_tx, mc_delivery_total
    eb_start_sh6 = 2 + len(subclasses_sh6) + n_mc_sh6
    sh6_eb: dict[str, str] = {
        sc: get_column_letter(eb_start_sh6 + i) for i, sc in enumerate(subclasses_sh6)
    }
    sh6 = "6. Aggregate Load Curves"
    sh5 = "5. Revenue Allocation"
    sh6_totals = 7  # section header row 5, col header row 6, totals row 7
    er_col = sh6_eb.get("electrical_resistance", "")
    ff_col = sh6_eb.get("fossil_fuel", "")
    hp_col = sh6_eb.get("heat_pump", "")
    gsheet_patches: dict[str, dict[str, str]] = {
        "4. Subclass Aggregation": {
            # HP row (row 8): customers from Sheet 5, EB from Sheet 6 heat_pump column
            "D8": f"='{sh5}'!B8",
            "E8": f"='{sh6}'!{hp_col}{sh6_totals}",
            # Non-HP row (row 9): customers from Sheet 5, EB = er + ff from Sheet 6
            "D9": f"='{sh5}'!B9",
            "E9": f"='{sh6}'!{er_col}{sh6_totals}+'{sh6}'!{ff_col}{sh6_totals}",
        },
        "5. Revenue Allocation": {
            # HP EB (row 8): same source as Sheet 4 E8 — load × MC from Sheet 6
            "C8": f"='{sh6}'!{hp_col}{sh6_totals}",
            # Non-HP EB (row 9): electrical resistance + fossil fuel from Sheet 6
            "C9": f"='{sh6}'!{er_col}{sh6_totals}+'{sh6}'!{ff_col}{sh6_totals}",
        },
    }
    return output_path, gsheet_patches


# ── Google Sheets upload ───────────────────────────────────────────────────────

_TAB_FORMATTING: dict[str, dict] = {
    "1. Overview": {
        "wrap_columns": ["B:B"],
        "column_widths_px": {"A": 240, "B": 560},
        "freeze_rows": 0,
        "bold_header": False,
    },
    "2. Marginal Costs": {
        "freeze_rows": 1,
        "bold_header": True,
        "auto_resize_columns": ["A:D"],
    },
    "3. ResStock Loads": {
        "freeze_rows": 1,
        "bold_header": True,
        "auto_resize_columns": ["A:D"],
    },
    "4. Economic Burden": {
        "freeze_rows": 1,
        "bold_header": True,
        "auto_resize_columns": ["A:E"],
    },
    "5. Subclass Aggregation": {
        "freeze_rows": 1,
        "bold_header": True,
        "auto_resize_columns": ["A:F"],
    },
    "6. Revenue Allocation": {
        "freeze_rows": 1,
        "bold_header": True,
        "auto_resize_columns": ["A:F"],
    },
    "7. Load Curves": {
        "freeze_rows": 1,
        "bold_header": True,
        "auto_resize_columns": ["A:H"],
    },
    "8. Validation vs. Testimony": {
        "wrap_columns": ["A:C"],
        "column_widths_px": {"A": 200},
        "auto_resize_columns": ["B:H"],
        "freeze_rows": 1,
        "bold_header": True,
    },
}


def upload_to_folder(
    xlsx_path: Path,
    folder_id: str,
    title: str,
    formula_patches: dict[str, dict[str, str]] | None = None,
) -> None:
    """Create (or replace) a Google Sheet in the given Drive folder.

    Searches the folder for any existing non-trashed file with the same name and
    trashes it before creating a fresh spreadsheet. Mirrors the workbook contents
    with live formulas via ``xlsx_to_gsheet``, then applies tab formatting.

    ``formula_patches`` is an optional ``{sheet_name: {cell_addr: formula}}`` map of
    cross-sheet formulas to re-write *after* the full upload.  Because
    ``xlsx_to_gsheet`` writes sheets one at a time, cross-sheet references written
    early are flagged as broken until all tabs exist.  Writing them again here, after
    every tab is present, ensures they evaluate correctly.
    """
    from lib.data.gsheets import (
        apply_sheet_formatting,
        create_sheet_in_folder,
        write_values_with_formulas,
        xlsx_to_gsheet,
    )

    print(f"Uploading '{title}' to Drive folder {folder_id} ...", flush=True)
    spreadsheet = create_sheet_in_folder(title, folder_id)
    xlsx_to_gsheet(xlsx_path, spreadsheet.id, delete_other_tabs=True)

    # Re-write any cross-sheet formulas now that all tabs exist.
    if formula_patches:
        print("Patching cross-sheet formulas ...", flush=True)
        for sheet_name, patches in formula_patches.items():
            ws = spreadsheet.worksheet(sheet_name)
            for cell_addr, formula in patches.items():
                write_values_with_formulas(ws, [[formula]], start=cell_addr)

    print("Applying formatting ...", flush=True)
    for ws in spreadsheet.worksheets():
        spec = _TAB_FORMATTING.get(ws.title)
        if spec:
            apply_sheet_formatting(ws, **spec)
    print(
        f"Done. View at https://docs.google.com/spreadsheets/d/{spreadsheet.id}/edit",
        flush=True,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0] if __doc__ else "")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("cache/schedule_jpv3_cost_allocation.xlsx"),
        help="Output .xlsx path. Default: cache/schedule_jpv3_cost_allocation.xlsx",
    )
    parser.add_argument(
        "--folder-id",
        default=DEFAULT_FOLDER_ID,
        help=f"Google Drive folder ID to upload into. Default: {DEFAULT_FOLDER_ID}",
    )
    parser.add_argument(
        "--title",
        default=DEFAULT_TITLE,
        help=f"Name for the Google Sheet. Default: '{DEFAULT_TITLE}'",
    )
    parser.add_argument(
        "--no-upload",
        action="store_true",
        help="Build the .xlsx locally without uploading to Google Drive.",
    )
    args = parser.parse_args(argv)

    out, patches = build_workbook(args.output)
    if not args.no_upload:
        upload_to_folder(out, args.folder_id, args.title, formula_patches=patches)
    return 0


if __name__ == "__main__":
    sys.exit(main())
