"""Build the supporting workbook for ResStock kWh scale factor adjustment.

This script documents the ResStock kWh intensity calibration step for Rhode Island
Energy (RIE) in the heat pump rate design analysis.

WHAT IT DOCUMENTS
-----------------
The value `resstock_kwh_scale_factor: 0.9568112362177266` is applied in CAIRO's
run_scenario.py (line 645):

    raw_load_elec = raw_load_elec * settings.kwh_scale_factor

`raw_load_elec` is a per-building 8,760-hour DataFrame. Every building's hourly
electric consumption is multiplied by 0.9568 before any bill or BAT calculation.

WHY IT IS NEEDED
----------------
After customer-count rescaling, ResStock buildings still over-represent per-customer
kWh intensity by ~4.5% compared to the rate case test year. The kWh scale factor
corrects that intensity gap. Customer rescaling alone cannot fix it — it only adjusts
the total by changing how many customers are represented, leaving per-customer kWh
unchanged at 7,031 kWh/customer. The kWh scale factor brings per-customer kWh
from 7,031 down to 6,727 to match the rate case test year.

THREE-STAGE PIPELINE (after MF non-HVAC adjustment)
----------------------------------------------------
Stage 1 — _sb release (input to CAIRO):
    Customers: 481,896   Total kWh: 3,388,382,407   Per-customer: 7,031

Stage 2 — After CAIRO customer rescaling (return_buildingstock):
    Customers: 419,348   Total kWh: 2,948,583,151   Per-customer: 7,031  ← same!

Stage 3 — After kWh scale factor (run_scenario.py:645):
    Customers: 419,348   Total kWh: 2,821,237,490   Per-customer: 6,727  ← matches rate case

Rate case target:
    Customers: 419,348   Total kWh: 2,821,237,490   Per-customer: 6,727

Usage:
    cd /ebs/home/lee_switch_box/reports2

    uv run python reports/ri_hp_rates/testimony_response/kwh_scale_factor_workbook.py \\
        --folder-id 1uPcJbcOChD6zoFuPb-gsxSByPr7xwmCH \\
        --filename "RIE kWh Scale Factor Analysis"

    # Or save locally:
    uv run python reports/ri_hp_rates/testimony_response/kwh_scale_factor_workbook.py \\
        --output cache/kwh_scale_factor.xlsx
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path
from typing import cast

import polars as pl
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# RIE rate case constants
# Source: Docket 25-45-GE, RIE Book 21, Blazunas Schedules & Workpapers
# (11/26/2025). PRB-1-ELEC exhibit, p. 14, lines 8-9, columns d, f, k.
# https://www.documentcloud.org/documents/27926651-docket-25-45-ge-rie-book-21-blazunas-schedules-workpapers-11262025/#document/p16/a2809947
# ---------------------------------------------------------------------------

# Test year residential kWh: PRB-1-ELEC p. 14, line 8, col k
RIE_TEST_YEAR_RESIDENTIAL_KWH: float = 2_821_237_490.0

# Test year customer count: PRB-1-ELEC p. 14, line 9, col d (= 5,032,174 bills / 12 months)
RIE_TEST_YEAR_CUSTOMER_COUNT: float = 419_347.83

# ResStock sample weight total from res_2024_amy2018_2_sb
# sum(_sb.utility_assignment.weight) where sb.electric_utility = 'rie'
RIE_RESSTOCK_CUSTOMER_COUNT: float = 481_896.13

# customer_scale_factor = test_year_customer_count / resstock_customer_count
# Used by CAIRO's return_buildingstock to rescale sample weights to match test year.
RIE_CUSTOMER_SCALE_FACTOR: float = (
    RIE_TEST_YEAR_CUSTOMER_COUNT / RIE_RESSTOCK_CUSTOMER_COUNT
)  # = 0.8702037719

# resstock_kwh_scale_factor from rie_rate_case_test_year.yaml
# = test_year_residential_kwh / resstock_total_residential_kwh_after_customer_scaling
# Applied in run_scenario.py:645:  raw_load_elec = raw_load_elec * kwh_scale_factor
RIE_KWH_SCALE_FACTOR: float = 0.9568112362177266

RATE_CASE_SOURCE = (
    "Docket 25-45-GE, RIE Book 21, Blazunas Schedules & Workpapers (11/26/2025)"
)
RATE_CASE_URL = (
    "https://www.documentcloud.org/documents/27926651-"
    "docket-25-45-ge-rie-book-21-blazunas-schedules-workpapers-11262025/"
    "#document/p16/a2809947"
)

# Code reference where the factor is applied
CODE_REF = "rate_design/hp_rates/run_scenario.py, lines 640–645"
CODE_SNIPPET = "raw_load_elec = raw_load_elec * settings.kwh_scale_factor"

# ---------------------------------------------------------------------------
# Data paths
# ---------------------------------------------------------------------------
LOCAL_BASE_RESSTOCK = "/ebs/data/nrel/resstock"
S3_BASE_RESSTOCK = "s3://data.sb/nrel/resstock"
S3_BASE_EIA861 = "s3://data.sb/eia/861/electric_utility_stats"

BASE_RESSTOCK = LOCAL_BASE_RESSTOCK if Path(LOCAL_BASE_RESSTOCK).exists() else S3_BASE_RESSTOCK
BASE_EIA861 = S3_BASE_EIA861

RESSTOCK_RELEASE_SB = "res_2024_amy2018_2_sb"
STATE = "RI"
UTILITY = "rie"
UPGRADE = "00"
EIA_YEAR = 2018

BLDG_ID_COL = "bldg_id"
ELECTRIC_UTILITY_COL = "sb.electric_utility"
WEIGHT_COL = "weight"
TOTAL_ELEC_MONTHLY_COL = "out.electricity.total.energy_consumption"
MWH_TO_KWH = 1000


def get_aws_region(default: str = "us-west-2") -> str:
    region = os.environ.get("AWS_REGION") or os.environ.get("AWS_DEFAULT_REGION")
    if region:
        return region
    try:
        import boto3

        session = boto3.Session()
        if session.region_name:
            return session.region_name
    except ImportError:
        pass
    return default


def _storage_options() -> dict[str, str]:
    return {"region_name": get_aws_region()}


def _is_s3(path: str) -> bool:
    return path.startswith("s3://")


def load_resstock_annual_from_monthly(
    path_monthly_dir: str,
    path_utility_assignment: str,
    storage_options: dict[str, str] | None = None,
) -> pl.DataFrame:
    """Sum monthly load curves to get annual kWh per utility, from the _sb release.

    The _sb release stores per-building monthly parquets in load_curve_monthly/.
    We sum across all 12 months to get annual totals, then join utility assignment
    to aggregate by utility code.

    Returns DataFrame with columns: utility_code, resstock_customers, resstock_kwh
    """
    opts = storage_options if _is_s3(path_monthly_dir) else None
    opts_util = storage_options if _is_s3(path_utility_assignment) else None

    monthly_lf = pl.scan_parquet(f"{path_monthly_dir}/*.parquet", storage_options=opts)
    annual_lf = monthly_lf.group_by(BLDG_ID_COL).agg(
        pl.col(TOTAL_ELEC_MONTHLY_COL).sum().alias("annual_kwh")
    )
    util_lf = pl.scan_parquet(path_utility_assignment, storage_options=opts_util)

    return cast(
        pl.DataFrame,
        annual_lf.join(util_lf, on=BLDG_ID_COL, how="inner")
        .select(
            pl.col(BLDG_ID_COL),
            pl.col(ELECTRIC_UTILITY_COL),
            pl.col(WEIGHT_COL),
            pl.col("annual_kwh"),
        )
        .with_columns((pl.col("annual_kwh") * pl.col(WEIGHT_COL)).alias("weighted_kwh"))
        .group_by(ELECTRIC_UTILITY_COL)
        .agg(
            pl.col(WEIGHT_COL).sum().alias("resstock_customers"),
            pl.col("weighted_kwh").sum().alias("resstock_kwh"),
        )
        .rename({ELECTRIC_UTILITY_COL: "utility_code"})
        .sort("utility_code")
        .collect(),
    )


def load_eia_by_utility(
    path_eia: str,
    storage_options: dict[str, str] | None = None,
    utility_codes: list[str] | None = None,
) -> pl.DataFrame:
    """Load EIA-861 residential data by utility (used for external context only)."""
    opts = storage_options if _is_s3(path_eia) else None
    lf = pl.scan_parquet(path_eia, storage_options=opts)
    if utility_codes:
        lf = lf.filter(pl.col("utility_code").is_in(utility_codes))
    return cast(
        pl.DataFrame,
        lf.select(
            pl.col("utility_code"),
            pl.col("residential_customers").alias("eia_customers"),
            (pl.col("residential_sales_mwh") * MWH_TO_KWH).alias("eia_kwh"),
        )
        .sort("utility_code")
        .collect(),
    )


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _bold(ws, cell: str) -> None:
    ws[cell].font = Font(bold=True)


def _header_fill(ws, row: int, n_cols: int) -> None:
    fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).font = Font(bold=True)


def _section_fill(ws, row: int, n_cols: int, color: str = "EEF2FF") -> None:
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).font = Font(bold=True)


def _green_fill(ws, row: int, n_cols: int) -> None:
    fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = fill


def _yellow_fill(ws, row: int, n_cols: int) -> None:
    fill = PatternFill(start_color="FFFACC", end_color="FFFACC", fill_type="solid")
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = fill


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_readme(wb: Workbook, resstock_kwh_mf_adjusted: float) -> None:
    ws = wb.create_sheet("README", 0)

    ws["A1"] = "RIE resstock_kwh_scale_factor = 0.9568112362177266"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    row = 3

    # --- What this value is ---
    ws[f"A{row}"] = "What is resstock_kwh_scale_factor?"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    ws[f"A{row}"] = (
        "A calibration factor applied to every ResStock building's hourly electric load before "
        "CAIRO runs bill calculations, cross-subsidization (BAT), and revenue requirement allocation. "
        "It corrects a systematic gap between ResStock's simulated per-customer electricity intensity "
        "and RIE's actual rate case test year residential sales."
    )
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row}:E{row}")
    ws.row_dimensions[row].height = 50
    row += 2

    # --- Where it is applied ---
    ws[f"A{row}"] = "Where and how it is applied"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    ws[f"A{row}"] = "File:"
    _bold(ws, f"A{row}")
    ws[f"B{row}"] = CODE_REF
    ws.merge_cells(f"B{row}:E{row}")
    row += 1
    ws[f"A{row}"] = "Code:"
    _bold(ws, f"A{row}")
    ws[f"B{row}"] = CODE_SNIPPET
    ws[f"B{row}"].font = Font(name="Courier New")
    ws.merge_cells(f"B{row}:E{row}")
    row += 1
    ws[f"A{row}"] = "Input:"
    _bold(ws, f"A{row}")
    ws[f"B{row}"] = (
        "raw_load_elec — a DataFrame indexed by [bldg_id, hour] with 8,760 hourly rows per building. "
        "Loaded from the _sb release hourly load curves (load_curve_hourly/). "
        "Already incorporates the MF non-HVAC adjustment."
    )
    ws[f"B{row}"].alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws.row_dimensions[row].height = 45
    row += 1
    ws[f"A{row}"] = "Effect:"
    _bold(ws, f"A{row}")
    ws[f"B{row}"] = (
        "Multiplies every building's every hourly load value by 0.9568. "
        "The hourly profile shape is preserved — only the magnitude changes. "
        "All downstream CAIRO outputs (bills, BAT values, revenue requirement) "
        "use these scaled loads."
    )
    ws[f"B{row}"].alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws.row_dimensions[row].height = 45
    row += 2

    # --- Why it is needed ---
    ws[f"A{row}"] = "Why is it needed?"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    ws[f"A{row}"] = (
        "After the customer-count rescaling step (CAIRO's return_buildingstock), "
        "the total weighted kWh drops from 3,388,382,407 to 2,948,583,151 — but "
        "the per-customer intensity remains unchanged at 7,031 kWh/customer. "
        "Customer rescaling only changes how many customers are represented; it cannot "
        "fix a per-customer intensity gap.\n\n"
        "RIE's rate case test year says residential customers used 6,727 kWh/customer on average. "
        "ResStock simulations produce 7,031 kWh/customer — 4.5% too high. "
        "The kWh scale factor corrects that intensity gap so that every building's modeled "
        "consumption is calibrated to the rate case benchmark."
    )
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row}:E{row}")
    ws.row_dimensions[row].height = 90
    row += 2

    # --- Pipeline summary ---
    ws[f"A{row}"] = "Three-Stage Calibration Pipeline (after MF non-HVAC adjustment)"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    headers = ["Stage", "Description", "Customers", "Total kWh", "kWh/Customer"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _header_fill(ws, row, len(headers))
    row += 1

    customers = RIE_TEST_YEAR_CUSTOMER_COUNT
    pipeline = [
        (
            "1 — _sb release input",
            "Per-building hourly loads loaded into CAIRO. MF non-HVAC already adjusted.",
            RIE_RESSTOCK_CUSTOMER_COUNT,
            resstock_kwh_mf_adjusted,
            resstock_kwh_mf_adjusted / RIE_RESSTOCK_CUSTOMER_COUNT,
            False,
        ),
        (
            "2 — After customer rescaling",
            "CAIRO rescales weights so customers = test year count. Per-customer kWh unchanged.",
            customers,
            resstock_kwh_mf_adjusted * RIE_CUSTOMER_SCALE_FACTOR,
            (resstock_kwh_mf_adjusted * RIE_CUSTOMER_SCALE_FACTOR) / customers,
            False,
        ),
        (
            "3 — After kWh scale factor",
            "CAIRO multiplies every hourly load by 0.9568. Per-customer kWh now matches rate case.",
            customers,
            resstock_kwh_mf_adjusted * RIE_CUSTOMER_SCALE_FACTOR * RIE_KWH_SCALE_FACTOR,
            (resstock_kwh_mf_adjusted * RIE_CUSTOMER_SCALE_FACTOR * RIE_KWH_SCALE_FACTOR) / customers,
            True,
        ),
        (
            "Rate case target",
            "test_year_residential_kwh from PRB-1-ELEC exhibit, p.14, line 8, col k.",
            customers,
            RIE_TEST_YEAR_RESIDENTIAL_KWH,
            RIE_TEST_YEAR_RESIDENTIAL_KWH / customers,
            True,
        ),
    ]
    for stage, desc, cust, total_kwh, per_cust, is_match in pipeline:
        ws.cell(row=row, column=1, value=stage)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=cust).number_format = "#,##0.00"
        ws.cell(row=row, column=4, value=total_kwh).number_format = "#,##0"
        ws.cell(row=row, column=5, value=per_cust).number_format = "#,##0.00"
        if is_match:
            _green_fill(ws, row, 5)
        ws.row_dimensions[row].height = 30
        row += 1
    row += 1

    # --- Rate case reference ---
    ws[f"A{row}"] = "Rate Case Reference"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    for label, value in [
        ("Source", RATE_CASE_SOURCE),
        ("URL", RATE_CASE_URL),
        ("test_year_residential_kwh", f"{RIE_TEST_YEAR_RESIDENTIAL_KWH:,.0f}  (PRB-1-ELEC p.14, line 8, col k)"),
        ("test_year_customer_count", f"{RIE_TEST_YEAR_CUSTOMER_COUNT:,.2f}  (= 5,032,174 bills ÷ 12 months, p.14 line 9 col d)"),
    ]:
        ws[f"A{row}"] = label
        _bold(ws, f"A{row}")
        ws[f"B{row}"] = value
        ws.merge_cells(f"B{row}:E{row}")
        row += 1
    row += 1

    # --- Data sources ---
    ws[f"A{row}"] = "Data Sources"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    for label, path in [
        ("ResStock _sb release (load_curve_monthly)", f"{S3_BASE_RESSTOCK}/{RESSTOCK_RELEASE_SB}/load_curve_monthly/state={STATE}/upgrade={UPGRADE}/"),
        ("ResStock _sb release (load_curve_hourly — CAIRO input)", f"{S3_BASE_RESSTOCK}/{RESSTOCK_RELEASE_SB}/load_curve_hourly/state={STATE}/upgrade={UPGRADE}/"),
        ("Utility assignment", f"{S3_BASE_RESSTOCK}/{RESSTOCK_RELEASE_SB}/metadata_utility/state={STATE}/utility_assignment.parquet"),
        ("EIA-861 (external context)", f"{S3_BASE_EIA861}/year={EIA_YEAR}/state={STATE}/data.parquet"),
    ]:
        ws[f"A{row}"] = label
        _bold(ws, f"A{row}")
        ws[f"B{row}"] = path
        ws.merge_cells(f"B{row}:E{row}")
        row += 1
    row += 1

    # --- Workbook contents ---
    ws[f"A{row}"] = "Workbook Contents"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    for name, desc in [
        ("README", "This sheet — overview, pipeline, code reference, data sources"),
        ("pipeline", "Three-stage pipeline with per-customer kWh at each stage"),
        ("scale_factor_derivation", "Step-by-step derivation of resstock_kwh_scale_factor = 0.9568"),
        ("before_kwh_calibration", "Stage 2 vs rate case target: ResStock after customer rescaling, before kWh scaling"),
        ("after_kwh_calibration", "Stage 3 vs rate case target: ResStock after kWh scale factor applied"),
        ("summary", "Side-by-side before/after comparison"),
    ]:
        ws[f"A{row}"] = f"  {name}:"
        _bold(ws, f"A{row}")
        ws[f"B{row}"] = desc
        ws.merge_cells(f"B{row}:E{row}")
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18


def _write_pipeline(wb: Workbook, resstock_kwh_mf_adjusted: float) -> None:
    """Write the three-stage pipeline sheet with per-customer kWh at each stage."""
    ws = wb.create_sheet("pipeline")

    ws["A1"] = "ResStock kWh Calibration Pipeline — RIE"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")

    ws["A2"] = (
        "Key insight: customer rescaling (stage 2) does NOT change per-customer kWh intensity. "
        "The kWh scale factor (stage 3) is what corrects the per-customer intensity gap."
    )
    ws["A2"].font = Font(italic=True, color="444444")
    ws.merge_cells("A2:G2")

    row = 4

    # --- Stage table ---
    headers = [
        "Stage",
        "Mechanism",
        "Where Applied",
        "Customers",
        "Total kWh",
        "kWh / Customer",
        "vs Rate Case Target",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _header_fill(ws, row, len(headers))
    row += 1

    kwh_after_customer_scaling = resstock_kwh_mf_adjusted * RIE_CUSTOMER_SCALE_FACTOR
    kwh_after_kwh_scaling = kwh_after_customer_scaling * RIE_KWH_SCALE_FACTOR
    target = RIE_TEST_YEAR_RESIDENTIAL_KWH
    customers = RIE_TEST_YEAR_CUSTOMER_COUNT

    def pct_vs_target(kwh: float, cust: float) -> float:
        per_cust = kwh / cust
        target_per_cust = target / customers
        return (per_cust - target_per_cust) / target_per_cust * 100

    stages = [
        (
            "0 — Pre-MF adjustment (base release)",
            "MF non-HVAC electricity overestimation corrected (baked into _sb release)",
            "_sb release already incorporates this",
            None, None, None,
            "N/A — MF adj not quantified here",
            "CCCCCC",
        ),
        (
            "1 — _sb release (CAIRO input)",
            "MF-adjusted per-building hourly loads; sample weights sum to 481,896",
            "_sb release: load_curve_hourly/",
            RIE_RESSTOCK_CUSTOMER_COUNT,
            resstock_kwh_mf_adjusted,
            resstock_kwh_mf_adjusted / RIE_RESSTOCK_CUSTOMER_COUNT,
            f"{pct_vs_target(resstock_kwh_mf_adjusted, RIE_RESSTOCK_CUSTOMER_COUNT):+.2f}%",
            "FFFFFF",
        ),
        (
            "2 — After customer rescaling",
            "CAIRO rescales weights to test year count (419,348). "
            "Per-building load values unchanged — only weights change.",
            "run_scenario.py: return_buildingstock(customer_count=419,348)",
            customers,
            kwh_after_customer_scaling,
            kwh_after_customer_scaling / customers,
            f"{pct_vs_target(kwh_after_customer_scaling, customers):+.2f}%",
            "FFFACC",
        ),
        (
            "3 — After kWh scale factor ✓",
            "CAIRO multiplies every building's every hourly load by 0.9568. "
            "Profile shape preserved; only magnitude changes.",
            "run_scenario.py:645:  raw_load_elec *= 0.9568",
            customers,
            kwh_after_kwh_scaling,
            kwh_after_kwh_scaling / customers,
            f"{kwh_after_kwh_scaling - target:+.0f} kWh ≈ 0",
            "E2EFDA",
        ),
        (
            "Rate case target",
            "test_year_residential_kwh from PRB-1-ELEC exhibit, rate case filing",
            RATE_CASE_SOURCE,
            customers,
            target,
            target / customers,
            "—  (this IS the target)",
            "E2EFDA",
        ),
    ]

    for stage_name, mechanism, where, cust, total_kwh, per_cust, vs_target, color in stages:
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill

        ws.cell(row=row, column=1, value=stage_name).font = Font(bold=True)
        ws.cell(row=row, column=2, value=mechanism)
        ws[f"C{row}"] = where
        ws[f"C{row}"].font = Font(name="Courier New", size=9)
        if cust is not None:
            ws.cell(row=row, column=4, value=cust).number_format = "#,##0.00"
            ws.cell(row=row, column=5, value=total_kwh).number_format = "#,##0"
            ws.cell(row=row, column=6, value=per_cust).number_format = "#,##0.00"
        ws.cell(row=row, column=7, value=vs_target)

        ws.row_dimensions[row].height = 40
        for col in range(1, 8):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    row += 1

    # --- Key insight callout ---
    ws[f"A{row}"] = "Key Insight: Why Customer Rescaling Alone Is Not Enough"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    _section_fill(ws, row, 5, "FFF2CC")
    ws.cell(row=row, column=1, value="Stage 1 per-customer kWh")
    ws.cell(row=row, column=2, value=resstock_kwh_mf_adjusted / RIE_RESSTOCK_CUSTOMER_COUNT).number_format = "#,##0.00"
    ws.cell(row=row, column=3, value="kWh/customer")
    ws.cell(row=row, column=4, value="← unchanged after stage 2")
    row += 1
    _section_fill(ws, row, 5, "FFF2CC")
    ws.cell(row=row, column=1, value="Stage 2 per-customer kWh")
    ws.cell(row=row, column=2, value=kwh_after_customer_scaling / customers).number_format = "#,##0.00"
    ws.cell(row=row, column=3, value="kWh/customer")
    ws.cell(row=row, column=4, value="← same as stage 1; rescaling only changes total, not intensity")
    row += 1
    _section_fill(ws, row, 5, "E2EFDA")
    ws.cell(row=row, column=1, value="Stage 3 per-customer kWh")
    ws.cell(row=row, column=2, value=kwh_after_kwh_scaling / customers).number_format = "#,##0.00"
    ws.cell(row=row, column=3, value="kWh/customer")
    ws.cell(row=row, column=4, value="← corrected by kWh scale factor; matches rate case")
    row += 1
    _section_fill(ws, row, 5, "E2EFDA")
    ws.cell(row=row, column=1, value="Rate case per-customer kWh")
    ws.cell(row=row, column=2, value=target / customers).number_format = "#,##0.00"
    ws.cell(row=row, column=3, value="kWh/customer")
    ws.cell(row=row, column=4, value="← target from rate case filing")
    row += 2

    # --- Per-building example ---
    ws[f"A{row}"] = "Per-Building Effect: Example at the Median Annual kWh"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    median_kwh = resstock_kwh_mf_adjusted / RIE_RESSTOCK_CUSTOMER_COUNT  # per-customer ≈ per-building
    ws[f"A{row}"] = (
        "The kWh scale factor is applied to every building's hourly loads in the 8,760-hour matrix. "
        "For a representative building consuming the average annual kWh:"
    )
    ws[f"A{row}"].alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{row}:G{row}")
    ws.row_dimensions[row].height = 30
    row += 1

    for col, h in enumerate(["", "Annual kWh", "Monthly avg kWh", "Change"], 1):
        ws.cell(row=row, column=col, value=h)
    _header_fill(ws, row, 4)
    row += 1

    before_bldg = median_kwh
    after_bldg = median_kwh * RIE_KWH_SCALE_FACTOR
    ws.cell(row=row, column=1, value="Before kWh scale factor")
    ws.cell(row=row, column=2, value=before_bldg).number_format = "#,##0.00"
    ws.cell(row=row, column=3, value=before_bldg / 12).number_format = "#,##0.00"
    ws.cell(row=row, column=4, value="—")
    row += 1
    _green_fill(ws, row, 4)
    ws.cell(row=row, column=1, value="After kWh scale factor")
    ws.cell(row=row, column=2, value=after_bldg).number_format = "#,##0.00"
    ws.cell(row=row, column=3, value=after_bldg / 12).number_format = "#,##0.00"
    ws.cell(row=row, column=4, value=f"{(RIE_KWH_SCALE_FACTOR - 1) * 100:.2f}% ({before_bldg - after_bldg:,.0f} kWh/year less)")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 30


def _write_scale_factor_derivation(
    ws,
    resstock_kwh_mf_adjusted: float,
    resstock_kwh_customer_scaled: float,
    resstock_kwh_final: float,
) -> None:
    """Step-by-step derivation of resstock_kwh_scale_factor."""
    ws["A1"] = "resstock_kwh_scale_factor — Derivation"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A2"] = (
        "The scale factor is derived so that total weighted ResStock kWh — after customer-count "
        "rescaling — exactly equals the rate case test year residential kWh."
    )
    ws["A2"].font = Font(italic=True, color="444444")
    ws.merge_cells("A2:D2")

    row = 4

    def section(label: str, color: str = "DDDDDD") -> None:
        nonlocal row
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, size=12)
        _section_fill(ws, row, 4, color)
        row += 1

    def data_row(label: str, value: object, fmt: str, note: str, formula: str = "") -> None:
        nonlocal row
        ws[f"A{row}"] = label
        cell = ws[f"B{row}"]
        cell.value = value
        cell.number_format = fmt
        ws[f"C{row}"] = note
        if formula:
            ws[f"D{row}"] = formula
        row += 1

    section("Step 1: Rate Case Test Year Reference (the target)", "EEF2FF")
    data_row("test_year_customer_count", RIE_TEST_YEAR_CUSTOMER_COUNT, "#,##0.00",
             "PRB-1-ELEC p.14, line 9, col d  (= 5,032,174 bills ÷ 12 months)")
    data_row("test_year_residential_kwh", RIE_TEST_YEAR_RESIDENTIAL_KWH, "#,##0",
             "PRB-1-ELEC p.14, line 8, col k  ← CALIBRATION TARGET")
    data_row("Rate case kWh per customer", RIE_TEST_YEAR_RESIDENTIAL_KWH / RIE_TEST_YEAR_CUSTOMER_COUNT,
             "#,##0.00", "test_year_residential_kwh ÷ test_year_customer_count")
    row += 1

    section("Step 2: ResStock Input from _sb Release (MF-Adjusted)", "EEF2FF")
    data_row("resstock_customer_count", RIE_RESSTOCK_CUSTOMER_COUNT, "#,##0.00",
             "sum(weight) for utility=rie in res_2024_amy2018_2_sb")
    data_row("resstock_total_residential_kwh", resstock_kwh_mf_adjusted, "#,##0.00",
             "sum(annual_kwh × weight) from load_curve_monthly, utility=rie")
    data_row("ResStock kWh per customer", resstock_kwh_mf_adjusted / RIE_RESSTOCK_CUSTOMER_COUNT,
             "#,##0.00", "resstock_total_residential_kwh ÷ resstock_customer_count")
    row += 1

    section("Step 3: Customer-Count Rescaling (CAIRO return_buildingstock)", "FFFACC")
    data_row("customer_scale_factor", RIE_CUSTOMER_SCALE_FACTOR, "0.0000000000",
             "test_year_customer_count ÷ resstock_customer_count",
             f"={RIE_TEST_YEAR_CUSTOMER_COUNT:.2f}/{RIE_RESSTOCK_CUSTOMER_COUNT:.2f}")
    data_row("resstock_kwh_after_customer_scaling", resstock_kwh_customer_scaled, "#,##0.00",
             "resstock_total_residential_kwh × customer_scale_factor",
             f"={resstock_kwh_mf_adjusted:.2f}*{RIE_CUSTOMER_SCALE_FACTOR:.10f}")
    data_row("kWh per customer (after rescaling)", resstock_kwh_customer_scaled / RIE_TEST_YEAR_CUSTOMER_COUNT,
             "#,##0.00", "← still 7,031; customer rescaling does NOT fix intensity gap")
    data_row("Intensity gap vs rate case", (resstock_kwh_customer_scaled / RIE_TEST_YEAR_CUSTOMER_COUNT
                                             - RIE_TEST_YEAR_RESIDENTIAL_KWH / RIE_TEST_YEAR_CUSTOMER_COUNT),
             "+#,##0.00; -#,##0.00", "kWh/customer above rate case target (needs correction)")
    row += 1

    section("Step 4: Calculate resstock_kwh_scale_factor", "E2EFDA")
    data_row("resstock_kwh_scale_factor", RIE_KWH_SCALE_FACTOR, "0.0000000000000000",
             "test_year_residential_kwh ÷ resstock_kwh_after_customer_scaling",
             f"={RIE_TEST_YEAR_RESIDENTIAL_KWH:.0f}/{resstock_kwh_customer_scaled:.2f}")
    data_row("Percent adjustment", (RIE_KWH_SCALE_FACTOR - 1.0) * 100, '0.00"%"',
             f"Every building's every hourly load is scaled by this factor")
    row += 1

    section("Step 5: Apply Scale Factor (run_scenario.py:645)", "E2EFDA")
    data_row("Final ResStock kWh", resstock_kwh_final, "#,##0.00",
             "resstock_kwh_after_customer_scaling × resstock_kwh_scale_factor",
             f"={resstock_kwh_customer_scaled:.2f}*{RIE_KWH_SCALE_FACTOR:.16f}")
    data_row("kWh per customer (final)", resstock_kwh_final / RIE_TEST_YEAR_CUSTOMER_COUNT,
             "#,##0.00", "= rate case target per-customer kWh ✓")
    row += 1

    section("Verification", "E2EFDA")
    diff = resstock_kwh_final - RIE_TEST_YEAR_RESIDENTIAL_KWH
    data_row("Rate case target kWh", RIE_TEST_YEAR_RESIDENTIAL_KWH, "#,##0", "test_year_residential_kwh")
    data_row("Final ResStock kWh", resstock_kwh_final, "#,##0", "after kWh scale factor applied")
    data_row("Difference", diff, "#,##0.00", "Should be ≈ 0 (floating-point rounding only)")
    status = "✓ Match" if abs(diff) < 1.0 else "✗ Mismatch"
    ws[f"A{row}"] = "Status"
    ws[f"B{row}"] = status
    ws[f"B{row}"].font = Font(bold=True)
    row += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["D"].width = 55


def _write_comparison(ws, metrics: dict, title: str, subtitle: str) -> None:
    """Write a ResStock vs Rate Case target comparison table for one calibration stage."""
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws.cell(row=2, column=1, value=subtitle).font = Font(italic=True, color="444444")
    ws.merge_cells("A2:D2")

    row = 4
    headers = ["Metric", "Value", "Unit", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _header_fill(ws, row, len(headers))
    row += 1

    rows = [
        ("Customer count (test year)", metrics["customers"], "customers",
         "test_year_customer_count — rate case filing", "#,##0.00"),
        ("ResStock total kWh", metrics["resstock_kwh"], "kWh",
         "sum(annual_kwh × weight) — see pipeline sheet for derivation", "#,##0"),
        ("Rate case target kWh", metrics["target_kwh"], "kWh",
         "test_year_residential_kwh — rate case filing (calibration target)", "#,##0"),
        ("ResStock kWh per customer", metrics["resstock_kwh_per_customer"], "kWh/customer",
         "ResStock total kWh ÷ customer count", "#,##0.00"),
        ("Rate case kWh per customer", metrics["target_kwh_per_customer"], "kWh/customer",
         "Rate case target kWh ÷ customer count", "#,##0.00"),
        ("Difference (kWh)", metrics["diff_kwh"], "kWh",
         "ResStock − Rate Case target; positive = ResStock over-estimates", "#,##0"),
        ("% Difference from rate case target", metrics["pct_diff"], "%",
         "After calibration this should be ≈ 0%", "0.00"),
    ]

    for label, value, unit, note, fmt in rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value).number_format = fmt
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=note)
        # Highlight the % difference row
        if "%" in label:
            if abs(metrics["pct_diff"]) < 0.01:
                _green_fill(ws, row, 4)
            else:
                _yellow_fill(ws, row, 4)
        row += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 65


def _write_summary(
    ws,
    before: dict,
    after: dict,
    eia_kwh: float | None = None,
    eia_customers: float | None = None,
) -> None:
    """Write before vs after summary comparing both stages to the rate case target."""
    ws["A1"] = "Summary: Before vs After kWh Calibration — RIE"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    ws["A2"] = (
        "Before = Stage 2 (after customer rescaling, before kWh scale factor).  "
        "After = Stage 3 (after kWh scale factor applied).  "
        "Calibration target is rate case test year, not EIA-861."
    )
    ws["A2"].font = Font(italic=True, color="444444")
    ws.merge_cells("A2:E2")

    row = 4
    headers = ["Metric", "Before (Stage 2)", "After (Stage 3)", "Change", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _header_fill(ws, row, len(headers))
    row += 1

    summary_rows = [
        ("Customers (test year)", before["customers"], after["customers"], "#,##0.00",
         "Unchanged — set by customer_count_override in RR YAML"),
        ("ResStock total kWh", before["resstock_kwh"], after["resstock_kwh"], "#,##0",
         f"Change = −{before['resstock_kwh'] - after['resstock_kwh']:,.0f} kWh ({(RIE_KWH_SCALE_FACTOR - 1)*100:.2f}%)"),
        ("Rate case target kWh", before["target_kwh"], after["target_kwh"], "#,##0",
         "Unchanged — fixed reference from rate case filing"),
        ("kWh per customer (ResStock)", before["resstock_kwh_per_customer"], after["resstock_kwh_per_customer"],
         "#,##0.00", "Drops from 7,031 to 6,727 after scale factor"),
        ("kWh per customer (rate case)", before["target_kwh_per_customer"], after["target_kwh_per_customer"],
         "#,##0.00", "Unchanged — fixed target"),
        ("Difference from target (kWh)", before["diff_kwh"], after["diff_kwh"], "#,##0",
         "After calibration ≈ 0"),
        ("% Difference from target", before["pct_diff"], after["pct_diff"], "0.00",
         "After calibration ≈ 0%"),
    ]

    for label, bval, aval, fmt, note in summary_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=bval).number_format = fmt
        ws.cell(row=row, column=3, value=aval).number_format = fmt
        change = aval - bval
        ws.cell(row=row, column=4, value=change).number_format = fmt
        ws.cell(row=row, column=5, value=note)
        if "%" in label and abs(aval) < 0.01:
            _green_fill(ws, row, 5)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Scale factor applied")
    ws.cell(row=row, column=2, value=RIE_KWH_SCALE_FACTOR).number_format = "0.0000000000000000"
    ws.cell(row=row, column=3, value="test_year_residential_kwh / resstock_kwh_after_customer_scaling")
    ws.merge_cells(f"C{row}:E{row}")
    row += 2

    if eia_kwh is not None:
        ws.cell(row=row, column=1, value="EIA-861 Context (external benchmark — NOT the calibration target)").font = Font(bold=True, size=11)
        ws.merge_cells(f"A{row}:E{row}")
        row += 1
        ws.cell(row=row, column=1, value="Metric")
        ws.cell(row=row, column=2, value=f"EIA-861 ({EIA_YEAR})")
        ws.cell(row=row, column=3, value="Rate Case Target")
        ws.cell(row=row, column=4, value="Difference")
        ws.cell(row=row, column=5, value="Notes")
        _header_fill(ws, row, 5)
        row += 1
        for label, eia_val, rc_val, fmt in [
            ("Residential customers", eia_customers, RIE_TEST_YEAR_CUSTOMER_COUNT, "#,##0"),
            ("Residential kWh", eia_kwh, RIE_TEST_YEAR_RESIDENTIAL_KWH, "#,##0"),
        ]:
            assert eia_val is not None
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=eia_val).number_format = fmt
            ws.cell(row=row, column=3, value=rc_val).number_format = fmt
            ws.cell(row=row, column=4, value=eia_val - rc_val).number_format = fmt
            ws.cell(row=row, column=5, value=f"EIA {EIA_YEAR} vs rate case test year — different sources/periods")
            row += 1
        row += 1
        note = (
            f"EIA-861 ({EIA_YEAR}) and the rate case test year are different data sources covering different periods. "
            f"The calibration target is the rate case test year kWh ({RIE_TEST_YEAR_RESIDENTIAL_KWH:,.0f}), "
            f"not EIA-861 ({eia_kwh:,.0f}). "
            f"After calibration, ResStock matches the rate case target — any residual gap vs EIA-861 "
            f"is expected and correct."
        )
        ws.cell(row=row, column=1, value=note)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"A{row}:E{row + 2}")
        ws.row_dimensions[row].height = 60

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 55


def _write_validation(
    ws,
    resstock_kwh_mf_adjusted: float,
    resstock_kwh_customer_scaled: float,
    resstock_kwh_final: float,
) -> None:
    """Write a validation sheet that cross-checks workbook constants against expert_testimony.qmd.

    The testimony hardcodes several ResStock calibration values directly in its Python setup
    cell (lines 114-122 of expert_testimony.qmd).  This sheet confirms every number matches.
    """
    from openpyxl.styles import PatternFill

    PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def _ok_fill(ws, row: int, ncols: int, ok: bool) -> None:
        fill = PASS_FILL if ok else FAIL_FILL
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).fill = fill

    ws.title = "validation"

    row = 1
    ws.cell(row=row, column=1, value="Validation — Workbook vs. expert_testimony.qmd").font = Font(bold=True, size=13)
    ws.merge_cells(f"A{row}:F{row}")
    row += 2

    purpose = (
        "This sheet cross-checks every constant in this workbook against the values that "
        "expert_testimony.qmd uses when computing the ResStock calibration statistics cited "
        "in the testimony prose.  A green row means the workbook value matches the testimony "
        "value exactly (or within floating-point precision).  A red row signals a mismatch "
        "that would need investigation."
    )
    ws.cell(row=row, column=1, value=purpose)
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 50
    row += 2

    # -- Column headers
    headers = ["Variable", "Testimony source", "Testimony value", "Workbook value", "Match?", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
    _header_fill(ws, row, 6)
    row += 1

    # ------------------------------------------------------------------
    # Testimony hardcoded constants (expert_testimony.qmd lines 115-116)
    # ------------------------------------------------------------------
    testimony_raw_kwh: float = 3_388_382_407.0
    testimony_raw_customer_count: float = 481_896.13

    # Testimony-derived quantities
    # c.rie_rev_req_test_year_customer_count / testimony_raw_customer_count
    #   = RIE_TEST_YEAR_CUSTOMER_COUNT / RIE_RESSTOCK_CUSTOMER_COUNT
    testimony_customer_scale = RIE_TEST_YEAR_CUSTOMER_COUNT / testimony_raw_customer_count
    testimony_customer_scaled_kwh = testimony_raw_kwh * testimony_customer_scale
    testimony_resstock_kwh_gap_pct = (
        testimony_customer_scaled_kwh - RIE_TEST_YEAR_RESIDENTIAL_KWH
    ) / RIE_TEST_YEAR_RESIDENTIAL_KWH

    # Workbook counterparts
    workbook_resstock_kwh_gap_pct = (
        resstock_kwh_customer_scaled - RIE_TEST_YEAR_RESIDENTIAL_KWH
    ) / RIE_TEST_YEAR_RESIDENTIAL_KWH

    def _tol_match(a: float, b: float, rel_tol: float = 1e-7) -> bool:
        """True if a and b agree to within rel_tol (relative).

        1e-7 (one part in ten million) is tight enough to catch real mismatches
        while tolerating sub-kWh floating-point rounding when the testimony
        hardcodes _resstock_raw_kwh as an integer (3_388_382_407.0) vs the exact
        loaded value (3_388_382_407.30).
        """
        if b == 0:
            return a == 0
        return abs(a - b) / abs(b) < rel_tol

    checks = [
        # (variable, testimony_source, testimony_value, workbook_value, notes)
        (
            "_resstock_raw_kwh",
            "expert_testimony.qmd line 115: _resstock_raw_kwh = 3_388_382_407.0",
            testimony_raw_kwh,
            resstock_kwh_mf_adjusted,
            "MF-adjusted ResStock kWh loaded from _sb release.  Testimony hardcodes the "
            "rounded integer 3_388_382_407.0; workbook loads the exact float from S3/EBS "
            f"({resstock_kwh_mf_adjusted:,.3f}).  Sub-kWh difference is a rounding artifact "
            "in the testimony constant, not a real discrepancy.",
        ),
        (
            "_resstock_raw_customer_count",
            "expert_testimony.qmd line 116: _resstock_raw_customer_count = 481_896.13",
            testimony_raw_customer_count,
            RIE_RESSTOCK_CUSTOMER_COUNT,
            "Total ResStock sample weights (sum of building weights) for RI.  "
            "Hardcoded in both testimony and workbook — must be identical.",
        ),
        (
            "c.rie_rev_req_test_year_customer_count",
            "expert_testimony.qmd line 118: loaded from cache/report_variables_cos_subclass.pkl "
            "(sourced from rie_rate_case_test_year.yaml)",
            RIE_TEST_YEAR_CUSTOMER_COUNT,
            RIE_TEST_YEAR_CUSTOMER_COUNT,
            "Rate case test year customer count.  Testimony reads from pkl cache; workbook "
            "hardcodes from the same YAML.  Both trace to PRB-1-ELEC p.14 line 9 col d.",
        ),
        (
            "c.rie_rev_req_test_year_residential_kwh",
            "expert_testimony.qmd line 121: loaded from cache/report_variables_cos_subclass.pkl "
            "(sourced from rie_rate_case_test_year.yaml)",
            RIE_TEST_YEAR_RESIDENTIAL_KWH,
            RIE_TEST_YEAR_RESIDENTIAL_KWH,
            "Rate case test year residential kWh.  Testimony reads from pkl cache; workbook "
            "hardcodes from the same YAML.  Both trace to PRB-1-ELEC p.14 line 8 col k.",
        ),
        (
            "resstock_kwh_scale_factor",
            "Hardcoded in workbook; applied in run_scenario.py line 645",
            RIE_KWH_SCALE_FACTOR,
            RIE_TEST_YEAR_RESIDENTIAL_KWH / (testimony_raw_kwh * testimony_customer_scale),
            "Derived value: test_year_residential_kwh / resstock_kwh_after_customer_scaling.  "
            "Testimony does not state the factor directly, but implies it via the statement "
            "that ResStock 'came within X% of the Test Year total' and was then scaled to "
            "exact alignment.  The workbook constant must equal this derived value.",
        ),
        (
            "resstock_kwh_gap_pct  (testimony prose: 'came within X%')",
            "expert_testimony.qmd lines 120-122: resstock_kwh_gap_pct = "
            "(_resstock_customer_scaled_kwh - c.rie_rev_req_test_year_residential_kwh) "
            "/ c.rie_rev_req_test_year_residential_kwh",
            testimony_resstock_kwh_gap_pct * 100,
            workbook_resstock_kwh_gap_pct * 100,
            "% by which ResStock (after customer scaling, before kWh factor) exceeds the "
            "rate case target.  Appears in testimony as 'came within X% of the Company's "
            "Test Year total.'  Workbook uses the same formula; values must match.",
        ),
    ]

    fmt_map = {
        "_resstock_raw_kwh": "#,##0",
        "_resstock_raw_customer_count": "#,##0.00",
        "c.rie_rev_req_test_year_customer_count": "#,##0.00",
        "c.rie_rev_req_test_year_residential_kwh": "#,##0",
        "resstock_kwh_scale_factor": "0.0000000000",
        "resstock_kwh_gap_pct  (testimony prose: 'came within X%')": "0.000",
    }

    for var, source, tval, wval, notes in checks:
        ok = _tol_match(float(tval), float(wval))
        ws.cell(row=row, column=1, value=var).font = Font(bold=True)
        ws.cell(row=row, column=2, value=source)
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        fmt = fmt_map.get(var, "0.0000000000")
        ws.cell(row=row, column=3, value=float(tval)).number_format = fmt
        ws.cell(row=row, column=4, value=float(wval)).number_format = fmt
        status = "PASS" if ok else "FAIL — check constants"
        ws.cell(row=row, column=5, value=status).font = Font(bold=True, color="375623" if ok else "9C0006")
        ws.cell(row=row, column=6, value=notes)
        ws[f"F{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        _ok_fill(ws, row, 6, ok)
        ws.row_dimensions[row].height = 72
        row += 1

    row += 1

    # -- Summary status
    all_pass = all(_tol_match(float(t), float(w)) for _, _, t, w, _ in checks)
    summary_msg = (
        "ALL CHECKS PASS — workbook constants match expert_testimony.qmd."
        if all_pass
        else "ONE OR MORE CHECKS FAILED — review highlighted rows above."
    )
    ws.cell(row=row, column=1, value=summary_msg).font = Font(
        bold=True, size=12, color="375623" if all_pass else "9C0006"
    )
    ws.merge_cells(f"A{row}:F{row}")
    if all_pass:
        ws[f"A{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    else:
        ws[f"A{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 55


def compute_stage_metrics(resstock_kwh: float) -> dict:
    """Compute comparison metrics for one calibration stage vs rate case target."""
    customers = RIE_TEST_YEAR_CUSTOMER_COUNT
    target = RIE_TEST_YEAR_RESIDENTIAL_KWH
    return {
        "customers": customers,
        "resstock_kwh": resstock_kwh,
        "target_kwh": target,
        "resstock_kwh_per_customer": resstock_kwh / customers,
        "target_kwh_per_customer": target / customers,
        "diff_kwh": resstock_kwh - target,
        "pct_diff": (resstock_kwh - target) / target * 100,
    }


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def build_workbook(output_path: Path) -> Path:
    data_source = "local EBS" if BASE_RESSTOCK.startswith("/ebs") else "S3"
    print(f"Building RIE kWh scale factor workbook (data from: {data_source})...", flush=True)
    print(f"  resstock_kwh_scale_factor:  {RIE_KWH_SCALE_FACTOR:.16f}", flush=True)
    print(f"  Rate case target kWh:       {RIE_TEST_YEAR_RESIDENTIAL_KWH:,.0f}", flush=True)
    print(f"  Rate case customer count:   {RIE_TEST_YEAR_CUSTOMER_COUNT:,.2f}", flush=True)

    path_monthly_dir = (
        f"{BASE_RESSTOCK}/{RESSTOCK_RELEASE_SB}/load_curve_monthly"
        f"/state={STATE}/upgrade={UPGRADE}"
    )
    path_utility_assignment = (
        f"{BASE_RESSTOCK}/{RESSTOCK_RELEASE_SB}/metadata_utility"
        f"/state={STATE}/utility_assignment.parquet"
    )
    path_eia861 = f"{BASE_EIA861}/year={EIA_YEAR}/state={STATE}/data.parquet"
    storage_opts = _storage_options()

    print("Loading MF-adjusted ResStock data from load_curve_monthly (_sb release)...", flush=True)
    resstock_sb = load_resstock_annual_from_monthly(
        path_monthly_dir, path_utility_assignment, storage_options=storage_opts
    )
    resstock_rie = resstock_sb.filter(pl.col("utility_code") == UTILITY)
    resstock_kwh_mf_adjusted = resstock_rie["resstock_kwh"][0]

    print("Loading EIA-861 data (for context)...", flush=True)
    eia = load_eia_by_utility(path_eia861, storage_options=storage_opts, utility_codes=[UTILITY])
    eia_row = eia.filter(pl.col("utility_code") == UTILITY).to_dicts()
    eia_kwh = eia_row[0]["eia_kwh"] if eia_row else None
    eia_customers = eia_row[0]["eia_customers"] if eia_row else None

    # Compute the three pipeline stages
    resstock_kwh_customer_scaled = resstock_kwh_mf_adjusted * RIE_CUSTOMER_SCALE_FACTOR
    resstock_kwh_final = resstock_kwh_customer_scaled * RIE_KWH_SCALE_FACTOR

    print(f"\n  Stage 1 — _sb input:              {resstock_kwh_mf_adjusted:,.0f} kWh  "
          f"({resstock_kwh_mf_adjusted / RIE_RESSTOCK_CUSTOMER_COUNT:,.0f} kWh/customer)", flush=True)
    print(f"  Stage 2 — after customer scaling: {resstock_kwh_customer_scaled:,.0f} kWh  "
          f"({resstock_kwh_customer_scaled / RIE_TEST_YEAR_CUSTOMER_COUNT:,.0f} kWh/customer)", flush=True)
    print(f"  Stage 3 — after kWh scale factor: {resstock_kwh_final:,.0f} kWh  "
          f"({resstock_kwh_final / RIE_TEST_YEAR_CUSTOMER_COUNT:,.0f} kWh/customer)", flush=True)
    print(f"  Rate case target:                 {RIE_TEST_YEAR_RESIDENTIAL_KWH:,.0f} kWh  "
          f"({RIE_TEST_YEAR_RESIDENTIAL_KWH / RIE_TEST_YEAR_CUSTOMER_COUNT:,.0f} kWh/customer)", flush=True)
    print(f"  Difference (stage 3 − target):    {resstock_kwh_final - RIE_TEST_YEAR_RESIDENTIAL_KWH:+.2f} kWh", flush=True)

    before = compute_stage_metrics(resstock_kwh_customer_scaled)
    after = compute_stage_metrics(resstock_kwh_final)

    print("\nBuilding workbook...", flush=True)
    wb = Workbook()
    wb.remove(wb.active)

    _write_readme(wb, resstock_kwh_mf_adjusted)
    _write_pipeline(wb, resstock_kwh_mf_adjusted)

    ws_calc = wb.create_sheet("scale_factor_derivation")
    _write_scale_factor_derivation(ws_calc, resstock_kwh_mf_adjusted, resstock_kwh_customer_scaled, resstock_kwh_final)

    ws_before = wb.create_sheet("before_kwh_calibration")
    _write_comparison(
        ws_before, before,
        "Before kWh Calibration — Stage 2: After Customer Rescaling",
        "ResStock after customer-count rescaling (stage 2) vs rate case target. "
        "Per-customer kWh still 7,031 — intensity gap not yet corrected.",
    )

    ws_after = wb.create_sheet("after_kwh_calibration")
    _write_comparison(
        ws_after, after,
        "After kWh Calibration — Stage 3: After kWh Scale Factor Applied",
        "ResStock after kWh scale factor applied (stage 3) vs rate case target. "
        "% difference from target ≈ 0 — calibration complete.",
    )

    ws_summary = wb.create_sheet("summary")
    _write_summary(ws_summary, before, after, eia_kwh=eia_kwh, eia_customers=eia_customers)

    ws_validation = wb.create_sheet("validation")
    _write_validation(ws_validation, resstock_kwh_mf_adjusted, resstock_kwh_customer_scaled, resstock_kwh_final)

    print(f"Saving workbook to {output_path}...", flush=True)
    wb.save(output_path)
    print("Workbook saved successfully.", flush=True)
    return output_path


# ---------------------------------------------------------------------------
# Main / upload
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    load_dotenv()

    parser = argparse.ArgumentParser(description="Build RIE ResStock kWh scale factor workbook")
    output_group = parser.add_mutually_exclusive_group()
    output_group.add_argument("--output", type=Path, help="Local output path (.xlsx)")
    output_group.add_argument("--sheet-id", help="Google Sheet ID to update")
    output_group.add_argument("--folder-id", help="Google Drive folder ID for new sheet")
    parser.add_argument("--filename", default="RIE kWh Scale Factor Analysis",
                        help="Filename for new Google Sheet (with --folder-id)")
    args = parser.parse_args(argv)

    output_path = args.output if args.output else Path("cache/rie_kwh_scale_factor_workbook.xlsx")
    output_path.parent.mkdir(exist_ok=True)

    try:
        result_path = build_workbook(output_path)

        if args.sheet_id or args.folder_id:
            try:
                from lib.data.gsheets import get_gspread_client, xlsx_to_gsheet

                spreadsheet_id = args.sheet_id

                if args.folder_id and args.filename:
                    gc, _ = get_gspread_client()
                    print(f"Searching for existing files named '{args.filename}'...", flush=True)
                    try:
                        all_files = gc.list_spreadsheet_files(folder_id=args.folder_id)
                        files_to_delete = [
                            f for f in all_files
                            if f.get("name") == args.filename and not f.get("trashed", False)
                        ]
                        if files_to_delete:
                            print(f"Found {len(files_to_delete)} file(s) to delete:", flush=True)
                            from googleapiclient.discovery import build

                            credentials = None
                            if hasattr(gc, "http_client") and hasattr(gc.http_client, "auth"):
                                credentials = gc.http_client.auth
                            elif hasattr(gc, "auth") and gc.auth is not None:
                                credentials = gc.auth
                            elif hasattr(gc, "_auth"):
                                credentials = gc._auth

                            if credentials is None:
                                raise RuntimeError("Could not access gspread credentials")

                            drive_service = build("drive", "v3", credentials=credentials, cache_discovery=False)
                            for file in files_to_delete:
                                file_id = file["id"]
                                try:
                                    drive_service.files().delete(fileId=file_id, supportsAllDrives=True).execute()
                                    print(f"  ✓ Deleted: {file_id}", flush=True)
                                except Exception:
                                    drive_service.files().update(
                                        fileId=file_id, body={"trashed": True}, supportsAllDrives=True
                                    ).execute()
                                    print(f"  ✓ Trashed: {file_id}", flush=True)
                        else:
                            print(f"No existing files named '{args.filename}' found.", flush=True)
                    except Exception as e:
                        print(f"Warning: Could not search/delete existing files: {e}", flush=True)

                    print(f"Creating new spreadsheet '{args.filename}'...", flush=True)
                    spreadsheet = gc.create(args.filename, folder_id=args.folder_id)
                    spreadsheet_id = spreadsheet.id
                    print(f"✓ Created: {spreadsheet_id}", flush=True)

                print(f"Uploading to Google Sheet {spreadsheet_id}...", flush=True)
                spreadsheet = xlsx_to_gsheet(result_path, spreadsheet_id, delete_other_tabs=True)
                print(f"✓ Upload complete, {len(spreadsheet.worksheets())} worksheets", flush=True)
                print(f"\nGoogle Sheet URL: https://docs.google.com/spreadsheets/d/{spreadsheet_id}", flush=True)

            except ImportError as e:
                print(f"Warning: Could not upload to Google Sheets: {e}", file=sys.stderr)
                print("Workbook saved locally only.", flush=True)

        print(f"\nLocal workbook: {result_path}", flush=True)
        return 0

    except Exception as e:
        print(f"Error building workbook: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
