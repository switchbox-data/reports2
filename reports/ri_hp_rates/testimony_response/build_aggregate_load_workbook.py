"""Build aggregate ResStock load workbooks for RI buildings by heating-type subclass.

Produces workbooks from two sources, written to separate Drive subfolders:

**"Raw ResStock load curve"** — upgrade=0 and upgrade=2 workbooks sourced from
individual per-building ResStock hourly parquet files on local disk. Each workbook has:

  - README (first sheet): data provenance, subclass table, column descriptions
  - One sheet per heating-type subclass: 8,760-row aggregate load (weighted sum)
  - Validation sheet

Columns included are all ``out.*.energy_consumption``, ``out.*.energy_delivered.kbtu``,
and ``out.*.co2e_kg`` columns. Intensity (per-sqft) and temperature columns are excluded.

**"CAIRO calculation load curve"** — upgrade=0 only, sourced from CAIRO's
``billing_kwh_8760.parquet`` (the same kWh series CAIRO uses internally for economic
burden and EPMC allocation). Weights and kWh are rescaled to match the rate-case test
year. Each workbook has:

  - README (first sheet): data provenance, CAIRO source description
  - One sheet per heating-type subclass: 8,760-row weighted-sum kWh

Subclasses are taken from ``postprocess_group.heating_type_v2`` in the ResStock
metadata. Weighted-sum aggregation uses ResStock sample weights from
``metadata_utility/state=RI/utility_assignment.parquet``.

Run from the report directory::

    uv run python -m testimony_response.build_aggregate_load_workbook
    uv run python -m testimony_response.build_aggregate_load_workbook --no-upload
    uv run python -m testimony_response.build_aggregate_load_workbook --no-upload \\
        --output-dir cache/aggregate_loads
    uv run python -m testimony_response.build_aggregate_load_workbook --upgrades 0
    uv run python -m testimony_response.build_aggregate_load_workbook --source raw
    uv run python -m testimony_response.build_aggregate_load_workbook --source cairo
"""

from __future__ import annotations

import argparse
import pickle
import sys
from pathlib import Path
from typing import Any

import numpy as np
import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Constants ──────────────────────────────────────────────────────────────────

REPORT_DIR = Path(__file__).resolve().parents[1]

UTILITY = "rie"
RESSTOCK_DATASET = "res_2024_amy2018_2_sb"
RESSTOCK_S3_BASE = f"s3://data.sb/nrel/resstock/{RESSTOCK_DATASET}"
LOCAL_RESSTOCK_BASE = Path(f"/ebs/data/nrel/resstock/{RESSTOCK_DATASET}")
LOCAL_METADATA_UTILITY = LOCAL_RESSTOCK_BASE / "metadata_utility" / "state=RI" / "utility_assignment.parquet"
LOCAL_METADATA_DIR = LOCAL_RESSTOCK_BASE / "metadata" / "state=RI"
LOCAL_LOADS_DIR = LOCAL_RESSTOCK_BASE / "load_curve_hourly" / "state=RI"

SUBCLASS_COL = "postprocess_group.heating_type_v2"
UPGRADE_DIR: dict[int, str] = {0: "upgrade=00", 2: "upgrade=02"}
UPGRADE_LABEL: dict[int, str] = {
    0: "Upgrade 0 — Baseline (as-built, no heat pump)",
    2: "Upgrade 2 — Heat Pump Retrofit (ccASHP)",
}
UPGRADE_TITLE: dict[int, str] = {
    0: "RI Aggregate Loads — Upgrade 0 (Baseline)",
    2: "RI Aggregate Loads — Upgrade 2 (Heat Pump)",
}
UPGRADE_FILE: dict[int, str] = {
    0: "ri_aggregate_loads_upgrade0.xlsx",
    2: "ri_aggregate_loads_upgrade2.xlsx",
}

# Preferred display order for subclass sheets; unknowns appended alphabetically.
SUBCLASS_PREFERRED_ORDER = [
    "heat_pump",
    "electrical_resistance",
    "natgas",
    "delivered_fuels",
    "other",
]
SUBCLASS_DISPLAY: dict[str, str] = {
    "heat_pump": "Heat Pump",
    "electrical_resistance": "Elec. Resistance",
    "natgas": "Natural Gas",
    "delivered_fuels": "Delivered Fuels",
    "other": "Other",
}
SUBCLASS_DESCRIPTION: dict[str, str] = {
    "heat_pump": "Buildings heated primarily by electric heat pumps (ccASHP or GSHP).",
    "electrical_resistance": "Buildings heated by electric resistance (baseboard heaters, etc.).",
    "natgas": "Buildings heated by natural gas (furnace, boiler, etc.).",
    "delivered_fuels": "Buildings heated by delivered fuels (fuel oil or propane).",
    "other": "Buildings with unclassified or mixed heating systems.",
}

# CAIRO billing kWh sources — update together with build_RIE_1_7_workbook.py when batch changes.
_CAIRO_BATCH = "ri_20260507_r1-2_grid_cons_fix"
_CAIRO_S3_BASE = "s3://data.sb/switchbox/cairo/outputs/hp_rates"
S3_BILLING_KWH_8760: dict[int, str] = {
    0: (
        f"{_CAIRO_S3_BASE}/ri/rie/{_CAIRO_BATCH}"
        "/20260507_213944_ri_rie_run1_up00_precalc__default/billing_kwh_8760.parquet"
    ),
    2: (
        f"{_CAIRO_S3_BASE}/ri/rie/{_CAIRO_BATCH}"
        "/20260507_214049_ri_rie_run3_up02_default__default/billing_kwh_8760.parquet"
    ),
}
BILLING_KWH_COL = "grid_cons_kwh"  # "grid_cons_kwh" = floored at 0; "load_data_kwh" = raw

CAIRO_UPGRADE_TITLE: dict[int, str] = {
    0: "RI CAIRO Aggregate Loads — Upgrade 0 (Baseline)",
    2: "RI CAIRO Aggregate Loads — Upgrade 2 (Heat Pump)",
}
CAIRO_UPGRADE_FILE: dict[int, str] = {
    0: "ri_cairo_aggregate_loads_upgrade0.xlsx",
    2: "ri_cairo_aggregate_loads_upgrade2.xlsx",
}

# Drive
DEFAULT_PARENT_FOLDER_ID = "1uPcJbcOChD6zoFuPb-gsxSByPr7xwmCH"
SUBFOLDER_NAME_RAW = "Raw ResStock load curve"
SUBFOLDER_NAME_CAIRO = "CAIRO calculation load curve"

# Styling — match other testimony workbooks
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PASS_FONT = Font(bold=True, color="276221")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FAIL_FONT = Font(bold=True, color="9C0006")
INFO_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
INFO_FONT = Font(bold=True, color="9C6500")

# ── Scale-factor constants (from rie_rate_case_test_year.yaml @ e9e5088) ───────
# Applied by the RIE 1-7 workbook to normalise ResStock loads to the rate-case
# basis; used here only for the upgrade=0 testimony cross-check.
_RIE17_KWH_SCALE_FACTOR = 0.9568112362177266
_RIE17_TEST_YEAR_CUSTOMER_COUNT = 419347.83
_RIE17_TEST_YEAR_KWH = 2_821_237_490.0

# Testimony expected annual electricity (GWh) by subclass, upgrade=0.
# Source: cache/report_variables_cos_subclass.pkl, key "testimony_subclass_delivery_rows".
# Updated whenever the cost_of_service_by_subclass.qmd notebook is re-run.
_TESTIMONY_SUBCLASS_DISPLAY_TO_KEY: dict[str, str] = {
    "Heat pump": "heat_pump",
    "Electric resistance": "electrical_resistance",
    "Natural gas": "natgas",
    "Delivered fuels": "delivered_fuels",
    "Other": "other",
}
_TESTIMONY_GWH_TOLERANCE_PCT = 0.3  # ±0.3 %


# ── Column helpers ─────────────────────────────────────────────────────────────


def _aggregate_columns(sample_file: Path) -> list[str]:
    """Return the columns that should be aggregated (weighted sum).

    Includes: ``out.*.energy_consumption``, ``out.*.energy_delivered.kbtu``,
    ``out.*.co2e_kg``.  Excludes: ``_intensity`` (per-sqft), temperature
    columns (``.c`` suffix), and metadata columns (timestamp, bldg_id, etc.).
    """
    schema = pl.read_parquet_schema(sample_file)
    excluded_exact = {"timestamp", "bldg_id", "year", "month", "day", "hour"}
    result = []
    for col in schema:
        if col in excluded_exact:
            continue
        if "intensity" in col:
            continue
        # Temperature columns end in .c (e.g. out.outdoor_air_dryblub_temp.c)
        if col.endswith(".c"):
            continue
        result.append(col)
    return result


def _ordered_subclasses(subclasses: list[str]) -> list[str]:
    """Sort subclasses in preferred display order; unknowns at end."""
    known = [sc for sc in SUBCLASS_PREFERRED_ORDER if sc in subclasses]
    unknown = sorted(sc for sc in subclasses if sc not in SUBCLASS_PREFERRED_ORDER)
    return known + unknown


def _sheet_name(sheet_num: int, subclass: str) -> str:
    display = SUBCLASS_DISPLAY.get(subclass, subclass.replace("_", " ").title())
    return f"{sheet_num}. {display}"


# ── Data loading ───────────────────────────────────────────────────────────────


def load_metadata_for_upgrade(upgrade_id: int) -> pl.DataFrame:
    """Load RI/RIE building metadata for the given upgrade.

    Joins ``utility_assignment.parquet`` (bldg_id, weight, utility) with
    the upgrade-specific metadata table (postprocess_group.heating_type_v2).
    Returns only buildings assigned to the RIE service territory.
    """
    util_meta = (
        pl.read_parquet(LOCAL_METADATA_UTILITY)
        .filter(pl.col("sb.electric_utility") == UTILITY)
        .select("bldg_id", "weight")
    )
    upgrade_meta = pl.read_parquet(LOCAL_METADATA_DIR / UPGRADE_DIR[upgrade_id]).select("bldg_id", SUBCLASS_COL)
    return util_meta.join(upgrade_meta, on="bldg_id", how="inner")


def _build_file_lookup(upgrade_id: int) -> dict[int, Path]:
    """Map bldg_id → Path for every load file in the given upgrade directory."""
    load_dir = LOCAL_LOADS_DIR / UPGRADE_DIR[upgrade_id]
    return {int(f.stem.split("-")[0]): f for f in sorted(load_dir.iterdir())}


def aggregate_subclass_loads(
    buildings: list[dict[str, Any]],
    agg_cols: list[str],
    file_lookup: dict[int, Path],
) -> pl.DataFrame:
    """Weighted-sum aggregate for one subclass (8,760 rows if no leap day).

    Reads each building's load file one at a time to minimise peak memory.
    If the data contains Feb 29 (leap year), those rows are dropped.
    ``buildings`` is a list of ``{"bldg_id": int, "weight": float}`` dicts.
    Returns a DataFrame with columns: hour_of_year, timestamp, year, month,
    day, hour, then all agg_cols.
    """
    acc: dict[str, np.ndarray] | None = None
    time_df: pl.DataFrame | None = None
    n = len(buildings)

    for i, bldg in enumerate(buildings):
        if i % 100 == 0:
            print(f"    [{i}/{n}] ...", flush=True)
        bldg_id: int = bldg["bldg_id"]
        weight: float = bldg["weight"]
        path = file_lookup.get(bldg_id)
        if path is None:
            print(f"    Warning: no load file for bldg_id={bldg_id}, skipping", flush=True)
            continue
        df = pl.read_parquet(path).sort("timestamp")
        if df["month"][0] is not None:
            df = df.filter(~((pl.col("month") == 2) & (pl.col("day") == 29)))
        if time_df is None:
            time_df = df.select(["timestamp", "year", "month", "day", "hour"])
            acc = {c: np.zeros(df.height, dtype=np.float64) for c in agg_cols}
        for c in agg_cols:
            if c in df.columns:
                acc[c] += df[c].cast(pl.Float64).fill_null(0.0).to_numpy() * weight

    if time_df is None or acc is None:
        raise RuntimeError("No load files found for subclass")

    n_rows = time_df.height
    return time_df.with_columns(
        [pl.Series(name=c, values=acc[c]) for c in agg_cols]
        + [
            pl.col("timestamp").dt.strftime("%Y-%m-%d %H:%M:%S"),
            (pl.int_range(1, n_rows + 1, eager=True)).alias("hour_of_year"),
        ]
    ).select(["hour_of_year", "timestamp", "year", "month", "day", "hour", *agg_cols])


def load_all_subclass_aggregates(
    upgrade_id: int,
    meta: pl.DataFrame,
) -> tuple[dict[str, pl.DataFrame], list[str]]:
    """Aggregate all subclass load curves for the given upgrade.

    Returns ``(agg_by_subclass, agg_cols)`` where ``agg_by_subclass`` maps
    each subclass key to its 8,760-row weighted-sum DataFrame, and ``agg_cols``
    is the ordered list of energy/emissions columns included.
    """
    file_lookup = _build_file_lookup(upgrade_id)
    sample_path = next(iter(file_lookup.values()))
    agg_cols = _aggregate_columns(sample_path)

    # Group buildings by subclass
    subclass_groups: dict[str, list[dict[str, Any]]] = {}
    for row in meta.iter_rows(named=True):
        sc = row[SUBCLASS_COL]
        subclass_groups.setdefault(sc, []).append({"bldg_id": row["bldg_id"], "weight": row["weight"]})

    ordered = _ordered_subclasses(list(subclass_groups.keys()))
    result: dict[str, pl.DataFrame] = {}
    for sc in ordered:
        buildings = subclass_groups[sc]
        print(f"  Aggregating '{sc}' ({len(buildings)} buildings) ...", flush=True)
        result[sc] = aggregate_subclass_loads(buildings, agg_cols, file_lookup)

    return result, agg_cols


# ── CAIRO data loading ─────────────────────────────────────────────────────────


def load_cairo_aggregate_loads(upgrade_id: int) -> tuple[dict[str, pl.DataFrame], pl.DataFrame]:
    """Load CAIRO billing_kwh_8760 from S3 and aggregate by heating_type_v2 subclass.

    Uses the ``billing_kwh_8760.parquet`` produced by CAIRO for the given upgrade
    (0 = baseline, 2 = heat pump retrofit).  The kWh column used is ``BILLING_KWH_COL``
    (``grid_cons_kwh``, floored at 0).

    Only the weight rescaling is applied — **not** the kWh scale factor:
      - Weights are rescaled so their sum equals the rate-case test-year customer count.
      - No additional kWh scale is applied because CAIRO has already normalized
        ``grid_cons_kwh`` during its internal calibration run.  The raw weighted sum
        of ``grid_cons_kwh`` therefore equals the test-year residential kWh
        (2,821,237,490 kWh) directly.  Applying ``_RIE17_KWH_SCALE_FACTOR`` again
        (as done for raw ResStock load files) would double-deflate the totals.

    Subclass assignments come from the upgrade-specific ResStock metadata, so upgrade=2
    buildings reflect post-retrofit heating types (all buildings appear as heat_pump).

    Returns ``(agg_by_subclass, meta)`` where:
      - ``agg_by_subclass`` maps each ``heating_type_v2`` subclass key to a 3-column
        DataFrame: ``hour_of_year``, ``timestamp``, ``weighted_kwh``.
      - ``meta`` is the unscaled building-level metadata (bldg_id, subclass, weight)
        used for building-count validation.
    """
    meta = load_metadata_for_upgrade(upgrade_id)

    # Rescale weights to match CAIRO's normalization to the rate-case customer count.
    raw_total = float(meta["weight"].sum())
    weight_scale = _RIE17_TEST_YEAR_CUSTOMER_COUNT / raw_total
    meta = meta.with_columns(pl.col("weight") * weight_scale)

    s3_path = S3_BILLING_KWH_8760[upgrade_id]
    print(f"Loading CAIRO billing kWh 8760 from S3: {s3_path}", flush=True)
    print(f"  Using column: {BILLING_KWH_COL}", flush=True)
    billing = pl.read_parquet(s3_path).select("bldg_id", "timestamp", BILLING_KWH_COL)

    loads = billing.join(meta, on="bldg_id")
    print(
        f"  Matched {loads['bldg_id'].n_unique():,} buildings, {loads.height:,} hourly rows",
        flush=True,
    )

    # Weighted sum per subclass per hour, pivoted to wide format.
    agg = (
        loads.group_by(["timestamp", SUBCLASS_COL])
        .agg((pl.col(BILLING_KWH_COL) * pl.col("weight")).sum().alias("weighted_kwh"))
        .pivot(on=SUBCLASS_COL, index="timestamp", values="weighted_kwh")
        .sort("timestamp")
        .head(8760)
    )
    agg = agg.with_columns((pl.int_range(1, agg.height + 1, eager=True)).alias("hour_of_year"))

    # Fill any missing subclass-hours with 0. No additional kWh scale factor is
    # applied: CAIRO's internal calibration already normalizes grid_cons_kwh so the
    # weighted total matches the test-year residential kWh.
    load_cols = [c for c in agg.columns if c not in ("timestamp", "hour_of_year")]
    agg = agg.with_columns([pl.col(c).fill_null(0.0) for c in load_cols])

    # Return one 3-column DataFrame per subclass (matching the interface of
    # load_all_subclass_aggregates so build_workbook_cairo can reuse sheet writers).
    result: dict[str, pl.DataFrame] = {}
    for sc in _ordered_subclasses(load_cols):
        if sc in agg.columns:
            result[sc] = agg.select(
                "hour_of_year",
                "timestamp",
                pl.col(sc).alias("weighted_kwh"),
            )
    return result, meta


# ── Workbook construction ──────────────────────────────────────────────────────


def _add_readme_sheet(
    wb: Workbook,
    upgrade_id: int,
    meta: pl.DataFrame,
    agg_by_subclass: dict[str, pl.DataFrame],
    agg_cols: list[str],
) -> None:
    """Add the README sheet as the first tab in the workbook."""
    ws = wb.create_sheet("README")

    # ── Title ──────────────────────────────────────────────────────────────────
    ws["A1"] = UPGRADE_TITLE[upgrade_id]
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")

    # ── Purpose ───────────────────────────────────────────────────────────────
    row = 3
    ws[f"A{row}"] = "Purpose"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    ws[f"A{row}"] = (
        f"This workbook contains hourly aggregate electric load curves for Rhode Island "
        f"Narragansett Electric (RIE) residential buildings from NREL's ResStock dataset, "
        f"grouped by heating-fuel subclass. Each tab shows the 8,760-hour weighted-sum "
        f"aggregate load for one subclass under {UPGRADE_LABEL[upgrade_id]}. "
        f"Subclasses are assigned by the '{SUBCLASS_COL}' column in the ResStock metadata. "
        f"Columns are aggregated as S(building_value x sample_weight) so that totals "
        f"represent the estimated RIE residential customer population."
    )
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row}:H{row}")
    ws.row_dimensions[row].height = 65
    row += 2

    # ── Data Sources ──────────────────────────────────────────────────────────
    ws[f"A{row}"] = "Data Sources"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    s3_loads = f"{RESSTOCK_S3_BASE}/load_curve_hourly/state=RI/{UPGRADE_DIR[upgrade_id]}/"
    s3_meta = f"{RESSTOCK_S3_BASE}/metadata/state=RI/{UPGRADE_DIR[upgrade_id]}/"
    s3_util = f"{RESSTOCK_S3_BASE}/metadata_utility/state=RI/utility_assignment.parquet"

    for col, label in [(1, "Dataset"), (2, "S3 Path"), (3, "Description")]:
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = Font(bold=True)
    ws.merge_cells(f"C{row}:H{row}")
    row += 1

    sources = [
        (
            "ResStock load curves",
            s3_loads,
            f"Hourly per-building electricity and fuel load profiles (8,760 rows x "
            f"~{len(agg_cols)} columns per building). Dataset: {RESSTOCK_DATASET}. "
            f"Upgrade: {UPGRADE_DIR[upgrade_id]}. Workbook built from local disk at "
            f"/ebs/data/nrel/resstock/{RESSTOCK_DATASET}/.",
        ),
        (
            "ResStock metadata",
            s3_meta,
            f"Building characteristics including '{SUBCLASS_COL}' (heating-type "
            f"subclass classification for upgrade {upgrade_id}).",
        ),
        (
            "Utility assignment",
            s3_util,
            "Maps bldg_id to electric utility (sb.electric_utility) and provides "
            "ResStock sample weights. Only upgrade=0 rows are present; bldg_ids are "
            "the same across upgrades.",
        ),
    ]
    for name, path, desc in sources:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=path)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=desc)
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"C{row}:H{row}")
        ws.row_dimensions[row].height = 55
        row += 1
    row += 1

    # ── Subclasses ────────────────────────────────────────────────────────────
    ws[f"A{row}"] = "Subclasses"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    for col, label in [
        (1, "Subclass key"),
        (2, "Display name"),
        (3, "Sheet"),
        (4, "# Buildings"),
        (5, "Description"),
    ]:
        ws.cell(row=row, column=col, value=label).font = Font(bold=True)
    ws.merge_cells(f"E{row}:H{row}")
    row += 1

    for sheet_num, sc in enumerate(agg_by_subclass.keys(), start=2):
        n_bldgs = meta.filter(pl.col(SUBCLASS_COL) == sc).height
        ws.cell(row=row, column=1, value=sc)
        ws.cell(row=row, column=2, value=SUBCLASS_DISPLAY.get(sc, sc))
        ws.cell(row=row, column=3, value=_sheet_name(sheet_num, sc))
        ws.cell(row=row, column=4, value=n_bldgs)
        ws.cell(row=row, column=5, value=SUBCLASS_DESCRIPTION.get(sc, ""))
        ws.merge_cells(f"E{row}:H{row}")
        row += 1
    row += 1

    # ── Column Descriptions ───────────────────────────────────────────────────
    ws[f"A{row}"] = "Column Descriptions"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    for col, label in [
        (1, "Column"),
        (2, "Type"),
        (3, "Unit"),
        (4, "Description"),
    ]:
        ws.cell(row=row, column=col, value=label).font = Font(bold=True)
    ws.merge_cells(f"D{row}:H{row}")
    row += 1

    time_col_rows = [
        ("hour_of_year", "Integer", "-", "Hour index 1-8,760 (chronological within AMY 2018)."),
        (
            "timestamp",
            "String",
            "-",
            "Datetime string YYYY-MM-DD HH:MM:SS. Weather year: AMY 2018 (actual 2018 weather).",
        ),
        ("year", "Integer", "-", "Calendar year (2018)."),
        ("month", "Integer", "-", "Calendar month (1-12)."),
        ("day", "Integer", "-", "Calendar day of month."),
        ("hour", "Integer", "-", "Hour of day (0-23)."),
    ]
    for col_name, col_type, unit, desc in time_col_rows:
        ws.cell(row=row, column=1, value=col_name)
        ws.cell(row=row, column=2, value=col_type)
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=desc)
        ws.merge_cells(f"D{row}:H{row}")
        row += 1

    n_elec = sum(1 for c in agg_cols if c.startswith("out.electricity."))
    n_gas = sum(1 for c in agg_cols if c.startswith("out.natural_gas."))
    n_oil = sum(1 for c in agg_cols if c.startswith("out.fuel_oil."))
    n_propane = sum(1 for c in agg_cols if c.startswith("out.propane."))
    n_site = sum(1 for c in agg_cols if c.startswith("out.site_energy."))
    n_load = sum(1 for c in agg_cols if c.startswith("out.load."))
    n_co2 = sum(1 for c in agg_cols if c.endswith(".co2e_kg"))

    energy_col_rows = [
        (
            f"out.electricity.* ({n_elec} cols)",
            "Float",
            "kWh",
            "Weighted-sum aggregate electricity consumption by end use "
            "(heating, cooling, hot water, lighting, appliances, etc.). "
            "Intensity (per-sqft) columns are excluded. "
            "Aggregation: S(bldg_kWh x sample_weight).",
        ),
        (
            f"out.natural_gas.* ({n_gas} cols)\nout.fuel_oil.* ({n_oil} cols)\nout.propane.* ({n_propane} cols)",
            "Float",
            "kWh",
            "Weighted-sum aggregate fuel energy consumption by fuel type and end use. "
            "Original ResStock units converted to kWh equivalent.",
        ),
        (
            f"out.site_energy.* ({n_site} cols)",
            "Float",
            "kWh",
            "Weighted-sum aggregate site energy (net and total, all fuels combined).",
        ),
        (
            f"out.load.* ({n_load} cols)",
            "Float",
            "kBtu",
            "Weighted-sum aggregate thermal load delivered by HVAC and hot-water systems. Units: kBtu.",
        ),
        (
            f"out.total.lrmer_* ({n_co2} cols)",
            "Float",
            "kg CO₂e",
            "Weighted-sum aggregate lifecycle CO2-equivalent emissions under various NREL "
            "Cambium LRMER scenarios (MidCase, HighRECost, LowRECost; 15- and 25-year).",
        ),
    ]
    for col_name, col_type, unit, desc in energy_col_rows:
        ws.cell(row=row, column=1, value=col_name)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=2, value=col_type)
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=desc)
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"D{row}:H{row}")
        ws.row_dimensions[row].height = 50
        row += 1
    row += 1

    # ── Notes ─────────────────────────────────────────────────────────────────
    ws[f"A{row}"] = "Notes"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    notes = [
        (
            "Aggregation methodology",
            "Each column is S(building_value x sample_weight) across all buildings "
            "in the subclass. Weights come from the ResStock sample and are NOT "
            "renormalized to the RIE test-year customer count in this workbook (unlike "
            "the RIE 1-7 cost-allocation workbook, which applies a resstock_kwh_scale_factor).",
        ),
        (
            "Weather year",
            "AMY 2018 (Actual Meteorological Year 2018). If the source data includes "
            "Feb 29 (leap year), those 24 hours are removed to produce a standard "
            "8,760-hour year. Hours are sorted chronologically.",
        ),
        (
            "Upgrade definition",
            f"{UPGRADE_LABEL[upgrade_id]}. "
            "Upgrade 0 = as-built baseline with each building's actual heating system. "
            "Upgrade 2 = all buildings assumed to have cold-climate air-source heat pumps "
            "(ccASHP) installed. Subclass labels reflect the post-upgrade heating system.",
        ),
        (
            "RIE filter",
            "Only buildings assigned to the Narragansett Electric (RIE) service territory "
            "are included, as determined by utility_assignment.parquet.",
        ),
        (
            "Local data source",
            f"Workbook built from local disk: /ebs/data/nrel/resstock/{RESSTOCK_DATASET}/. "
            "The canonical S3 source paths are listed in the Data Sources section above.",
        ),
    ]
    for label, note in notes:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=note).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"B{row}:H{row}")
        ws.row_dimensions[row].height = 45
        row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 45
    for letter in ["E", "F", "G", "H"]:
        ws.column_dimensions[letter].width = 10


def _add_subclass_sheet(
    wb: Workbook,
    sheet_num: int,
    subclass: str,
    df: pl.DataFrame,
) -> None:
    """Add one subclass data sheet (8,760 rows x all columns) to the workbook."""
    ws = wb.create_sheet(_sheet_name(sheet_num, subclass))
    headers = list(df.columns)

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # Data rows — use ws.append() for speed
    for row_data in df.iter_rows():
        ws.append(list(row_data))

    ws.freeze_panes = "A2"

    # Column widths
    width_map = {"A": 14, "B": 22, "C": 7, "D": 8, "E": 6, "F": 6}
    for letter, width in width_map.items():
        ws.column_dimensions[letter].width = width
    for c in range(len(width_map) + 1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13


# ── CAIRO workbook ─────────────────────────────────────────────────────────────


def _add_readme_sheet_cairo(
    wb: Workbook,
    upgrade_id: int,
    agg_by_subclass: dict[str, pl.DataFrame],
) -> None:
    """README sheet for the CAIRO calculation load curve workbook."""
    ws = wb.create_sheet("README")

    ws["A1"] = CAIRO_UPGRADE_TITLE[upgrade_id]
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")

    row = 3
    ws[f"A{row}"] = "Purpose"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    ws[f"A{row}"] = (
        "This workbook contains hourly aggregate electricity load curves for Rhode Island "
        "Energy (RIE) residential buildings, grouped by heating-fuel subclass. "
        f"Upgrade {upgrade_id}: {UPGRADE_LABEL[upgrade_id]}. "
        "Unlike the 'Raw ResStock load curve' workbook (which aggregates directly from "
        "per-building ResStock parquet files), this workbook uses the kWh series produced "
        f"by CAIRO's billing calculation (column: '{BILLING_KWH_COL}'). "
        "This is the same kWh series CAIRO uses internally when computing economic burden "
        "and EPMC cost allocation. Weights are rescaled to the rate-case test-year customer "
        "count and kWh is scaled by the rate-case resstock_kwh_scale_factor, so totals "
        "match the test-year residential kWh exactly."
    )
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row}:H{row}")
    ws.row_dimensions[row].height = 80
    row += 2

    ws[f"A{row}"] = "Data Sources"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    for col, label in [(1, "Dataset"), (2, "S3 Path / Value"), (3, "Description")]:
        ws.cell(row=row, column=col, value=label).font = Font(bold=True)
    ws.merge_cells(f"C{row}:H{row}")
    row += 1

    s3_util = f"{RESSTOCK_S3_BASE}/metadata_utility/state=RI/utility_assignment.parquet"
    s3_meta = f"{RESSTOCK_S3_BASE}/metadata/state=RI/{UPGRADE_DIR[0]}/"
    sources = [
        (
            "CAIRO billing kWh 8760",
            S3_BILLING_KWH_8760[upgrade_id],
            f"Per-building 8,760-hour electricity consumption produced by CAIRO's billing "
            f"calculation for upgrade={upgrade_id}. Column used: '{BILLING_KWH_COL}' "
            f"(floored at 0; use 'load_data_kwh' for the raw signal). "
            f"One row per building per hour. Batch: {_CAIRO_BATCH}.",
        ),
        (
            "Utility assignment",
            s3_util,
            "Maps bldg_id to electric utility and provides ResStock sample weights. "
            "Filtered to sb.electric_utility='rie'. Weights rescaled so their sum equals "
            f"the test-year customer count ({_RIE17_TEST_YEAR_CUSTOMER_COUNT:,.2f}).",
        ),
        (
            "ResStock metadata (upgrade=0)",
            s3_meta,
            f"Building characteristics including '{SUBCLASS_COL}' (heating-type subclass). "
            "Joined to utility assignment to assign each building its subclass.",
        ),
        (
            "kWh scale factor",
            f"{_RIE17_KWH_SCALE_FACTOR:.10f}",
            "Applied to all subclass kWh totals after weight rescaling so that the "
            f"aggregate matches the rate-case test-year residential kWh "
            f"({_RIE17_TEST_YEAR_KWH:,.0f} kWh). "
            "Source: rie_rate_case_test_year.yaml, Docket 25-45-GE PRB-1-ELEC exhibit.",
        ),
    ]
    for name, path, desc in sources:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=path)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=desc)
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"C{row}:H{row}")
        ws.row_dimensions[row].height = 55
        row += 1
    row += 1

    ws[f"A{row}"] = "Subclasses"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    for col, label in [
        (1, "Subclass key"),
        (2, "Display name"),
        (3, "Sheet"),
        (4, "Description"),
    ]:
        ws.cell(row=row, column=col, value=label).font = Font(bold=True)
    ws.merge_cells(f"D{row}:H{row}")
    row += 1

    for sheet_num, sc in enumerate(agg_by_subclass.keys(), start=2):
        ws.cell(row=row, column=1, value=sc)
        ws.cell(row=row, column=2, value=SUBCLASS_DISPLAY.get(sc, sc))
        ws.cell(row=row, column=3, value=_sheet_name(sheet_num, sc))
        ws.cell(row=row, column=4, value=SUBCLASS_DESCRIPTION.get(sc, ""))
        ws.merge_cells(f"D{row}:H{row}")
        row += 1
    row += 1

    ws[f"A{row}"] = "Column Descriptions"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    for col, label in [(1, "Column"), (2, "Type"), (3, "Unit"), (4, "Description")]:
        ws.cell(row=row, column=col, value=label).font = Font(bold=True)
    ws.merge_cells(f"D{row}:H{row}")
    row += 1

    for col_name, col_type, unit, desc in [
        ("hour_of_year", "Integer", "-", "Hour index 1-8,760 (chronological within AMY 2018)."),
        (
            "timestamp",
            "String",
            "-",
            "Datetime string YYYY-MM-DD HH:MM:SS. Weather year: AMY 2018.",
        ),
        (
            "weighted_kwh",
            "Float",
            "kWh",
            f"Weighted-sum aggregate electricity consumption: sum(building_{BILLING_KWH_COL} x "
            f"rescaled_weight) x {_RIE17_KWH_SCALE_FACTOR:.6f}. "
            "Represents the estimated aggregate load for the subclass as used in CAIRO's "
            "economic burden and EPMC cost allocation calculation.",
        ),
    ]:
        ws.cell(row=row, column=1, value=col_name)
        ws.cell(row=row, column=2, value=col_type)
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=desc)
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"D{row}:H{row}")
        ws.row_dimensions[row].height = 40
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 45
    for letter in ["E", "F", "G", "H"]:
        ws.column_dimensions[letter].width = 10


def build_workbook_cairo(upgrade_id: int, output_path: Path) -> Path:
    """Load CAIRO billing kWh for the given upgrade, build the xlsx, and write to disk."""
    title = CAIRO_UPGRADE_TITLE[upgrade_id]
    print(f"\nBuilding CAIRO workbook: {title}", flush=True)
    print(f"Output: {output_path}", flush=True)

    agg_by_subclass, meta = load_cairo_aggregate_loads(upgrade_id)
    print(f"Aggregated {len(agg_by_subclass)} subclasses", flush=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    _add_readme_sheet_cairo(wb, upgrade_id, agg_by_subclass)
    for sheet_num, (subclass, df) in enumerate(agg_by_subclass.items(), start=2):
        print(f"  Writing sheet: {_sheet_name(sheet_num, subclass)}", flush=True)
        _add_subclass_sheet(wb, sheet_num, subclass, df)
    _add_validation_sheet_cairo(wb, upgrade_id, meta, agg_by_subclass)

    wb.save(str(output_path))
    size_kb = output_path.stat().st_size / 1024
    print(f"Wrote {output_path} ({size_kb:.1f} KB)", flush=True)
    return output_path


# ── Validation helpers ─────────────────────────────────────────────────────────


def _load_testimony_expected_gwh() -> dict[str, float]:
    """Load per-subclass annual electricity (GWh) from the testimony pickle.

    Reads ``cache/report_variables_cos_subclass.pkl`` (produced by
    cost_of_service_by_subclass.qmd) and returns a dict mapping heating_type_v2
    subclass keys to the testimony consumption_gwh values for upgrade=0.
    Returns an empty dict if the pickle is missing.
    """
    cache_path = REPORT_DIR / "cache" / "report_variables_cos_subclass.pkl"
    if not cache_path.exists():
        print(f"Warning: testimony cache not found at {cache_path}", flush=True)
        return {}
    with cache_path.open("rb") as fh:
        rv: dict[str, Any] = pickle.load(fh)
    rows: list[dict[str, Any]] = rv.get("testimony_subclass_delivery_rows", [])
    result: dict[str, float] = {}
    for row in rows:
        key = _TESTIMONY_SUBCLASS_DISPLAY_TO_KEY.get(str(row.get("subclass", "")))
        if key and "consumption_gwh" in row:
            result[key] = float(row["consumption_gwh"])
    return result


def _add_validation_sheet(
    wb: Workbook,
    upgrade_id: int,
    meta: pl.DataFrame,
    agg_by_subclass: dict[str, pl.DataFrame],
) -> None:
    """Add a Validation sheet comparing aggregate totals against reference values.

    **Upgrade=0** — Cross-checks our raw weighted hourly-load sums (scaled by
    combined_scale) against the testimony's per-subclass annual electricity.
    The testimony value comes from a completely independent code path:
    ``cost_of_service_by_subclass.qmd`` loads CAIRO's master BAT output and
    *derives* per-building annual kWh by inverting the calibrated tariff
    (``annual_kwh = (annual_bill_delivery - fixed_charges) / volumetric_rate``).
    It then weights and sums by subclass.  An assertion in that notebook
    guarantees the weighted total matches the YAML
    ``test_year_residential_kwh`` (2,821,237,490 kWh) to within 1e-6.  Because
    the testimony arrives at annual kWh from *bills* and we arrive from *hourly
    loads*, a match validates our aggregation pipeline end-to-end.

    **Upgrade=2** — No independent reference for total electricity exists in the
    testimony or analysis notebooks (they only state per-building multipliers and
    bill-change metrics, not aggregate totals).  The validation sheet for
    upgrade=2 therefore contains only a building-count check.
    """
    ws = wb.create_sheet("Validation")

    # ── Title ─────────────────────────────────────────────────────────────────
    ws["A1"] = f"Validation — {UPGRADE_TITLE[upgrade_id]}"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:H1")

    if upgrade_id == 0:
        ws["A2"] = (
            "Each check below compares this workbook's aggregate totals against a reference value. "
            "The reference is independently computed: cost_of_service_by_subclass.qmd derives "
            "per-building annual kWh from CAIRO delivery bills (not from load files), then weights "
            "and sums by subclass. PASS = within tolerance. FAIL = investigate before use."
        )
    else:
        ws["A2"] = (
            "No independent reference for total electricity consumption exists for upgrade=2 "
            "(the testimony and analysis notebooks only state per-building multipliers and "
            "bill-change metrics, not aggregate totals). This sheet verifies building counts only."
        )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 42

    # ── Legend ────────────────────────────────────────────────────────────────
    row = 4
    for lbl, fill, font, meaning in [
        ("PASS", PASS_FILL, PASS_FONT, "Within tolerance."),
        ("FAIL", FAIL_FILL, FAIL_FONT, "Outside tolerance — investigate before use."),
        ("INFO", INFO_FILL, INFO_FONT, "Reference value shown for context; no pass/fail threshold."),
        ("N/A", None, Font(bold=True), "Check not applicable for this upgrade."),
    ]:
        cell = ws.cell(row=row, column=1, value=lbl)
        cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=meaning)
        ws.merge_cells(f"B{row}:H{row}")
        row += 1
    row += 1

    # Helper to write a PASS/FAIL row
    def _check_row(
        row: int,
        check: str,
        our_val: str,
        ref_val: str,
        diff_str: str,
        tolerance: str,
        status: str,
        note: str = "",
    ) -> int:
        ws.cell(row=row, column=1, value=check)
        ws.cell(row=row, column=2, value=our_val)
        ws.cell(row=row, column=3, value=ref_val)
        ws.cell(row=row, column=4, value=diff_str)
        ws.cell(row=row, column=5, value=tolerance)
        st_cell = ws.cell(row=row, column=6, value=status)
        if status == "PASS":
            st_cell.fill = PASS_FILL
            st_cell.font = PASS_FONT
        elif status == "FAIL":
            st_cell.fill = FAIL_FILL
            st_cell.font = FAIL_FONT
        elif status == "INFO":
            st_cell.fill = INFO_FILL
            st_cell.font = INFO_FONT
        st_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=note)
        ws.merge_cells(f"G{row}:H{row}")
        ws.row_dimensions[row].height = 30
        return row + 1

    def _section(row: int, title: str, note: str = "") -> int:
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(f"A{row}:H{row}")
        row += 1
        if note:
            ws.cell(row=row, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(f"A{row}:H{row}")
            ws.row_dimensions[row].height = max(55, 15 * (note.count("\n") + 1))
            row += 1
        for col, label in [
            (1, "Check"),
            (2, "This workbook"),
            (3, "Reference value"),
            (4, "Difference"),
            (5, "Tolerance"),
            (6, "Status"),
            (7, "Notes"),
        ]:
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = Font(bold=True)
        ws.merge_cells(f"G{row}:H{row}")
        return row + 1

    # ── Check 1: Building counts ──────────────────────────────────────────────
    row = _section(row, "Check 1: Building Counts")

    total_bldgs = meta.height
    expected_total = 1910
    status = "PASS" if total_bldgs == expected_total else "FAIL"
    row = _check_row(
        row,
        "Total RIE buildings",
        str(total_bldgs),
        str(expected_total),
        str(total_bldgs - expected_total),
        "Exact",
        status,
        "utility_assignment.parquet filtered to sb.electric_utility='rie'.",
    )

    for sc in _ordered_subclasses(list(agg_by_subclass.keys())):
        n = meta.filter(pl.col(SUBCLASS_COL) == sc).height
        row = _check_row(
            row,
            f"  {SUBCLASS_DISPLAY.get(sc, sc)} buildings",
            str(n),
            "—",
            "—",
            "—",
            "INFO",
        )
    row += 1

    # ── Check 2 (upgrade=0): Annual electricity vs testimony ──────────────────
    elec_col = "out.electricity.total.energy_consumption"

    if upgrade_id == 0:
        raw_weight_total = float(meta["weight"].sum())
        weight_scale = _RIE17_TEST_YEAR_CUSTOMER_COUNT / raw_weight_total
        combined_scale = weight_scale * _RIE17_KWH_SCALE_FACTOR

        row = _section(
            row,
            "Check 2: Scaled Annual Electricity vs. Testimony",
            note=(
                "This workbook sums hourly out.electricity.total.energy_consumption from the "
                "ResStock load files and applies sample weights. The testimony's reference values "
                "come from an independent code path: cost_of_service_by_subclass.qmd derives "
                "per-building annual kWh from CAIRO delivery bills "
                "(annual_kwh = (bill - fixed_charges) / vol_rate), then weights and sums by "
                "subclass. That notebook asserts the weighted total = test_year_residential_kwh "
                "(2,821,237,490 kWh from YAML) within 1e-6. Because the two paths arrive at "
                "annual kWh from different inputs (hourly loads vs. bills), agreement validates "
                "the aggregation pipeline.\n"
                f"Scale factors applied to our raw sums: weight_scale = "
                f"{_RIE17_TEST_YEAR_CUSTOMER_COUNT:,.2f} / {raw_weight_total:,.2f} = "
                f"{weight_scale:.6f}; kWh_scale = {_RIE17_KWH_SCALE_FACTOR:.8f} "
                f"(from rie_hp_vs_nonhp_rate_case_test_year.yaml); "
                f"combined = {combined_scale:.6f}."
            ),
        )

        testimony = _load_testimony_expected_gwh()

        our_total_raw_kwh = 0.0
        testimony_total_gwh = 0.0
        for sc in _ordered_subclasses(list(agg_by_subclass.keys())):
            df = agg_by_subclass[sc]
            if elec_col not in df.columns:
                continue
            raw_kwh = float(df[elec_col].sum())
            scaled_gwh = raw_kwh * combined_scale / 1e6
            expected_gwh = testimony.get(sc)
            our_total_raw_kwh += raw_kwh
            if expected_gwh is not None:
                testimony_total_gwh += expected_gwh
            if expected_gwh is None:
                status = "N/A"
                diff_str = "—"
                pct_str = "—"
            else:
                pct_diff = (scaled_gwh - expected_gwh) / expected_gwh * 100
                diff_str = f"{pct_diff:+.3f}%"
                status = "PASS" if abs(pct_diff) <= _TESTIMONY_GWH_TOLERANCE_PCT else "FAIL"
                pct_str = f"±{_TESTIMONY_GWH_TOLERANCE_PCT}%"
            row = _check_row(
                row,
                f"  {SUBCLASS_DISPLAY.get(sc, sc)} — annual elec. (GWh)",
                f"{scaled_gwh:.4f} GWh",
                f"{expected_gwh:.4f} GWh" if expected_gwh is not None else "—",
                diff_str,
                pct_str,
                status,
                f"Raw: {raw_kwh / 1e6:.4f} GWh x {combined_scale:.6f} = {scaled_gwh:.4f} GWh",
            )

        total_scaled = our_total_raw_kwh * combined_scale / 1e6
        if testimony_total_gwh > 0:
            pct_diff_total = (total_scaled - testimony_total_gwh) / testimony_total_gwh * 100
            total_status = "PASS" if abs(pct_diff_total) <= _TESTIMONY_GWH_TOLERANCE_PCT else "FAIL"
            row = _check_row(
                row,
                "TOTAL — annual electricity (GWh)",
                f"{total_scaled:.4f} GWh",
                f"{testimony_total_gwh:.4f} GWh",
                f"{pct_diff_total:+.3f}%",
                f"±{_TESTIMONY_GWH_TOLERANCE_PCT}%",
                total_status,
                "Testimony total = test_year_residential_kwh from YAML (2,821,237,490 kWh).",
            )
        row += 1

    # upgrade=2: building counts above are the only meaningful check

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 38
    ws.column_dimensions["H"].width = 12
    ws.freeze_panes = "A3"


def _add_validation_sheet_cairo(
    wb: Workbook,
    upgrade_id: int,
    meta: pl.DataFrame,
    agg_by_subclass: dict[str, pl.DataFrame],
) -> None:
    """Add a Validation sheet for a CAIRO workbook.

    **Upgrade=0** — Compares this workbook's per-subclass and total annual kWh
    (summed directly from ``billing_kwh_8760.parquet`` after weight rescaling and
    the kWh scale factor) against the per-subclass testimony values from the
    report-variables pickle produced by ``cost_of_service_by_subclass.qmd``.

    The two paths share the same CAIRO run but aggregate differently:
    - This workbook: ``sum(grid_cons_kwh * weight) * kWh_scale`` across all 8760 hours.
    - Testimony notebook: inverts the calibrated tariff per building
      (``annual_kwh = (delivery_bill - fixed_charges) / vol_rate``), then
      weights and sums by subclass.
    Agreement validates that the billing-kWh aggregation pipeline is consistent
    with the testimony cost-of-service figures.

    **Upgrade=2** — No independent reference for total electricity exists for the
    heat-pump upgrade; the validation sheet contains only a building-count check.
    """
    ws = wb.create_sheet("Validation")

    # ── Title ─────────────────────────────────────────────────────────────────
    ws["A1"] = f"Validation - {CAIRO_UPGRADE_TITLE[upgrade_id]}"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:H1")

    if upgrade_id == 0:
        ws["A2"] = (
            "Each check below compares this workbook's aggregate totals against a reference value. "
            "This workbook sums grid_cons_kwh from billing_kwh_8760.parquet weighted by rescaled "
            "sample weights (no additional kWh scale factor - CAIRO's calibration run already "
            "normalized grid_cons_kwh so no further adjustment is needed). "
            "The reference is from cost_of_service_by_subclass.qmd, which inverts the calibrated "
            "tariff to recover per-building annual kWh from delivery bills, then weights and sums "
            "by subclass. Both paths use the same CAIRO run but aggregate kWh differently. "
            "PASS = within tolerance. FAIL = investigate before use."
        )
    else:
        ws["A2"] = (
            "No independent reference for total electricity consumption exists for upgrade=2 "
            "(the testimony and analysis notebooks only state per-building multipliers and "
            "bill-change metrics, not aggregate totals). This sheet verifies building counts only."
        )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 56

    # ── Legend ────────────────────────────────────────────────────────────────
    row = 4
    for lbl, fill, font, meaning in [
        ("PASS", PASS_FILL, PASS_FONT, "Within tolerance."),
        ("FAIL", FAIL_FILL, FAIL_FONT, "Outside tolerance - investigate before use."),
        ("INFO", INFO_FILL, INFO_FONT, "Reference value shown for context; no pass/fail threshold."),
        ("N/A", None, Font(bold=True), "Check not applicable for this upgrade."),
    ]:
        cell = ws.cell(row=row, column=1, value=lbl)
        cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=meaning)
        ws.merge_cells(f"B{row}:H{row}")
        row += 1
    row += 1

    def _check_row(
        row: int,
        check: str,
        our_val: str,
        ref_val: str,
        diff_str: str,
        tolerance: str,
        status: str,
        note: str = "",
    ) -> int:
        ws.cell(row=row, column=1, value=check)
        ws.cell(row=row, column=2, value=our_val)
        ws.cell(row=row, column=3, value=ref_val)
        ws.cell(row=row, column=4, value=diff_str)
        ws.cell(row=row, column=5, value=tolerance)
        st_cell = ws.cell(row=row, column=6, value=status)
        if status == "PASS":
            st_cell.fill = PASS_FILL
            st_cell.font = PASS_FONT
        elif status == "FAIL":
            st_cell.fill = FAIL_FILL
            st_cell.font = FAIL_FONT
        elif status == "INFO":
            st_cell.fill = INFO_FILL
            st_cell.font = INFO_FONT
        st_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=note)
        ws.merge_cells(f"G{row}:H{row}")
        ws.row_dimensions[row].height = 30
        return row + 1

    def _section(row: int, title: str, note: str = "") -> int:
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(f"A{row}:H{row}")
        row += 1
        if note:
            ws.cell(row=row, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(f"A{row}:H{row}")
            ws.row_dimensions[row].height = max(55, 15 * (note.count("\n") + 1))
            row += 1
        for col, label in [
            (1, "Check"),
            (2, "This workbook"),
            (3, "Reference value"),
            (4, "Difference"),
            (5, "Tolerance"),
            (6, "Status"),
            (7, "Notes"),
        ]:
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = Font(bold=True)
        ws.merge_cells(f"G{row}:H{row}")
        return row + 1

    # ── Check 1: Building counts ───────────────────────────────────────────────
    row = _section(row, "Check 1: Building Counts")

    total_bldgs = meta.height
    expected_total = 1910
    status = "PASS" if total_bldgs == expected_total else "FAIL"
    row = _check_row(
        row,
        "Total RIE buildings",
        str(total_bldgs),
        str(expected_total),
        str(total_bldgs - expected_total),
        "Exact",
        status,
        "utility_assignment.parquet filtered to sb.electric_utility='rie'.",
    )

    for sc in _ordered_subclasses(list(agg_by_subclass.keys())):
        n = meta.filter(pl.col(SUBCLASS_COL) == sc).height
        row = _check_row(
            row,
            f"  {SUBCLASS_DISPLAY.get(sc, sc)} buildings",
            str(n),
            "-",
            "-",
            "-",
            "INFO",
        )
    row += 1

    # ── Check 2 (upgrade=0): Annual electricity vs testimony ──────────────────
    if upgrade_id == 0:
        row = _section(
            row,
            "Check 2: Annual Electricity vs. Testimony",
            note=(
                "This workbook's kWh = sum(grid_cons_kwh * rescaled_weight) across all 8760 hours "
                "(weight_scale only; no additional kWh scale factor because CAIRO's calibration run "
                "already normalized grid_cons_kwh). The testimony reference comes from "
                "cost_of_service_by_subclass.qmd which inverts the calibrated tariff per building "
                "(annual_kwh = (delivery_bill - fixed_charges) / vol_rate) then weights and sums "
                "by subclass. That notebook asserts the weighted total = test_year_residential_kwh "
                "(2,821,237,490 kWh from YAML) within 1e-6. Agreement here validates that "
                "billing_kwh_8760.parquet aggregation is consistent with the testimony figures."
            ),
        )

        testimony = _load_testimony_expected_gwh()

        our_total_gwh = 0.0
        testimony_total_gwh = 0.0
        for sc in _ordered_subclasses(list(agg_by_subclass.keys())):
            df = agg_by_subclass[sc]
            our_gwh = float(df["weighted_kwh"].sum()) / 1e6
            our_total_gwh += our_gwh
            expected_gwh = testimony.get(sc)
            if expected_gwh is not None:
                testimony_total_gwh += expected_gwh
            if expected_gwh is None:
                status = "N/A"
                diff_str = "-"
                pct_str = "-"
            else:
                pct_diff = (our_gwh - expected_gwh) / expected_gwh * 100
                diff_str = f"{pct_diff:+.3f}%"
                status = "PASS" if abs(pct_diff) <= _TESTIMONY_GWH_TOLERANCE_PCT else "FAIL"
                pct_str = f"+-{_TESTIMONY_GWH_TOLERANCE_PCT}%"
            row = _check_row(
                row,
                f"  {SUBCLASS_DISPLAY.get(sc, sc)} - annual elec. (GWh)",
                f"{our_gwh:.4f} GWh",
                f"{expected_gwh:.4f} GWh" if expected_gwh is not None else "-",
                diff_str,
                pct_str,
                status,
                "Scale factors applied inside load_cairo_aggregate_loads.",
            )

        if testimony_total_gwh > 0:
            pct_diff_total = (our_total_gwh - testimony_total_gwh) / testimony_total_gwh * 100
            total_status = "PASS" if abs(pct_diff_total) <= _TESTIMONY_GWH_TOLERANCE_PCT else "FAIL"
            row = _check_row(
                row,
                "TOTAL - annual electricity (GWh)",
                f"{our_total_gwh:.4f} GWh",
                f"{testimony_total_gwh:.4f} GWh",
                f"{pct_diff_total:+.3f}%",
                f"+-{_TESTIMONY_GWH_TOLERANCE_PCT}%",
                total_status,
                "Testimony total = test_year_residential_kwh from YAML (2,821,237,490 kWh).",
            )
        row += 1

    # upgrade=2: building counts above are the only meaningful check

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 38
    ws.column_dimensions["H"].width = 12
    ws.freeze_panes = "A3"


# ── Orchestration ──────────────────────────────────────────────────────────────


def build_workbook(upgrade_id: int, output_path: Path) -> Path:
    """Load all data, build the xlsx workbook, and write it to disk."""
    print(f"\nBuilding workbook: {UPGRADE_TITLE[upgrade_id]}", flush=True)
    print(f"Output: {output_path}", flush=True)

    meta = load_metadata_for_upgrade(upgrade_id)
    print(f"Loaded {meta.height} RIE buildings for upgrade={upgrade_id}", flush=True)
    print("Subclass counts:")
    for row in meta.group_by(SUBCLASS_COL).len().sort("len", descending=True).iter_rows():
        print(f"  {row[0]}: {row[1]}", flush=True)

    agg_by_subclass, agg_cols = load_all_subclass_aggregates(upgrade_id, meta)
    print(
        f"Aggregated {len(agg_by_subclass)} subclasses x {len(agg_cols)} energy columns",
        flush=True,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # remove default Sheet1

    _add_readme_sheet(wb, upgrade_id, meta, agg_by_subclass, agg_cols)
    for sheet_num, (subclass, df) in enumerate(agg_by_subclass.items(), start=2):
        print(f"  Writing sheet: {_sheet_name(sheet_num, subclass)}", flush=True)
        _add_subclass_sheet(wb, sheet_num, subclass, df)
    print("  Writing sheet: Validation", flush=True)
    _add_validation_sheet(wb, upgrade_id, meta, agg_by_subclass)

    wb.save(str(output_path))
    size_kb = output_path.stat().st_size / 1024
    print(f"Wrote {output_path} ({size_kb:.1f} KB)", flush=True)
    return output_path


# ── Google Drive / Sheets upload ───────────────────────────────────────────────


def _get_gspread_session() -> Any:
    """Return the authenticated requests Session from gspread (shared across helpers)."""
    from lib.data.gsheets import get_gspread_client

    gc, _ = get_gspread_client()
    return gc.http_client.session


def delete_drive_folder_if_exists(parent_folder_id: str, folder_name: str) -> None:
    """Trash all Drive folders named ``folder_name`` under ``parent_folder_id``.

    If no matching folder exists, does nothing.  If multiple copies exist (e.g. from a
    previous partial run), all are trashed.  Uses Drive v3 PATCH with ``trashed=true``
    (permanent deletion is not supported on Shared Drives without Organizer role).
    """
    session = _get_gspread_session()
    drive_url = "https://www.googleapis.com/drive/v3/files"
    safe_name = folder_name.replace("'", "\\'")
    query = (
        f"name = '{safe_name}' and '{parent_folder_id}' in parents "
        "and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    )
    resp = session.get(
        drive_url,
        params={
            "q": query,
            "fields": "files(id, name)",
            "supportsAllDrives": "true",
            "includeItemsFromAllDrives": "true",
        },
    )
    if not resp.ok:
        return
    files = resp.json().get("files", [])
    for f in files:
        fid = f["id"]
        print(f"Trashing existing folder '{folder_name}' ({fid}) ...", flush=True)
        patch_resp = session.patch(
            f"{drive_url}/{fid}",
            json={"trashed": True},
            params={"supportsAllDrives": "true"},
        )
        patch_resp.raise_for_status()
        print(f"Trashed '{folder_name}' ({fid}).", flush=True)


def get_or_create_drive_subfolder(parent_folder_id: str, subfolder_name: str) -> str:
    """Return the Drive folder ID of a named subfolder, creating it if needed."""
    session = _get_gspread_session()
    drive_url = "https://www.googleapis.com/drive/v3/files"
    safe_name = subfolder_name.replace("'", "\\'")
    query = (
        f"name = '{safe_name}' and '{parent_folder_id}' in parents "
        "and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    )
    resp = session.get(
        drive_url,
        params={
            "q": query,
            "fields": "files(id, name)",
            "supportsAllDrives": "true",
            "includeItemsFromAllDrives": "true",
        },
    )
    if resp.ok:
        files = resp.json().get("files", [])
        if files:
            folder_id: str = files[0]["id"]
            print(
                f"Found existing subfolder '{subfolder_name}': {folder_id}",
                flush=True,
            )
            return folder_id

    print(f"Creating subfolder '{subfolder_name}' in {parent_folder_id} ...", flush=True)
    resp = session.post(
        drive_url,
        json={
            "name": subfolder_name,
            "mimeType": "application/vnd.google-apps.folder",
            "parents": [parent_folder_id],
        },
        params={"supportsAllDrives": "true"},
    )
    resp.raise_for_status()
    folder_id = resp.json()["id"]
    print(f"Created subfolder '{subfolder_name}': {folder_id}", flush=True)
    return folder_id


_TAB_FORMATTING: dict[str, dict[str, Any]] = {
    "README": {
        "wrap_columns": ["A:D"],
        "column_widths_px": {"A": 220, "B": 400, "C": 120, "D": 330},
        "freeze_rows": 0,
        "bold_header": False,
    },
    "Validation": {
        "wrap_columns": ["A:C", "G:H"],
        "column_widths_px": {"A": 350, "B": 230, "C": 230, "D": 105, "E": 90, "F": 60, "G": 280},
        "freeze_rows": 2,
        "bold_header": False,
    },
}


def upload_to_folder(xlsx_path: Path, folder_id: str, title: str) -> None:
    """Create (or replace) a Google Sheet in ``folder_id`` from ``xlsx_path``."""
    from lib.data.gsheets import (
        apply_sheet_formatting,
        create_sheet_in_folder,
        xlsx_to_gsheet,
    )

    print(f"Uploading '{title}' to Drive folder {folder_id} ...", flush=True)
    spreadsheet = create_sheet_in_folder(title, folder_id)
    xlsx_to_gsheet(xlsx_path, spreadsheet.id, delete_other_tabs=True)

    print("Applying formatting ...", flush=True)
    for ws in spreadsheet.worksheets():
        spec = _TAB_FORMATTING.get(ws.title)
        if spec:
            apply_sheet_formatting(ws, **spec)
        else:
            # Data sheets: freeze header row, bold it, auto-resize first 6 columns
            apply_sheet_formatting(
                ws,
                freeze_rows=1,
                bold_header=True,
                auto_resize_columns=["A:F"],
            )

    print(
        f"Done. View at https://docs.google.com/spreadsheets/d/{spreadsheet.id}/edit",
        flush=True,
    )


# ── Entry point ────────────────────────────────────────────────────────────────


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__.splitlines()[0] if __doc__ else "",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("cache/aggregate_loads"),
        help="Directory for output .xlsx files (default: cache/aggregate_loads).",
    )
    parser.add_argument(
        "--folder-id",
        default=DEFAULT_PARENT_FOLDER_ID,
        help=f"Google Drive parent folder ID (default: {DEFAULT_PARENT_FOLDER_ID}).",
    )
    parser.add_argument(
        "--no-upload",
        action="store_true",
        help="Build .xlsx files locally without uploading to Google Drive.",
    )
    parser.add_argument(
        "--upgrades",
        type=int,
        nargs="+",
        default=[0, 2],
        choices=list(UPGRADE_TITLE.keys()),
        help="Which upgrades to build for the raw ResStock source (default: 0 2).",
    )
    parser.add_argument(
        "--source",
        choices=["raw", "cairo", "both"],
        default="both",
        help=(
            "Which load curve source to build: 'raw' (Raw ResStock load curve), "
            "'cairo' (CAIRO calculation load curve, upgrade=0 only), or 'both' (default)."
        ),
    )
    args = parser.parse_args(argv)

    build_raw = args.source in ("raw", "both")
    build_cairo = args.source in ("cairo", "both")

    # All workbooks go under: parent folder → aggregate_loads/ → <source subfolder>/
    # Delete any existing aggregate_loads folder first so the upload is a clean slate,
    # then recreate it (only needed for upload).
    aggregate_loads_id: str | None = None
    if not args.no_upload:
        delete_drive_folder_if_exists(args.folder_id, "aggregate_loads")
        aggregate_loads_id = get_or_create_drive_subfolder(args.folder_id, "aggregate_loads")

    # ── Raw ResStock workbooks ─────────────────────────────────────────────────
    if build_raw:
        raw_subfolder_id: str | None = None
        if not args.no_upload:
            assert aggregate_loads_id is not None
            raw_subfolder_id = get_or_create_drive_subfolder(aggregate_loads_id, SUBFOLDER_NAME_RAW)

        for upgrade_id in args.upgrades:
            out = args.output_dir / "raw" / UPGRADE_FILE[upgrade_id]
            build_workbook(upgrade_id, out)
            if not args.no_upload:
                assert raw_subfolder_id is not None
                upload_to_folder(out, raw_subfolder_id, UPGRADE_TITLE[upgrade_id])

    # ── CAIRO calculation workbooks ────────────────────────────────────────────
    if build_cairo:
        cairo_subfolder_id: str | None = None
        if not args.no_upload:
            assert aggregate_loads_id is not None
            cairo_subfolder_id = get_or_create_drive_subfolder(aggregate_loads_id, SUBFOLDER_NAME_CAIRO)

        for upgrade_id in args.upgrades:
            out_cairo = args.output_dir / "cairo" / CAIRO_UPGRADE_FILE[upgrade_id]
            build_workbook_cairo(upgrade_id, out_cairo)
            if not args.no_upload:
                assert cairo_subfolder_id is not None
                upload_to_folder(out_cairo, cairo_subfolder_id, CAIRO_UPGRADE_TITLE[upgrade_id])

    return 0


if __name__ == "__main__":
    sys.exit(main())
