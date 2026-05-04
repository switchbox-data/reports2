"""Build the supporting workbook for the revenue-neutrality claim (page 56).

Pre-Filed Direct Testimony of Juan-Pablo Velez, Page 56, Lines 8-14:

    Q. Is the heat pump rate revenue-neutral?
    A. Yes, at the residential class level.

This script produces an ``.xlsx`` proving that the HP and non-HP subclass
revenue requirements sum to the total residential delivery revenue
requirement.

Run from the report directory::

    uv run python -m testimony_response.build_RIE_1_12_workbook \\
        --output cache/revenue_neutrality.xlsx
    uv run python -m testimony_response.build_RIE_1_12_workbook --upload

See ``cost_of_service_by_subclass.qmd`` for the published-side analysis
logic that this workbook reproduces with formulas.
"""

from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from lib.rdp import fetch_rdp_file

# ---------------------------------------------------------------------------
# Constants.
# ---------------------------------------------------------------------------
RDP_REF = "e9e5088"
RDP_REV_YAML_PATH = "rate_design/hp_rates/ri/config/rev_requirement/rie_rate_case_test_year.yaml"
RDP_HPVS_YAML_PATH = "rate_design/hp_rates/ri/config/rev_requirement/rie_hp_vs_nonhp_rate_case_test_year.yaml"
RDP_GITHUB_BASE = "https://github.com/switchbox-data/rate-design-platform/blob"
REPORTS2_GITHUB_BASE = "https://github.com/switchbox-data/reports2/blob"

DEFAULT_SPREADSHEET_ID = "1JlSDvgS6H70OCIJ4Q8LRQGNFS6AJcaeh7ccNxaab08A"

DELIVERY_METHOD = "epmc"


# ---------------------------------------------------------------------------
# Permalink helpers.
# ---------------------------------------------------------------------------
def _rdp_permalink(rel_path: str) -> str:
    return f"{RDP_GITHUB_BASE}/{RDP_REF}/{rel_path}"


def _reports2_head_sha() -> str:
    if not hasattr(_reports2_head_sha, "_cached"):
        repo_root = Path(__file__).resolve().parents[3]
        sha = subprocess.check_output(
            ["git", "-C", str(repo_root), "rev-parse", "HEAD"],
            text=True,
        ).strip()
        _reports2_head_sha._cached = sha  # type: ignore[attr-defined]
    return _reports2_head_sha._cached  # type: ignore[attr-defined]


def _reports2_permalink(rel_path: str) -> str:
    return f"{REPORTS2_GITHUB_BASE}/{_reports2_head_sha()}/{rel_path}"


# ---------------------------------------------------------------------------
# Data loading (YAMLs only — no BAT data, no tariff JSONs).
# ---------------------------------------------------------------------------
def load_inputs() -> dict:
    """Pull total delivery RR and subclass RR splits from the two YAMLs."""
    raw_rev = fetch_rdp_file(RDP_REV_YAML_PATH, RDP_REF)
    rev = yaml.safe_load(raw_rev)

    raw_hpvs = fetch_rdp_file(RDP_HPVS_YAML_PATH, RDP_REF)
    hpvs = yaml.safe_load(raw_hpvs)

    return {
        "total_delivery_revenue_requirement": float(rev["total_delivery_revenue_requirement"]),
        "subclass_revenue_requirements": hpvs["subclass_revenue_requirements"],
    }


# ---------------------------------------------------------------------------
# Formatting helpers.
# ---------------------------------------------------------------------------
def _header_fill(ws, row: int, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="E8E8E8")
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = fill


def _autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Sheet writers.
# ---------------------------------------------------------------------------
def _write_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README", 0)

    ws["A1"] = "Revenue-neutrality supporting workbook (RIE residential HP rate, Docket 2545GE)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    ws["A3"] = "Claim"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = (
        'Q. "Is the heat pump rate revenue-neutral?" '
        'A. "Yes, at the residential class level." '
        "The HP and non-HP subclass delivery revenue requirements sum to the "
        "total residential delivery revenue requirement."
    )
    ws["B3"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 45

    ws["A5"] = "Sources"
    ws["A5"].font = Font(bold=True)

    ws["A6"] = "Revenue-requirement YAML"
    ws["B6"] = _rdp_permalink(RDP_REV_YAML_PATH)

    ws["A7"] = "Subclass RR YAML (HP vs non-HP)"
    ws["B7"] = _rdp_permalink(RDP_HPVS_YAML_PATH)

    ws["A8"] = "Analysis notebook"
    ws["B8"] = _reports2_permalink("reports/ri_hp_rates/notebooks/cost_of_service_by_subclass.qmd")

    ws["A9"] = "This workbook's build script"
    ws["B9"] = _reports2_permalink("reports/ri_hp_rates/testimony_response/build_RIE_1_12_workbook.py")

    _autosize(ws, {"A": 30, "B": 100})
    ws.sheet_view.showGridLines = False


def _write_subclass_revenue_proof(wb: Workbook, inputs: dict) -> None:
    """Single sheet proving subclass RRs sum to total delivery RR."""
    ws = wb.create_sheet("subclass_revenue_proof")

    total_rr = inputs["total_delivery_revenue_requirement"]
    subs = inputs["subclass_revenue_requirements"]

    ws["A1"] = "Subclass revenue-requirement proof"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    vals = subs["delivery"][DELIVERY_METHOD]

    hdr_row = 3
    for ci, h in enumerate(["", "Revenue requirement ($)", "Notes"], start=1):
        ws.cell(row=hdr_row, column=ci, value=h)
    _header_fill(ws, hdr_row, 3)

    hp_row = 4
    nonhp_row = 5
    sum_row = 6
    total_row = 7
    delta_row = 8

    ws.cell(row=hp_row, column=1, value="HP subclass RR")
    ws.cell(row=hp_row, column=2, value=float(vals["hp"]))
    ws.cell(row=hp_row, column=3, value="EPMC cost-of-service allocation")

    ws.cell(row=nonhp_row, column=1, value="Non-HP subclass RR")
    ws.cell(row=nonhp_row, column=2, value=float(vals["non-hp"]))
    ws.cell(row=nonhp_row, column=3, value="EPMC cost-of-service allocation")

    ws.cell(row=sum_row, column=1, value="Sum of subclass RRs")
    ws.cell(row=sum_row, column=2, value=f"=B{hp_row}+B{nonhp_row}")
    ws.cell(row=sum_row, column=3, value="Should equal total RR")
    ws.cell(row=sum_row, column=1).font = Font(bold=True)
    ws.cell(row=sum_row, column=2).font = Font(bold=True)

    ws.cell(row=total_row, column=1, value="Total delivery RR")
    ws.cell(row=total_row, column=2, value=total_rr)
    ws.cell(row=total_row, column=3, value="From rate-case YAML")

    ws.cell(row=delta_row, column=1, value="Difference")
    ws.cell(row=delta_row, column=2, value=f"=B{sum_row}-B{total_row}")
    ws.cell(row=delta_row, column=3, value="Should be ~$0 (float rounding)")

    for r in range(hp_row, delta_row + 1):
        ws[f"B{r}"].number_format = '"$"#,##0.00'

    _autosize(ws, {"A": 28, "B": 24, "C": 36})
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Orchestration.
# ---------------------------------------------------------------------------
def build_workbook(output_path: Path) -> Path:
    print("Loading revenue-requirement YAMLs ...", flush=True)
    inputs = load_inputs()
    print(
        f"  total_delivery_RR = ${inputs['total_delivery_revenue_requirement']:,.0f}",
        flush=True,
    )

    vals = inputs["subclass_revenue_requirements"]["delivery"][DELIVERY_METHOD]
    hp_rr = float(vals["hp"])
    nonhp_rr = float(vals["non-hp"])
    print(f"  HP subclass RR = ${hp_rr:,.2f}", flush=True)
    print(f"  non-HP subclass RR = ${nonhp_rr:,.2f}", flush=True)
    print(f"  sum = ${hp_rr + nonhp_rr:,.2f}", flush=True)

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    _write_readme(wb)
    _write_subclass_revenue_proof(wb, inputs)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(
        f"Wrote {output_path} ({output_path.stat().st_size / 1024:.1f} KB)",
        flush=True,
    )
    return output_path


# ---------------------------------------------------------------------------
# Google Sheets upload formatting.
# ---------------------------------------------------------------------------
_TAB_FORMATTING: dict[str, dict] = {
    "README": {
        "wrap_columns": ["B:B"],
        "column_widths_px": {"A": 240, "B": 640},
        "freeze_rows": 0,
        "bold_header": False,
    },
    "subclass_revenue_proof": {
        "column_number_formats": {"B": '"$"#,##0.00'},
        "wrap_columns": ["C:C"],
        "column_widths_px": {"A": 220, "B": 180, "C": 280},
        "freeze_rows": 0,
        "bold_header": False,
    },
}


def upload_to_sheet(xlsx_path: Path, spreadsheet_id: str) -> None:
    from lib.data.gsheets import apply_sheet_formatting, xlsx_to_gsheet

    print(
        f"Uploading {xlsx_path} -> Google Sheet {spreadsheet_id} ...",
        flush=True,
    )
    spreadsheet = xlsx_to_gsheet(xlsx_path, spreadsheet_id, delete_other_tabs=True)
    print("Applying formatting ...", flush=True)
    for ws in spreadsheet.worksheets():
        spec = _TAB_FORMATTING.get(ws.title)
        if spec:
            apply_sheet_formatting(ws, **spec)
    print(
        f"Done. View at https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit",
        flush=True,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0] if __doc__ else "")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("cache/revenue_neutrality.xlsx"),
        help="Output .xlsx path. Default: cache/revenue_neutrality.xlsx",
    )
    parser.add_argument(
        "--upload",
        action="store_true",
        help="Upload to Google Sheet after building.",
    )
    parser.add_argument(
        "--spreadsheet-id",
        default=DEFAULT_SPREADSHEET_ID,
        help=f"Override upload target. Default: {DEFAULT_SPREADSHEET_ID}",
    )
    args = parser.parse_args(argv)

    out = build_workbook(args.output)
    if args.upload:
        upload_to_sheet(out, args.spreadsheet_id)
    return 0


if __name__ == "__main__":
    sys.exit(main())
