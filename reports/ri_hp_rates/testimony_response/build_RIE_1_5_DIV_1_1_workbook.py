"""Build the supporting workbook for Figure 2 (RIE residential subclass delivery revenue).

Figure 2 in the Pre-Filed Direct Testimony of Juan-Pablo Velez (page 10,
``tbl-testimony-subclass-delivery`` in ``expert_testimony.qmd``) shows the
breakdown of customers, consumption, and delivery revenue by residential
heating subclass for the test year (9/1/24 to 8/31/25).

This script reproduces every published number from per-building CAIRO outputs
plus a small set of revenue-requirement and tariff inputs, with **live
formulas** in every aggregation cell. The output is an ``.xlsx`` that opens
identically in Excel and Google Sheets; with ``--upload`` the same workbook is
mirrored into a target Google Sheet, preserving formulas via
``value_input_option="USER_ENTERED"``.

Run from the report directory::

    uv run python -m testimony_response.build_RIE_1_5_DIV_1_1_workbook --output cache/fig2_subclass_delivery.xlsx
    uv run python -m testimony_response.build_RIE_1_5_DIV_1_1_workbook --upload

See ``cost_of_service_by_subclass.qmd`` (cell ``tbl-testimony-subclass-delivery``)
for the published-side aggregation logic that this workbook recreates with
formulas instead of polars group-by/sum.
"""

from __future__ import annotations

import argparse
import pickle
import subprocess
import sys
from pathlib import Path
from types import SimpleNamespace

import polars as pl
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName

from lib.rdp import fetch_rdp_file, parse_urdb_json

# Cross-sheet A1 references used inside formulas. We use these everywhere in
# place of named ranges, because Google Sheets does not import workbook-level
# defined names from an .xlsx pushed via the Sheets API: only cell values move,
# so any formula referencing `=ws_weight` would become `#NAME?`. Explicit A1
# references survive the upload unchanged and evaluate identically in Excel
# and Sheets.
REF_TOTAL_RR = "inputs_revenue_requirement!$B$2"
REF_N_CUSTOMERS = "inputs_revenue_requirement!$B$3"
REF_TY_KWH = "inputs_revenue_requirement!$B$4"
REF_CUSTOMER_CHARGE = "inputs_revenue_requirement!$B$5"
REF_CORE_DELIVERY = "inputs_revenue_requirement!$B$6"
REF_ANNUAL_FIXED_PER_CUSTOMER = "inputs_revenue_requirement!$B$7"
REF_DISPLAY_TOTAL = "inputs_revenue_requirement!$B$8"
REF_DEFAULT_VOL = "inputs_tariffs!$B$2"

# Same constants as cost_of_service_by_subclass.qmd; if the testimony rebases
# onto a new batch or RDP ref, update these in lock-step with the notebook.
UTILITY = "rie"
BATCH = "ri_20260331_r1-20_rate_case_test_year"
STATE_LOWER = "ri"
S3_BASE = "s3://data.sb/switchbox/cairo/outputs/hp_rates"
PATH_MASTER_BAT_12 = f"{S3_BASE}/{STATE_LOWER}/all_utilities/{BATCH}/run_1+2/cross_subsidization_BAT_values/"
RDP_REF = "e9e5088"
RDP_REV_YAML_PATH = "rate_design/hp_rates/ri/config/rev_requirement/rie_rate_case_test_year.yaml"
RDP_TARIFF_DIR = "rate_design/hp_rates/ri/config/tariffs/electric"
RDP_GITHUB_BASE = "https://github.com/switchbox-data/rate-design-platform/blob"


def _rdp_permalink(rel_path: str) -> str:
    """SHA-pinned GitHub permalink for a rate-design-platform file."""
    return f"{RDP_GITHUB_BASE}/{RDP_REF}/{rel_path}"


REPORTS2_GITHUB_BASE = "https://github.com/switchbox-data/reports2/blob"


def _reports2_head_sha() -> str:
    """Current HEAD sha of the reports2 repo (this script's repo). Cached."""
    if not hasattr(_reports2_head_sha, "_cached"):
        repo_root = Path(__file__).resolve().parents[3]
        sha = subprocess.check_output(
            ["git", "-C", str(repo_root), "rev-parse", "HEAD"],
            text=True,
        ).strip()
        _reports2_head_sha._cached = sha  # type: ignore[attr-defined]
    return _reports2_head_sha._cached  # type: ignore[attr-defined]


def _reports2_permalink(rel_path: str, *, line_range: tuple[int, int] | None = None) -> str:
    """SHA-pinned GitHub permalink for a file in this reports2 repo."""
    url = f"{REPORTS2_GITHUB_BASE}/{_reports2_head_sha()}/{rel_path}"
    if line_range is not None:
        start, end = line_range
        url += f"#L{start}-L{end}"
    return url


# Default upload target: RIE 1-5 / DIV 1-1 discovery response Sheet (Figure 2).
DEFAULT_SPREADSHEET_ID = "10pe7gh9FWWkGKrru7hr6BWXjS3DGzgGe_gKBX68kZ6A"

HT_V2_ORDER = (
    "heat_pump",
    "electrical_resistance",
    "natgas",
    "delivered_fuels",
    "other",
)
HT_V2_LABELS: dict[str, str] = {
    "heat_pump": "Heat pump",
    "electrical_resistance": "Electric resistance",
    "natgas": "Natural gas",
    "delivered_fuels": "Delivered fuels",
    "other": "Other",
}


def load_master_bat() -> pl.DataFrame:
    """Load per-building BAT data for RIE (delivery bills and weights only)."""
    df = (
        pl.scan_parquet(PATH_MASTER_BAT_12, hive_partitioning=True)
        .filter(pl.col("sb.electric_utility") == UTILITY)
        .select(
            "bldg_id",
            "weight",
            "postprocess_group.heating_type_v2",
            "annual_bill_delivery",
        )
        .collect()
    )
    assert isinstance(df, pl.DataFrame)
    return df


def load_inputs() -> dict:
    """Pull revenue-requirement YAML + calibrated tariff JSON from rate-design-platform."""
    raw_yaml = fetch_rdp_file(RDP_REV_YAML_PATH, RDP_REF)
    rev = yaml.safe_load(raw_yaml)
    total_rr = float(rev["total_delivery_revenue_requirement"])
    n_customers = float(rev["test_year_customer_count"])
    ty_kwh = float(rev["test_year_residential_kwh"])

    drr = rev["delivery_revenue_requirement"]
    customer_charge_total = float(drr["customer_charge"]["total_budget"])
    core_delivery_total = float(drr["core_delivery_rate"]["total_budget"])

    def vol_rate(rel_filename: str) -> float:
        path = f"{RDP_TARIFF_DIR}/{rel_filename}"
        doc = parse_urdb_json(fetch_rdp_file(path, RDP_REF))
        return float(doc["items"][0]["energyratestructure"][0][0]["rate"])

    default_vol = vol_rate("rie_default_calibrated.json")

    annual_fixed_per_customer = (total_rr - default_vol * ty_kwh) / n_customers

    return {
        "total_delivery_revenue_requirement": total_rr,
        "test_year_customer_count": n_customers,
        "test_year_residential_kwh": ty_kwh,
        "customer_charge_total": customer_charge_total,
        "core_delivery_rate_total": core_delivery_total,
        "default_vol_usd_per_kwh": default_vol,
        "annual_fixed_per_customer": annual_fixed_per_customer,
    }


def _bold(ws, cell: str) -> None:
    ws[cell].font = Font(bold=True)


def _header_fill(ws, row: int, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="E8E8E8")
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = fill


def _autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Sheet writers.
# ---------------------------------------------------------------------------
def _write_readme(wb: Workbook, inputs: dict) -> None:
    ws = wb.create_sheet("README", 0)
    s3_bat = f"{S3_BASE}/{STATE_LOWER}/all_utilities/{BATCH}/run_1+2/cross_subsidization_BAT_values/"
    rows: list[list] = [
        ["Figure 2 supporting workbook (RIE residential subclass delivery revenue)", "", ""],
        ["", "", ""],
        ["Item", "Source", "Notes"],
        [
            "Per-building CAIRO outputs",
            s3_bat,
            "One row per residential building under the status-quo uniform default delivery + supply rate for all customers. Used as the per-building basis for all weighted aggregates.",
        ],
        [
            "Revenue-requirement YAML",
            _rdp_permalink(RDP_REV_YAML_PATH),
            "Test-year customer count, total delivery revenue requirement, test-year residential kWh, customer charge total, core delivery rate total.",
        ],
        [
            "Calibrated default tariff JSON",
            _rdp_permalink(f"{RDP_TARIFF_DIR}/rie_default_calibrated.json"),
            "energyratestructure[0][0].rate is the default uniform $/kWh used to back out annual kWh per building from delivery bill.",
        ],
        [
            "Notebook that produces the published table",
            _reports2_permalink("reports/ri_hp_rates/notebooks/cost_of_service_by_subclass.qmd"),
            "Cell tbl-testimony-subclass-delivery. This workbook reproduces its aggregation logic as live formulas.",
        ],
        [
            "Testimony embed",
            _reports2_permalink("reports/ri_hp_rates/expert_testimony.qmd", line_range=(266, 268)),
            "Quarto embed shortcode that pulls tbl-testimony-subclass-delivery into the testimony as Figure 2 (page 10).",
        ],
        ["", "", ""],
        ["Sheet", "What it contains", ""],
        [
            "inputs_revenue_requirement",
            "Test-year customer count, total delivery revenue requirement, test-year residential kWh, customer charge total, core delivery rate total, derived annual fixed delivery $/customer.",
            "",
        ],
        [
            "inputs_tariffs",
            "Calibrated default volumetric delivery $/kWh.",
            "",
        ],
        [
            "bill_per_building",
            (
                "One row per residential building. Columns: bldg_id, weight (calibrated to test_year_customer_count), "
                "heating_type_v2, annual_kwh, annual_bill_delivery (= annual_kwh * vol_rate + annual_fixed_per_customer), "
                "w_revenue (= weight * annual_bill_delivery), w_kwh (= weight * annual_kwh)."
            ),
            "",
        ],
        [
            "subclass_aggregates",
            "Five subclass rows + 'All customers' total via live SUMIFS / SUMPRODUCT over bill_per_building. Includes consumption_gwh, pct_all_consumption, and pct_all_revenue alongside the largest-remainder customer display columns.",
            "",
        ],
        [
            "fig2_published",
            "Final published layout: customers, % of customers, consumption (GWh), % of consumption, delivery revenue, % of revenue.",
            "",
        ],
        ["", "", ""],
        ["fig2_published column", "Formula (descriptive; live values are in fig2_published rows 5-10)", ""],
        [
            "Customers",
            "subclass_aggregates!n_customers_display (largest-remainder rounding to test_year_customer_count).",
            "",
        ],
        ["% of customers", "n_customers_display / DISPLAY_CUSTOMER_TOTAL.", ""],
        [
            "Consumption",
            "SUMIFS(weight x annual_kwh, ...) / 1,000,000 for the subclass (GWh).",
            "",
        ],
        [
            "% of consumption",
            "subclass total_kwh / total kWh across all subclasses.",
            "",
        ],
        [
            "Delivery revenue",
            "SUMIFS(weight x annual_bill_delivery, ...) for the subclass.",
            "",
        ],
        [
            "% of revenue",
            "subclass revenue_delivery / total revenue across all subclasses.",
            "",
        ],
        ["", "", ""],
        ["Key inputs (also live in inputs_revenue_requirement / inputs_tariffs)", "Value", "Source"],
        [
            "total_delivery_revenue_requirement ($)",
            inputs["total_delivery_revenue_requirement"],
            (
                "Rhode Island Energy, Application for Approval of a Change in Electric and Gas "
                "Base Distribution Rates, Docket 25-45-GE, PRB-1-ELEC exhibit, p. 14, lines 8-9, "
                "columns f & m."
            ),
        ],
        [
            "test_year_customer_count",
            inputs["test_year_customer_count"],
            "PRB-1-ELEC exhibit, p. 14, lines 8-9, column d. 5,032,174 bills / 12 = 419,347.83 customers.",
        ],
        [
            "test_year_residential_kwh",
            inputs["test_year_residential_kwh"],
            "PRB-1-ELEC exhibit, p. 14, lines 8-9, column k.",
        ],
        [
            "default_vol_usd_per_kwh",
            inputs["default_vol_usd_per_kwh"],
            _rdp_permalink(f"{RDP_TARIFF_DIR}/rie_default_calibrated.json"),
        ],
        [
            "annual_fixed_per_customer ($)",
            inputs["annual_fixed_per_customer"],
            (
                "Test-year revenue equation: (total_RR - vol_rate * total_kWh) / customers. "
                "Inputs from RIE Docket 25-45-GE, Book 21, PRB-1-ELEC p. 14, lines 8-9 "
                "(Blazunas Schedules, Nov 2025)."
            ),
        ],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    # Section headers: item=3, sheets=10, columns=17, key inputs=25.
    for header_row in (3, 10, 17, 25):
        _header_fill(ws, header_row, 3)
    for label_row in range(26, 31):
        _bold(ws, f"A{label_row}")
    _autosize(ws, {"A": 42, "B": 70, "C": 80})
    ws.sheet_view.showGridLines = False


def _write_inputs_revenue_requirement(wb: Workbook, inputs: dict) -> None:
    ws = wb.create_sheet("inputs_revenue_requirement")
    yaml_ref = _rdp_permalink(RDP_REV_YAML_PATH)
    rows = [
        ["key", "value", "source", "notes"],
        [
            "total_delivery_revenue_requirement",
            inputs["total_delivery_revenue_requirement"],
            yaml_ref,
            "YAML field: total_delivery_revenue_requirement. PRB-1-ELEC exhibit, p. 14, lines 8-9, columns f & m. Total RIE delivery revenue requirement for the test year ($).",
        ],
        [
            "test_year_customer_count",
            inputs["test_year_customer_count"],
            yaml_ref,
            "YAML field: test_year_customer_count. PRB-1-ELEC exhibit, p. 14, lines 8-9, column d. Total RIE residential customers in the test year.",
        ],
        [
            "test_year_residential_kwh",
            inputs["test_year_residential_kwh"],
            yaml_ref,
            "YAML field: test_year_residential_kwh. PRB-1-ELEC exhibit, p. 14, lines 8-9, column k. Total RIE residential delivered kWh in the test year.",
        ],
        [
            "customer_charge_total",
            inputs["customer_charge_total"],
            yaml_ref,
            "YAML field: delivery_revenue_requirement.customer_charge.total_budget. Portion of revenue requirement recovered through fixed customer charges.",
        ],
        [
            "core_delivery_rate_total",
            inputs["core_delivery_rate_total"],
            yaml_ref,
            "YAML field: delivery_revenue_requirement.core_delivery_rate.total_budget. Portion of revenue requirement recovered through volumetric delivery rates.",
        ],
        [
            "annual_fixed_per_customer",
            f"=({REF_TOTAL_RR} - {REF_DEFAULT_VOL} * {REF_TY_KWH}) / {REF_N_CUSTOMERS}",
            "Test-year revenue equation (RIE Docket 25-45-GE, Book 21, PRB-1-ELEC p. 14)",
            "Annual non-volumetric delivery charges per customer. Per-customer share of delivery revenue not recovered through volumetric rates.",
        ],
        [
            "DISPLAY_CUSTOMER_TOTAL",
            f"=ROUND({REF_N_CUSTOMERS}, 0)",
            "Derived in this workbook",
            "Integer total used for largest-remainder customer display rounding.",
        ],
    ]
    for r in rows:
        ws.append(r)
    _header_fill(ws, 1, 4)
    _autosize(ws, {"A": 36, "B": 22, "C": 70, "D": 70})
    for row, name in [
        (2, "total_delivery_revenue_requirement"),
        (3, "test_year_customer_count"),
        (4, "test_year_residential_kwh"),
        (5, "customer_charge_total"),
        (6, "core_delivery_rate_total"),
        (7, "annual_fixed_per_customer"),
        (8, "DISPLAY_CUSTOMER_TOTAL"),
    ]:
        wb.defined_names[name] = DefinedName(
            name=name,
            attr_text=f"inputs_revenue_requirement!$B${row}",
        )
    ws.sheet_view.showGridLines = False


def _write_inputs_tariffs(wb: Workbook, inputs: dict) -> None:
    ws = wb.create_sheet("inputs_tariffs")
    rows = [
        ["key", "value", "source", "notes"],
        [
            "default_vol_usd_per_kwh",
            inputs["default_vol_usd_per_kwh"],
            _rdp_permalink(f"{RDP_TARIFF_DIR}/rie_default_calibrated.json"),
            "Field: energyratestructure[0][0].rate. Status-quo uniform default delivery $/kWh.",
        ],
    ]
    for r in rows:
        ws.append(r)
    _header_fill(ws, 1, 4)
    _autosize(ws, {"A": 32, "B": 18, "C": 80, "D": 70})
    for row, name in [
        (2, "default_vol_usd_per_kwh"),
    ]:
        wb.defined_names[name] = DefinedName(name=name, attr_text=f"inputs_tariffs!$B${row}")
    ws.sheet_view.showGridLines = False


def _write_bill_per_building(wb: Workbook, bat: pl.DataFrame) -> int:
    """Write per-building rows + formula columns. Returns last data row index."""
    ws = wb.create_sheet("bill_per_building")
    headers = [
        "bldg_id",  # A
        "weight",  # B
        "heating_type_v2",  # C
        "annual_kwh",  # D  data
        "annual_bill_delivery",  # E  formula = D * vol_rate + fixed
        "w_revenue",  # F  formula = B * E
        "w_kwh",  # G  formula = B * D
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    n = bat.height
    rows = list(
        bat.select(
            "bldg_id",
            "weight",
            "postprocess_group.heating_type_v2",
            "annual_kwh",
        ).iter_rows()
    )
    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row[0])  # A: bldg_id
        ws.cell(row=i, column=2, value=float(row[1]))  # B: weight
        ws.cell(row=i, column=3, value=row[2])  # C: heating_type_v2
        ws.cell(row=i, column=4, value=float(row[3]))  # D: annual_kwh
        ws.cell(row=i, column=5, value=f"=D{i}*{REF_DEFAULT_VOL}+{REF_ANNUAL_FIXED_PER_CUSTOMER}")
        ws.cell(row=i, column=6, value=f"=B{i}*E{i}")  # F: w_revenue
        ws.cell(row=i, column=7, value=f"=B{i}*D{i}")  # G: w_kwh

    last_row = 1 + n
    _autosize(ws, {"A": 10, "B": 10, "C": 22, "D": 14, "E": 18, "F": 16, "G": 16})

    col_to_name = {
        "B": "ws_weight",
        "C": "ws_heating_type",
        "D": "ws_annual_kwh",
        "E": "ws_annual_bill",
        "F": "ws_w_revenue",
        "G": "ws_w_kwh",
    }
    for col, name in col_to_name.items():
        wb.defined_names[name] = DefinedName(
            name=name,
            attr_text=f"bill_per_building!${col}$2:${col}${last_row}",
        )
    return last_row


def _write_subclass_aggregates(wb: Workbook, last_bat_row: int) -> None:
    ws = wb.create_sheet("subclass_aggregates")
    headers = [
        "subclass_key",  # A
        "subclass",  # B
        "n_customers",  # C
        "revenue_delivery",  # D
        "total_kwh",  # E
        "raw_display_count",  # F
        "floor_display_count",  # G
        "remainder",  # H
        "rank_remainder",  # I
        "n_customers_display",  # J
        "pct_customers_display",  # K
        "consumption_gwh",  # L
        "pct_all_consumption",  # M
        "pct_all_revenue",  # N
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))

    n_sub = len(HT_V2_ORDER)

    last = last_bat_row
    rng_weight = f"bill_per_building!$B$2:$B${last}"
    rng_heating = f"bill_per_building!$C$2:$C${last}"
    rng_w_revenue = f"bill_per_building!$F$2:$F${last}"
    rng_w_kwh = f"bill_per_building!$G$2:$G${last}"
    sub_last = 1 + n_sub
    total_row = 2 + n_sub

    for idx, key in enumerate(HT_V2_ORDER):
        row = idx + 2
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=HT_V2_LABELS[key])
        ws.cell(row=row, column=3, value=f"=SUMIFS({rng_weight}, {rng_heating}, A{row})")
        ws.cell(row=row, column=4, value=f"=SUMIFS({rng_w_revenue}, {rng_heating}, A{row})")
        ws.cell(row=row, column=5, value=f"=SUMIFS({rng_w_kwh}, {rng_heating}, A{row})")
        ws.cell(row=row, column=6, value=f"=C{row}*{REF_DISPLAY_TOTAL}/SUM($C$2:$C${sub_last})")
        ws.cell(row=row, column=7, value=f"=INT(F{row})")
        ws.cell(row=row, column=8, value=f"=F{row}-G{row}")
        ws.cell(row=row, column=9, value=f"=RANK(H{row},$H$2:$H${sub_last})")
        ws.cell(row=row, column=10, value=f"=G{row}+IF(I{row}<={REF_DISPLAY_TOTAL}-SUM($G$2:$G${sub_last}),1,0)")
        ws.cell(row=row, column=11, value=f"=J{row}/{REF_DISPLAY_TOTAL}")
        ws.cell(row=row, column=12, value=f"=E{row}/1000000")
        ws.cell(row=row, column=13, value=f"=E{row}/$E${total_row}")
        ws.cell(row=row, column=14, value=f"=D{row}/$D${total_row}")

    ws.cell(row=total_row, column=1, value="all_customers")
    ws.cell(row=total_row, column=2, value="All customers")
    ws.cell(row=total_row, column=3, value=f"=SUM({rng_weight})")
    ws.cell(row=total_row, column=4, value=f"=SUM({rng_w_revenue})")
    ws.cell(row=total_row, column=5, value=f"=SUM({rng_w_kwh})")
    ws.cell(row=total_row, column=6, value=f"={REF_DISPLAY_TOTAL}")
    ws.cell(row=total_row, column=7, value=f"={REF_DISPLAY_TOTAL}")
    ws.cell(row=total_row, column=8, value=0)
    ws.cell(row=total_row, column=9, value="")
    ws.cell(row=total_row, column=10, value=f"={REF_DISPLAY_TOTAL}")
    ws.cell(row=total_row, column=11, value=1)
    ws.cell(row=total_row, column=12, value=f"=E{total_row}/1000000")
    ws.cell(row=total_row, column=13, value=1)
    ws.cell(row=total_row, column=14, value=1)

    _bold(ws, f"A{total_row}")
    _bold(ws, f"B{total_row}")

    for r in range(2, total_row + 1):
        ws[f"C{r}"].number_format = "#,##0.0"
        ws[f"D{r}"].number_format = '"$"#,##0'
        ws[f"E{r}"].number_format = "#,##0"
        ws[f"F{r}"].number_format = "#,##0.00"
        ws[f"G{r}"].number_format = "#,##0"
        ws[f"H{r}"].number_format = "#,##0.00"
        ws[f"J{r}"].number_format = "#,##0"
        ws[f"K{r}"].number_format = "0.0%"
        ws[f"L{r}"].number_format = "#,##0.0"
        ws[f"M{r}"].number_format = "0.0%"
        ws[f"N{r}"].number_format = "0.0%"

    _autosize(
        ws,
        {
            "A": 22,
            "B": 22,
            "C": 16,
            "D": 18,
            "E": 18,
            "F": 14,
            "G": 14,
            "H": 12,
            "I": 8,
            "J": 14,
            "K": 14,
            "L": 14,
            "M": 16,
            "N": 14,
        },
    )
    ws.freeze_panes = "C2"

    wb.defined_names["agg_first_row"] = DefinedName(name="agg_first_row", attr_text="subclass_aggregates!$A$2")
    wb.defined_names["agg_total_row"] = DefinedName(
        name="agg_total_row",
        attr_text=f"subclass_aggregates!$A${total_row}",
    )


def _write_fig2_published(wb: Workbook) -> None:
    """Final published layout, mirroring the GT in tbl-testimony-subclass-delivery."""
    ws = wb.create_sheet("fig2_published")
    title = "Customers, consumption, and delivery revenue by residential subclass"
    subtitle = (
        "RIE Test Year (9/1/2024 to 8/31/2025). Source: cost_of_service_by_subclass.qmd "
        "(tbl-testimony-subclass-delivery) and the per-building CAIRO outputs documented in README. "
        "Embedded as Figure 2 on page 10 of the Pre-Filed Direct Testimony of Juan-Pablo Velez."
    )
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")
    ws["A2"] = subtitle
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 45

    headers = [
        "Subclass",
        "Customers",
        "% of customers",
        "Consumption",
        "% of consumption",
        "Delivery revenue",
        "% of revenue",
    ]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    _header_fill(ws, header_row, len(headers))

    rows = [*HT_V2_ORDER, "all_customers"]
    for i, _key in enumerate(rows, start=1):
        agg_row = 1 + i
        out_row = header_row + i
        ws.cell(row=out_row, column=1, value=f"=subclass_aggregates!B{agg_row}")
        ws.cell(row=out_row, column=2, value=f"=subclass_aggregates!J{agg_row}")
        ws.cell(row=out_row, column=3, value=f"=subclass_aggregates!K{agg_row}")
        ws.cell(row=out_row, column=4, value=f"=subclass_aggregates!L{agg_row}")
        ws.cell(row=out_row, column=5, value=f"=subclass_aggregates!M{agg_row}")
        ws.cell(row=out_row, column=6, value=f"=subclass_aggregates!D{agg_row}")
        ws.cell(row=out_row, column=7, value=f"=subclass_aggregates!N{agg_row}")

    n_rows = len(rows)
    last_data_row = header_row + n_rows
    for r in range(header_row + 1, last_data_row + 1):
        ws[f"B{r}"].number_format = "#,##0"
        ws[f"C{r}"].number_format = "0.0%"
        ws[f"D{r}"].number_format = '#,##0.0" GWh"'
        ws[f"E{r}"].number_format = "0.0%"
        ws[f"F{r}"].number_format = '"$"#,##0'
        ws[f"G{r}"].number_format = "0.0%"

    for c in range(1, 8):
        ws.cell(row=last_data_row, column=c).font = Font(bold=True)

    for c in range(2, 8):
        ws.cell(row=header_row, column=c).alignment = Alignment(horizontal="center")
        for r in range(header_row + 1, last_data_row + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    _autosize(ws, {"A": 24, "B": 14, "C": 16, "D": 16, "E": 18, "F": 18, "G": 14})
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Runtime validation against published Figure 2.
# ---------------------------------------------------------------------------
def _validate_against_published(bat: pl.DataFrame, inputs: dict) -> None:
    """Assert that weighted aggregates match report_variables.pkl (Figure 2)."""
    total_w = float(bat["weight"].sum())
    total_rev = float((bat["weight"] * bat["annual_bill_delivery"]).sum())
    total_kwh = float((bat["weight"] * bat["annual_kwh"]).sum())

    # Aggregate totals.
    assert abs(total_w - inputs["test_year_customer_count"]) < 0.5, (
        f"sum(weight) = {total_w:,.2f} vs test_year_customer_count = {inputs['test_year_customer_count']:,.2f}"
    )
    assert abs(total_rev - inputs["total_delivery_revenue_requirement"]) < 2_000, (
        f"sum(w*bill) = ${total_rev:,.0f} vs total_delivery_RR = ${inputs['total_delivery_revenue_requirement']:,.0f}"
    )
    assert abs(total_kwh - inputs["test_year_residential_kwh"]) < 1.0, (
        f"sum(w*kwh) = {total_kwh:,.0f} vs test_year_residential_kwh = {inputs['test_year_residential_kwh']:,.0f}"
    )
    print("  Aggregate totals: PASS", flush=True)

    # Per-subclass validation against the published Figure 2.
    pkl_path = Path(__file__).resolve().parents[1] / "cache" / "report_variables_cos_subclass.pkl"
    assert pkl_path.exists(), f"Missing {pkl_path}. Render cost_of_service_by_subclass.qmd first."
    v = SimpleNamespace(**pickle.loads(pkl_path.read_bytes()))
    published_rows: list[dict] = v.testimony_subclass_delivery_rows

    by_ht = bat.group_by("postprocess_group.heating_type_v2").agg(
        pl.col("weight").sum().alias("n_customers"),
        (pl.col("weight") * pl.col("annual_bill_delivery")).sum().alias("revenue_delivery"),
        (pl.col("weight") * pl.col("annual_kwh")).sum().alias("total_kwh"),
    )
    actual = {row["postprocess_group.heating_type_v2"]: row for row in by_ht.iter_rows(named=True)}

    import numpy as np

    # Reproduce largest-remainder rounding for customer display counts.
    display_total = round(inputs["test_year_customer_count"])
    sub_weights = [actual[k]["n_customers"] for k in HT_V2_ORDER]
    raw_counts = np.array(sub_weights) * display_total / total_w
    floors = np.floor(raw_counts).astype(np.int64)
    remainder = display_total - int(floors.sum())
    order = np.argsort(-(raw_counts - floors))
    for k in range(remainder):
        floors[order[k]] += 1

    for pub_row in published_rows:
        sub_name = pub_row["subclass"]
        if sub_name == "All customers":
            continue
        key = next(k for k, v_label in HT_V2_LABELS.items() if v_label == sub_name)
        idx = list(HT_V2_ORDER).index(key)
        a = actual[key]

        wb_customers = int(floors[idx])
        pub_customers = int(pub_row["n_customers_display"])
        assert wb_customers == pub_customers, f"{sub_name}: customers {wb_customers} != published {pub_customers}"

        wb_gwh = a["total_kwh"] / 1_000_000.0
        pub_gwh = float(pub_row["consumption_gwh"])
        assert abs(wb_gwh - pub_gwh) < 0.05, f"{sub_name}: consumption {wb_gwh:.1f} GWh != published {pub_gwh:.1f} GWh"

        wb_rev = a["revenue_delivery"]
        pub_rev = float(pub_row["revenue_delivery"])
        assert abs(wb_rev - pub_rev) < 1.0, f"{sub_name}: revenue ${wb_rev:,.0f} != published ${pub_rev:,.0f}"

        wb_pct_cust = float(floors[idx]) / display_total
        pub_pct_cust = float(pub_row["pct_customers_display"])
        assert abs(wb_pct_cust - pub_pct_cust) < 0.001, (
            f"{sub_name}: pct_customers {wb_pct_cust:.4f} != published {pub_pct_cust:.4f}"
        )

        wb_pct_cons = a["total_kwh"] / total_kwh
        pub_pct_cons = float(pub_row["pct_all_consumption"])
        assert abs(wb_pct_cons - pub_pct_cons) < 0.001, (
            f"{sub_name}: pct_consumption {wb_pct_cons:.4f} != published {pub_pct_cons:.4f}"
        )

        wb_pct_rev = a["revenue_delivery"] / total_rev
        pub_pct_rev = float(pub_row["pct_all_revenue"])
        assert abs(wb_pct_rev - pub_pct_rev) < 0.001, (
            f"{sub_name}: pct_revenue {wb_pct_rev:.4f} != published {pub_pct_rev:.4f}"
        )

    print("  Figure 2 replication: PASS (all subclass values match published)", flush=True)


# ---------------------------------------------------------------------------
# Orchestration.
# ---------------------------------------------------------------------------
def build_workbook(output_path: Path) -> Path:
    """Build and save the .xlsx workbook. Returns the output path."""
    print(f"Loading per-building data from {PATH_MASTER_BAT_12} ...", flush=True)
    bat = load_master_bat()
    print(f"  {bat.height:,} rows", flush=True)

    print("Loading revenue-requirement YAML and tariff JSON from rate-design-platform ...", flush=True)
    inputs = load_inputs()
    print(f"  total_delivery_revenue_requirement = ${inputs['total_delivery_revenue_requirement']:,.0f}", flush=True)
    print(f"  test_year_customer_count = {inputs['test_year_customer_count']:,.0f}", flush=True)
    print(f"  test_year_residential_kwh = {inputs['test_year_residential_kwh']:,.0f}", flush=True)
    print(f"  default_vol_usd_per_kwh = {inputs['default_vol_usd_per_kwh']:.6f}", flush=True)
    print(f"  annual_fixed_per_customer = ${inputs['annual_fixed_per_customer']:,.2f}", flush=True)

    bat = bat.with_columns(
        (
            (pl.col("annual_bill_delivery") - inputs["annual_fixed_per_customer"]) / inputs["default_vol_usd_per_kwh"]
        ).alias("annual_kwh")
    )

    print("Validating against published Figure 2 ...", flush=True)
    _validate_against_published(bat, inputs)

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    _write_readme(wb, inputs)
    _write_inputs_tariffs(wb, inputs)
    _write_inputs_revenue_requirement(wb, inputs)
    last_bat_row = _write_bill_per_building(wb, bat)
    _write_subclass_aggregates(wb, last_bat_row)
    _write_fig2_published(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Wrote {output_path} ({output_path.stat().st_size / 1024:.1f} KB)", flush=True)
    return output_path


# ---------------------------------------------------------------------------
# Google Sheets upload.
# ---------------------------------------------------------------------------
_TAB_FORMATTING: dict[str, dict] = {
    "README": {
        "wrap_columns": ["A:C"],
        "column_widths_px": {"A": 280, "B": 480, "C": 480},
        "freeze_rows": 1,
        "bold_header": True,
        "bold_rows": [3, 10, 17, 25],
    },
    "inputs_revenue_requirement": {
        "column_number_formats": {"B": "#,##0.00"},
        "wrap_columns": ["C:D"],
        "column_widths_px": {"A": 240, "B": 140, "C": 480, "D": 480},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "inputs_tariffs": {
        "column_number_formats": {"B": "0.0000"},
        "wrap_columns": ["C:D"],
        "column_widths_px": {"A": 240, "B": 130, "C": 520, "D": 480},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "bill_per_building": {
        "column_number_formats": {
            "B": "#,##0.00",
            "D": "#,##0.00",
            "E": '"$"#,##0.00',
            "F": "#,##0.00",
            "G": "#,##0.00",
        },
        "auto_resize_columns": ["A:G"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "subclass_aggregates": {
        "column_number_formats": {
            "C": "#,##0.00",
            "D": '"$"#,##0.00',
            "E": "#,##0.00",
            "F": "#,##0.00",
            "G": "#,##0",
            "H": "#,##0.00",
            "I": "#,##0",
            "J": "#,##0",
            "K": "0.0%",
            "L": "#,##0.0",
            "M": "0.0%",
            "N": "0.0%",
        },
        "auto_resize_columns": ["A:N"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "fig2_published": {
        "column_number_formats": {
            "B": "#,##0",
            "C": "0.0%",
            "D": '#,##0.0" GWh"',
            "E": "0.0%",
            "F": '"$"#,##0',
            "G": "0.0%",
        },
        "auto_resize_columns": ["A:G"],
        "freeze_rows": 4,
        "bold_header": True,
    },
}


def upload_to_sheet(xlsx_path: Path, spreadsheet_id: str) -> None:
    """Mirror the workbook into the target Google Sheet, preserving formulas."""
    from lib.data.gsheets import apply_sheet_formatting, xlsx_to_gsheet

    print(f"Uploading {xlsx_path} -> Google Sheet {spreadsheet_id} ...", flush=True)
    spreadsheet = xlsx_to_gsheet(xlsx_path, spreadsheet_id, delete_other_tabs=True)
    print("Applying number / wrap / width formatting ...", flush=True)
    for ws in spreadsheet.worksheets():
        spec = _TAB_FORMATTING.get(ws.title)
        if spec:
            apply_sheet_formatting(ws, **spec)
    print(
        f"Done. View at https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit",
        flush=True,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0] if __doc__ else "")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("cache/fig2_subclass_delivery.xlsx"),
        help="Output .xlsx path (relative to the report directory). Default: cache/fig2_subclass_delivery.xlsx",
    )
    parser.add_argument(
        "--upload",
        action="store_true",
        help="Upload to the default Figure 2 Google Sheet after building.",
    )
    parser.add_argument(
        "--spreadsheet-id",
        default=DEFAULT_SPREADSHEET_ID,
        help=f"Override the upload target Sheet id. Default: {DEFAULT_SPREADSHEET_ID}",
    )
    args = parser.parse_args(argv)

    out = build_workbook(args.output)
    if args.upload:
        upload_to_sheet(out, args.spreadsheet_id)
    return 0


if __name__ == "__main__":
    sys.exit(main())
