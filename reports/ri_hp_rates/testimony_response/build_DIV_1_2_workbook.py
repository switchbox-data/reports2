"""Build the supporting workbook for DIV 1-2 data request (Figure 3).

DIV 1-2: Direct Testimony of Juan-Pablo Velez, page 12, Figure 3:

    "a. Please provide in excel format the source files for the data in
    Figure 3, including base input data and all intermediary calculations.
    b. If not provided in response to a. above, please provide
        (1) test year (9/1/24 to 8/31/25) 24-hour interval load profile demand
            data in excel format for each of the RIE residential subclasses
            listed in Figure 3.
        (2) the date and hour of the test year system peak demand."

Figure 3 is ``tbl-cos-by-subclass`` from
``notebooks/cost_of_service_by_subclass.qmd``, showing delivery revenue,
cost-of-service, and cross-subsidy for five residential heating subclasses.

Run from the report directory::

    uv run python -m testimony_response.build_DIV_1_2_workbook
    uv run python -m testimony_response.build_DIV_1_2_workbook --no-upload
    uv run python -m testimony_response.build_DIV_1_2_workbook --title "DIV 1-2 (revised)"
"""

from __future__ import annotations

import argparse
import pickle
import subprocess
import sys
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────

REPORT_DIR = Path(__file__).resolve().parents[1]

UTILITY = "rie"
STATE = "ri"
BATCH = "ri_20260331_r1-20_rate_case_test_year"
RUN_DELIVERY = "1"
RUN_SUPPLY = "2"
S3_BASE = "s3://data.sb/switchbox/cairo/outputs/hp_rates"
RESSTOCK_BASE = "s3://data.sb/nrel/resstock/res_2024_amy2018_2_sb"
LOCAL_RESSTOCK_BASE = Path("/ebs/data/nrel/resstock/res_2024_amy2018_2_sb")
LOCAL_RESSTOCK_METADATA = LOCAL_RESSTOCK_BASE / "metadata_utility" / "state=RI" / "utility_assignment.parquet"
LOCAL_RESSTOCK_LOADS_UPGRADE0 = LOCAL_RESSTOCK_BASE / "load_curve_hourly" / "state=RI" / "upgrade=00"
S3_MC_DIST_SUB_TX = "s3://data.sb/switchbox/marginal_costs/ri/dist_and_sub_tx/utility=rie/year=2025/data.parquet"
S3_MC_BULK_TX = "s3://data.sb/switchbox/marginal_costs/ri/bulk_tx/utility=rie/year=2025/data.parquet"
S3_MASTER_BAT = (
    f"{S3_BASE}/{STATE}/all_utilities/{BATCH}/run_{RUN_DELIVERY}+{RUN_SUPPLY}/cross_subsidization_BAT_values/"
)
ELEC_TOTAL_COL = "out.electricity.total.energy_consumption"

RDP_REF = "e9e5088"
RDP_GITHUB_BASE = "https://github.com/switchbox-data/rate-design-platform/blob"
REPORTS2_GITHUB_BASE = "https://github.com/switchbox-data/reports2/blob"

# Revenue-requirement constants — sourced from rate-design-platform @ e9e5088:
#   rate_design/hp_rates/ri/config/rev_requirement/rie_rate_case_test_year.yaml
REV_REQ: dict = {
    "total_delivery_revenue_requirement": 446463143.03,
    "test_year_customer_count": 419347.83,
    "test_year_residential_kwh": 2821237490.0,
    "resstock_kwh_scale_factor": 0.9568112362177266,
}

DEFAULT_FOLDER_ID = "1uPcJbcOChD6zoFuPb-gsxSByPr7xwmCH"
DEFAULT_TITLE = "DIV 1-2"

# heating_type_v2 subclasses — matches cost_of_service_by_subclass.qmd
HT_V2_ORDER = ("heat_pump", "electrical_resistance", "natgas", "delivered_fuels", "other")
HT_V2_LABELS: dict[str, str] = {
    "heat_pump": "Heat pump",
    "electrical_resistance": "Electric resistance",
    "natgas": "Natural gas",
    "delivered_fuels": "Delivered fuels",
    "other": "Other",
}

# Styling
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
NUMBER_FORMAT_CURRENCY = "$#,##0"
NUMBER_FORMAT_KWH = "#,##0"
NUMBER_FORMAT_PCT = "0.0%"
NUMBER_FORMAT_MCPKWH = "0.000000"
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PASS_FONT = Font(color="006100")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FAIL_FONT = Font(color="9C0006", bold=True)
INFO_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
INFO_FONT = Font(color="003366")
NA_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
NA_FONT = Font(color="808080")


# ── Permalink helpers ─────────────────────────────────────────────────────────


def _rdp_permalink(rel_path: str) -> str:
    return f"{RDP_GITHUB_BASE}/{RDP_REF}/{rel_path}"


def _reports2_head_sha() -> str:
    if not hasattr(_reports2_head_sha, "_cached"):
        repo_root = Path(__file__).resolve().parents[3]
        sha = subprocess.check_output(
            ["git", "-C", str(repo_root), "rev-parse", "HEAD"],
            text=True,
        ).strip()
        _reports2_head_sha._cached = sha  # type: ignore[attr-defined]
    return _reports2_head_sha._cached  # type: ignore[attr-defined]


def _reports2_permalink(rel_path: str) -> str:
    return f"{REPORTS2_GITHUB_BASE}/{_reports2_head_sha()}/{rel_path}"


# ── Data loading ──────────────────────────────────────────────────────────────


def load_master_bat() -> pl.DataFrame:
    """Load master BAT data from S3 for the default run, filtered to RIE."""
    print(f"Loading master BAT from: {S3_MASTER_BAT}", flush=True)
    df = (
        pl.scan_parquet(S3_MASTER_BAT, hive_partitioning=True)
        .filter(pl.col("sb.electric_utility") == UTILITY)
        .collect()
    )
    assert isinstance(df, pl.DataFrame)
    print(f"Loaded {df.height:,} buildings for utility={UTILITY}", flush=True)
    return df


def load_aggregate_load_curves_v2(
    bat_df: pl.DataFrame,
) -> tuple[pl.DataFrame, pl.DataFrame]:
    """Load 8760-hour ResStock load curves aggregated by heating_type_v2 subclass.

    Uses the bldg_id → heating_type_v2 mapping from the master BAT (which has the
    5-class column) and joins it onto ResStock metadata (which only has the 3-class
    heating_type).

    Returns (agg_by_subclass, mc_delivery) where agg_by_subclass has columns
    for each heating_type_v2 label (Heat pump, Electric resistance, etc.) plus
    timestamp and hour_of_year.
    """
    kwh_scale_factor = REV_REQ["resstock_kwh_scale_factor"]
    test_year_customer_count = REV_REQ["test_year_customer_count"]

    # bldg_id → heating_type_v2 from master BAT
    ht_v2_map = bat_df.select("bldg_id", "postprocess_group.heating_type_v2").unique()

    print("Loading ResStock metadata ...", flush=True)
    meta = (
        pl.read_parquet(LOCAL_RESSTOCK_METADATA)
        .filter((pl.col("sb.electric_utility") == UTILITY) & (pl.col("upgrade") == 0))
        .select("bldg_id", "weight")
        .join(ht_v2_map, on="bldg_id", how="left")
    )

    raw_total = float(meta["weight"].sum())
    weight_scale = test_year_customer_count / raw_total
    meta = meta.with_columns(pl.col("weight") * weight_scale)

    print(f"Loading {meta.height} ResStock load curves from local disk ...", flush=True)
    frames: list[pl.DataFrame] = []
    load_dir = LOCAL_RESSTOCK_LOADS_UPGRADE0
    bldg_set = set(meta["bldg_id"].to_list())
    for fname in sorted(load_dir.iterdir()):
        bldg_id = int(fname.stem.split("-")[0])
        if bldg_id not in bldg_set:
            continue
        lf = pl.read_parquet(fname).select("timestamp", ELEC_TOTAL_COL)
        lf = lf.with_columns(pl.lit(bldg_id).alias("bldg_id"))
        frames.append(lf)

    loads = pl.concat(frames)
    loads = loads.join(meta, on="bldg_id")

    agg = (
        loads.group_by(["timestamp", "postprocess_group.heating_type_v2"])
        .agg((pl.col(ELEC_TOTAL_COL) * pl.col("weight")).sum().alias("weighted_kwh"))
        .pivot(on="postprocess_group.heating_type_v2", index="timestamp", values="weighted_kwh")
        .sort("timestamp")
        .head(8760)
    )
    agg = agg.with_columns((pl.int_range(1, agg.height + 1, eager=True)).alias("hour_of_year"))

    # Rename raw subclass codes to display labels and fill nulls
    for code, label in HT_V2_LABELS.items():
        if code in agg.columns:
            agg = agg.rename({code: label})
    for label in HT_V2_LABELS.values():
        if label not in agg.columns:
            agg = agg.with_columns(pl.lit(0.0).alias(label))
        else:
            agg = agg.with_columns(pl.col(label).fill_null(0.0))

    # Apply CAIRO kWh scale factor
    if kwh_scale_factor != 1.0:
        load_cols = list(HT_V2_LABELS.values())
        agg = agg.with_columns([pl.col(c) * kwh_scale_factor for c in load_cols])

    # Add "All customers" total column
    load_cols = list(HT_V2_LABELS.values())
    agg = agg.with_columns(
        pl.sum_horizontal([pl.col(c) for c in load_cols]).alias("All customers"),
    )

    # Load delivery MC parquets
    print("Loading delivery marginal cost parquets from S3 ...", flush=True)
    mc_dist = _load_mc(S3_MC_DIST_SUB_TX, "mc_dist_sub_tx")
    mc_bulk = _load_mc(S3_MC_BULK_TX, "mc_bulk_tx")
    mc = mc_dist.join(mc_bulk.drop("timestamp"), on="hour_of_year").with_columns(
        (pl.col("mc_dist_sub_tx") + pl.col("mc_bulk_tx")).alias("mc_delivery_total"),
    )
    return agg, mc


def _load_mc(s3_path: str, name: str) -> pl.DataFrame:
    df = pl.read_parquet(s3_path)
    if df["timestamp"].dtype.time_zone is not None:  # type: ignore[union-attr]
        df = df.with_columns(pl.col("timestamp").dt.replace_time_zone(None))
    if "mc_total_per_kwh" in df.columns:
        mc_col = "mc_total_per_kwh"
    elif "bulk_tx_cost_enduse" in df.columns:
        mc_col = "bulk_tx_cost_enduse"
    else:
        float_cols = [c for c in df.columns if c != "timestamp" and df[c].dtype.is_float()]
        mc_col = (
            float_cols[0]
            if float_cols
            else [c for c in df.columns if c != "timestamp" and df[c].dtype.is_numeric()][-1]
        )
    df = df.select("timestamp", pl.col(mc_col).alias(name)).sort("timestamp").head(8760)
    return df.with_columns((pl.int_range(1, df.height + 1, eager=True)).alias("hour_of_year"))


def _get_default_vol_rate() -> float:
    """Derive the default volumetric delivery rate ($/kWh) from the calibrated tariff.

    Same approach as cost_of_service_by_subclass.qmd: parse the URDB JSON.
    We hardcode the value to avoid a runtime dependency on rate-design-platform.
    The calibrated rate for rie_default is 0.05039586... $/kWh (from e9e5088).
    """
    from lib.rdp import fetch_rdp_file, parse_urdb_json

    path = "rate_design/hp_rates/ri/config/tariffs/electric/rie_default_calibrated.json"
    doc = parse_urdb_json(fetch_rdp_file(path, RDP_REF))
    return float(doc["items"][0]["energyratestructure"][0][0]["rate"])


# ── Workbook construction ─────────────────────────────────────────────────────


def create_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "README"
    return wb


def add_overview_sheet(wb: Workbook) -> None:
    """README sheet: overview and data source references."""
    ws = wb["README"]

    ws["A1"] = "DIV 1-2: Source Data for Figure 3 (Cost-of-Service by Residential Subclass)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    ws["A3"] = "Data Request"
    ws["A3"].font = SECTION_FONT
    ws["A3"].fill = SECTION_FILL
    ws.merge_cells("A3:F3")

    ws["A4"] = (
        "DIV 1-2: Please refer to the Direct Testimony of Juan-Pablo Velez, page 12, Figure 3.\n"
        "a. Please provide in excel format the source files for the data in Figure 3, "
        "including base input data and all intermediary calculations.\n"
        "b. If not provided in response to a. above, please provide\n"
        "   (1) test year (9/1/24 to 8/31/25) 24-hour interval load profile demand data "
        "in excel format for each of the RIE residential subclasses listed in Figure 3.\n"
        "   (2) the date and hour of the test year system peak demand."
    )
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A4:F4")
    ws.row_dimensions[4].height = 100

    ws["A6"] = "Response"
    ws["A6"].font = SECTION_FONT
    ws["A6"].fill = SECTION_FILL
    ws.merge_cells("A6:F6")

    ws["A7"] = (
        "This workbook provides the source data and all intermediary calculations "
        "behind Figure 3.\n\n"
        "Sheet 1: 8,760-hour load profiles for each residential subclass.\n"
        "Sheet 2: Consumption summary (total kWh and GWh per subclass via SUM formulas over Sheet 1).\n"
        "Sheet 3: Hourly delivery marginal costs.\n"
        "Sheet 4: Per-building CAIRO BAT data — bldg_id, heating_type_v2, weight, "
        "annual_bill_delivery, economic_burden_delivery, with formula columns for "
        "weight*bill and weight*EB. Used as the source for SUMIF aggregations in Sheet 5.\n"
        "Sheet 5: COS by subclass — parameters block (revenue requirement, sum of weighted "
        "economic burdens, EPMC rate) followed by the subclass table. Customers, Revenue, "
        "and Wtd EB are SUMIF formulas over Sheet 4. EPMC Residual, COS, Cross-Subsidy, "
        "and all percentages are derived Excel formulas.\n"
        "Sheet 6: Figure 3 reproduction (all values reference Sheet 5 via formulas).\n"
        "Sheet 7: System peak demand.\n"
        "Sheet 8: Validation against testimony cache values."
    )
    ws["A7"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A7:F7")
    ws.row_dimensions[7].height = 180

    ws["A9"] = "Data Sources (S3 Paths)"
    ws["A9"].font = SECTION_FONT
    ws["A9"].fill = SECTION_FILL
    ws.merge_cells("A9:F9")

    sources = [
        ("Master BAT (CAIRO outputs)", S3_MASTER_BAT),
        ("ResStock metadata", f"{RESSTOCK_BASE}/metadata_utility/state=RI/utility_assignment.parquet"),
        ("ResStock load curves", f"{RESSTOCK_BASE}/load_curve_hourly/state=RI/upgrade=00/"),
        ("Sub-TX/distribution MC", S3_MC_DIST_SUB_TX),
        ("Bulk transmission MC", S3_MC_BULK_TX),
        (
            "Revenue requirement YAML",
            _rdp_permalink("rate_design/hp_rates/ri/config/rev_requirement/rie_rate_case_test_year.yaml"),
        ),
        ("COS notebook", _reports2_permalink("reports/ri_hp_rates/notebooks/cost_of_service_by_subclass.qmd")),
    ]

    row = 10
    for label, path in sources:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = path
        ws.merge_cells(f"B{row}:F{row}")
        row += 1

    row += 1
    ws[f"A{row}"] = "Parameters"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    params = [
        ("Utility", UTILITY),
        ("Batch", BATCH),
        ("Run (delivery + supply)", f"{RUN_DELIVERY}+{RUN_SUPPLY}"),
        ("Total Delivery Revenue Requirement", f"${REV_REQ['total_delivery_revenue_requirement']:,.2f}"),
        ("Test Year Customer Count", f"{REV_REQ['test_year_customer_count']:,.2f}"),
        ("Test Year Residential kWh", f"{REV_REQ['test_year_residential_kwh']:,.0f}"),
        ("ResStock kWh Scale Factor", f"{REV_REQ['resstock_kwh_scale_factor']:.16f}"),
        ("Test Year", "9/1/2024 - 8/31/2025"),
    ]
    for label, val in params:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = val
        ws.merge_cells(f"B{row}:F{row}")
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40


def add_load_curves_sheet(wb: Workbook, agg: pl.DataFrame) -> None:
    """Sheet 1: 8,760-hour load profiles by subclass (DIV 1-2 part b.1)."""
    ws = wb.create_sheet("1. Load Profiles (8760)")

    ws["A1"] = "8,760-Hour Load Profiles by Residential Subclass (kWh)"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        "Aggregated from NREL ResStock building simulations, weighted to Test Year "
        "customer count and scaled by resstock_kwh_scale_factor."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 30

    subclass_cols = [*list(HT_V2_LABELS.values()), "All customers"]
    headers = ["Hour", "Timestamp", *subclass_cols]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, r in enumerate(agg.iter_rows(named=True)):
        row = 5 + i
        ws.cell(row=row, column=1, value=r["hour_of_year"])
        ws.cell(row=row, column=2, value=str(r["timestamp"]))
        for col_idx, sc in enumerate(subclass_cols, start=3):
            ws.cell(row=row, column=col_idx, value=round(r.get(sc, 0.0), 2)).number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    for col_idx in range(3, 3 + len(subclass_cols)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10
    ws.freeze_panes = "A5"


def add_consumption_summary_sheet(wb: Workbook, n_load_rows: int) -> None:
    """Sheet 2: Consumption summary — SUM of load profiles by subclass, via formulas."""
    ws = wb.create_sheet("2. Consumption Summary")

    ws["A1"] = "Consumption by Subclass (from Load Profiles)"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:D1")

    headers = ["Subclass", "Total kWh", "GWh", "% of Consumption"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    load_sheet = "'1. Load Profiles (8760)'"
    last_data_row = 4 + n_load_rows
    subclass_cols = [*list(HT_V2_LABELS.values()), "All customers"]
    load_col_letters = [get_column_letter(i) for i in range(3, 3 + len(subclass_cols))]
    n_sub = len(HT_V2_ORDER)
    total_row = 4 + n_sub

    for idx, (label, col_letter) in enumerate(zip(subclass_cols, load_col_letters, strict=True)):
        row = 4 + idx
        ws.cell(row=row, column=1, value=label)
        ws.cell(
            row=row, column=2, value=f"=SUM({load_sheet}!{col_letter}5:{col_letter}{last_data_row})"
        ).number_format = "#,##0"
        ws.cell(row=row, column=3, value=f"=B{row}/1000000").number_format = "#,##0.0"
        if label == "All customers":
            ws.cell(row=row, column=4, value=100.0).number_format = "0.0"
            for c in range(1, 5):
                ws.cell(row=row, column=c).font = Font(bold=True)
        else:
            ws.cell(row=row, column=4, value=f"=ROUND(B{row}/$B${total_row}*100,1)").number_format = "0.0"

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.freeze_panes = "A4"


def add_marginal_cost_sheet(wb: Workbook, mc: pl.DataFrame) -> None:
    """Sheet 3: Hourly delivery marginal costs."""
    ws = wb.create_sheet("3. Delivery Marginal Costs")

    ws["A1"] = "Hourly Delivery Marginal Costs ($/kWh)"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:E1")

    headers = ["Hour", "Timestamp", "Sub-TX/Dist MC", "Bulk TX MC", "Total Del MC"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, r in enumerate(mc.iter_rows(named=True)):
        row = 4 + i
        ws.cell(row=row, column=1, value=r["hour_of_year"])
        ws.cell(row=row, column=2, value=str(r["timestamp"]))
        ws.cell(row=row, column=3, value=r["mc_dist_sub_tx"]).number_format = NUMBER_FORMAT_MCPKWH
        ws.cell(row=row, column=4, value=r["mc_bulk_tx"]).number_format = NUMBER_FORMAT_MCPKWH
        ws.cell(row=row, column=5, value=r["mc_delivery_total"]).number_format = NUMBER_FORMAT_MCPKWH

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    for c in ("C", "D", "E"):
        ws.column_dimensions[c].width = 12
    ws.freeze_panes = "A4"


def add_bat_building_data_sheet(wb: Workbook, bat_df: pl.DataFrame) -> None:
    """Sheet 4: Per-building CAIRO BAT data — source for SUMIF aggregations in sheet 5.

    Columns:
      A  bldg_id
      B  heating_type_v2 (display label — used as SUMIF criteria in sheet 5)
      C  weight
      D  annual_bill_delivery ($)
      E  economic_burden_delivery ($)
      F  =C*D  weight * bill  (formula)
      G  =C*E  weight * EB   (formula)
    """
    ws = wb.create_sheet("4. BAT Building Data")

    ws["A1"] = "CAIRO Master BAT - Per-Building Data (RIE Residential, Test Year)"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:G1")

    ws["A2"] = f"S3 source: {S3_MASTER_BAT}"
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:G2")

    col_headers = [
        "bldg_id",
        "heating_type_v2",
        "weight",
        "annual_bill_delivery ($)",
        "economic_burden_delivery ($)",
        "weight * bill",
        "weight * EB",
    ]
    HDR_ROW = 3
    for ci, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Sort by heating_type_v2 order then bldg_id for readability
    order_map = {k: i for i, k in enumerate(HT_V2_ORDER)}
    sorted_df = (
        bat_df.select(
            "bldg_id",
            "postprocess_group.heating_type_v2",
            "weight",
            "annual_bill_delivery",
            "economic_burden_delivery",
        )
        .with_columns(pl.col("postprocess_group.heating_type_v2").replace(order_map).alias("_sort_order"))
        .sort(["_sort_order", "bldg_id"])
        .drop("_sort_order")
    )

    DATA_START = HDR_ROW + 1
    n_bldg = sorted_df.height
    for i, row_data in enumerate(sorted_df.iter_rows(named=True)):
        row = DATA_START + i
        label = HT_V2_LABELS[row_data["postprocess_group.heating_type_v2"]]
        ws.cell(row=row, column=1, value=row_data["bldg_id"])
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=float(row_data["weight"])).number_format = "#,##0.000"
        ws.cell(row=row, column=4, value=float(row_data["annual_bill_delivery"])).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=5, value=float(row_data["economic_burden_delivery"])).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=6, value=f"=C{row}*D{row}").number_format = '"$"#,##0.00'
        ws.cell(row=row, column=7, value=f"=C{row}*E{row}").number_format = '"$"#,##0.00'

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = f"A{DATA_START}"

    # Store range info for SUMIF references in sheet 5
    wb._div12_bat_sheet = "4. BAT Building Data"  # type: ignore[attr-defined]
    wb._div12_bat_data_start = DATA_START  # type: ignore[attr-defined]
    wb._div12_bat_last_row = DATA_START + n_bldg - 1  # type: ignore[attr-defined]


def add_cos_by_subclass_sheet(wb: Workbook, bat_df: pl.DataFrame) -> None:
    """Sheet 5: COS by subclass — labeled parameter block then formula table.

    Customers, Revenue, and Wtd EB are SUMIF aggregations from '4. BAT Building Data'.
    All other columns (EPMC Residual, COS, Cross-Subsidy, percentages) are Excel formulas.
    """
    ws = wb.create_sheet("5. COS by Subclass")

    ws["A1"] = "Cost of Service by Residential Subclass"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:H1")

    # Define table layout constants up-front so formulas in the parameters
    # block can reference the table cells (e.g. sum_weighted_eb = SUM(D col)).
    TBL_HDR = 10  # Section header row for the subclass table
    HDR_ROW = TBL_HDR + 1  # Column header row
    DATA_START = HDR_ROW + 1  # First subclass data row
    n_sub = len(HT_V2_ORDER)
    TOTAL_ROW = DATA_START + n_sub  # "All customers" totals row

    # BAT building data sheet range (set by add_bat_building_data_sheet)
    bat_sheet = f"'{wb._div12_bat_sheet}'"  # type: ignore[attr-defined]
    bat_ds = wb._div12_bat_data_start  # type: ignore[attr-defined]
    bat_lr = wb._div12_bat_last_row  # type: ignore[attr-defined]
    bat_label_range = f"{bat_sheet}!$B${bat_ds}:$B${bat_lr}"
    bat_weight_range = f"{bat_sheet}!$C${bat_ds}:$C${bat_lr}"
    bat_wtd_bill_range = f"{bat_sheet}!$F${bat_ds}:$F${bat_lr}"
    bat_wtd_eb_range = f"{bat_sheet}!$G${bat_ds}:$G${bat_lr}"

    # ── Parameters block ──────────────────────────────────────────────────────

    ws["A3"] = "Parameters"
    ws["A3"].font = SECTION_FONT
    ws["A3"].fill = SECTION_FILL
    ws.merge_cells("A3:H3")

    for ci, h in enumerate(["Parameter", "Value", "Source", "Notes"], start=1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    total_rr = REV_REQ["total_delivery_revenue_requirement"]
    sum_weighted_eb = float((bat_df["weight"] * bat_df["economic_burden_delivery"]).sum())
    yaml_ref = _rdp_permalink("rate_design/hp_rates/ri/config/rev_requirement/rie_rate_case_test_year.yaml")
    pct_mc = 100.0 * sum_weighted_eb / total_rr
    pct_residual = 100.0 - pct_mc
    epmc_rate_val = (total_rr - sum_weighted_eb) / sum_weighted_eb

    # Row indices (1-based) for cells used in formulas below
    RR_ROW = 5
    SWE_ROW = 6
    RESIDUAL_ROW = 7
    EPMC_ROW = 8

    # sum_weighted_eb is written as a formula that sums the Wtd EB column in the
    # table below (column D, rows DATA_START through DATA_START+n_sub-1).
    swe_formula = f"=SUM(D{DATA_START}:D{DATA_START + n_sub - 1})"

    epmc_note = (
        f"Equi-Proportional Marginal Cost (EPMC) residual allocation rate, in $/$ of economic burden. "
        f"epmc_rate = total_residual / sum_weighted_eb = "
        f"(B{RR_ROW} - B{SWE_ROW}) / B{SWE_ROW}. "
        f"Marginal costs (Wtd EB) recover only {pct_mc:.1f}% of the revenue requirement; "
        f"the remaining {pct_residual:.1f}% is the 'residual' allocated back to customers "
        f"in proportion to their economic burden. "
        f"A rate of {epmc_rate_val:.4f} means that for every $1.00 of marginal cost a subclass "
        f"causes, ${epmc_rate_val:.4f} more is charged to cover the residual."
    )

    params = [
        (
            RR_ROW,
            "total_delivery_revenue_requirement",
            total_rr,
            yaml_ref,
            "Total test-year delivery revenue requirement ($). Source: RIE rate case filing, PRB-1-ELEC.",
        ),
        (
            SWE_ROW,
            "sum_weighted_eb",
            swe_formula,
            f"Formula: =SUM(D{DATA_START}:D{DATA_START + n_sub - 1}) — sum of 'Wtd EB ($)' column",
            "Sum of (weight * economic_burden_delivery) across all RIE buildings = total weighted marginal cost ($)",
        ),
        (
            RESIDUAL_ROW,
            "total_residual",
            f"=B{RR_ROW}-B{SWE_ROW}",
            "Formula: =B5-B6",
            "Revenue requirement not recovered by marginal costs alone ($)",
        ),
        (EPMC_ROW, "epmc_rate", f"=B{RESIDUAL_ROW}/B{SWE_ROW}", "Formula: =B7/B6", epmc_note),
    ]
    for row, key, val, src, notes in params:
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=3, value=src)
        ws.cell(row=row, column=4, value=notes)

    ws[f"B{RR_ROW}"].number_format = '"$"#,##0.00'
    ws[f"B{SWE_ROW}"].number_format = '"$"#,##0.00'
    ws[f"B{RESIDUAL_ROW}"].number_format = '"$"#,##0.00'
    ws[f"B{EPMC_ROW}"].number_format = "0.0000000000"

    # ── Column source notes ───────────────────────────────────────────────────
    # (Replaced by '4. BAT Building Data' sheet — formulas in this sheet
    #  directly reference that sheet's columns via SUMIF.)

    # ── Subclass data table ───────────────────────────────────────────────────

    ws.cell(row=TBL_HDR, column=1, value="Subclass Calculations").font = SECTION_FONT
    ws.cell(row=TBL_HDR, column=1).fill = SECTION_FILL
    ws.merge_cells(f"A{TBL_HDR}:L{TBL_HDR}")

    col_headers = [
        "Subclass",  # A
        "Customers",  # B  SUMIF(bat!B, label, bat!C)
        "Revenue ($)",  # C  SUMIF(bat!B, label, bat!F = weight*bill)
        "Wtd EB ($)",  # D  SUMIF(bat!B, label, bat!G = weight*EB)
        "EPMC Residual ($)",  # E  =D * epmc_rate  (formula)
        "COS ($)",  # F  =D + E  (formula)
        "Cross-Subsidy ($)",  # G  =C - F  (formula: revenue minus COS)
        "GWh",  # H  = from '2. Consumption Summary'
        "% Cust",  # I  (formula)
        "% Cons",  # J  (formula ref consumption summary)
        "% Rev",  # K  (formula)
        "% COS",  # L  (formula)
    ]
    for ci, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Consumption summary sheet row alignment:
    # '2. Consumption Summary' data rows start at row 4 (heat_pump = row 4, ..., all = row 9)
    cons_sheet = "'2. Consumption Summary'"
    cons_data_start = 4

    for idx, code in enumerate(HT_V2_ORDER):
        row = DATA_START + idx
        cons_row = cons_data_start + idx
        ws.cell(row=row, column=1, value=HT_V2_LABELS[code])
        ws.cell(row=row, column=2, value=f"=SUMIF({bat_label_range},A{row},{bat_weight_range})").number_format = "#,##0"
        ws.cell(
            row=row, column=3, value=f"=SUMIF({bat_label_range},A{row},{bat_wtd_bill_range})"
        ).number_format = '"$"#,##0'
        ws.cell(
            row=row, column=4, value=f"=SUMIF({bat_label_range},A{row},{bat_wtd_eb_range})"
        ).number_format = '"$"#,##0'
        ws.cell(row=row, column=5, value=f"=D{row}*$B${EPMC_ROW}").number_format = '"$"#,##0'
        ws.cell(row=row, column=6, value=f"=D{row}+E{row}").number_format = '"$"#,##0'
        ws.cell(row=row, column=7, value=f"=C{row}-F{row}").number_format = '"$"#,##0'
        ws.cell(row=row, column=8, value=f"={cons_sheet}!C{cons_row}").number_format = "#,##0.0"
        ws.cell(row=row, column=9, value=f"=ROUND(B{row}/$B${TOTAL_ROW}*100,1)").number_format = "0.0"
        ws.cell(row=row, column=10, value=f"={cons_sheet}!D{cons_row}").number_format = "0.0"
        ws.cell(row=row, column=11, value=f"=ROUND(C{row}/$C${TOTAL_ROW}*100,1)").number_format = "0.0"
        ws.cell(row=row, column=12, value=f"=ROUND(F{row}/$F${TOTAL_ROW}*100,1)").number_format = "0.0"

    # All customers total row
    cons_total_row = cons_data_start + n_sub
    ws.cell(row=TOTAL_ROW, column=1, value="All customers")
    ws.cell(row=TOTAL_ROW, column=2, value=f"=SUM(B{DATA_START}:B{TOTAL_ROW - 1})").number_format = "#,##0"
    ws.cell(row=TOTAL_ROW, column=3, value=f"=SUM(C{DATA_START}:C{TOTAL_ROW - 1})").number_format = '"$"#,##0'
    ws.cell(row=TOTAL_ROW, column=4, value=f"=SUM(D{DATA_START}:D{TOTAL_ROW - 1})").number_format = '"$"#,##0'
    ws.cell(row=TOTAL_ROW, column=5, value=f"=SUM(E{DATA_START}:E{TOTAL_ROW - 1})").number_format = '"$"#,##0'
    ws.cell(row=TOTAL_ROW, column=6, value=f"=SUM(F{DATA_START}:F{TOTAL_ROW - 1})").number_format = '"$"#,##0'
    ws.cell(row=TOTAL_ROW, column=7, value=f"=SUM(G{DATA_START}:G{TOTAL_ROW - 1})").number_format = '"$"#,##0'
    ws.cell(row=TOTAL_ROW, column=8, value=f"={cons_sheet}!C{cons_total_row}").number_format = "#,##0.0"
    ws.cell(row=TOTAL_ROW, column=9, value=100.0).number_format = "0.0"
    ws.cell(row=TOTAL_ROW, column=10, value=100.0).number_format = "0.0"
    ws.cell(row=TOTAL_ROW, column=11, value=100.0).number_format = "0.0"
    ws.cell(row=TOTAL_ROW, column=12, value=100.0).number_format = "0.0"
    for c in range(1, 13):
        ws.cell(row=TOTAL_ROW, column=c).font = Font(bold=True)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    for c in "CDEFGH":
        ws.column_dimensions[c].width = 14
    for c in "IJKL":
        ws.column_dimensions[c].width = 8
    ws.freeze_panes = f"A{HDR_ROW + 1}"

    # Store row constants so callers can reference this sheet
    wb._div12_cos_data_start = DATA_START  # type: ignore[attr-defined]
    wb._div12_cos_total_row = TOTAL_ROW  # type: ignore[attr-defined]
    wb._div12_cos_n_sub = n_sub  # type: ignore[attr-defined]


def add_figure3_sheet(wb: Workbook) -> None:
    """Sheet 6: Figure 3 reproduction — all values reference '5. COS by Subclass'."""
    ws = wb.create_sheet("6. Figure 3 (COS by Subclass)")

    ws["A1"] = "Figure 3: Delivery revenue, cost of service, and cross-subsidy"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:K1")
    ws["A2"] = "Test Year: 9/1/2024 - 8/31/2025. All values reference '5. COS by Subclass'."
    ws.merge_cells("A2:K2")

    headers = [
        "Subclass",
        "# Customers",
        "% Cust",
        "GWh",
        "% Cons",
        "Revenue",
        "% Rev",
        "Cost of Service",
        "% COS",
        "Cross-Subsidy",
        "XS/COS %",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    cos_sheet = "'5. COS by Subclass'"
    data_start = wb._div12_cos_data_start  # type: ignore[attr-defined]
    all_labels = [*[HT_V2_LABELS[k] for k in HT_V2_ORDER], "All customers"]

    for idx, label in enumerate(all_labels):
        row = 5 + idx
        cos_row = data_start + idx
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=f"=ROUND({cos_sheet}!B{cos_row},0)").number_format = "#,##0"
        ws.cell(row=row, column=3, value=f"={cos_sheet}!I{cos_row}").number_format = "0.0"
        ws.cell(row=row, column=4, value=f"={cos_sheet}!H{cos_row}").number_format = "#,##0.0"
        ws.cell(row=row, column=5, value=f"={cos_sheet}!J{cos_row}").number_format = "0.0"
        ws.cell(row=row, column=6, value=f"=ROUND({cos_sheet}!C{cos_row},0)").number_format = "$#,##0"
        ws.cell(row=row, column=7, value=f"={cos_sheet}!K{cos_row}").number_format = "0.0"
        ws.cell(row=row, column=8, value=f"=ROUND({cos_sheet}!F{cos_row},0)").number_format = "$#,##0"
        ws.cell(row=row, column=9, value=f"={cos_sheet}!L{cos_row}").number_format = "0.0"
        ws.cell(row=row, column=10, value=f"=ROUND({cos_sheet}!G{cos_row},0)").number_format = "$#,##0"
        # XS/COS from formula (not pulled from COS sheet to keep this sheet self-explanatory)
        ws.cell(row=row, column=11, value=f"=IF(H{row}<>0,ROUND(J{row}/H{row}*100,1),0)").number_format = "0.0"
        if label == "All customers":
            for c in range(1, 12):
                ws.cell(row=row, column=c).font = Font(bold=True)

    ws.column_dimensions["A"].width = 16
    for col in ("B", "D", "F", "H", "J"):
        ws.column_dimensions[col].width = 14
    for col in ("C", "E", "G", "I", "K"):
        ws.column_dimensions[col].width = 8
    ws.freeze_panes = "A5"


def add_system_peak_sheet(wb: Workbook, mc: pl.DataFrame) -> None:
    """Sheet 7: System peak demand (DIV 1-2 part b.2)."""
    ws = wb.create_sheet("7. System Peak Demand")

    ws["A1"] = "System Peak Demand - Test Year (9/1/2024 to 8/31/2025)"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:E1")

    # Sub-TX/Dist peak
    ws["A3"] = "Sub-TX/Distribution Peak (RIE)"
    ws["A3"].font = SECTION_FONT
    ws["A3"].fill = SECTION_FILL
    ws.merge_cells("A3:E3")

    peak_dist = mc.sort("mc_dist_sub_tx", descending=True).head(1)
    ws["A4"] = "Timestamp"
    ws["B4"] = str(peak_dist["timestamp"][0])
    ws["A5"] = "Hour of year"
    ws["B5"] = int(peak_dist["hour_of_year"][0])
    ws["A6"] = "MC ($/kWh)"
    ws["B6"] = float(peak_dist["mc_dist_sub_tx"][0])
    ws["B6"].number_format = NUMBER_FORMAT_MCPKWH

    # Bulk TX peak
    ws["A8"] = "Bulk TX Peak (New England)"
    ws["A8"].font = SECTION_FONT
    ws["A8"].fill = SECTION_FILL
    ws.merge_cells("A8:E8")

    peak_bulk = mc.sort("mc_bulk_tx", descending=True).head(1)
    ws["A9"] = "Timestamp"
    ws["B9"] = str(peak_bulk["timestamp"][0])
    ws["A10"] = "Hour of year"
    ws["B10"] = int(peak_bulk["hour_of_year"][0])
    ws["A11"] = "MC ($/kWh)"
    ws["B11"] = float(peak_bulk["mc_bulk_tx"][0])
    ws["B11"].number_format = NUMBER_FORMAT_MCPKWH

    # Top 10 tables
    ws["A13"] = "Top 10 - Sub-TX/Dist (RIE)"
    ws["A13"].font = SECTION_FONT
    ws["A13"].fill = SECTION_FILL
    ws.merge_cells("A13:D13")

    for col_idx, h in enumerate(["Rank", "Timestamp", "Hour", "MC ($/kWh)"], start=1):
        cell = ws.cell(row=14, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    top10_dist = mc.sort("mc_dist_sub_tx", descending=True).head(10)
    for i, r in enumerate(top10_dist.iter_rows(named=True)):
        row = 15 + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=str(r["timestamp"]))
        ws.cell(row=row, column=3, value=r["hour_of_year"])
        ws.cell(row=row, column=4, value=r["mc_dist_sub_tx"]).number_format = NUMBER_FORMAT_MCPKWH

    ws["A26"] = "Top 10 - Bulk TX (New England)"
    ws["A26"].font = SECTION_FONT
    ws["A26"].fill = SECTION_FILL
    ws.merge_cells("A26:D26")

    for col_idx, h in enumerate(["Rank", "Timestamp", "Hour", "MC ($/kWh)"], start=1):
        cell = ws.cell(row=27, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    top10_bulk = mc.sort("mc_bulk_tx", descending=True).head(10)
    for i, r in enumerate(top10_bulk.iter_rows(named=True)):
        row = 28 + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=str(r["timestamp"]))
        ws.cell(row=row, column=3, value=r["hour_of_year"])
        ws.cell(row=row, column=4, value=r["mc_bulk_tx"]).number_format = NUMBER_FORMAT_MCPKWH

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 10


def add_validation_sheet(wb: Workbook) -> None:
    """Sheet 8: Validate workbook values against testimony cache."""
    ws = wb.create_sheet("8. Validation vs. Testimony")

    ws["A1"] = "Validation: Workbook vs. Expert Testimony"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:G1")

    hdr_row = 3
    headers = ["Metric", "Workbook", "Testimony", "Diff", "% Diff", "Tol", "Status"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    cos_pkl = REPORT_DIR / "cache" / "report_variables_cos_subclass.pkl"
    t_vars: dict[str, Any] = {}
    if cos_pkl.exists():
        t_vars = pickle.loads(cos_pkl.read_bytes())

    # Cell refs point to '5. COS by Subclass'
    cos = "'5. COS by Subclass'"
    data_start = wb._div12_cos_data_start  # type: ignore[attr-defined]
    total_cos_row = wb._div12_cos_total_row  # type: ignore[attr-defined]
    checks: list[tuple[str, str | float, float | None, float]] = []

    checks.append(
        (
            "Total Delivery RR (param)",
            REV_REQ["total_delivery_revenue_requirement"],
            t_vars.get("rie_rev_req_total_delivery_rr"),
            0.01,
        )
    )
    checks.append(
        (
            "Test Year Customers (param)",
            REV_REQ["test_year_customer_count"],
            t_vars.get("rie_rev_req_test_year_customer_count"),
            0.01,
        )
    )

    for code, label in HT_V2_LABELS.items():
        prefix_map = {"heat_pump": "hp", "electrical_resistance": "er", "natgas": "ng", "delivered_fuels": "df"}
        prefix = prefix_map.get(code)
        if prefix is None:
            continue
        idx = list(HT_V2_ORDER).index(code)
        cos_row = data_start + idx
        # COS sheet cols: B=customers, C=revenue, D=weighted_eb, E=epmc_residual, F=COS, G=cross_sub
        checks.append((f"{label} - COS", f"={cos}!F{cos_row}", t_vars.get(f"cos_default_{prefix}_group_cos"), 1.0))
        checks.append((f"{label} - Revenue", f"={cos}!C{cos_row}", t_vars.get(f"cos_default_{prefix}_group_rev"), 1.0))
        checks.append((f"{label} - Cross-Sub", f"={cos}!G{cos_row}", t_vars.get(f"cos_default_{prefix}_group_xs"), 1.0))

    checks.append(("Total customers", f"={cos}!B{total_cos_row}", t_vars.get("cos_subclass_total_customers"), 1.0))
    checks.append(("Total revenue", f"={cos}!C{total_cos_row}", t_vars.get("cos_subclass_total_delivery_rev"), 100.0))
    checks.append(("Total COS", f"={cos}!F{total_cos_row}", t_vars.get("cos_subclass_total_cos"), 100.0))

    row = hdr_row + 1
    for metric, wb_val, t_val, tol in checks:
        ws.cell(row=row, column=1, value=metric)

        if (isinstance(wb_val, str) and wb_val.startswith("=")) or wb_val is not None:
            ws.cell(row=row, column=2, value=wb_val).number_format = "#,##0.00"
        else:
            ws.cell(row=row, column=2, value="N/A")

        if t_val is not None:
            ws.cell(row=row, column=3, value=t_val).number_format = "#,##0.00"
            ws.cell(row=row, column=4, value=f"=B{row}-C{row}").number_format = "#,##0.00"
            ws.cell(row=row, column=5, value=f"=IF(C{row}<>0,ABS(D{row}/C{row}),0)").number_format = "0.00%"
            ws.cell(row=row, column=6, value=tol).number_format = "#,##0.00"
            ws.cell(row=row, column=7, value=f'=IF(ABS(D{row})<={tol},"PASS","FAIL")')
        else:
            ws.cell(row=row, column=3, value="N/A")
            for c in (4, 5, 6):
                ws.cell(row=row, column=c, value="")
            ws.cell(row=row, column=7, value="N/A")

        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 6
    ws.column_dimensions["G"].width = 6
    ws.freeze_panes = f"A{hdr_row + 1}"


# ── Orchestration ─────────────────────────────────────────────────────────────


def build_workbook(output_path: Path) -> Path:
    """Load all inputs, build every sheet, write the .xlsx."""
    bat_df = load_master_bat()

    print("Loading 8760-hour load curves and marginal costs ...", flush=True)
    agg_loads, mc_delivery = load_aggregate_load_curves_v2(bat_df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = create_workbook()
    add_overview_sheet(wb)
    add_load_curves_sheet(wb, agg_loads)
    add_consumption_summary_sheet(wb, n_load_rows=agg_loads.height)
    add_marginal_cost_sheet(wb, mc_delivery)
    add_bat_building_data_sheet(wb, bat_df)
    add_cos_by_subclass_sheet(wb, bat_df)
    add_figure3_sheet(wb)
    add_system_peak_sheet(wb, mc_delivery)
    add_validation_sheet(wb)
    wb.save(str(output_path))
    print(f"Wrote {output_path} ({output_path.stat().st_size / 1024:.1f} KB)", flush=True)
    return output_path


# ── Google Sheets upload ──────────────────────────────────────────────────────

_TAB_FORMATTING: dict[str, dict] = {
    "README": {
        "wrap_columns": ["B:B"],
        "column_widths_px": {"A": 160, "B": 450},
        "freeze_rows": 0,
        "bold_header": False,
    },
    "1. Load Profiles (8760)": {
        "freeze_rows": 4,
        "bold_header": True,
        "column_widths_px": {"A": 45, "B": 140, "C": 90, "D": 90, "E": 90, "F": 90, "G": 90, "H": 90},
    },
    "2. Consumption Summary": {
        "freeze_rows": 3,
        "bold_header": True,
        "column_widths_px": {"A": 130, "B": 110, "C": 70, "D": 80},
    },
    "3. Delivery Marginal Costs": {
        "freeze_rows": 3,
        "bold_header": True,
        "column_widths_px": {"A": 45, "B": 140, "C": 110, "D": 110, "E": 110},
    },
    "4. BAT Building Data": {
        "freeze_rows": 3,
        "bold_header": True,
        "column_widths_px": {
            "A": 75,
            "B": 140,
            "C": 80,
            "D": 110,
            "E": 110,
            "F": 100,
            "G": 100,
        },
    },
    "5. COS by Subclass": {
        "freeze_rows": 0,
        "bold_header": True,
        "column_widths_px": {
            "A": 140,
            "B": 95,
            "C": 100,
            "D": 100,
            "E": 110,
            "F": 95,
            "G": 115,
            "H": 65,
            "I": 60,
            "J": 60,
            "K": 60,
            "L": 60,
        },
    },
    "6. Figure 3 (COS by Subclass)": {
        "freeze_rows": 4,
        "bold_header": True,
        "column_widths_px": {
            "A": 130,
            "B": 100,
            "C": 60,
            "D": 100,
            "E": 60,
            "F": 100,
            "G": 60,
            "H": 110,
            "I": 60,
            "J": 110,
            "K": 65,
        },
    },
    "7. System Peak Demand": {
        "freeze_rows": 0,
        "bold_header": False,
        "column_widths_px": {"A": 110, "B": 140, "C": 55, "D": 90},
    },
    "8. Validation vs. Testimony": {
        "freeze_rows": 3,
        "bold_header": True,
        "column_widths_px": {"A": 160, "B": 100, "C": 100, "D": 90, "E": 70, "F": 55, "G": 60},
    },
}


def upload_to_folder(xlsx_path: Path, folder_id: str, title: str) -> None:
    """Create (or replace) a Google Sheet in the given Drive folder."""
    from lib.data.gsheets import (
        apply_sheet_formatting,
        create_sheet_in_folder,
        xlsx_to_gsheet,
    )

    print(f"Uploading '{title}' to Drive folder {folder_id} ...", flush=True)
    spreadsheet = create_sheet_in_folder(title, folder_id)
    xlsx_to_gsheet(xlsx_path, spreadsheet.id, delete_other_tabs=True)

    print("Applying formatting ...", flush=True)
    for ws in spreadsheet.worksheets():
        spec = _TAB_FORMATTING.get(ws.title)
        if spec:
            apply_sheet_formatting(ws, **spec)
    print(
        f"Done. View at https://docs.google.com/spreadsheets/d/{spreadsheet.id}/edit",
        flush=True,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0] if __doc__ else "")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("cache/div_1_2_figure3.xlsx"),
        help="Output .xlsx path. Default: cache/div_1_2_figure3.xlsx",
    )
    parser.add_argument(
        "--folder-id",
        default=DEFAULT_FOLDER_ID,
        help=f"Google Drive folder ID to upload into. Default: {DEFAULT_FOLDER_ID}",
    )
    parser.add_argument(
        "--title",
        default=DEFAULT_TITLE,
        help=f"Name for the Google Sheet. Default: '{DEFAULT_TITLE}'",
    )
    parser.add_argument(
        "--no-upload",
        action="store_true",
        help="Build the .xlsx locally without uploading to Google Drive.",
    )
    args = parser.parse_args(argv)

    out = build_workbook(args.output)
    if not args.no_upload:
        upload_to_folder(out, args.folder_id, args.title)
    return 0


if __name__ == "__main__":
    sys.exit(main())
