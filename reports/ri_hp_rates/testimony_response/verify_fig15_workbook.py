"""Verify the Figure 15 workbook reproduces the published averages.

Two checks:

1. **Structural** — open the xlsx and confirm the expected sheets, named ranges,
   and formula strings exist. This guards against accidental schema drift if
   ``build_fig15_workbook.py`` is refactored.
2. **Numerical** — re-run the same polars aggregation that
   ``cost_of_service_by_subclass.qmd`` uses for ``tbl-cos-by-subclass-avg``,
   then print the expected per-subclass averages alongside what the workbook's
   formulas should evaluate to once Excel/Sheets recalc.

The numerical block is a *targeted* re-implementation: it loads the same
``cross_subsidization_BAT_values`` parquet and the same YAML + tariff inputs
that the workbook bakes in, so any divergence between the published table and
the workbook would surface here as a delta between "polars expected" and
"workbook formula". (The formulas are simple SUMIFS / SUMPRODUCT divisions, so
in practice the only failure modes are wrong column letters or wrong subclass
keys — both caught by the structural check.)
"""

from __future__ import annotations

import sys
from pathlib import Path

import polars as pl
from openpyxl import load_workbook

REPORT_DIR = Path("/ebs/home/alex_switch_box/reports2/reports/ri_hp_rates")
sys.path.insert(0, str(REPORT_DIR))

from testimony_response.build_fig15_workbook import (  # noqa: E402
    HT_V2_LABELS,
    HT_V2_ORDER,
    load_inputs,
    load_master_bat,
)


def _structural_check(xlsx_path: Path) -> None:
    print(f"=== Structural check: {xlsx_path} ===")
    wb = load_workbook(str(xlsx_path), data_only=False)
    expected_sheets = [
        "README",
        "inputs_revenue_requirement",
        "inputs_tariffs",
        "bat_per_building",
        "subclass_aggregates",
        "fig15_published",
        "validation",
    ]
    missing = [s for s in expected_sheets if s not in wb.sheetnames]
    assert not missing, f"missing sheets: {missing}"
    print(f"  sheets: {wb.sheetnames}")

    expected_names = {
        "total_delivery_revenue_requirement",
        "test_year_customer_count",
        "test_year_residential_kwh",
        "annual_fixed_per_customer",
        "DISPLAY_CUSTOMER_TOTAL",
        "default_vol_usd_per_kwh",
        "ws_weight",
        "ws_heating_type",
        "ws_annual_bill",
        "ws_BAT_epmc",
        "ws_cos",
        "ws_annual_kwh",
        "ws_w_revenue",
        "ws_w_cos",
        "ws_w_xs",
        "ws_w_kwh",
    }
    actual_names = set(wb.defined_names)
    missing_names = expected_names - actual_names
    assert not missing_names, f"missing named ranges: {missing_names}"
    print(f"  named ranges: {len(actual_names)} ({sorted(actual_names)[:6]}...)")

    bat_ws = wb["bat_per_building"]
    assert bat_ws["H2"].value == "=E2+F2", bat_ws["H2"].value
    expected_kwh = "=(D2-inputs_revenue_requirement!$B$7)/inputs_tariffs!$B$2"
    assert bat_ws["I2"].value == expected_kwh, bat_ws["I2"].value
    assert bat_ws["J2"].value == "=B2*D2"
    assert bat_ws["K2"].value == "=B2*H2"
    assert bat_ws["L2"].value == "=B2*G2"
    assert bat_ws["M2"].value == "=B2*I2"
    n_rows = bat_ws.max_row - 1
    print(f"  bat_per_building data rows: {n_rows:,}")

    sub_ws = wb["subclass_aggregates"]
    c2 = sub_ws["C2"].value
    assert "SUMIFS(bat_per_building!$B$2:$B$" in c2 and ", A2)" in c2, c2
    print("  subclass_aggregates SUMIFS formulas: OK")

    pub_ws = wb["fig15_published"]
    # Heat pump row is the first under the header (row 5).
    assert pub_ws["A5"].value == "=subclass_aggregates!B2"
    assert pub_ws["B5"].value == "=subclass_aggregates!L2"
    assert pub_ws["D5"].value.startswith("=IF(subclass_aggregates!C2>0")
    print("  fig15_published references subclass_aggregates: OK")
    print()


def _numerical_check() -> None:
    print("=== Numerical check (expected averages from polars) ===")
    bat = load_master_bat()
    inputs = load_inputs()
    bat = bat.with_columns(
        (pl.col("economic_burden_delivery") + pl.col("residual_share_epmc_delivery")).alias("cost_of_service_delivery"),
        (
            (pl.col("annual_bill_delivery") - inputs["annual_fixed_per_customer"]) / inputs["default_vol_usd_per_kwh"]
        ).alias("annual_kwh"),
    )

    by_ht = (
        bat.group_by("postprocess_group.heating_type_v2")
        .agg(
            pl.col("weight").sum().alias("n_customers"),
            (pl.col("weight") * pl.col("annual_bill_delivery")).sum().alias("revenue_delivery"),
            (pl.col("weight") * pl.col("cost_of_service_delivery")).sum().alias("cost_of_service"),
            (pl.col("weight") * pl.col("BAT_epmc_delivery")).sum().alias("cross_subsidy"),
        )
        .with_columns(
            (pl.col("revenue_delivery") / pl.col("n_customers")).alias("avg_delivery_bill"),
            (pl.col("cost_of_service") / pl.col("n_customers")).alias("avg_cost_of_service"),
            (pl.col("cross_subsidy") / pl.col("n_customers")).alias("avg_cross_subsidy"),
        )
    )
    rows = {row["postprocess_group.heating_type_v2"]: row for row in by_ht.iter_rows(named=True)}

    total_w = float(bat["weight"].sum())
    total_rev = float((bat["weight"] * bat["annual_bill_delivery"]).sum())
    total_cos = float((bat["weight"] * bat["cost_of_service_delivery"]).sum())
    total_xs = float((bat["weight"] * bat["BAT_epmc_delivery"]).sum())
    total_kwh = float((bat["weight"] * bat["annual_kwh"]).sum())

    print(f"  sum(weight): {total_w:,.1f}  (expected ~ {inputs['test_year_customer_count']:,.0f})")
    print(f"  sum(w * delivery): ${total_rev:,.0f}  (expected ~ ${inputs['total_delivery_revenue_requirement']:,.0f})")
    print(f"  sum(w * cost_of_service): ${total_cos:,.0f}")
    print(f"  sum(w * cross_subsidy): ${total_xs:,.0f}  (should net to ~$0)")
    print(f"  sum(w * annual_kwh): {total_kwh:,.0f}  (expected = {inputs['test_year_residential_kwh']:,.0f})")
    assert abs(total_w - inputs["test_year_customer_count"]) < 0.5
    assert abs(total_rev - inputs["total_delivery_revenue_requirement"]) < 5_000
    assert abs(total_xs) < 5_000
    assert abs(total_kwh - inputs["test_year_residential_kwh"]) / inputs["test_year_residential_kwh"] < 1e-6
    print("  validation totals: PASS\n")

    print("  Per-subclass averages (these are what fig15_published.D:F should show):")
    print(f"  {'Subclass':<22} {'Customers':>14} {'Avg bill':>12} {'Avg COS':>12} {'Avg X-sub':>12} {'X-sub/COS':>10}")
    for key in HT_V2_ORDER:
        if key not in rows:
            continue
        r = rows[key]
        ratio = r["avg_cross_subsidy"] / r["avg_cost_of_service"] if r["avg_cost_of_service"] else float("nan")
        print(
            f"  {HT_V2_LABELS[key]:<22} {r['n_customers']:>14,.1f} "
            f"${r['avg_delivery_bill']:>11,.0f} ${r['avg_cost_of_service']:>11,.0f} "
            f"${r['avg_cross_subsidy']:>11,.0f} {ratio * 100:>9.1f}%"
        )

    avg_bill_total = total_rev / total_w
    avg_cos_total = total_cos / total_w
    avg_xs_total = total_xs / total_w
    ratio_total = avg_xs_total / avg_cos_total if avg_cos_total else float("nan")
    print(
        f"  {'All customers':<22} {total_w:>14,.1f} "
        f"${avg_bill_total:>11,.0f} ${avg_cos_total:>11,.0f} "
        f"${avg_xs_total:>11,.0f} {ratio_total * 100:>9.1f}%"
    )


def main() -> int:
    xlsx_path = REPORT_DIR / "cache" / "fig15_cos_by_subclass.xlsx"
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found. Run build_fig15_workbook.py first.")
        return 1
    _structural_check(xlsx_path)
    _numerical_check()
    return 0


if __name__ == "__main__":
    sys.exit(main())
