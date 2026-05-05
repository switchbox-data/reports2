"""Build the supporting workbook for Figure 15 (RIE COS by subclass).

Figure 15 in the Pre-Filed Direct Testimony of Juan-Pablo Velez (page 33,
``tbl-cos-by-subclass-avg`` in ``expert_testimony.qmd``) shows the average
delivery bill, cost of service, and cross-subsidy by residential heating
subclass for the test year (9/1/24 to 8/31/25).

This script reproduces every published number from per-building BAT outputs
plus a small set of revenue-requirement and tariff inputs, with **live
formulas** in every aggregation cell. The output is an ``.xlsx`` that opens
identically in Excel and Google Sheets; with ``--upload`` the same workbook is
mirrored into a target Google Sheet, preserving formulas via
``value_input_option="USER_ENTERED"``.

Run from the report directory::

    uv run python -m testimony_response.build_RIE_1_11_DIV_7_workbook --output cache/fig15_cos_by_subclass.xlsx
    uv run python -m testimony_response.build_RIE_1_11_DIV_7_workbook --upload

See ``cost_of_service_by_subclass.qmd`` for the published-side aggregation
logic that this workbook recreates with formulas instead of polars
group-by/sum.
"""

from __future__ import annotations

import argparse
import pickle
import subprocess
import sys
from pathlib import Path
from types import SimpleNamespace

import polars as pl
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName

from lib.rdp import fetch_rdp_file, parse_urdb_json

# Cross-sheet A1 references used inside formulas. We use these everywhere in
# place of named ranges, because Google Sheets does not import workbook-level
# defined names from an .xlsx pushed via the Sheets API: only cell values move,
# so any formula referencing `=ws_weight` would become `#NAME?`. Explicit A1
# references survive the upload unchanged and evaluate identically in Excel
# and Sheets. The named ranges are still defined in the .xlsx (for usability
# when opened directly in Excel) but no formula depends on them.
REF_TOTAL_RR = "inputs_revenue_requirement!$B$2"
REF_N_CUSTOMERS = "inputs_revenue_requirement!$B$3"
REF_TY_KWH = "inputs_revenue_requirement!$B$4"
REF_CUSTOMER_CHARGE = "inputs_revenue_requirement!$B$5"
REF_CORE_DELIVERY = "inputs_revenue_requirement!$B$6"
REF_ANNUAL_FIXED_PER_CUSTOMER = "inputs_revenue_requirement!$B$7"
REF_DISPLAY_TOTAL = "inputs_revenue_requirement!$B$8"
REF_SUM_WEIGHTED_EB = "inputs_revenue_requirement!$B$9"
REF_TOTAL_RESIDUAL = "inputs_revenue_requirement!$B$10"
REF_EPMC_RATE = "inputs_revenue_requirement!$B$11"
REF_DEFAULT_VOL = "inputs_tariffs!$B$2"

# Same constants as cost_of_service_by_subclass.qmd; if the testimony rebases
# onto a new batch or RDP ref, update these in lock-step with the notebook.
UTILITY = "rie"
BATCH = "ri_20260331_r1-20_rate_case_test_year"
STATE_LOWER = "ri"
S3_BASE = "s3://data.sb/switchbox/cairo/outputs/hp_rates"
PATH_MASTER_BAT_12 = f"{S3_BASE}/{STATE_LOWER}/all_utilities/{BATCH}/run_1+2/cross_subsidization_BAT_values/"
RDP_REF = "e9e5088"
RDP_REV_YAML_PATH = "rate_design/hp_rates/ri/config/rev_requirement/rie_rate_case_test_year.yaml"
RDP_TARIFF_DIR = "rate_design/hp_rates/ri/config/tariffs/electric"
RDP_GITHUB_BASE = "https://github.com/switchbox-data/rate-design-platform/blob"

FIXED_CHARGE_PER_MONTH = 10.01
FIXED_CHARGE_PER_YEAR = FIXED_CHARGE_PER_MONTH * 12

KWH_BATCH = "ri_20260504_kwh_export_v2"
_KWH_BASE = f"{S3_BASE}/{STATE_LOWER}/{UTILITY}/{KWH_BATCH}"
PATH_KWH_U0 = f"{_KWH_BASE}/20260505_011359_ri_rie_run1_up00_precalc__default/billing_kwh_annual.parquet"


def _rdp_permalink(rel_path: str) -> str:
    """SHA-pinned GitHub permalink for a rate-design-platform file."""
    return f"{RDP_GITHUB_BASE}/{RDP_REF}/{rel_path}"


REPORTS2_GITHUB_BASE = "https://github.com/switchbox-data/reports2/blob"


def _reports2_head_sha() -> str:
    """Current HEAD sha of the reports2 repo (this script's repo). Cached."""
    if not hasattr(_reports2_head_sha, "_cached"):
        repo_root = Path(__file__).resolve().parents[3]
        sha = subprocess.check_output(
            ["git", "-C", str(repo_root), "rev-parse", "HEAD"],
            text=True,
        ).strip()
        _reports2_head_sha._cached = sha  # type: ignore[attr-defined]
    return _reports2_head_sha._cached  # type: ignore[attr-defined]


def _reports2_permalink(rel_path: str, *, line_range: tuple[int, int] | None = None) -> str:
    """SHA-pinned GitHub permalink for a file in this reports2 repo.

    `rel_path` is repo-relative (e.g. ``reports/ri_hp_rates/index.qmd``).
    Optional ``line_range`` appends a ``#L<start>-L<end>`` fragment.
    """
    url = f"{REPORTS2_GITHUB_BASE}/{_reports2_head_sha()}/{rel_path}"
    if line_range is not None:
        start, end = line_range
        url += f"#L{start}-L{end}"
    return url


# Default upload target: RIE 1-11 / DIV-7 discovery response Sheet.
DEFAULT_SPREADSHEET_ID = "12uMyGBkQ5yVffmr9Xc_23Q1o9xhYe_muHsqH4NdQlw4"

HT_V2_ORDER = (
    "heat_pump",
    "electrical_resistance",
    "natgas",
    "delivered_fuels",
    "other",
)
HT_V2_LABELS: dict[str, str] = {
    "heat_pump": "Heat pump",
    "electrical_resistance": "Electric resistance",
    "natgas": "Natural gas",
    "delivered_fuels": "Delivered fuels",
    "other": "Other",
}


def load_master_bat() -> pl.DataFrame:
    """Mirror ``load_master_bat`` in ``cost_of_service_by_subclass.qmd``.

    Loads ``residual_share_epmc_delivery`` and ``BAT_epmc_delivery`` from the
    parquet for runtime validation only — these columns are derived via formulas
    in the workbook (EPMC allocation from ``economic_burden_delivery``).
    """
    df = (
        pl.scan_parquet(PATH_MASTER_BAT_12, hive_partitioning=True)
        .filter(pl.col("sb.electric_utility") == UTILITY)
        .select(
            "bldg_id",
            "weight",
            "postprocess_group.heating_type_v2",
            "annual_bill_delivery",
            "economic_burden_delivery",
            "residual_share_epmc_delivery",
            "BAT_epmc_delivery",
        )
        .collect()
    )
    assert isinstance(df, pl.DataFrame)
    return df


def _load_billing_kwh(path: str) -> pl.DataFrame:
    """Load exported billing kWh (annual grid-consumed electricity)."""
    return pl.read_parquet(path).select("bldg_id", "annual_kwh_grid")


def load_inputs() -> dict:
    """Pull revenue-requirement YAML + calibrated tariff JSONs from rate-design-platform."""
    raw_yaml = fetch_rdp_file(RDP_REV_YAML_PATH, RDP_REF)
    rev = yaml.safe_load(raw_yaml)
    total_rr = float(rev["total_delivery_revenue_requirement"])
    n_customers = float(rev["test_year_customer_count"])
    ty_kwh = float(rev["test_year_residential_kwh"])

    drr = rev["delivery_revenue_requirement"]
    customer_charge_total = float(drr["customer_charge"]["total_budget"])
    core_delivery_total = float(drr["core_delivery_rate"]["total_budget"])

    def vol_rate(rel_filename: str) -> float:
        path = f"{RDP_TARIFF_DIR}/{rel_filename}"
        doc = parse_urdb_json(fetch_rdp_file(path, RDP_REF))
        return float(doc["items"][0]["energyratestructure"][0][0]["rate"])

    default_vol = vol_rate("rie_default_calibrated.json")

    return {
        "total_delivery_revenue_requirement": total_rr,
        "test_year_customer_count": n_customers,
        "test_year_residential_kwh": ty_kwh,
        "customer_charge_total": customer_charge_total,
        "core_delivery_rate_total": core_delivery_total,
        "default_vol_usd_per_kwh": default_vol,
        "annual_fixed_per_customer": FIXED_CHARGE_PER_YEAR,
    }


def _bold(ws, cell: str) -> None:
    ws[cell].font = Font(bold=True)


def _header_fill(ws, row: int, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="E8E8E8")
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = fill


def _autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _add_named_range(wb: Workbook, name: str, sheet: str, cell: str) -> None:
    """Workbook-scoped named range so cross-sheet formulas read clearly."""
    wb.defined_names[name] = DefinedName(name=name, attr_text=f"'{sheet}'!${cell.replace('$', '').replace(':', ':$')}")


def _write_readme(wb: Workbook, inputs: dict) -> None:
    ws = wb.create_sheet("README", 0)
    s3_bat = f"{S3_BASE}/{STATE_LOWER}/all_utilities/{BATCH}/run_1+2/cross_subsidization_BAT_values/"
    rows: list[list] = [
        ["Figure 15 supporting workbook (RIE residential COS by subclass)", "", ""],
        ["", "", ""],
        ["Item", "Source", "Notes"],
        [
            "Per-building CAIRO outputs",
            s3_bat,
            (
                "One row per residential building. "
                "RIE test-year status-quo run (run_1+2 = uniform default delivery + supply for all customers)."
            ),
        ],
        [
            "Revenue-requirement YAML",
            _rdp_permalink(RDP_REV_YAML_PATH),
            "Test-year customer count, total delivery revenue requirement, test-year residential kWh, customer charge total, core delivery rate total.",
        ],
        [
            "Calibrated default tariff JSON",
            _rdp_permalink(f"{RDP_TARIFF_DIR}/rie_default_calibrated.json"),
            "energyratestructure[0][0].rate is the default uniform $/kWh.",
        ],
        [
            "Notebook that produces the published table",
            _reports2_permalink("reports/ri_hp_rates/notebooks/cost_of_service_by_subclass.qmd"),
            "Cell tbl-cos-by-subclass-avg. This workbook reproduces its aggregation logic as live formulas.",
        ],
        ["", "", ""],
        ["Sheet", "What it contains", ""],
        [
            "inputs_revenue_requirement",
            (
                "Test-year customer count, total delivery revenue requirement, test-year residential kWh, "
                "customer charge total, core delivery rate total, derived annual fixed delivery $/customer, "
                "sum_weighted_eb (sum of weight * economic_burden_delivery), "
                "total_residual (= total_RR - sum_weighted_eb), "
                "and epmc_rate (= total_residual / sum_weighted_eb)."
            ),
            "",
        ],
        [
            "inputs_tariffs",
            "Calibrated default volumetric delivery $/kWh.",
            "",
        ],
        [
            "bat_per_building",
            (
                "One row per residential building. Data columns: bldg_id, weight, heating_type_v2, "
                "annual_bill_delivery, economic_burden_delivery (from CAIRO), and annual_kwh "
                "(CAIRO billing kWh export: grid-consumed kWh after PV clipping and kwh_scale_factor). "
                "Formula columns: residual_share_epmc_delivery (= EB * epmc_rate), "
                "BAT_epmc_delivery (= bill - COS), cost_of_service_delivery (= EB + residual_share), "
                "bill_formula (= kwh * vol + fixed, verifies bill), bill_residual (= bill - bill_formula), "
                "and weighted_* products for aggregation."
            ),
            "",
        ],
        [
            "subclass_aggregates",
            "Five subclass rows + 'All customers' total via live SUMIFS / SUMPRODUCT over bat_per_building.",
            "",
        ],
        [
            "fig15_published",
            "Final published layout: customers, % of customers, average delivery bill, average cost of service, average cross-subsidy, and avg cross-subsidy / avg COS.",
            "",
        ],
        ["", "", ""],
        ["fig15_published column", "Formula (descriptive; live values are in fig15_published rows 5-10)", ""],
        [
            "Customers",
            "subclass_aggregates!n_customers_display (largest-remainder rounding to test_year_customer_count).",
            "",
        ],
        ["% of customers", "n_customers_display / DISPLAY_CUSTOMER_TOTAL.", ""],
        [
            "Avg. delivery bill",
            "SUMPRODUCT(weight, annual_bill_delivery) / SUMIFS(weight, ...) for the subclass.",
            "",
        ],
        [
            "Avg. cost of service",
            "SUMPRODUCT(weight, cost_of_service_delivery) / SUMIFS(weight, ...). COS = EB * epmc_multiplier.",
            "",
        ],
        [
            "Avg. cross-subsidy",
            "SUMPRODUCT(weight, BAT_epmc_delivery) / SUMIFS(weight, ...). BAT = bill - COS.",
            "",
        ],
        [
            "Avg. cross-subsidy / Avg. COS",
            "avg_cross_subsidy / avg_cost_of_service.",
            "",
        ],
        ["", "", ""],
        ["Key inputs (also live in inputs_revenue_requirement / inputs_tariffs)", "Value", "Source"],
        [
            "total_delivery_revenue_requirement ($)",
            inputs["total_delivery_revenue_requirement"],
            (
                "Rhode Island Energy, Application for Approval of a Change in Electric and Gas "
                "Base Distribution Rates, Docket 25-45-GE, PRB-1-ELEC exhibit, p. 14, lines 8-9, "
                "columns f & m."
            ),
        ],
        [
            "test_year_customer_count",
            inputs["test_year_customer_count"],
            ("PRB-1-ELEC exhibit, p. 14, lines 8-9, column d. 5,032,174 bills / 12 = 419,347.83 customers."),
        ],
        [
            "test_year_residential_kwh",
            inputs["test_year_residential_kwh"],
            "PRB-1-ELEC exhibit, p. 14, lines 8-9, column k.",
        ],
        [
            "default_vol_usd_per_kwh",
            inputs["default_vol_usd_per_kwh"],
            _rdp_permalink(f"{RDP_TARIFF_DIR}/rie_default_calibrated.json"),
        ],
        [
            "annual_fixed_per_customer ($)",
            inputs["annual_fixed_per_customer"],
            (
                f"Calibrated tariff: fixedchargefirstmeter = ${FIXED_CHARGE_PER_MONTH}/mo x 12. "
                "Customer charge ($6.00) + RE Growth ($3.22) + LIHEAP Enhancement ($0.79). "
                "Matches rie_default_calibrated.json."
            ),
        ],
        [
            "billing kWh (upgrade 0)",
            PATH_KWH_U0,
            (
                "CAIRO billing pipeline export: annual grid-consumed kWh per building "
                "(max(electricity_net, 0) per hour, after kwh_scale_factor and 48h timeshift)."
            ),
        ],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    # Section-header rows: adjust if rows above change.
    for header_row in (3, 9, 16, 24):
        _header_fill(ws, header_row, 3)
    for label_row in range(25, 31):
        _bold(ws, f"A{label_row}")
    _autosize(ws, {"A": 42, "B": 70, "C": 80})
    ws.sheet_view.showGridLines = False


def _write_inputs_revenue_requirement(wb: Workbook, inputs: dict) -> None:
    ws = wb.create_sheet("inputs_revenue_requirement")
    yaml_ref = _rdp_permalink(RDP_REV_YAML_PATH)
    rows = [
        ["key", "value", "source", "notes"],
        [
            "total_delivery_revenue_requirement",
            inputs["total_delivery_revenue_requirement"],
            yaml_ref,
            "YAML field: total_delivery_revenue_requirement. PRB-1-ELEC exhibit, p. 14, lines 8-9, columns f & m. Total RIE delivery revenue requirement for the test year ($).",
        ],
        [
            "test_year_customer_count",
            inputs["test_year_customer_count"],
            yaml_ref,
            "YAML field: test_year_customer_count. PRB-1-ELEC exhibit, p. 14, lines 8-9, column d. Total RIE residential customers in the test year.",
        ],
        [
            "test_year_residential_kwh",
            inputs["test_year_residential_kwh"],
            yaml_ref,
            "YAML field: test_year_residential_kwh. PRB-1-ELEC exhibit, p. 14, lines 8-9, column k. Total RIE residential delivered kWh in the test year.",
        ],
        [
            "customer_charge_total",
            inputs["customer_charge_total"],
            yaml_ref,
            "YAML field: delivery_revenue_requirement.customer_charge.total_budget. Portion of revenue requirement recovered through fixed customer charges.",
        ],
        [
            "core_delivery_rate_total",
            inputs["core_delivery_rate_total"],
            yaml_ref,
            "YAML field: delivery_revenue_requirement.core_delivery_rate.total_budget. Portion of revenue requirement recovered through volumetric delivery rates.",
        ],
        [
            "annual_fixed_per_customer",
            FIXED_CHARGE_PER_YEAR,
            _rdp_permalink(f"{RDP_TARIFF_DIR}/rie_default_calibrated.json"),
            f"Calibrated tariff: fixedchargefirstmeter = ${FIXED_CHARGE_PER_MONTH}/mo x 12 = ${FIXED_CHARGE_PER_YEAR}/yr. "
            "Customer charge ($6.00) + RE Growth ($3.22) + LIHEAP Enhancement ($0.79).",
        ],
        [
            "DISPLAY_CUSTOMER_TOTAL",
            f"=ROUND({REF_N_CUSTOMERS}, 0)",
            "Derived in this workbook",
            "Integer total used for largest-remainder customer display rounding.",
        ],
        [
            "sum_weighted_eb",
            inputs["sum_weighted_eb"],
            "Computed from per-building data",
            "Sum of (weight * economic_burden_delivery) across all buildings. Total weighted marginal cost recovered if every customer paid exactly their economic burden.",
        ],
        [
            "total_residual",
            f"={REF_TOTAL_RR} - {REF_SUM_WEIGHTED_EB}",
            "Derived in this workbook",
            "Revenue requirement not recovered by marginal costs alone. This residual pool is allocated back to each building proportionally to their economic burden (EPMC method).",
        ],
        [
            "epmc_rate",
            f"={REF_TOTAL_RESIDUAL} / {REF_SUM_WEIGHTED_EB}",
            "Derived in this workbook",
            "Residual allocation rate ($/$ of EB). Each building's residual share = economic_burden * epmc_rate. Mirrors CAIRO: epmc_rate = Residual Costs / sum(EB * weight).",
        ],
    ]
    for r in rows:
        ws.append(r)
    _header_fill(ws, 1, 4)
    _autosize(ws, {"A": 36, "B": 22, "C": 70, "D": 70})
    for row, name in [
        (2, "total_delivery_revenue_requirement"),
        (3, "test_year_customer_count"),
        (4, "test_year_residential_kwh"),
        (5, "customer_charge_total"),
        (6, "core_delivery_rate_total"),
        (7, "annual_fixed_per_customer"),
        (8, "DISPLAY_CUSTOMER_TOTAL"),
        (9, "sum_weighted_eb"),
        (10, "total_residual"),
        (11, "epmc_rate"),
    ]:
        wb.defined_names[name] = DefinedName(
            name=name,
            attr_text=f"inputs_revenue_requirement!$B${row}",
        )
    ws.sheet_view.showGridLines = False


def _write_inputs_tariffs(wb: Workbook, inputs: dict) -> None:
    ws = wb.create_sheet("inputs_tariffs")
    rows = [
        ["key", "value", "source", "notes"],
        [
            "default_vol_usd_per_kwh",
            inputs["default_vol_usd_per_kwh"],
            _rdp_permalink(f"{RDP_TARIFF_DIR}/rie_default_calibrated.json"),
            "Field: energyratestructure[0][0].rate. Status-quo uniform default delivery $/kWh.",
        ],
    ]
    for r in rows:
        ws.append(r)
    _header_fill(ws, 1, 4)
    _autosize(ws, {"A": 32, "B": 18, "C": 80, "D": 70})
    for row, name in [
        (2, "default_vol_usd_per_kwh"),
    ]:
        wb.defined_names[name] = DefinedName(name=name, attr_text=f"inputs_tariffs!$B${row}")
    ws.sheet_view.showGridLines = False


def _write_bat_per_building(wb: Workbook, bat: pl.DataFrame) -> int:
    """Write per-building rows with EPMC formula columns. Returns last data row."""
    ws = wb.create_sheet("bat_per_building")
    headers = [
        "bldg_id",  # A
        "weight",  # B
        "heating_type_v2",  # C
        "annual_bill_delivery",  # D  DATA
        "economic_burden_delivery",  # E  DATA
        "residual_share_epmc_delivery",  # F  FORMULA
        "BAT_epmc_delivery",  # G  FORMULA
        "cost_of_service_delivery",  # H  FORMULA
        "annual_kwh",  # I  DATA (CAIRO export)
        "bill_formula",  # J  FORMULA (kwh * vol + fixed)
        "bill_residual",  # K  FORMULA (bill - bill_formula)
        "weighted_bill",  # L  FORMULA
        "weighted_cos",  # M  FORMULA
        "weighted_cross_subsidy",  # N  FORMULA
        "weighted_kwh",  # O  FORMULA
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    n = bat.height
    rows = list(
        bat.select(
            "bldg_id",
            "weight",
            "postprocess_group.heating_type_v2",
            "annual_bill_delivery",
            "economic_burden_delivery",
            "annual_kwh",
        ).iter_rows()
    )
    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row[0])  # A: bldg_id
        ws.cell(row=i, column=2, value=float(row[1]))  # B: weight
        ws.cell(row=i, column=3, value=row[2])  # C: heating_type_v2
        ws.cell(row=i, column=4, value=float(row[3]))  # D: annual_bill_delivery
        ws.cell(row=i, column=5, value=float(row[4]))  # E: economic_burden_delivery
        ws.cell(row=i, column=6, value=f"=E{i}*{REF_EPMC_RATE}")  # F: residual_share
        ws.cell(row=i, column=7, value=f"=D{i}-(E{i}+F{i})")  # G: BAT_epmc
        ws.cell(row=i, column=8, value=f"=E{i}+F{i}")  # H: COS
        ws.cell(row=i, column=9, value=float(row[5]))  # I: annual_kwh
        ws.cell(row=i, column=10, value=f"=I{i}*{REF_DEFAULT_VOL}+{REF_ANNUAL_FIXED_PER_CUSTOMER}")  # J: bill_formula
        ws.cell(row=i, column=11, value=f"=D{i}-J{i}")  # K: bill_residual
        ws.cell(row=i, column=12, value=f"=B{i}*D{i}")  # L: weighted_bill
        ws.cell(row=i, column=13, value=f"=B{i}*H{i}")  # M: weighted_cos
        ws.cell(row=i, column=14, value=f"=B{i}*G{i}")  # N: weighted_cross_subsidy
        ws.cell(row=i, column=15, value=f"=B{i}*I{i}")  # O: weighted_kwh

    last_row = 1 + n
    _autosize(
        ws,
        {
            "A": 10,
            "B": 10,
            "C": 22,
            "D": 18,
            "E": 22,
            "F": 26,
            "G": 18,
            "H": 22,
            "I": 14,
            "J": 16,
            "K": 16,
            "L": 16,
            "M": 16,
            "N": 22,
            "O": 16,
        },
    )

    col_to_name = {
        "B": "ws_weight",
        "C": "ws_heating_type",
        "D": "ws_annual_bill",
        "G": "ws_BAT_epmc",
        "H": "ws_cos",
        "I": "ws_annual_kwh",
        "L": "ws_weighted_bill",
        "M": "ws_weighted_cos",
        "N": "ws_weighted_cross_subsidy",
        "O": "ws_weighted_kwh",
    }
    for col, name in col_to_name.items():
        wb.defined_names[name] = DefinedName(
            name=name,
            attr_text=f"bat_per_building!${col}$2:${col}${last_row}",
        )
    return last_row


def _write_subclass_aggregates(wb: Workbook, last_bat_row: int) -> None:
    ws = wb.create_sheet("subclass_aggregates")
    headers = [
        "subclass_key",
        "subclass",
        "n_customers",
        "revenue_delivery",
        "cost_of_service",
        "cross_subsidy",
        "total_kwh",
        "raw_display_count",
        "floor_display_count",
        "remainder",
        "rank_remainder",
        "n_customers_display",
        "pct_customers_display",
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))

    n_sub = len(HT_V2_ORDER)

    # Per-column ranges into bat_per_building. Using explicit A1 references
    # rather than named ranges so formulas survive the gspread upload.
    last = last_bat_row
    rng_weight = f"bat_per_building!$B$2:$B${last}"
    rng_heating = f"bat_per_building!$C$2:$C${last}"
    rng_w_revenue = f"bat_per_building!$L$2:$L${last}"
    rng_w_cos = f"bat_per_building!$M$2:$M${last}"
    rng_w_xs = f"bat_per_building!$N$2:$N${last}"
    rng_w_kwh = f"bat_per_building!$O$2:$O${last}"
    sub_last = 1 + n_sub  # Last subclass data row index.

    # Subclass rows.
    for idx, key in enumerate(HT_V2_ORDER):
        row = idx + 2
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=HT_V2_LABELS[key])
        ws.cell(row=row, column=3, value=f"=SUMIFS({rng_weight}, {rng_heating}, A{row})")
        ws.cell(row=row, column=4, value=f"=SUMIFS({rng_w_revenue}, {rng_heating}, A{row})")
        ws.cell(row=row, column=5, value=f"=SUMIFS({rng_w_cos}, {rng_heating}, A{row})")
        ws.cell(row=row, column=6, value=f"=SUMIFS({rng_w_xs}, {rng_heating}, A{row})")
        ws.cell(row=row, column=7, value=f"=SUMIFS({rng_w_kwh}, {rng_heating}, A{row})")
        # Largest-remainder rounding so the integer customer counts sum to
        # DISPLAY_CUSTOMER_TOTAL exactly. Helpers are visible so the math is
        # auditable in the workbook.
        ws.cell(
            row=row,
            column=8,
            value=f"=C{row}*{REF_DISPLAY_TOTAL}/SUM($C$2:$C${sub_last})",
        )
        ws.cell(row=row, column=9, value=f"=INT(H{row})")
        ws.cell(row=row, column=10, value=f"=H{row}-I{row}")
        ws.cell(row=row, column=11, value=f"=RANK(J{row},$J$2:$J${sub_last})")
        ws.cell(
            row=row,
            column=12,
            value=f"=I{row}+IF(K{row}<={REF_DISPLAY_TOTAL}-SUM($I$2:$I${sub_last}),1,0)",
        )
        ws.cell(row=row, column=13, value=f"=L{row}/{REF_DISPLAY_TOTAL}")

    # Total row ("All customers"): direct sums over per-building columns.
    total_row = 2 + n_sub
    ws.cell(row=total_row, column=1, value="all_customers")
    ws.cell(row=total_row, column=2, value="All customers")
    ws.cell(row=total_row, column=3, value=f"=SUM({rng_weight})")
    ws.cell(row=total_row, column=4, value=f"=SUM({rng_w_revenue})")
    ws.cell(row=total_row, column=5, value=f"=SUM({rng_w_cos})")
    ws.cell(row=total_row, column=6, value=f"=SUM({rng_w_xs})")
    ws.cell(row=total_row, column=7, value=f"=SUM({rng_w_kwh})")
    ws.cell(row=total_row, column=8, value=f"={REF_DISPLAY_TOTAL}")
    ws.cell(row=total_row, column=9, value=f"={REF_DISPLAY_TOTAL}")
    ws.cell(row=total_row, column=10, value=0)
    ws.cell(row=total_row, column=11, value="")
    ws.cell(row=total_row, column=12, value=f"={REF_DISPLAY_TOTAL}")
    ws.cell(row=total_row, column=13, value=1)

    _bold(ws, f"A{total_row}")
    _bold(ws, f"B{total_row}")

    # Number formats.
    money_cols = ("D", "E", "F")
    for c in money_cols:
        for r in range(2, total_row + 1):
            ws[f"{c}{r}"].number_format = '"$"#,##0'
    for r in range(2, total_row + 1):
        ws[f"C{r}"].number_format = "#,##0.0"
        ws[f"G{r}"].number_format = "#,##0"
        ws[f"H{r}"].number_format = "#,##0.00"
        ws[f"I{r}"].number_format = "#,##0"
        ws[f"J{r}"].number_format = "#,##0.00"
        ws[f"L{r}"].number_format = "#,##0"
        ws[f"M{r}"].number_format = "0.0%"

    _autosize(
        ws,
        {
            "A": 22,
            "B": 22,
            "C": 16,
            "D": 18,
            "E": 18,
            "F": 16,
            "G": 18,
            "H": 14,
            "I": 14,
            "J": 12,
            "K": 8,
            "L": 14,
            "M": 14,
        },
    )
    ws.freeze_panes = "C2"

    # Named ranges for the published view.
    wb.defined_names["agg_first_row"] = DefinedName(name="agg_first_row", attr_text="subclass_aggregates!$A$2")
    wb.defined_names["agg_total_row"] = DefinedName(
        name="agg_total_row",
        attr_text=f"subclass_aggregates!$A${total_row}",
    )


def _write_fig15_published(wb: Workbook) -> None:
    """Final published layout, mirroring the GT in tbl-cos-by-subclass-avg."""
    ws = wb.create_sheet("fig15_published")
    title = "Average annual delivery bill, cost of service, and cross-subsidy by residential subclass"
    subtitle = (
        "RIE Test Year (9/1/2024 to 8/31/2025). Source: cost_of_service_by_subclass.qmd "
        "(tbl-cos-by-subclass-avg) and the per-building BAT outputs documented in README."
    )
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")
    ws["A2"] = subtitle
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 30

    headers = [
        "Subclass",
        "Customers",
        "% of customers",
        "Avg. delivery bill",
        "Avg. cost of service",
        "Avg. cross-subsidy",
        "Avg. cross-subsidy / Avg. COS",
    ]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    _header_fill(ws, header_row, len(headers))

    rows = [*HT_V2_ORDER, "all_customers"]
    for i, _key in enumerate(rows, start=1):
        agg_row = 1 + i  # subclass_aggregates row index for this subclass.
        out_row = header_row + i
        ws.cell(row=out_row, column=1, value=f"=subclass_aggregates!B{agg_row}")
        ws.cell(row=out_row, column=2, value=f"=subclass_aggregates!L{agg_row}")
        ws.cell(row=out_row, column=3, value=f"=subclass_aggregates!M{agg_row}")
        ws.cell(
            row=out_row,
            column=4,
            value=f"=IF(subclass_aggregates!C{agg_row}>0,subclass_aggregates!D{agg_row}/subclass_aggregates!C{agg_row},NA())",
        )
        ws.cell(
            row=out_row,
            column=5,
            value=f"=IF(subclass_aggregates!C{agg_row}>0,subclass_aggregates!E{agg_row}/subclass_aggregates!C{agg_row},NA())",
        )
        ws.cell(
            row=out_row,
            column=6,
            value=f"=IF(subclass_aggregates!C{agg_row}>0,subclass_aggregates!F{agg_row}/subclass_aggregates!C{agg_row},NA())",
        )
        ws.cell(row=out_row, column=7, value=f"=IF(E{out_row}>0,F{out_row}/E{out_row},NA())")

    # Number formats matching GT layout (currency $0, percent 0.0%).
    n_rows = len(rows)
    last_data_row = header_row + n_rows
    for r in range(header_row + 1, last_data_row + 1):
        ws[f"B{r}"].number_format = "#,##0"
        ws[f"C{r}"].number_format = "0.0%"
        for c in ("D", "E", "F"):
            ws[f"{c}{r}"].number_format = '"$"#,##0'
        ws[f"G{r}"].number_format = "0.0%"

    # Bold the All customers row (last).
    for c in range(1, 8):
        ws.cell(row=last_data_row, column=c).font = Font(bold=True)

    # Centered column headers and body.
    for c in range(2, 8):
        ws.cell(row=header_row, column=c).alignment = Alignment(horizontal="center")
        for r in range(header_row + 1, last_data_row + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    _autosize(
        ws,
        {
            "A": 24,
            "B": 14,
            "C": 16,
            "D": 20,
            "E": 22,
            "F": 22,
            "G": 26,
        },
    )
    ws.sheet_view.showGridLines = False


def _validate_against_published(bat: pl.DataFrame, inputs: dict) -> None:
    """Assert that weighted aggregates match published Figure 15 and EPMC formulas."""
    import numpy as np

    total_w = float(bat["weight"].sum())
    total_rev = float((bat["weight"] * bat["annual_bill_delivery"]).sum())

    # --- Aggregate totals ---
    assert abs(total_w - inputs["test_year_customer_count"]) < 0.5, (
        f"sum(weight) = {total_w:,.2f} vs test_year_customer_count = {inputs['test_year_customer_count']:,.2f}"
    )
    assert abs(total_rev - inputs["total_delivery_revenue_requirement"]) < 2_000, (
        f"sum(w*bill) = ${total_rev:,.0f} vs total_delivery_RR = ${inputs['total_delivery_revenue_requirement']:,.0f}"
    )
    print("  Aggregate totals: PASS", flush=True)

    # --- Bill formula verification ---
    # annual_kwh * vol_rate + annual_fixed should exactly reproduce annual_bill_delivery.
    bill_check = bat["annual_kwh"] * inputs["default_vol_usd_per_kwh"] + FIXED_CHARGE_PER_YEAR
    bill_err = (bat["annual_bill_delivery"] - bill_check).abs()
    max_bill_err = float(bill_err.max())
    assert max_bill_err < 0.01, f"Bill formula: max per-building error = ${max_bill_err:.6f} (tol = $0.01)"
    print(f"  Bill formula: PASS (max err = ${max_bill_err:.2e})", flush=True)

    # --- EPMC formula validation ---
    # Verify that the EPMC formula (EB * (RR/sum_w_EB - 1)) reproduces the
    # parquet's residual_share_epmc_delivery per building.
    sum_weighted_eb = float((bat["weight"] * bat["economic_burden_delivery"]).sum())
    epmc_mult = inputs["total_delivery_revenue_requirement"] / sum_weighted_eb
    derived_residual = bat["economic_burden_delivery"] * (epmc_mult - 1)
    parquet_residual = bat["residual_share_epmc_delivery"]
    max_residual_err = float((derived_residual - parquet_residual).abs().max())
    assert max_residual_err < 0.01, f"EPMC residual: max per-building error = {max_residual_err:.6f} (tol = 0.01)"

    derived_cos = bat["economic_burden_delivery"] + derived_residual
    derived_bat = bat["annual_bill_delivery"] - derived_cos
    parquet_bat = bat["BAT_epmc_delivery"]
    max_bat_err = float((derived_bat - parquet_bat).abs().max())
    assert max_bat_err < 0.01, f"EPMC BAT: max per-building error = {max_bat_err:.6f} (tol = 0.01)"
    print(
        f"  EPMC formula: PASS (residual max err = {max_residual_err:.2e}, BAT max err = {max_bat_err:.2e})",
        flush=True,
    )

    # --- Published Figure 15 values ---
    pkl_path = Path(__file__).resolve().parents[1] / "cache" / "report_variables_cos_subclass.pkl"
    assert pkl_path.exists(), f"Missing {pkl_path}. Render cost_of_service_by_subclass.qmd first."
    v = SimpleNamespace(**pickle.loads(pkl_path.read_bytes()))

    # Per-subclass total COS, revenue, cross-subsidy.
    bat_v = bat.with_columns(
        derived_cos.alias("_derived_cos"),
        derived_bat.alias("_derived_bat"),
    )
    by_ht = bat_v.group_by("postprocess_group.heating_type_v2").agg(
        pl.col("weight").sum().alias("n_customers"),
        (pl.col("weight") * pl.col("annual_bill_delivery")).sum().alias("revenue_delivery"),
        (pl.col("weight") * pl.col("annual_kwh")).sum().alias("total_kwh"),
        (pl.col("weight") * pl.col("_derived_cos")).sum().alias("cost_of_service"),
        (pl.col("weight") * pl.col("_derived_bat")).sum().alias("cross_subsidy"),
    )
    actual = {row["postprocess_group.heating_type_v2"]: row for row in by_ht.iter_rows(named=True)}

    pub_map = {"hp": "heat_pump", "ng": "natgas", "df": "delivered_fuels", "er": "electrical_resistance"}
    for short, key in pub_map.items():
        pub_cos = getattr(v, f"cos_default_{short}_group_cos")
        pub_rev = getattr(v, f"cos_default_{short}_group_rev")
        pub_xs = getattr(v, f"cos_default_{short}_group_xs")
        a = actual[key]
        label = HT_V2_LABELS[key]
        assert abs(a["cost_of_service"] - pub_cos) < 500, (
            f"{label}: COS ${a['cost_of_service']:,.0f} != published ${pub_cos:,.0f}"
        )
        assert abs(a["revenue_delivery"] - pub_rev) < 500, (
            f"{label}: revenue ${a['revenue_delivery']:,.0f} != published ${pub_rev:,.0f}"
        )
        assert abs(a["cross_subsidy"] - pub_xs) < 500, (
            f"{label}: cross-subsidy ${a['cross_subsidy']:,.0f} != published ${pub_xs:,.0f}"
        )

    # Customer display counts (largest-remainder rounding).
    published_rows: list[dict] = v.testimony_subclass_delivery_rows
    display_total = round(inputs["test_year_customer_count"])
    sub_weights = [actual[k]["n_customers"] for k in HT_V2_ORDER]
    raw_counts = np.array(sub_weights) * display_total / total_w
    floors = np.floor(raw_counts).astype(np.int64)
    remainder = display_total - int(floors.sum())
    order = np.argsort(-(raw_counts - floors))
    for k in range(remainder):
        floors[order[k]] += 1

    for pub_row in published_rows:
        sub_name = pub_row["subclass"]
        if sub_name == "All customers":
            continue
        key = next(k for k, v_label in HT_V2_LABELS.items() if v_label == sub_name)
        idx = list(HT_V2_ORDER).index(key)
        wb_customers = int(floors[idx])
        pub_customers = int(pub_row["n_customers_display"])
        assert wb_customers == pub_customers, f"{sub_name}: customers {wb_customers} != published {pub_customers}"

    print("  Figure 15 replication: PASS (all subclass values match published)", flush=True)


def build_workbook(output_path: Path) -> Path:
    """Build and save the .xlsx workbook. Returns the output path."""
    print(f"Loading per-building BAT from {PATH_MASTER_BAT_12} ...", flush=True)
    bat = load_master_bat()
    print(f"  {bat.height:,} rows", flush=True)

    print("Loading revenue-requirement YAML and tariff JSONs from rate-design-platform ...", flush=True)
    inputs = load_inputs()
    print(f"  total_delivery_revenue_requirement = ${inputs['total_delivery_revenue_requirement']:,.0f}", flush=True)
    print(f"  test_year_customer_count = {inputs['test_year_customer_count']:,.0f}", flush=True)
    print(f"  test_year_residential_kwh = {inputs['test_year_residential_kwh']:,.0f}", flush=True)
    print(f"  default_vol_usd_per_kwh = {inputs['default_vol_usd_per_kwh']:.6f}", flush=True)
    print(f"  annual_fixed_per_customer = ${inputs['annual_fixed_per_customer']:,.2f}", flush=True)

    print(f"Loading billing kWh from {PATH_KWH_U0} ...", flush=True)
    kwh = _load_billing_kwh(PATH_KWH_U0)
    print(f"  {kwh.height:,} buildings, median {kwh['annual_kwh_grid'].median():,.0f} kWh", flush=True)

    bat = bat.join(
        kwh.select("bldg_id", pl.col("annual_kwh_grid").alias("annual_kwh")),
        on="bldg_id",
        how="left",
    )
    assert bat["annual_kwh"].null_count() == 0, "Some buildings missing billing kWh"

    inputs["sum_weighted_eb"] = float((bat["weight"] * bat["economic_burden_delivery"]).sum())
    print(f"  sum_weighted_eb = ${inputs['sum_weighted_eb']:,.0f}", flush=True)

    print("Validating against published Figure 15 and EPMC formulas ...", flush=True)
    _validate_against_published(bat, inputs)

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    # Sheet creation order is also the upload order. Put inputs_tariffs before
    # inputs_revenue_requirement so that cross-sheet references (e.g.
    # bat_per_building formulas referencing inputs_tariffs!$B$2) resolve.
    _write_readme(wb, inputs)
    _write_inputs_tariffs(wb, inputs)
    _write_inputs_revenue_requirement(wb, inputs)
    last_bat_row = _write_bat_per_building(wb, bat)
    _write_subclass_aggregates(wb, last_bat_row)
    _write_fig15_published(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Wrote {output_path} ({output_path.stat().st_size / 1024:.1f} KB)", flush=True)
    return output_path


_TAB_FORMATTING: dict[str, dict] = {
    "README": {
        "wrap_columns": ["A:C"],
        "column_widths_px": {"A": 280, "B": 480, "C": 480},
        "freeze_rows": 1,
        "bold_header": True,
        # Section-header rows inside the README; must stay in sync with the row
        # offsets used by _header_fill in _write_readme.
        "bold_rows": [3, 9, 16, 24],
    },
    "inputs_revenue_requirement": {
        "column_number_formats": {"B": "#,##0.00"},
        "wrap_columns": ["C:D"],
        "column_widths_px": {"A": 240, "B": 140, "C": 480, "D": 480},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "inputs_tariffs": {
        # Volumetric rates are sub-cent precision; 4 dp keeps them readable.
        "column_number_formats": {"B": "0.0000"},
        "wrap_columns": ["C:D"],
        "column_widths_px": {"A": 240, "B": 130, "C": 520, "D": 480},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "bat_per_building": {
        "column_number_formats": {
            "B": "#,##0.00",
            "D": '"$"#,##0.00',
            "E": '"$"#,##0.00',
            "F": '"$"#,##0.00',
            "G": '"$"#,##0.00',
            "H": '"$"#,##0.00',
            "I": "#,##0.00",
            "J": '"$"#,##0.00',
            "K": '"$"#,##0.00',
            "L": "#,##0.00",
            "M": "#,##0.00",
            "N": "#,##0.00",
            "O": "#,##0.00",
        },
        "auto_resize_columns": ["A:O"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "subclass_aggregates": {
        "column_number_formats": {
            "C": "#,##0.00",
            "D": '"$"#,##0.00',
            "E": '"$"#,##0.00',
            "F": '"$"#,##0.00',
            "G": "#,##0.00",
            "H": "#,##0.00",
            "I": "#,##0",
            "J": "#,##0.00",
            "K": "#,##0",
            "L": "#,##0",
            "M": "0.0%",
        },
        "auto_resize_columns": ["A:M"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "fig15_published": {
        "column_number_formats": {
            "B": "#,##0",
            "C": "0.0%",
            "D": '"$"#,##0',
            "E": '"$"#,##0',
            "F": '"$"#,##0',
            "G": "0.0%",
        },
        "auto_resize_columns": ["A:G"],
        "freeze_rows": 4,
        "bold_header": True,
    },
}


def upload_to_sheet(xlsx_path: Path, spreadsheet_id: str) -> None:
    """Mirror the workbook into the target Google Sheet, preserving formulas."""
    from lib.data.gsheets import apply_sheet_formatting, xlsx_to_gsheet

    print(f"Uploading {xlsx_path} -> Google Sheet {spreadsheet_id} ...", flush=True)
    # Remove any pre-existing tabs (e.g. stale `Sheet1`) so the discovery
    # response shows exactly the workbook contents and nothing else.
    spreadsheet = xlsx_to_gsheet(xlsx_path, spreadsheet_id, delete_other_tabs=True)
    print("Applying number / wrap / width formatting ...", flush=True)
    for ws in spreadsheet.worksheets():
        spec = _TAB_FORMATTING.get(ws.title)
        if spec:
            apply_sheet_formatting(ws, **spec)
    print(
        f"Done. View at https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit",
        flush=True,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0] if __doc__ else "")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("cache/fig15_cos_by_subclass.xlsx"),
        help="Output .xlsx path (relative to the report directory). Default: cache/fig15_cos_by_subclass.xlsx",
    )
    parser.add_argument(
        "--upload",
        action="store_true",
        help="Upload to the default Google Sheet (RIE 1-11/DIV-7) after building.",
    )
    parser.add_argument(
        "--spreadsheet-id",
        default=DEFAULT_SPREADSHEET_ID,
        help=f"Override the upload target Sheet id. Default: {DEFAULT_SPREADSHEET_ID}",
    )
    args = parser.parse_args(argv)

    out = build_workbook(args.output)
    if args.upload:
        upload_to_sheet(out, args.spreadsheet_id)
    return 0


if __name__ == "__main__":
    sys.exit(main())
