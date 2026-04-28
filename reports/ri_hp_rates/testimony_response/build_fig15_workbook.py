"""Build the supporting workbook for Figure 15 (RIE COS by subclass).

Figure 15 in the Pre-Filed Direct Testimony of Juan-Pablo Velez (page 33,
``tbl-cos-by-subclass-avg`` in ``expert_testimony.qmd``) shows the average
delivery bill, cost of service, and cross-subsidy by residential heating
subclass for the test year (9/1/24 to 8/31/25).

This script reproduces every published number from per-building BAT outputs
plus a small set of revenue-requirement and tariff inputs, with **live
formulas** in every aggregation cell. The output is an ``.xlsx`` that opens
identically in Excel and Google Sheets; with ``--upload`` the same workbook is
mirrored into a target Google Sheet, preserving formulas via
``value_input_option="USER_ENTERED"``.

Run from the report directory::

    uv run python -m testimony_response.build_fig15_workbook --output cache/fig15_cos_by_subclass.xlsx
    uv run python -m testimony_response.build_fig15_workbook --upload

See ``cost_of_service_by_subclass.qmd`` for the published-side aggregation
logic that this workbook recreates with formulas instead of polars
group-by/sum.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import polars as pl
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName

from lib.rdp import fetch_rdp_file, parse_urdb_json

# Cross-sheet A1 references used inside formulas. We use these everywhere in
# place of named ranges, because Google Sheets does not import workbook-level
# defined names from an .xlsx pushed via the Sheets API: only cell values move,
# so any formula referencing `=ws_weight` would become `#NAME?`. Explicit A1
# references survive the upload unchanged and evaluate identically in Excel
# and Sheets. The named ranges are still defined in the .xlsx (for usability
# when opened directly in Excel) but no formula depends on them.
REF_TOTAL_RR = "inputs_revenue_requirement!$B$2"
REF_N_CUSTOMERS = "inputs_revenue_requirement!$B$3"
REF_TY_KWH = "inputs_revenue_requirement!$B$4"
REF_CUSTOMER_CHARGE = "inputs_revenue_requirement!$B$5"
REF_CORE_DELIVERY = "inputs_revenue_requirement!$B$6"
REF_ANNUAL_FIXED_PER_CUSTOMER = "inputs_revenue_requirement!$B$7"
REF_DISPLAY_TOTAL = "inputs_revenue_requirement!$B$8"
REF_DEFAULT_VOL = "inputs_tariffs!$B$2"

# Same constants as cost_of_service_by_subclass.qmd; if the testimony rebases
# onto a new batch or RDP ref, update these in lock-step with the notebook.
UTILITY = "rie"
BATCH = "ri_20260331_r1-20_rate_case_test_year"
STATE_LOWER = "ri"
S3_BASE = "s3://data.sb/switchbox/cairo/outputs/hp_rates"
PATH_MASTER_BAT_12 = f"{S3_BASE}/{STATE_LOWER}/all_utilities/{BATCH}/run_1+2/cross_subsidization_BAT_values/"
RDP_REF = "e9e5088"
RDP_REV_YAML_PATH = "rate_design/hp_rates/ri/config/rev_requirement/rie_rate_case_test_year.yaml"
RDP_TARIFF_DIR = "rate_design/hp_rates/ri/config/tariffs/electric"

# Default upload target: RIE 1-11 / DIV-7 discovery response Sheet.
DEFAULT_SPREADSHEET_ID = "12uMyGBkQ5yVffmr9Xc_23Q1o9xhYe_muHsqH4NdQlw4"

HT_V2_ORDER = (
    "heat_pump",
    "electrical_resistance",
    "natgas",
    "delivered_fuels",
    "other",
)
HT_V2_LABELS: dict[str, str] = {
    "heat_pump": "Heat pump",
    "electrical_resistance": "Electric resistance",
    "natgas": "Natural gas",
    "delivered_fuels": "Delivered fuels",
    "other": "Other",
}


def load_master_bat() -> pl.DataFrame:
    """Mirror ``load_master_bat`` in ``cost_of_service_by_subclass.qmd``."""
    df = (
        pl.scan_parquet(PATH_MASTER_BAT_12, hive_partitioning=True)
        .filter(pl.col("sb.electric_utility") == UTILITY)
        .select(
            "bldg_id",
            "weight",
            "postprocess_group.heating_type_v2",
            "annual_bill_delivery",
            "economic_burden_delivery",
            "residual_share_epmc_delivery",
            "BAT_epmc_delivery",
        )
        .collect()
    )
    assert isinstance(df, pl.DataFrame)
    return df


def load_inputs() -> dict:
    """Pull revenue-requirement YAML + calibrated tariff JSONs from rate-design-platform."""
    raw_yaml = fetch_rdp_file(RDP_REV_YAML_PATH, RDP_REF)
    rev = yaml.safe_load(raw_yaml)
    total_rr = float(rev["total_delivery_revenue_requirement"])
    n_customers = float(rev["test_year_customer_count"])
    ty_kwh = float(rev["test_year_residential_kwh"])

    drr = rev["delivery_revenue_requirement"]
    customer_charge_total = float(drr["customer_charge"]["total_budget"])
    core_delivery_total = float(drr["core_delivery_rate"]["total_budget"])

    def vol_rate(rel_filename: str) -> float:
        path = f"{RDP_TARIFF_DIR}/{rel_filename}"
        doc = parse_urdb_json(fetch_rdp_file(path, RDP_REF))
        return float(doc["items"][0]["energyratestructure"][0][0]["rate"])

    default_vol = vol_rate("rie_default_calibrated.json")
    hp_flat_vol = vol_rate("rie_hp_flat_calibrated.json")
    nonhp_default_vol = vol_rate("rie_nonhp_default_calibrated.json")

    annual_fixed_per_customer = (total_rr - default_vol * ty_kwh) / n_customers

    return {
        "total_delivery_revenue_requirement": total_rr,
        "test_year_customer_count": n_customers,
        "test_year_residential_kwh": ty_kwh,
        "customer_charge_total": customer_charge_total,
        "core_delivery_rate_total": core_delivery_total,
        "default_vol_usd_per_kwh": default_vol,
        "hp_flat_vol_usd_per_kwh": hp_flat_vol,
        "nonhp_default_vol_usd_per_kwh": nonhp_default_vol,
        "annual_fixed_per_customer": annual_fixed_per_customer,
    }


def _bold(ws, cell: str) -> None:
    ws[cell].font = Font(bold=True)


def _header_fill(ws, row: int, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="E8E8E8")
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = fill


def _autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _add_named_range(wb: Workbook, name: str, sheet: str, cell: str) -> None:
    """Workbook-scoped named range so cross-sheet formulas read clearly."""
    wb.defined_names[name] = DefinedName(name=name, attr_text=f"'{sheet}'!${cell.replace('$', '').replace(':', ':$')}")


def _write_readme(wb: Workbook, inputs: dict) -> None:
    ws = wb.create_sheet("README", 0)
    s3_bat = f"{S3_BASE}/{STATE_LOWER}/all_utilities/{BATCH}/run_1+2/cross_subsidization_BAT_values/"
    rdp = f"rate-design-platform@{RDP_REF}"
    rows: list[list] = [
        ["Figure 15 supporting workbook (RIE residential COS by subclass)", "", ""],
        ["", "", ""],
        ["Item", "Source", "Notes"],
        [
            "Per-building BAT outputs",
            s3_bat,
            "CAIRO batch outputs (one row per RIE residential building) used as the per-building basis for all weighted aggregates.",
        ],
        [
            "CAIRO batch",
            BATCH,
            "RIE test-year status-quo batch (run_1+2 = uniform default delivery + supply for all customers).",
        ],
        [
            "Revenue-requirement YAML",
            f"{rdp}: {RDP_REV_YAML_PATH}",
            "Test-year customer count, total delivery revenue requirement, test-year residential kWh, customer charge total, core delivery rate total.",
        ],
        [
            "Calibrated default tariff JSON",
            f"{rdp}: {RDP_TARIFF_DIR}/rie_default_calibrated.json",
            "energyratestructure[0][0].rate is the default uniform $/kWh used to back out annual kWh per building from delivery bill.",
        ],
        [
            "Calibrated HP-flat tariff JSON",
            f"{rdp}: {RDP_TARIFF_DIR}/rie_hp_flat_calibrated.json",
            "Reference only. Figure 15 is computed under the status-quo uniform default rate, not the HP-flat rate.",
        ],
        [
            "Calibrated non-HP-default tariff JSON",
            f"{rdp}: {RDP_TARIFF_DIR}/rie_nonhp_default_calibrated.json",
            "Reference only. Companion to HP-flat; included for context.",
        ],
        [
            "Notebook that produces the published table",
            "reports/ri_hp_rates/notebooks/cost_of_service_by_subclass.qmd",
            "Cell tbl-cos-by-subclass-avg. This workbook reproduces its aggregation logic as live formulas.",
        ],
        [
            "Testimony embed",
            "reports/ri_hp_rates/expert_testimony.qmd lines 577-579",
            "Quarto embed shortcode that pulls tbl-cos-by-subclass-avg into the testimony as Figure 15.",
        ],
        ["", "", ""],
        ["Sheet", "What it contains", ""],
        [
            "inputs_revenue_requirement",
            "Test-year customer count, total delivery revenue requirement, test-year residential kWh, customer charge total, core delivery rate total, derived annual fixed delivery $/customer.",
            "",
        ],
        [
            "inputs_tariffs",
            "Calibrated default volumetric delivery $/kWh (also HP-flat and non-HP-default for context). Used to derive per-building annual kWh from the delivery bill.",
            "",
        ],
        [
            "bat_per_building",
            "One row per RIE residential building. Columns from CAIRO BAT parquet, plus formula columns: cost_of_service_delivery, annual_kwh, weighted aggregates.",
            "",
        ],
        [
            "subclass_aggregates",
            "Five subclass rows + 'All customers' total via live SUMIFS / SUMPRODUCT over bat_per_building.",
            "",
        ],
        [
            "fig15_published",
            "Final published layout: customers, % of customers, average delivery bill, average cost of service, average cross-subsidy, and avg cross-subsidy / avg COS.",
            "",
        ],
        [
            "validation",
            "Formula-level checks that mirror the polars asserts in the notebook (sum of weights, total revenue, cross-subsidy nets to zero, derived total kWh).",
            "",
        ],
        ["", "", ""],
        ["fig15_published column", "Formula (descriptive; live values are in fig15_published rows 5-10)", ""],
        [
            "Customers",
            "subclass_aggregates!n_customers_display (largest-remainder rounding to test_year_customer_count).",
            "",
        ],
        ["% of customers", "n_customers_display / DISPLAY_CUSTOMER_TOTAL.", ""],
        [
            "Avg. delivery bill",
            "SUMPRODUCT(weight, annual_bill_delivery) / SUMIFS(weight, ...) for the subclass.",
            "",
        ],
        [
            "Avg. cost of service",
            "SUMPRODUCT(weight, economic_burden_delivery + residual_share_epmc_delivery) / SUMIFS(weight, ...).",
            "",
        ],
        [
            "Avg. cross-subsidy",
            "SUMPRODUCT(weight, BAT_epmc_delivery) / SUMIFS(weight, ...).",
            "",
        ],
        [
            "Avg. cross-subsidy / Avg. COS",
            "avg_cross_subsidy / avg_cost_of_service.",
            "",
        ],
        ["", "", ""],
        ["Non-goal", "", ""],
        [
            "Per-building EB / EPMC residual / BAT_epmc reconstruction",
            (
                "Out of scope. These are produced upstream by CAIRO from ResStock hourly loads + marginal "
                "cost shapes; reconstructing them in spreadsheet formulas across ~131k buildings x 8760 hours "
                "is infeasible. See rate-design-platform/context/methods/bat_mc_residual/epmc_and_supply_allocation.md "
                "and utils/post/build_master_bat.py for the upstream methodology."
            ),
            "",
        ],
        ["", "", ""],
        ["Key inputs (also live in inputs_revenue_requirement / inputs_tariffs)", "Value", ""],
        ["total_delivery_revenue_requirement ($)", inputs["total_delivery_revenue_requirement"], ""],
        ["test_year_customer_count", inputs["test_year_customer_count"], ""],
        ["test_year_residential_kwh", inputs["test_year_residential_kwh"], ""],
        ["default_vol_usd_per_kwh", inputs["default_vol_usd_per_kwh"], ""],
        ["annual_fixed_per_customer ($)", inputs["annual_fixed_per_customer"], ""],
    ]
    for r in rows:
        ws.append(r)
    _bold(ws, "A1")
    ws["A1"].font = Font(bold=True, size=14)
    for header_row in (3, 13, 21, 29, 32):
        _header_fill(ws, header_row, 3)
    _autosize(ws, {"A": 42, "B": 70, "C": 80})
    ws.sheet_view.showGridLines = False


def _write_inputs_revenue_requirement(wb: Workbook, inputs: dict) -> None:
    ws = wb.create_sheet("inputs_revenue_requirement")
    yaml_ref = f"rate-design-platform@{RDP_REF}: {RDP_REV_YAML_PATH}"
    rows = [
        ["key", "value", "source", "notes"],
        [
            "total_delivery_revenue_requirement",
            inputs["total_delivery_revenue_requirement"],
            f"{yaml_ref} -> total_delivery_revenue_requirement",
            "Total RIE delivery revenue requirement for the test year ($).",
        ],
        [
            "test_year_customer_count",
            inputs["test_year_customer_count"],
            f"{yaml_ref} -> test_year_customer_count",
            "Total RIE residential customers in the test year.",
        ],
        [
            "test_year_residential_kwh",
            inputs["test_year_residential_kwh"],
            f"{yaml_ref} -> test_year_residential_kwh",
            "Total RIE residential delivered kWh in the test year.",
        ],
        [
            "customer_charge_total",
            inputs["customer_charge_total"],
            f"{yaml_ref} -> delivery_revenue_requirement.customer_charge.total_budget",
            "Portion of revenue requirement recovered through fixed customer charges.",
        ],
        [
            "core_delivery_rate_total",
            inputs["core_delivery_rate_total"],
            f"{yaml_ref} -> delivery_revenue_requirement.core_delivery_rate.total_budget",
            "Portion of revenue requirement recovered through volumetric delivery rates.",
        ],
        [
            "annual_fixed_per_customer",
            f"=({REF_TOTAL_RR} - {REF_DEFAULT_VOL} * {REF_TY_KWH}) / {REF_N_CUSTOMERS}",
            "Derived in this workbook",
            "Used to back out annual_kwh per building from delivery bill. Mirrors cost_of_service_by_subclass.qmd line ~159.",
        ],
        [
            "DISPLAY_CUSTOMER_TOTAL",
            f"=ROUND({REF_N_CUSTOMERS}, 0)",
            "Derived in this workbook",
            "Integer total used for largest-remainder customer display rounding.",
        ],
    ]
    for r in rows:
        ws.append(r)
    _header_fill(ws, 1, 4)
    _autosize(ws, {"A": 36, "B": 22, "C": 70, "D": 70})
    # Named ranges (workbook-scoped) referencing column B.
    for row, name in [
        (2, "total_delivery_revenue_requirement"),
        (3, "test_year_customer_count"),
        (4, "test_year_residential_kwh"),
        (5, "customer_charge_total"),
        (6, "core_delivery_rate_total"),
        (7, "annual_fixed_per_customer"),
        (8, "DISPLAY_CUSTOMER_TOTAL"),
    ]:
        wb.defined_names[name] = DefinedName(
            name=name,
            attr_text=f"inputs_revenue_requirement!$B${row}",
        )
    ws.sheet_view.showGridLines = False


def _write_inputs_tariffs(wb: Workbook, inputs: dict) -> None:
    ws = wb.create_sheet("inputs_tariffs")
    rdp = f"rate-design-platform@{RDP_REF}"
    rows = [
        ["key", "value", "source", "notes"],
        [
            "default_vol_usd_per_kwh",
            inputs["default_vol_usd_per_kwh"],
            f"{rdp}: {RDP_TARIFF_DIR}/rie_default_calibrated.json -> energyratestructure[0][0].rate",
            "Status-quo uniform default delivery $/kWh. Used in the BAT pipeline to back out annual kWh per building from the delivery bill.",
        ],
        [
            "hp_flat_vol_usd_per_kwh",
            inputs["hp_flat_vol_usd_per_kwh"],
            f"{rdp}: {RDP_TARIFF_DIR}/rie_hp_flat_calibrated.json -> energyratestructure[0][0].rate",
            "Reference only. Figure 15 is computed under the default uniform rate, not the HP-flat rate.",
        ],
        [
            "nonhp_default_vol_usd_per_kwh",
            inputs["nonhp_default_vol_usd_per_kwh"],
            f"{rdp}: {RDP_TARIFF_DIR}/rie_nonhp_default_calibrated.json -> energyratestructure[0][0].rate",
            "Reference only. Companion to HP-flat for the HP-flat scenario.",
        ],
    ]
    for r in rows:
        ws.append(r)
    _header_fill(ws, 1, 4)
    _autosize(ws, {"A": 32, "B": 18, "C": 80, "D": 70})
    for row, name in [
        (2, "default_vol_usd_per_kwh"),
        (3, "hp_flat_vol_usd_per_kwh"),
        (4, "nonhp_default_vol_usd_per_kwh"),
    ]:
        wb.defined_names[name] = DefinedName(name=name, attr_text=f"inputs_tariffs!$B${row}")
    ws.sheet_view.showGridLines = False


def _write_bat_per_building(wb: Workbook, bat: pl.DataFrame) -> int:
    """Write per-building BAT rows + formula columns. Returns last data row index."""
    ws = wb.create_sheet("bat_per_building")
    headers = [
        "bldg_id",
        "weight",
        "heating_type_v2",
        "annual_bill_delivery",
        "economic_burden_delivery",
        "residual_share_epmc_delivery",
        "BAT_epmc_delivery",
        "cost_of_service_delivery",
        "annual_kwh",
        "w_revenue",
        "w_cos",
        "w_xs",
        "w_kwh",
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    n = bat.height
    rows = list(
        bat.select(
            "bldg_id",
            "weight",
            "postprocess_group.heating_type_v2",
            "annual_bill_delivery",
            "economic_burden_delivery",
            "residual_share_epmc_delivery",
            "BAT_epmc_delivery",
        ).iter_rows()
    )
    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=float(row[1]))
        ws.cell(row=i, column=3, value=row[2])
        ws.cell(row=i, column=4, value=float(row[3]))
        ws.cell(row=i, column=5, value=float(row[4]))
        ws.cell(row=i, column=6, value=float(row[5]))
        ws.cell(row=i, column=7, value=float(row[6]))
        # cost_of_service_delivery = economic_burden + residual_share_epmc.
        ws.cell(row=i, column=8, value=f"=E{i}+F{i}")
        # annual_kwh = (annual_bill_delivery - annual_fixed_per_customer) / default_vol_usd_per_kwh.
        ws.cell(
            row=i,
            column=9,
            value=f"=(D{i}-{REF_ANNUAL_FIXED_PER_CUSTOMER})/{REF_DEFAULT_VOL}",
        )
        ws.cell(row=i, column=10, value=f"=B{i}*D{i}")
        ws.cell(row=i, column=11, value=f"=B{i}*H{i}")
        ws.cell(row=i, column=12, value=f"=B{i}*G{i}")
        ws.cell(row=i, column=13, value=f"=B{i}*I{i}")

    last_row = 1 + n
    widths = {
        "A": 10,
        "B": 10,
        "C": 22,
        "D": 18,
        "E": 22,
        "F": 22,
        "G": 18,
        "H": 22,
        "I": 14,
        "J": 16,
        "K": 16,
        "L": 14,
        "M": 16,
    }
    _autosize(ws, widths)

    # Workbook-scoped named ranges over each column for crisp aggregation formulas.
    col_to_name = {
        "B": "ws_weight",
        "C": "ws_heating_type",
        "D": "ws_annual_bill",
        "G": "ws_BAT_epmc",
        "H": "ws_cos",
        "I": "ws_annual_kwh",
        "J": "ws_w_revenue",
        "K": "ws_w_cos",
        "L": "ws_w_xs",
        "M": "ws_w_kwh",
    }
    for col, name in col_to_name.items():
        wb.defined_names[name] = DefinedName(
            name=name,
            attr_text=f"bat_per_building!${col}$2:${col}${last_row}",
        )
    return last_row


def _write_subclass_aggregates(wb: Workbook, last_bat_row: int) -> None:
    ws = wb.create_sheet("subclass_aggregates")
    headers = [
        "subclass_key",
        "subclass",
        "n_customers",
        "revenue_delivery",
        "cost_of_service",
        "cross_subsidy",
        "total_kwh",
        "raw_display_count",
        "floor_display_count",
        "remainder",
        "rank_remainder",
        "n_customers_display",
        "pct_customers_display",
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))

    n_sub = len(HT_V2_ORDER)

    # Per-column ranges into bat_per_building. Using explicit A1 references
    # rather than named ranges so formulas survive the gspread upload.
    last = last_bat_row
    rng_weight = f"bat_per_building!$B$2:$B${last}"
    rng_heating = f"bat_per_building!$C$2:$C${last}"
    rng_w_revenue = f"bat_per_building!$J$2:$J${last}"
    rng_w_cos = f"bat_per_building!$K$2:$K${last}"
    rng_w_xs = f"bat_per_building!$L$2:$L${last}"
    rng_w_kwh = f"bat_per_building!$M$2:$M${last}"
    sub_last = 1 + n_sub  # Last subclass data row index.

    # Subclass rows.
    for idx, key in enumerate(HT_V2_ORDER):
        row = idx + 2
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=HT_V2_LABELS[key])
        ws.cell(row=row, column=3, value=f"=SUMIFS({rng_weight}, {rng_heating}, A{row})")
        ws.cell(row=row, column=4, value=f"=SUMIFS({rng_w_revenue}, {rng_heating}, A{row})")
        ws.cell(row=row, column=5, value=f"=SUMIFS({rng_w_cos}, {rng_heating}, A{row})")
        ws.cell(row=row, column=6, value=f"=SUMIFS({rng_w_xs}, {rng_heating}, A{row})")
        ws.cell(row=row, column=7, value=f"=SUMIFS({rng_w_kwh}, {rng_heating}, A{row})")
        # Largest-remainder rounding so the integer customer counts sum to
        # DISPLAY_CUSTOMER_TOTAL exactly. Helpers are visible so the math is
        # auditable in the workbook.
        ws.cell(
            row=row,
            column=8,
            value=f"=C{row}*{REF_DISPLAY_TOTAL}/SUM($C$2:$C${sub_last})",
        )
        ws.cell(row=row, column=9, value=f"=INT(H{row})")
        ws.cell(row=row, column=10, value=f"=H{row}-I{row}")
        ws.cell(row=row, column=11, value=f"=RANK(J{row},$J$2:$J${sub_last})")
        ws.cell(
            row=row,
            column=12,
            value=f"=I{row}+IF(K{row}<={REF_DISPLAY_TOTAL}-SUM($I$2:$I${sub_last}),1,0)",
        )
        ws.cell(row=row, column=13, value=f"=L{row}/{REF_DISPLAY_TOTAL}")

    # Total row ("All customers"): direct sums over per-building columns.
    total_row = 2 + n_sub
    ws.cell(row=total_row, column=1, value="all_customers")
    ws.cell(row=total_row, column=2, value="All customers")
    ws.cell(row=total_row, column=3, value=f"=SUM({rng_weight})")
    ws.cell(row=total_row, column=4, value=f"=SUM({rng_w_revenue})")
    ws.cell(row=total_row, column=5, value=f"=SUM({rng_w_cos})")
    ws.cell(row=total_row, column=6, value=f"=SUM({rng_w_xs})")
    ws.cell(row=total_row, column=7, value=f"=SUM({rng_w_kwh})")
    ws.cell(row=total_row, column=8, value=f"={REF_DISPLAY_TOTAL}")
    ws.cell(row=total_row, column=9, value=f"={REF_DISPLAY_TOTAL}")
    ws.cell(row=total_row, column=10, value=0)
    ws.cell(row=total_row, column=11, value="")
    ws.cell(row=total_row, column=12, value=f"={REF_DISPLAY_TOTAL}")
    ws.cell(row=total_row, column=13, value=1)

    _bold(ws, f"A{total_row}")
    _bold(ws, f"B{total_row}")

    # Number formats.
    money_cols = ("D", "E", "F")
    for c in money_cols:
        for r in range(2, total_row + 1):
            ws[f"{c}{r}"].number_format = '"$"#,##0'
    for r in range(2, total_row + 1):
        ws[f"C{r}"].number_format = "#,##0.0"
        ws[f"G{r}"].number_format = "#,##0"
        ws[f"H{r}"].number_format = "#,##0.00"
        ws[f"I{r}"].number_format = "#,##0"
        ws[f"J{r}"].number_format = "#,##0.00"
        ws[f"L{r}"].number_format = "#,##0"
        ws[f"M{r}"].number_format = "0.0%"

    _autosize(
        ws,
        {
            "A": 22,
            "B": 22,
            "C": 16,
            "D": 18,
            "E": 18,
            "F": 16,
            "G": 18,
            "H": 14,
            "I": 14,
            "J": 12,
            "K": 8,
            "L": 14,
            "M": 14,
        },
    )
    ws.freeze_panes = "C2"

    # Named ranges for the published view.
    wb.defined_names["agg_first_row"] = DefinedName(name="agg_first_row", attr_text="subclass_aggregates!$A$2")
    wb.defined_names["agg_total_row"] = DefinedName(
        name="agg_total_row",
        attr_text=f"subclass_aggregates!$A${total_row}",
    )


def _write_fig15_published(wb: Workbook) -> None:
    """Final published layout, mirroring the GT in tbl-cos-by-subclass-avg."""
    ws = wb.create_sheet("fig15_published")
    title = "Average annual delivery bill, cost of service, and cross-subsidy by residential subclass"
    subtitle = (
        "RIE Test Year (9/1/2024 to 8/31/2025). Source: cost_of_service_by_subclass.qmd "
        "(tbl-cos-by-subclass-avg) and the per-building BAT outputs documented in README."
    )
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")
    ws["A2"] = subtitle
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 30

    headers = [
        "Subclass",
        "Customers",
        "% of customers",
        "Avg. delivery bill",
        "Avg. cost of service",
        "Avg. cross-subsidy",
        "Avg. cross-subsidy / Avg. COS",
    ]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    _header_fill(ws, header_row, len(headers))

    rows = [*HT_V2_ORDER, "all_customers"]
    for i, _key in enumerate(rows, start=1):
        agg_row = 1 + i  # subclass_aggregates row index for this subclass.
        out_row = header_row + i
        ws.cell(row=out_row, column=1, value=f"=subclass_aggregates!B{agg_row}")
        ws.cell(row=out_row, column=2, value=f"=subclass_aggregates!L{agg_row}")
        ws.cell(row=out_row, column=3, value=f"=subclass_aggregates!M{agg_row}")
        ws.cell(
            row=out_row,
            column=4,
            value=f"=IF(subclass_aggregates!C{agg_row}>0,subclass_aggregates!D{agg_row}/subclass_aggregates!C{agg_row},NA())",
        )
        ws.cell(
            row=out_row,
            column=5,
            value=f"=IF(subclass_aggregates!C{agg_row}>0,subclass_aggregates!E{agg_row}/subclass_aggregates!C{agg_row},NA())",
        )
        ws.cell(
            row=out_row,
            column=6,
            value=f"=IF(subclass_aggregates!C{agg_row}>0,subclass_aggregates!F{agg_row}/subclass_aggregates!C{agg_row},NA())",
        )
        ws.cell(row=out_row, column=7, value=f"=IF(E{out_row}>0,F{out_row}/E{out_row},NA())")

    # Number formats matching GT layout (currency $0, percent 0.0%).
    n_rows = len(rows)
    last_data_row = header_row + n_rows
    for r in range(header_row + 1, last_data_row + 1):
        ws[f"B{r}"].number_format = "#,##0"
        ws[f"C{r}"].number_format = "0.0%"
        for c in ("D", "E", "F"):
            ws[f"{c}{r}"].number_format = '"$"#,##0'
        ws[f"G{r}"].number_format = "0.0%"

    # Bold the All customers row (last).
    for c in range(1, 8):
        ws.cell(row=last_data_row, column=c).font = Font(bold=True)

    # Centered column headers and body.
    for c in range(2, 8):
        ws.cell(row=header_row, column=c).alignment = Alignment(horizontal="center")
        for r in range(header_row + 1, last_data_row + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    _autosize(
        ws,
        {
            "A": 24,
            "B": 14,
            "C": 16,
            "D": 20,
            "E": 22,
            "F": 22,
            "G": 26,
        },
    )
    ws.sheet_view.showGridLines = False


def _write_validation(wb: Workbook, last_bat_row: int) -> None:
    ws = wb.create_sheet("validation")
    headers = ["check", "actual", "expected", "abs_error", "tolerance", "ok"]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))

    last = last_bat_row
    rng_weight = f"bat_per_building!$B$2:$B${last}"
    rng_bill = f"bat_per_building!$D$2:$D${last}"
    rng_xs = f"bat_per_building!$G$2:$G${last}"
    rng_cos = f"bat_per_building!$H$2:$H${last}"
    rng_kwh = f"bat_per_building!$I$2:$I${last}"

    rows = [
        (
            "sum(weight) approx test_year_customer_count",
            f"=SUM({rng_weight})",
            f"={REF_N_CUSTOMERS}",
            None,
            0.05,
        ),
        (
            "sum(weight x delivery_bill) approx total_delivery_revenue_requirement",
            f"=SUMPRODUCT({rng_weight}, {rng_bill})",
            f"={REF_TOTAL_RR}",
            None,
            2000.0,
        ),
        (
            "sum(weight x BAT_epmc) approx 0 (cross-subsidy nets to zero)",
            f"=SUMPRODUCT({rng_weight}, {rng_xs})",
            "=0",
            None,
            5000.0,
        ),
        (
            "sum(weight x cost_of_service) approx total_delivery_revenue_requirement",
            f"=SUMPRODUCT({rng_weight}, {rng_cos})",
            f"={REF_TOTAL_RR}",
            None,
            5000.0,
        ),
        (
            "sum(weight x annual_kwh) approx test_year_residential_kwh",
            f"=SUMPRODUCT({rng_weight}, {rng_kwh})",
            f"={REF_TY_KWH}",
            None,
            1.0,
        ),
    ]
    for i, (name, actual, expected, _err, tol) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=actual)
        ws.cell(row=i, column=3, value=expected)
        ws.cell(row=i, column=4, value=f"=ABS(B{i}-C{i})")
        ws.cell(row=i, column=5, value=tol)
        ws.cell(row=i, column=6, value=f'=IF(D{i}<=E{i}, "OK", "FAIL")')

    _autosize(ws, {"A": 70, "B": 22, "C": 22, "D": 16, "E": 14, "F": 8})
    for r in range(2, 2 + len(rows)):
        ws[f"B{r}"].number_format = "#,##0.00"
        ws[f"C{r}"].number_format = "#,##0.00"
        ws[f"D{r}"].number_format = "#,##0.00"
    ws.sheet_view.showGridLines = False


def build_workbook(output_path: Path) -> Path:
    """Build and save the .xlsx workbook. Returns the output path."""
    print(f"Loading per-building BAT from {PATH_MASTER_BAT_12} ...", flush=True)
    bat = load_master_bat()
    print(f"  {bat.height:,} rows", flush=True)

    print("Loading revenue-requirement YAML and tariff JSONs from rate-design-platform ...", flush=True)
    inputs = load_inputs()
    print(f"  total_delivery_revenue_requirement = ${inputs['total_delivery_revenue_requirement']:,.0f}", flush=True)
    print(f"  test_year_customer_count = {inputs['test_year_customer_count']:,.0f}", flush=True)
    print(f"  test_year_residential_kwh = {inputs['test_year_residential_kwh']:,.0f}", flush=True)
    print(f"  default_vol_usd_per_kwh = {inputs['default_vol_usd_per_kwh']:.6f}", flush=True)
    print(f"  annual_fixed_per_customer = ${inputs['annual_fixed_per_customer']:,.2f}", flush=True)

    wb = Workbook()
    # Remove the default empty sheet; we re-create README at index 0.
    default = wb.active
    if default is not None:
        wb.remove(default)

    # Sheet creation order is also the upload order. Put inputs_tariffs before
    # inputs_revenue_requirement so that the latter's annual_fixed_per_customer
    # formula (which references inputs_tariffs!$B$2) resolves at upload time
    # rather than caching an "Unresolved sheet name" error in Sheets.
    _write_readme(wb, inputs)
    _write_inputs_tariffs(wb, inputs)
    _write_inputs_revenue_requirement(wb, inputs)
    last_bat_row = _write_bat_per_building(wb, bat)
    _write_subclass_aggregates(wb, last_bat_row)
    _write_fig15_published(wb)
    _write_validation(wb, last_bat_row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Wrote {output_path} ({output_path.stat().st_size / 1024:.1f} KB)", flush=True)
    return output_path


_TAB_FORMATTING: dict[str, dict] = {
    "README": {
        "wrap_columns": ["A:C"],
        "column_widths_px": {"A": 280, "B": 480, "C": 480},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "inputs_revenue_requirement": {
        "column_number_formats": {"B": "#,##0.00"},
        "wrap_columns": ["C:D"],
        "column_widths_px": {"A": 240, "B": 140, "C": 480, "D": 480},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "inputs_tariffs": {
        # Volumetric rates are sub-cent precision; 4 dp keeps them readable.
        "column_number_formats": {"B": "0.0000"},
        "wrap_columns": ["C:D"],
        "column_widths_px": {"A": 240, "B": 130, "C": 520, "D": 480},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "bat_per_building": {
        "column_number_formats": {
            "B": "#,##0.00",
            "D": '"$"#,##0.00',
            "E": '"$"#,##0.00',
            "F": '"$"#,##0.00',
            "G": '"$"#,##0.00',
            "H": '"$"#,##0.00',
            "I": "#,##0.00",
            "J": "#,##0.00",
            "K": "#,##0.00",
            "L": "#,##0.00",
            "M": "#,##0.00",
        },
        "auto_resize_columns": ["A:M"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "subclass_aggregates": {
        "column_number_formats": {
            "C": "#,##0.00",
            "D": '"$"#,##0.00',
            "E": '"$"#,##0.00',
            "F": '"$"#,##0.00',
            "G": "#,##0.00",
            "H": "#,##0.00",
            "I": "#,##0",
            "J": "#,##0.00",
            "K": "#,##0",
            "L": "#,##0",
            "M": "0.0%",
        },
        "auto_resize_columns": ["A:M"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "fig15_published": {
        "column_number_formats": {
            "B": "#,##0",
            "C": "0.0%",
            "D": '"$"#,##0',
            "E": '"$"#,##0',
            "F": '"$"#,##0',
            "G": "0.0%",
        },
        "auto_resize_columns": ["A:G"],
        "freeze_rows": 4,
        "bold_header": True,
    },
    "validation": {
        "column_number_formats": {"B": "#,##0.00", "C": "#,##0.00", "D": "#,##0.00"},
        "auto_resize_columns": ["A:F"],
        "freeze_rows": 1,
        "bold_header": True,
    },
}


def upload_to_sheet(xlsx_path: Path, spreadsheet_id: str) -> None:
    """Mirror the workbook into the target Google Sheet, preserving formulas."""
    from lib.data.gsheets import apply_sheet_formatting, xlsx_to_gsheet

    print(f"Uploading {xlsx_path} -> Google Sheet {spreadsheet_id} ...", flush=True)
    # Remove any pre-existing tabs (e.g. stale `Sheet1`) so the discovery
    # response shows exactly the workbook contents and nothing else.
    spreadsheet = xlsx_to_gsheet(xlsx_path, spreadsheet_id, delete_other_tabs=True)
    print("Applying number / wrap / width formatting ...", flush=True)
    for ws in spreadsheet.worksheets():
        spec = _TAB_FORMATTING.get(ws.title)
        if spec:
            apply_sheet_formatting(ws, **spec)
    print(
        f"Done. View at https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit",
        flush=True,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0] if __doc__ else "")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("cache/fig15_cos_by_subclass.xlsx"),
        help="Output .xlsx path (relative to the report directory). Default: cache/fig15_cos_by_subclass.xlsx",
    )
    parser.add_argument(
        "--upload",
        action="store_true",
        help="Upload to the default Google Sheet (RIE 1-11/DIV-7) after building.",
    )
    parser.add_argument(
        "--spreadsheet-id",
        default=DEFAULT_SPREADSHEET_ID,
        help=f"Override the upload target Sheet id. Default: {DEFAULT_SPREADSHEET_ID}",
    )
    args = parser.parse_args(argv)

    out = build_workbook(args.output)
    if args.upload:
        upload_to_sheet(out, args.spreadsheet_id)
    return 0


if __name__ == "__main__":
    sys.exit(main())
