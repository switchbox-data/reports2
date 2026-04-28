"""Google Sheets client with cached OAuth credentials.

Includes generic helpers for writing workbooks (lists of lists, polars/pandas
DataFrames, or full ``openpyxl`` workbooks) into a target Sheet while preserving
formulas via ``value_input_option="USER_ENTERED"``. Used by discovery-response
scripts that build an ``.xlsx`` locally and then mirror it to a shared Sheet.
"""

from __future__ import annotations

import json
import os
from pathlib import Path
from typing import TYPE_CHECKING, Any

import gspread
from dotenv import load_dotenv

if TYPE_CHECKING:
    import pandas as pd
    import polars as pl


def _gspread_token_path() -> Path:
    """Path for cached OAuth token; reuse to avoid browser prompt on every run."""
    if os.name == "nt":
        base = Path(os.environ.get("APPDATA", Path.home() / "AppData" / "Roaming"))
    else:
        base = Path.home() / ".config"
    return base / "gspread" / "authorized_user_reports2.json"


def load_cached_token() -> dict | None:
    """Load cached gspread OAuth token from disk, or None if missing or invalid."""
    token_path = _gspread_token_path()
    if not token_path.exists():
        return None
    try:
        return json.loads(token_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def save_cached_token(authorized_user: dict | str) -> None:
    """Persist gspread authorized_user token so the next run can skip the browser."""
    token_path = _gspread_token_path()
    token_path.parent.mkdir(parents=True, exist_ok=True)
    if isinstance(authorized_user, dict):
        token_path.write_text(json.dumps(authorized_user, indent=2), encoding="utf-8")
    else:
        token_path.write_text(str(authorized_user), encoding="utf-8")


def get_gspread_client():
    """Return a gspread client using env credentials and cached token when possible.

    Loads G_* env vars (optionally from .env via load_dotenv()). Uses a cached
    token at ~/.config/gspread/authorized_user_reports2.json when present so
    the browser is only opened on first run or when the token is expired.

    Returns:
        Tuple of (gc, authorized_user) where gc is the gspread Client and
        authorized_user is the token dict/str returned by gspread (for persistence).
    """
    load_dotenv()
    app_creds = {
        "installed": {
            "client_id": os.getenv("G_CLIENT_ID"),
            "client_secret": os.getenv("G_CLIENT_SECRET"),
            "project_id": os.getenv("G_PROJECT_ID"),
            "auth_uri": os.getenv("G_AUTH_URI"),
            "token_uri": os.getenv("G_TOKEN_URI"),
            "redirect_uris": ["http://localhost"],
        }
    }
    saved_token = load_cached_token()
    if saved_token is not None:
        gc, authorized_user = gspread.oauth_from_dict(credentials=app_creds, authorized_user_info=saved_token)
    else:
        gc, authorized_user = gspread.oauth_from_dict(credentials=app_creds)
    if authorized_user:
        save_cached_token(authorized_user)
    return gc, authorized_user


def open_sheet_by_id(spreadsheet_id: str) -> gspread.Spreadsheet:
    """Authenticate via cached OAuth and open the target Sheet by id."""
    gc, _ = get_gspread_client()
    return gc.open_by_key(spreadsheet_id)


def upsert_worksheet(
    spreadsheet: gspread.Spreadsheet,
    title: str,
    *,
    rows: int,
    cols: int,
) -> gspread.Worksheet:
    """Replace ``title`` if present, then return a freshly created worksheet.

    A delete-and-recreate keeps the grid sized exactly right and clears any
    stale formulas/formatting from a prior upload.
    """
    try:
        existing = spreadsheet.worksheet(title)
    except gspread.exceptions.WorksheetNotFound:
        existing = None
    if existing is not None:
        spreadsheet.del_worksheet(existing)
    return spreadsheet.add_worksheet(title=title, rows=max(rows, 1), cols=max(cols, 1))


def _normalize_value(v: Any) -> Any:
    """Coerce values to types Sheets accepts; preserve formula strings as-is."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float, str)):
        return v
    return str(v)


def _normalize_rows(values: list[list[Any]]) -> list[list[Any]]:
    return [[_normalize_value(v) for v in row] for row in values]


def write_values_with_formulas(
    worksheet: gspread.Worksheet,
    values: list[list[Any]],
    *,
    start: str = "A1",
    chunk_rows: int = 5000,
) -> None:
    """Write a list-of-lists to ``worksheet`` with formula evaluation enabled.

    Cells starting with ``=`` are interpreted as formulas. Large frames are
    written in row chunks of ``chunk_rows`` to stay well under per-request
    cell limits.
    """
    if not values:
        return
    norm = _normalize_rows(values)
    n_cols = max((len(row) for row in norm), default=0)
    if n_cols == 0:
        return
    from gspread.utils import ValueInputOption, a1_to_rowcol, rowcol_to_a1

    start_r, start_c = a1_to_rowcol(start)
    end_c = start_c + n_cols - 1
    for i in range(0, len(norm), chunk_rows):
        chunk = norm[i : i + chunk_rows]
        for row in chunk:
            if len(row) < n_cols:
                row.extend([""] * (n_cols - len(row)))
        r0 = start_r + i
        r1 = r0 + len(chunk) - 1
        a1 = f"{rowcol_to_a1(r0, start_c)}:{rowcol_to_a1(r1, end_c)}"
        worksheet.update(values=chunk, range_name=a1, value_input_option=ValueInputOption.user_entered)


def write_dataframe_with_formulas(
    worksheet: gspread.Worksheet,
    df: pl.DataFrame | pd.DataFrame,
    *,
    start: str = "A1",
    include_header: bool = True,
    chunk_rows: int = 5000,
) -> None:
    """Write a polars or pandas DataFrame to ``worksheet`` with formulas honored.

    Cell values that are strings starting with ``=`` are sent as formulas (the
    underlying ``USER_ENTERED`` option). All other values pass through. The
    DataFrame should already contain the formula strings in the relevant cells.
    """
    cols = list(df.columns)
    if hasattr(df, "to_pandas"):
        rows_iter = df.iter_rows()
    else:
        rows_iter = (tuple(row) for row in df.itertuples(index=False, name=None))
    values: list[list[Any]] = []
    if include_header:
        values.append(list(cols))
    values.extend([list(row) for row in rows_iter])
    write_values_with_formulas(worksheet, values, start=start, chunk_rows=chunk_rows)


def xlsx_to_gsheet(
    xlsx_path: str | Path,
    spreadsheet_id: str,
    *,
    tab_prefix: str = "",
    only_sheets: list[str] | None = None,
    delete_other_tabs: bool = False,
) -> gspread.Spreadsheet:
    """Mirror every sheet of an openpyxl workbook into the target Google Sheet.

    For each worksheet, an existing tab with the same (prefixed) title is
    replaced. Cell values starting with ``=`` are written as live formulas;
    cross-sheet references (``=other_tab!A1``) evaluate as expected as long as
    the target tab name matches what the formulas reference.

    Args:
        xlsx_path: Local path to the ``.xlsx`` file produced by openpyxl.
        spreadsheet_id: Google Sheet id (the long string in the Sheet URL).
        tab_prefix: Optional prefix prepended to every uploaded tab name.
        only_sheets: If given, only upload sheets whose title is in this list.
        delete_other_tabs: If True, remove any pre-existing tabs in the target
            Sheet that are not part of this upload (e.g. stale ``Sheet1``).
            A spreadsheet must have at least one tab, so the cleanup keeps any
            uploaded tabs even if all are slated for deletion.

    Returns:
        The opened ``gspread.Spreadsheet``.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(xlsx_path), data_only=False, read_only=True)
    spreadsheet = open_sheet_by_id(spreadsheet_id)
    titles = list(wb.sheetnames)
    if only_sheets is not None:
        titles = [t for t in titles if t in set(only_sheets)]
    uploaded: set[str] = set()
    for title in titles:
        ws_src = wb[title]
        rows = list(ws_src.iter_rows(values_only=True))
        n_rows = len(rows)
        n_cols = max((len(r) for r in rows), default=1)
        target_title = f"{tab_prefix}{title}" if tab_prefix else title
        ws_dst = upsert_worksheet(
            spreadsheet,
            target_title,
            rows=max(n_rows, 1),
            cols=max(n_cols, 1),
        )
        write_values_with_formulas(ws_dst, [list(r) for r in rows])
        uploaded.add(target_title)
    wb.close()
    if delete_other_tabs:
        # Re-fetch worksheets after upserts so we delete the right set.
        for ws in spreadsheet.worksheets():
            if ws.title not in uploaded and len(spreadsheet.worksheets()) > 1:
                spreadsheet.del_worksheet(ws)
    return spreadsheet


def _col_letter_to_index(letter: str) -> int:
    """A->1, B->2, ..., Z->26, AA->27, ..."""
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _col_range_to_indices(rng: str) -> tuple[int, int]:
    """Convert ``"A"`` or ``"C:D"`` to inclusive 1-based (start, end) indices."""
    if ":" in rng:
        a, b = rng.split(":")
        return _col_letter_to_index(a), _col_letter_to_index(b)
    i = _col_letter_to_index(rng)
    return i, i


def apply_sheet_formatting(
    worksheet: gspread.Worksheet,
    *,
    column_number_formats: dict[str, str] | None = None,
    wrap_columns: list[str] | None = None,
    column_widths_px: dict[str, int] | None = None,
    auto_resize_columns: list[str] | None = None,
    freeze_rows: int | None = None,
    bold_header: bool = False,
) -> None:
    """Apply common visual formatting to a Google Sheets worksheet.

    All operations are issued in a single ``batch_update`` so the call uses one
    Sheets API quota unit per worksheet.

    Args:
        worksheet: Target ``gspread.Worksheet``.
        column_number_formats: ``{"D:F": "#,##0.00", "G": "0.0%"}``. Uses Excel
            number-format codes; Sheets infers ``type`` (NUMBER / PERCENT /
            CURRENCY) from the pattern.
        wrap_columns: Columns (e.g. ``["C:D"]``) on which to set
            ``wrapStrategy=WRAP``.
        column_widths_px: ``{"A": 280, "B:C": 480}`` fixed pixel widths.
        auto_resize_columns: Columns to auto-fit to content. Applied AFTER any
            fixed widths so it can override them.
        freeze_rows: Number of frozen header rows (e.g. 1).
        bold_header: If True, bold the first row.
    """
    sheet_id = worksheet.id
    requests: list[dict[str, Any]] = []

    def _col_range(rng: str) -> dict[str, Any]:
        s, e = _col_range_to_indices(rng)
        return {
            "sheetId": sheet_id,
            "dimension": "COLUMNS",
            "startIndex": s - 1,
            "endIndex": e,
        }

    if column_number_formats:
        for rng, pattern in column_number_formats.items():
            s, e = _col_range_to_indices(rng)
            ntype = "PERCENT" if "%" in pattern else "CURRENCY" if "$" in pattern or "€" in pattern else "NUMBER"
            requests.append(
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startColumnIndex": s - 1,
                            "endColumnIndex": e,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "numberFormat": {"type": ntype, "pattern": pattern},
                            },
                        },
                        "fields": "userEnteredFormat.numberFormat",
                    }
                }
            )

    if wrap_columns:
        for rng in wrap_columns:
            s, e = _col_range_to_indices(rng)
            requests.append(
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startColumnIndex": s - 1,
                            "endColumnIndex": e,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "wrapStrategy": "WRAP",
                                "verticalAlignment": "TOP",
                            },
                        },
                        "fields": "userEnteredFormat.wrapStrategy,userEnteredFormat.verticalAlignment",
                    }
                }
            )

    if column_widths_px:
        for rng, px in column_widths_px.items():
            requests.append(
                {
                    "updateDimensionProperties": {
                        "range": _col_range(rng),
                        "properties": {"pixelSize": int(px)},
                        "fields": "pixelSize",
                    }
                }
            )

    if auto_resize_columns:
        for rng in auto_resize_columns:
            requests.append({"autoResizeDimensions": {"dimensions": _col_range(rng)}})

    if freeze_rows is not None:
        requests.append(
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": sheet_id,
                        "gridProperties": {"frozenRowCount": int(freeze_rows)},
                    },
                    "fields": "gridProperties.frozenRowCount",
                }
            }
        )

    if bold_header:
        requests.append(
            {
                "repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1},
                    "cell": {"userEnteredFormat": {"textFormat": {"bold": True}}},
                    "fields": "userEnteredFormat.textFormat.bold",
                }
            }
        )

    if requests:
        worksheet.spreadsheet.batch_update({"requests": requests})
