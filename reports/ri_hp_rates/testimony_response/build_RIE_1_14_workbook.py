"""Build RIE 1-14: Subclass-level cost-of-service study workbook.

Responds to: "Please provide, in Excel with all formulas intact, the
'cost-of-service study at the subclass level' performed by Witness Velez."

The workbook decomposes total delivery revenue into marginal costs (economic
burden) and EPMC residual for each heating-type subclass, using hourly
(8760) load profiles and the same marginal cost derivation documented in
RIE 1-10 (build_marginal_costs_workbook.py --delivery-only).

Tabs
----
README             Overview and source links.
inputs_scalars     Key parameters from the RR YAML and tariff JSON.
mc_combined        8760 delivery MC (dist/sub-TX + bulk TX), same as RIE 1-10.
subclass_loads     8760 weighted kWh by heating subclass (from billing_kwh_8760).
subclass_mc_alloc  8760 MC x load allocation by subclass -- live formulas.
subclass_coss      Summary: customers, kWh, marginal cost, EPMC residual,
                   cost of service, revenue, cross-subsidy per subclass.
validation         Formula-level checks.

Usage::

    cd /ebs/home/alex_switch_box/reports2

    uv run python -m testimony_response.build_RIE_1_14_workbook \
        --output cache/rie_1_14_subclass_coss.xlsx

    uv run python -m testimony_response.build_RIE_1_14_workbook --upload
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
from pathlib import Path

import polars as pl
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────

UTILITY = "rie"
STATE = "ri"
YEAR = 2025
BATCH = "ri_20260507_r1-2_grid_cons_fix"
RUN_DIR = "20260507_213944_ri_rie_run1_up00_precalc__default"

S3_CAIRO_BASE = "s3://data.sb/switchbox/cairo/outputs/hp_rates"
PATH_BILLING_KWH_8760 = f"{S3_CAIRO_BASE}/{STATE}/{UTILITY}/{BATCH}/{RUN_DIR}/billing_kwh_8760.parquet"
PATH_MASTER_BAT_12 = f"{S3_CAIRO_BASE}/{STATE}/all_utilities/{BATCH}/run_1+2/cross_subsidization_BAT_values/"

S3_DIST_SUB_TX = (
    f"s3://data.sb/switchbox/marginal_costs/{STATE}/dist_and_sub_tx/utility={UTILITY}/year={YEAR}/data.parquet"
)
S3_BULK_TX = f"s3://data.sb/switchbox/marginal_costs/{STATE}/bulk_tx/utility={UTILITY}/year={YEAR}/data.parquet"

RDP_REF = "0b203bc"
RDP_GITHUB_BASE = "https://github.com/switchbox-data/rate-design-platform/blob"
REPORTS2_GITHUB_BASE = "https://github.com/switchbox-data/reports2/blob"

RDP_RR_YAML_PATH = "rate_design/hp_rates/ri/config/rev_requirement/rie_rate_case_test_year.yaml"
RDP_TARIFF_PATH = "rate_design/hp_rates/ri/config/tariffs/electric/rie_default_calibrated.json"

SUBCLASS_ORDER = [
    "heat_pump",
    "electrical_resistance",
    "natgas",
    "delivered_fuels",
    "other",
]
SUBCLASS_LABELS = {
    "heat_pump": "Heat pump",
    "electrical_resistance": "Electric resistance",
    "natgas": "Natural gas",
    "delivered_fuels": "Delivered fuels",
    "other": "Other",
}

DEFAULT_SPREADSHEET_ID = "1wC4tH5jrOWfuDqUz_DPBORfxb_sHDv2SYBWnHEhFbVA"

# ── Permalink helpers ─────────────────────────────────────────────────────────


def _rdp_permalink(rel_path: str) -> str:
    return f"{RDP_GITHUB_BASE}/{RDP_REF}/{rel_path}"


def _reports2_head_sha() -> str:
    if not hasattr(_reports2_head_sha, "_cached"):
        repo_root = Path(__file__).resolve().parents[3]
        sha = subprocess.check_output(
            ["git", "-C", str(repo_root), "rev-parse", "HEAD"],
            text=True,
        ).strip()
        _reports2_head_sha._cached = sha  # type: ignore[attr-defined]
    return _reports2_head_sha._cached  # type: ignore[attr-defined]


def _reports2_permalink(rel_path: str) -> str:
    return f"{REPORTS2_GITHUB_BASE}/{_reports2_head_sha()}/{rel_path}"


# ── Formatting helpers ────────────────────────────────────────────────────────

_README_BOLD_ROWS: list[int] = []


def _header_fill(ws, row: int, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="E8E8E8")
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = fill


def _autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Data loading ──────────────────────────────────────────────────────────────


def _load_mc_parquet(s3_path: str) -> pl.DataFrame:
    df = pl.read_parquet(s3_path)
    numeric_cols = [c for c in df.columns if c != "timestamp" and df[c].dtype.is_numeric() and c != "year"]
    return df.select("timestamp", numeric_cols[0]).rename({numeric_cols[0]: "mc_value"})


def load_delivery_mc() -> pl.DataFrame:
    """Load and combine the two delivery MC parquets into one 8760 DataFrame."""
    dist = _load_mc_parquet(S3_DIST_SUB_TX).rename({"mc_value": "mc_dist_sub_tx"})
    bulk = _load_mc_parquet(S3_BULK_TX).rename({"mc_value": "mc_bulk_tx"})
    combined = dist.join(bulk, on="timestamp", how="inner").sort("timestamp")
    return combined.with_columns(
        (pl.col("mc_dist_sub_tx") + pl.col("mc_bulk_tx")).alias("mc_delivery_total"),
    )


def load_billing_8760() -> pl.DataFrame:
    """Load per-building hourly grid consumption from CAIRO output."""
    return pl.read_parquet(PATH_BILLING_KWH_8760)


def load_master_bat() -> pl.DataFrame:
    """Load master BAT for run 1+2, filtered to RIE."""
    df = (
        pl.scan_parquet(PATH_MASTER_BAT_12, hive_partitioning=True)
        .filter(pl.col("sb.electric_utility") == UTILITY)
        .select(
            "bldg_id",
            "weight",
            "postprocess_group.heating_type_v2",
            "annual_bill_delivery",
            "economic_burden_delivery",
            "residual_share_epmc_delivery",
            "BAT_epmc_delivery",
        )
        .collect()
    )
    assert isinstance(df, pl.DataFrame)
    return df


def load_rr_yaml() -> dict:
    """Load the revenue requirement YAML from the RDP repo at RDP_REF."""
    from lib.rdp import fetch_rdp_file

    raw = fetch_rdp_file(RDP_RR_YAML_PATH, RDP_REF)
    return yaml.safe_load(raw)


def load_calibrated_tariff() -> dict:
    """Load the calibrated tariff JSON from the RDP repo at RDP_REF."""
    from lib.rdp import fetch_rdp_file

    raw = fetch_rdp_file(RDP_TARIFF_PATH, RDP_REF)
    return json.loads(raw)


# ── Build subclass 8760 loads ─────────────────────────────────────────────────


def build_subclass_8760(billing_8760: pl.DataFrame, bat: pl.DataFrame) -> pl.DataFrame:
    """Aggregate per-building 8760 kWh into weighted subclass totals.

    Returns an 8760-row DataFrame with columns:
        timestamp, kWh_heat_pump, kWh_electrical_resistance, kWh_natgas,
        kWh_delivered_fuels, kWh_other, kWh_total
    """
    bldg_info = bat.select("bldg_id", "weight", "postprocess_group.heating_type_v2")
    joined = billing_8760.join(bldg_info, on="bldg_id", how="inner")

    # weighted kWh per building-hour
    joined = joined.with_columns(
        (pl.col("grid_cons_kwh") * pl.col("weight")).alias("w_kwh"),
    )

    subclass_dfs = []
    for sc_key in SUBCLASS_ORDER:
        sc_agg = (
            joined.filter(pl.col("postprocess_group.heating_type_v2") == sc_key)
            .group_by("timestamp")
            .agg(pl.col("w_kwh").sum().alias(f"kWh_{sc_key}"))
        )
        subclass_dfs.append(sc_agg)

    result = subclass_dfs[0]
    for sc_df in subclass_dfs[1:]:
        result = result.join(sc_df, on="timestamp", how="full", coalesce=True)

    for sc_key in SUBCLASS_ORDER:
        col_name = f"kWh_{sc_key}"
        result = result.with_columns(pl.col(col_name).fill_null(0.0))

    kwh_cols = [f"kWh_{sc}" for sc in SUBCLASS_ORDER]
    result = result.with_columns(
        pl.sum_horizontal([pl.col(c) for c in kwh_cols]).alias("kWh_total"),
    ).sort("timestamp")

    return result


# ── Tab writers ───────────────────────────────────────────────────────────────


def _write_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README", 0)
    rows: list[list] = [
        [
            "Subclass Cost-of-Service Workbook — RIE 2025 (delivery)",
            "",
            "",
        ],
        ["", "", ""],
        ["Item", "Source", "Notes"],
        [
            "RIE 1-10 (marginal costs derivation)",
            "https://docs.google.com/spreadsheets/d/1cYGVrLmDTKm9XXEw7jjLVrSqTr8xp7pVK1o0Kk8AOkw/edit",
            "Companion workbook with full 8760 MC derivation from raw ISO-NE/EIA data.",
        ],
        [
            "Revenue requirement YAML",
            _rdp_permalink(RDP_RR_YAML_PATH),
            "rie_rate_case_test_year.yaml: customer count, delivery RR, kWh scale factor.",
        ],
        [
            "Calibrated tariff JSON",
            _rdp_permalink(RDP_TARIFF_PATH),
            "rie_default_calibrated.json: volumetric rate and fixed charge after CAIRO precalc.",
        ],
        [
            "Billing kWh 8760",
            PATH_BILLING_KWH_8760,
            "CAIRO output: per-building hourly grid consumption (post-scaling, post-PV floor).",
        ],
        [
            "Master BAT (run 1+2)",
            PATH_MASTER_BAT_12,
            "Post-processed BAT: per-building delivery bill, economic burden, EPMC residual.",
        ],
        [
            "Cost-of-service notebook",
            _reports2_permalink("reports/ri_hp_rates/notebooks/cost_of_service_by_subclass.qmd"),
            "Quarto notebook that produces the subclass COSS tables in the expert testimony.",
        ],
        [
            "This workbook builder",
            _reports2_permalink("reports/ri_hp_rates/testimony_response/build_RIE_1_14_workbook.py"),
            "Script that generated this workbook.",
        ],
        ["", "", ""],
        ["Sheet", "What it contains", ""],
        [
            "inputs_scalars",
            "Revenue requirement, customer count, tariff parameters, kWh scale factor.",
            "",
        ],
        [
            "mc_combined",
            "8760 rows: delivery MC per kWh (dist/sub-TX + bulk TX). Same data as RIE 1-10.",
            "",
        ],
        [
            "subclass_loads",
            "8760 rows: weighted grid kWh by heating subclass + total.",
            "",
        ],
        [
            "subclass_mc_alloc",
            "8760 rows: MC x load = hourly marginal cost allocation by subclass. All formulas.",
            "",
        ],
        [
            "subclass_coss",
            "Summary: customers, consumption, economic burden, EPMC residual, COS, revenue, cross-subsidy by subclass.",
            "",
        ],
        ["validation", "Formula-level checks against YAML targets and BAT totals.", ""],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    bold_rows = []
    for row_idx, row_data in enumerate(rows, start=1):
        if row_data and row_data[0] in ("Item", "Sheet"):
            _header_fill(ws, row_idx, 3)
            bold_rows.append(row_idx)
    _README_BOLD_ROWS.clear()
    _README_BOLD_ROWS.extend(bold_rows)
    _autosize(ws, {"A": 40, "B": 80, "C": 80})
    ws.sheet_view.showGridLines = False


def _write_inputs_scalars(wb: Workbook, rr_yaml: dict, tariff: dict) -> None:
    ws = wb.create_sheet("inputs_scalars")
    vol_rate = float(tariff["items"][0]["energyratestructure"][0][0]["rate"])
    fixed_monthly = float(tariff["items"][0]["fixedchargefirstmeter"])
    fixed_annual = fixed_monthly * 12.0
    total_delivery_rr = float(rr_yaml["total_delivery_revenue_requirement"])
    ty_kwh = float(rr_yaml["test_year_residential_kwh"])
    ty_customers = float(rr_yaml["test_year_customer_count"])
    kwh_sf = float(rr_yaml["resstock_kwh_scale_factor"])

    rows = [
        ["key", "value", "source"],
        ["total_delivery_revenue_requirement", total_delivery_rr, "RR YAML"],
        ["test_year_residential_kwh", ty_kwh, "RR YAML"],
        ["test_year_customer_count", ty_customers, "RR YAML"],
        ["resstock_kwh_scale_factor", kwh_sf, "RR YAML"],
        ["volumetric_rate_usd_per_kwh", vol_rate, "Calibrated tariff JSON"],
        ["fixed_charge_usd_per_month", fixed_monthly, "Calibrated tariff JSON"],
        ["fixed_charge_usd_per_year", fixed_annual, "=B7*12"],
    ]
    for r in rows:
        ws.append(r)
    # Make the annual fixed charge a formula
    ws.cell(row=9, column=2, value="=B7*12")
    ws.cell(row=9, column=3, value="Derived: monthly x 12")
    _header_fill(ws, 1, 3)
    _autosize(ws, {"A": 40, "B": 24, "C": 40})
    ws.sheet_view.showGridLines = False


def _write_mc_combined(wb: Workbook, mc_df: pl.DataFrame) -> None:
    """mc_combined: 8760 rows of delivery MC (values, not formulas — same as RIE 1-10)."""
    ws = wb.create_sheet("mc_combined")
    headers = ["timestamp", "mc_dist_sub_tx", "mc_bulk_tx", "mc_delivery_total"]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    for i, row in enumerate(mc_df.iter_rows(named=True), start=2):
        ws.cell(row=i, column=1, value=row["timestamp"].strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=i, column=2, value=float(row["mc_dist_sub_tx"]))
        ws.cell(row=i, column=3, value=float(row["mc_bulk_tx"]))
        ws.cell(row=i, column=4, value=f"=B{i}+C{i}")

    _autosize(ws, {"A": 18, "B": 18, "C": 16, "D": 20})


def _write_subclass_loads(wb: Workbook, loads_8760: pl.DataFrame) -> None:
    """subclass_loads: 8760 rows of weighted kWh by subclass."""
    ws = wb.create_sheet("subclass_loads")

    sc_labels = [SUBCLASS_LABELS[sc] for sc in SUBCLASS_ORDER]
    headers = ["timestamp", *sc_labels, "Total"]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    n_sc = len(SUBCLASS_ORDER)
    for i, row in enumerate(loads_8760.iter_rows(named=True), start=2):
        ws.cell(row=i, column=1, value=row["timestamp"].strftime("%Y-%m-%d %H:%M"))
        for j, sc_key in enumerate(SUBCLASS_ORDER):
            ws.cell(row=i, column=2 + j, value=float(row[f"kWh_{sc_key}"]))
        # Total as formula: sum of subclass columns
        first_col = get_column_letter(2)
        last_col = get_column_letter(1 + n_sc)
        ws.cell(
            row=i,
            column=2 + n_sc,
            value=f"=SUM({first_col}{i}:{last_col}{i})",
        )

    widths = {"A": 18}
    for j in range(n_sc + 1):
        widths[get_column_letter(2 + j)] = 18
    _autosize(ws, widths)


def _write_subclass_mc_alloc(wb: Workbook, n_hours: int = 8760) -> None:
    """subclass_mc_alloc: MC x load for each subclass, all as cross-sheet formulas."""
    ws = wb.create_sheet("subclass_mc_alloc")

    sc_labels = [SUBCLASS_LABELS[sc] for sc in SUBCLASS_ORDER]
    headers = ["timestamp", "mc_delivery_per_kwh", *sc_labels, "Total"]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    n_sc = len(SUBCLASS_ORDER)
    for i in range(2, n_hours + 2):
        ws.cell(row=i, column=1, value=f"=mc_combined!A{i}")
        ws.cell(row=i, column=2, value=f"=mc_combined!D{i}")
        for j in range(n_sc):
            load_col = get_column_letter(2 + j)  # col in subclass_loads
            ws.cell(
                row=i,
                column=3 + j,
                value=f"=$B{i}*subclass_loads!{load_col}{i}",
            )
        # Total = sum of subclass MC allocations
        first = get_column_letter(3)
        last = get_column_letter(2 + n_sc)
        ws.cell(
            row=i,
            column=3 + n_sc,
            value=f"=SUM({first}{i}:{last}{i})",
        )

    widths = {"A": 18, "B": 20}
    for j in range(n_sc + 1):
        widths[get_column_letter(3 + j)] = 18
    _autosize(ws, widths)


# Scalar references in inputs_scalars (row indices match _write_inputs_scalars)
REF_DELIVERY_RR = "inputs_scalars!$B$2"
REF_TY_KWH = "inputs_scalars!$B$3"
REF_TY_CUSTOMERS = "inputs_scalars!$B$4"
REF_FIXED_ANNUAL = "inputs_scalars!$B$9"
REF_VOL_RATE = "inputs_scalars!$B$6"


def _write_subclass_coss(
    wb: Workbook,
    bat: pl.DataFrame,
    rr_yaml: dict,
    tariff: dict,
) -> None:
    """Summary table: customers, kWh, MC, residual, COS, revenue, cross-subsidy."""
    ws = wb.create_sheet("subclass_coss")

    headers = [
        "Subclass",
        "Customers",
        "% of customers",
        "Consumption (kWh)",
        "% of consumption",
        "Marginal cost (economic burden)",
        "% of MC",
        "EPMC residual",
        "Cost of service (MC + residual)",
        "% of COS",
        "Delivery revenue",
        "% of revenue",
        "Cross-subsidy (revenue - COS)",
        "Cross-subsidy / COS",
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    vol_rate = float(tariff["items"][0]["energyratestructure"][0][0]["rate"])
    fixed_monthly = float(tariff["items"][0]["fixedchargefirstmeter"])
    fixed_annual = fixed_monthly * 12.0

    n_sc = len(SUBCLASS_ORDER)
    all_rows_data: list[dict] = []

    for sc_key in SUBCLASS_ORDER:
        sc_bat = bat.filter(pl.col("postprocess_group.heating_type_v2") == sc_key)
        customers = float(sc_bat["weight"].sum())
        revenue = float((sc_bat["weight"] * sc_bat["annual_bill_delivery"]).sum())
        economic_burden = float((sc_bat["weight"] * sc_bat["economic_burden_delivery"]).sum())
        epmc_residual = float((sc_bat["weight"] * sc_bat["residual_share_epmc_delivery"]).sum())
        cos = economic_burden + epmc_residual
        cross_subsidy = revenue - cos
        # kWh derived from bills: (bill - fixed_annual) / vol_rate
        consumption = float((sc_bat["weight"] * ((sc_bat["annual_bill_delivery"] - fixed_annual) / vol_rate)).sum())
        all_rows_data.append(
            {
                "label": SUBCLASS_LABELS[sc_key],
                "customers": customers,
                "consumption": consumption,
                "mc": economic_burden,
                "residual": epmc_residual,
                "cos": cos,
                "revenue": revenue,
                "cross_subsidy": cross_subsidy,
            }
        )

    # Write subclass rows as values
    for r_idx, d in enumerate(all_rows_data, start=2):
        ws.cell(row=r_idx, column=1, value=d["label"])
        ws.cell(row=r_idx, column=2, value=d["customers"])
        ws.cell(row=r_idx, column=4, value=d["consumption"])
        ws.cell(row=r_idx, column=6, value=d["mc"])
        ws.cell(row=r_idx, column=8, value=d["residual"])
        ws.cell(row=r_idx, column=11, value=d["revenue"])

    total_row = 2 + n_sc
    first_data = 2
    last_data = first_data + n_sc - 1

    # Total row as SUM formulas
    ws.cell(row=total_row, column=1, value="All customers")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    for col_idx in (2, 4, 6, 8, 11):
        cl = get_column_letter(col_idx)
        ws.cell(
            row=total_row,
            column=col_idx,
            value=f"=SUM({cl}{first_data}:{cl}{last_data})",
        )

    # Formula columns for all rows (subclass + total)
    for r in range(first_data, total_row + 1):
        total_cust_ref = f"$B${total_row}"
        total_kwh_ref = f"$D${total_row}"
        total_mc_ref = f"$F${total_row}"
        total_cos_ref = f"$I${total_row}"
        total_rev_ref = f"$K${total_row}"

        # % of customers
        ws.cell(row=r, column=3, value=f"=B{r}/{total_cust_ref}")
        # % of consumption
        ws.cell(row=r, column=5, value=f"=D{r}/{total_kwh_ref}")
        # % of MC
        ws.cell(row=r, column=7, value=f"=F{r}/{total_mc_ref}")
        # COS = MC + residual
        ws.cell(row=r, column=9, value=f"=F{r}+H{r}")
        # % of COS
        ws.cell(row=r, column=10, value=f"=I{r}/{total_cos_ref}")
        # % of revenue
        ws.cell(row=r, column=12, value=f"=K{r}/{total_rev_ref}")
        # Cross-subsidy = revenue - COS
        ws.cell(row=r, column=13, value=f"=K{r}-I{r}")
        # Cross-subsidy / COS
        ws.cell(row=r, column=14, value=f"=IF(I{r}<>0,M{r}/I{r},0)")

    # Number formatting
    for r in range(first_data, total_row + 1):
        ws.cell(row=r, column=2).number_format = "#,##0.00"
        ws.cell(row=r, column=3).number_format = "0.0%"
        ws.cell(row=r, column=4).number_format = "#,##0"
        ws.cell(row=r, column=5).number_format = "0.0%"
        ws.cell(row=r, column=6).number_format = "$#,##0"
        ws.cell(row=r, column=7).number_format = "0.0%"
        ws.cell(row=r, column=8).number_format = "$#,##0"
        ws.cell(row=r, column=9).number_format = "$#,##0"
        ws.cell(row=r, column=10).number_format = "0.0%"
        ws.cell(row=r, column=11).number_format = "$#,##0"
        ws.cell(row=r, column=12).number_format = "0.0%"
        ws.cell(row=r, column=13).number_format = "$#,##0"
        ws.cell(row=r, column=14).number_format = "0.0%"

    _autosize(
        ws,
        {
            "A": 22,
            "B": 14,
            "C": 16,
            "D": 20,
            "E": 18,
            "F": 30,
            "G": 10,
            "H": 18,
            "I": 30,
            "J": 10,
            "K": 18,
            "L": 14,
            "M": 28,
            "N": 20,
        },
    )
    ws.sheet_view.showGridLines = False


def _write_validation(
    wb: Workbook,
    bat: pl.DataFrame,
    rr_yaml: dict,
    tariff: dict,
    loads_8760: pl.DataFrame,
) -> None:
    ws = wb.create_sheet("validation")
    headers = ["check", "actual", "expected", "abs_error", "tolerance", "ok"]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))

    total_delivery_rr = float(rr_yaml["total_delivery_revenue_requirement"])
    ty_kwh = float(rr_yaml["test_year_residential_kwh"])
    ty_customers = float(rr_yaml["test_year_customer_count"])

    bat_total_w = float(bat["weight"].sum())
    bat_total_rev = float((bat["weight"] * bat["annual_bill_delivery"]).sum())
    bat_total_eb = float((bat["weight"] * bat["economic_burden_delivery"]).sum())
    bat_total_res = float((bat["weight"] * bat["residual_share_epmc_delivery"]).sum())
    bat_total_cos = bat_total_eb + bat_total_res
    bat_total_xs = bat_total_rev - bat_total_cos
    loads_total_kwh = float(loads_8760["kWh_total"].sum())

    # MC from subclass_mc_alloc total column
    n_sc = len(SUBCLASS_ORDER)
    mc_total_col = get_column_letter(3 + n_sc)

    checks = [
        (
            "sum(weight) = test_year_customer_count",
            bat_total_w,
            ty_customers,
            1.0,
        ),
        (
            "sum(w * delivery_bill) = total_delivery_RR",
            bat_total_rev,
            total_delivery_rr,
            100.0,
        ),
        (
            "sum(w * COS) ~ total_delivery_RR",
            bat_total_cos,
            total_delivery_rr,
            5000.0,
        ),
        (
            "sum(cross-subsidy) nets to ~0",
            abs(bat_total_xs),
            0.0,
            5000.0,
        ),
        (
            "8760 weighted grid kWh = test_year_residential_kwh",
            loads_total_kwh,
            ty_kwh,
            1000.0,
        ),
        (
            "subclass_mc_alloc sum = BAT economic_burden total",
            f"=SUM(subclass_mc_alloc!{mc_total_col}$2:{mc_total_col}$8761)",
            bat_total_eb,
            1000.0,
        ),
    ]

    for i, (name, actual, expected, tol) in enumerate(checks, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=actual)
        ws.cell(row=i, column=3, value=expected)
        ws.cell(row=i, column=4, value=f"=ABS(B{i}-C{i})")
        ws.cell(row=i, column=5, value=tol)
        ws.cell(row=i, column=6, value=f'=IF(D{i}<=E{i},"OK","FAIL")')

    _autosize(ws, {"A": 52, "B": 20, "C": 20, "D": 14, "E": 14, "F": 8})
    for r in range(2, 2 + len(checks)):
        ws[f"B{r}"].number_format = "#,##0.00"
        ws[f"C{r}"].number_format = "#,##0.00"
        ws[f"D{r}"].number_format = "#,##0.00"
    ws.sheet_view.showGridLines = False


# ── Main build ────────────────────────────────────────────────────────────────


def build_workbook(output_path: Path) -> Path:
    print("Building RIE 1-14 subclass COSS workbook ...", flush=True)

    print("Loading RR YAML and calibrated tariff ...", flush=True)
    rr_yaml = load_rr_yaml()
    tariff = load_calibrated_tariff()

    print("Loading delivery MCs from S3 ...", flush=True)
    mc_df = load_delivery_mc()
    print(f"  MC combined: {mc_df.height} hours", flush=True)

    print("Loading billing_kwh_8760 ...", flush=True)
    billing_8760 = load_billing_8760()
    print(f"  Billing 8760: {billing_8760.height} rows ({billing_8760.height // 8760} buildings)", flush=True)

    print("Loading master BAT (run 1+2) ...", flush=True)
    bat = load_master_bat()
    print(f"  BAT: {bat.height} buildings", flush=True)

    print("Building subclass 8760 loads ...", flush=True)
    loads_8760 = build_subclass_8760(billing_8760, bat)
    print(f"  Subclass loads: {loads_8760.height} hours", flush=True)

    print("Assembling workbook tabs ...", flush=True)
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    _write_readme(wb)
    _write_inputs_scalars(wb, rr_yaml, tariff)
    _write_mc_combined(wb, mc_df)
    _write_subclass_loads(wb, loads_8760)
    _write_subclass_mc_alloc(wb)
    _write_subclass_coss(wb, bat, rr_yaml, tariff)
    _write_validation(wb, bat, rr_yaml, tariff, loads_8760)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    size_kb = output_path.stat().st_size / 1024
    print(f"Wrote {output_path} ({size_kb:.1f} KB)", flush=True)
    return output_path


_TAB_FORMATTING: dict[str, dict] = {
    "README": {
        "wrap_columns": ["A:C"],
        "column_widths_px": {"A": 300, "B": 540, "C": 540},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "inputs_scalars": {
        "column_number_formats": {"B": "#,##0.000000"},
        "wrap_columns": ["C"],
        "column_widths_px": {"A": 300, "B": 160, "C": 300},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "mc_combined": {
        "column_number_formats": {
            "B": "0.000000",
            "C": "0.000000",
            "D": "0.000000",
        },
        "auto_resize_columns": ["A:D"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "subclass_loads": {
        "column_number_formats": {
            "B": "#,##0.000",
            "C": "#,##0.000",
            "D": "#,##0.000",
            "E": "#,##0.000",
            "F": "#,##0.000",
            "G": "#,##0.000",
        },
        "auto_resize_columns": ["A:G"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "subclass_mc_alloc": {
        "column_number_formats": {
            "B": "0.000000",
            "C": "$#,##0.000000",
            "D": "$#,##0.000000",
            "E": "$#,##0.000000",
            "F": "$#,##0.000000",
            "G": "$#,##0.000000",
            "H": "$#,##0.000000",
        },
        "auto_resize_columns": ["A:H"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "subclass_coss": {
        "auto_resize_columns": ["A:N"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "validation": {
        "column_number_formats": {
            "B": "#,##0.00",
            "C": "#,##0.00",
            "D": "#,##0.00",
        },
        "auto_resize_columns": ["A:F"],
        "freeze_rows": 1,
        "bold_header": True,
    },
}


def upload_to_sheet(xlsx_path: Path, spreadsheet_id: str) -> None:
    from lib.data.gsheets import apply_sheet_formatting, xlsx_to_gsheet

    print(f"Uploading {xlsx_path} -> Google Sheet {spreadsheet_id} ...", flush=True)
    spreadsheet = xlsx_to_gsheet(xlsx_path, spreadsheet_id, delete_other_tabs=True)
    print("Applying formatting ...", flush=True)
    for ws in spreadsheet.worksheets():
        spec = _TAB_FORMATTING.get(ws.title)
        if spec:
            fmt = dict(spec)
            if ws.title == "README" and _README_BOLD_ROWS:
                fmt["bold_rows"] = list(_README_BOLD_ROWS)
            apply_sheet_formatting(ws, **fmt)
    print(
        f"Done. View at https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit",
        flush=True,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build RIE 1-14: subclass cost-of-service workbook.")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("cache/rie_1_14_subclass_coss.xlsx"),
    )
    parser.add_argument(
        "--upload",
        action="store_true",
        help="Upload to a Google Sheet after building.",
    )
    parser.add_argument(
        "--spreadsheet-id",
        default=DEFAULT_SPREADSHEET_ID,
        help="Target Google Sheet id for upload.",
    )
    args = parser.parse_args(argv)

    out = build_workbook(args.output)
    if args.upload:
        if not args.spreadsheet_id:
            print("ERROR: --spreadsheet-id is required for upload.", file=sys.stderr)
            return 1
        upload_to_sheet(out, args.spreadsheet_id)
    return 0


if __name__ == "__main__":
    sys.exit(main())
