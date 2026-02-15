"""
Fixtures for ny_aeba_savings model tests.

The core fixture copies the Excel workbook to a temp directory, optionally
modifies input cells via openpyxl, recalculates the workbook with LibreOffice
in headless mode, and reads back computed values with data_only=True.
"""

from __future__ import annotations

import shutil
import subprocess
import sys
from collections.abc import Generator
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
import pytest

# ---------------------------------------------------------------------------
# Ensure the model source directory is importable.
# model.py lives at reports/ny_aeba_savings/model.py, so we add its parent
# to sys.path. This allows `from model import ...` in test files.
# ---------------------------------------------------------------------------
_MODEL_DIR = Path(__file__).resolve().parent.parent
if str(_MODEL_DIR) not in sys.path:
    sys.path.insert(0, str(_MODEL_DIR))

# Path to the authoritative Excel workbook (has spaces in filename)
WORKBOOK_PATH = (
    _MODEL_DIR
    / "Switchbox - Heat Pump Cost Savings in New Buildings Model (New GSHP in New Construction Incentives).xlsx"
)

WORKBOOK_FILENAME = WORKBOOK_PATH.name


# Cells in the workbook that use Excel-specific array formulas (AVERAGEIFS,
# SUMPRODUCT-with-criteria, etc.) that LibreOffice cannot evaluate, returning
# #NAME?.  These formulas depend only on county data — not on model inputs —
# so their values are constant across all perturbations.  Before handing the
# workbook to LibreOffice we replace them with literal cached values so every
# downstream cell recalculates correctly.
_ARRAY_FORMULA_CELLS: list[tuple[str, str]] | None = None


def _get_array_formula_cells() -> list[tuple[str, str]]:
    """Return [(sheet!cell, ...), ...] for cells that need literal replacement.

    Discovered once by scanning the original workbook for ArrayFormula objects.
    """
    global _ARRAY_FORMULA_CELLS
    if _ARRAY_FORMULA_CELLS is None:
        _ARRAY_FORMULA_CELLS = []
        wb_formulas = openpyxl.load_workbook(WORKBOOK_PATH, data_only=False)
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if hasattr(cell, "value") and hasattr(cell.value, "__class__"):
                        from openpyxl.worksheet.formula import ArrayFormula

                        if isinstance(cell.value, ArrayFormula):
                            _ARRAY_FORMULA_CELLS.append((sheet_name, cell.coordinate))
        wb_formulas.close()
    return _ARRAY_FORMULA_CELLS


def _bake_array_formulas(wb_to_fix: openpyxl.Workbook) -> None:
    """Replace array-formula cells with their cached literal values.

    We read the cached values from the original workbook (data_only=True)
    and overwrite the formula cells in *wb_to_fix* so that LibreOffice sees
    plain numbers instead of formulas it cannot evaluate.
    """
    cached = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    for sheet_name, cell_ref in _get_array_formula_cells():
        val = cached[sheet_name][cell_ref].value
        if val is not None:
            wb_to_fix[sheet_name][cell_ref] = val
    cached.close()


@dataclass
class RecalculatedWorkbook:
    """Result of a LibreOffice recalculation. Provides cell-reading helpers."""

    path: Path
    modifications: dict[str, object] = field(default_factory=dict)
    _wb: openpyxl.Workbook | None = field(default=None, repr=False)

    @property
    def wb(self) -> openpyxl.Workbook:
        if self._wb is None:
            self._wb = openpyxl.load_workbook(self.path, data_only=True)
        return self._wb

    def cell_value(self, cell_ref: str, sheet: str = "Model") -> object:
        """Read a single cell value from the recalculated workbook.

        Args:
            cell_ref: Excel-style cell reference, e.g. "F40".
            sheet: Sheet name. Defaults to "Model".

        Returns:
            The cell value (typically float for numeric cells, str, or None).
        """
        ws = self.wb[sheet]
        return ws[cell_ref].value

    def row_values(
        self,
        row: int,
        columns: list[str] | None = None,
        sheet: str = "Model",
    ) -> list[object]:
        """Read values from multiple columns in a single row.

        Args:
            row: The row number (1-based).
            columns: Column letters to read. Defaults to ["E","F","G","H","I","J"].
            sheet: Sheet name. Defaults to "Model".

        Returns:
            List of cell values in the specified column order.
        """
        if columns is None:
            columns = ["E", "F", "G", "H", "I", "J"]
        ws = self.wb[sheet]
        return [ws[f"{col}{row}"].value for col in columns]

    def close(self) -> None:
        if self._wb is not None:
            self._wb.close()
            self._wb = None


def _recalculate_workbook(
    tmp_path: Path,
    modifications: dict[str, object] | None = None,
) -> RecalculatedWorkbook:
    """Copy the workbook to tmp_path, apply modifications, recalculate with LibreOffice.

    Args:
        tmp_path: Temporary directory for the working copy.
        modifications: Dict mapping "SheetName!CellRef" to new values.
            Example: {"Model!C8": 7000, "Model!D8": 6500}
            If the key has no "!" separator, the "Model" sheet is assumed.

    Returns:
        A RecalculatedWorkbook with the recalculated file path and helpers.
    """
    if not WORKBOOK_PATH.exists():
        pytest.fail(f"Excel workbook not found at {WORKBOOK_PATH}")

    # Use separate input and output directories so LibreOffice --convert-to
    # does not try to overwrite the source file. When the input and output
    # paths are identical, LibreOffice silently fails to save (returns rc=0
    # but emits an SfxBaseModel::impl_store error), leaving formula cells
    # without cached values. Splitting directories avoids this entirely.
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()

    dest = input_dir / WORKBOOK_FILENAME
    shutil.copy2(WORKBOOK_PATH, dest)

    mods = modifications or {}

    # Always open the workbook to bake array formulas (and optionally apply
    # cell modifications).  Array formulas are replaced with literal cached
    # values so LibreOffice does not encounter #NAME? errors.
    wb = openpyxl.load_workbook(dest)
    _bake_array_formulas(wb)
    for key, value in mods.items():
        if "!" in key:
            sheet_name, cell_ref = key.split("!", 1)
        else:
            sheet_name, cell_ref = "Model", key
        ws = wb[sheet_name]
        ws[cell_ref] = value
    wb.save(dest)
    wb.close()

    # Recalculate with LibreOffice.
    # --convert-to xlsx forces a full recalculation of all formulas.
    # The output goes to a separate directory to avoid the silent overwrite bug.
    result = subprocess.run(
        [
            "libreoffice",
            "--headless",
            "--calc",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(output_dir),
            str(dest),
        ],
        capture_output=True,
        text=True,
        timeout=120,
    )

    if result.returncode != 0:
        pytest.fail(
            f"LibreOffice recalculation failed (rc={result.returncode}).\n"
            f"stdout: {result.stdout}\n"
            f"stderr: {result.stderr}"
        )

    recalculated_path = output_dir / WORKBOOK_FILENAME
    if not recalculated_path.exists():
        pytest.fail(
            f"Recalculated workbook not found at {recalculated_path}.\n"
            f"LibreOffice stdout: {result.stdout}\n"
            f"LibreOffice stderr: {result.stderr}"
        )

    return RecalculatedWorkbook(
        path=recalculated_path,
        modifications=mods,
    )


@pytest.fixture
def recalculated_workbook(tmp_path: Path) -> Generator[RecalculatedWorkbook]:
    """Fixture: recalculate the workbook with no modifications (default inputs).

    Returns a RecalculatedWorkbook whose cell values reflect a full LibreOffice
    recalculation of the original Excel file.
    """
    result = _recalculate_workbook(tmp_path)
    yield result
    result.close()


@pytest.fixture
def recalculate(tmp_path: Path) -> Generator[Any]:
    """Factory fixture: recalculate the workbook with custom cell modifications.

    Usage in tests::

        def test_something(recalculate):
            wb = recalculate({"Model!C8": 7000})
            assert wb.cell_value("F40") == pytest.approx(expected, rel=1e-4)

    Each invocation creates an independent copy and recalculation, so multiple
    calls within a single test are safe (e.g., for comparing default vs
    perturbed results).
    """
    created: list[RecalculatedWorkbook] = []

    def _factory(modifications: dict[str, object] | None = None) -> RecalculatedWorkbook:
        # Each call gets its own subdirectory to avoid collisions
        subdir = tmp_path / f"run_{len(created)}"
        subdir.mkdir()
        result = _recalculate_workbook(subdir, modifications)
        created.append(result)
        return result

    yield _factory

    for wb in created:
        wb.close()
