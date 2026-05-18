"""Build RIE 1-14a: Subclass-level cost-of-service study workbook.

Responds to RIE 1-14(a): "Please provide, in Excel with all formulas intact,
the 'cost-of-service study at the subclass level' performed by Witness Velez."

Reference: Pre-Filed Direct Testimony of Juan-Pablo Velez, Page 76 of 85,
line 8: "I performed a cost-of-service study at the subclass level -- the
same kind of exercise the Company performs at the class level in its Allocated
Cost of Service Study ('ACOSS'), but applied one level deeper, within the
residential class."

The workbook decomposes total delivery revenue into marginal costs (economic
burden) and EPMC residual for each heating-type subclass, using hourly
(8760) load profiles and the same marginal cost derivation documented in
RIE 1-10 (build_marginal_costs_workbook.py --delivery-only).

Tabs
----
README             Overview and source links.
inputs_scalars     Key parameters from the RR YAML and tariff JSON, with notes
                   citing relevant expert testimony sections.
mc_combined        8760 delivery MC (dist/sub-TX + bulk TX), same as RIE 1-10.
subclass_loads     8760 weighted kWh by heating subclass (from billing_kwh_8760).
subclass_mc_alloc  8760 MC x load allocation by subclass -- live formulas.
subclass_coss      Summary: customers, kWh, marginal cost, EPMC residual,
                   cost of service, revenue, cross-subsidy per subclass.
validation         Formula-level checks with expert testimony cross-references.
validation_testimony  Per-subclass comparison vs. expert testimony cache.

Usage::

    cd reports/ri_hp_rates

    # Build locally only:
    uv run python -m testimony_response.build_RIE_1_14a_workbook --no-upload

    # Build and upload to default Drive folder (default behavior):
    uv run python -m testimony_response.build_RIE_1_14a_workbook
"""

from __future__ import annotations

import argparse
import json
import pickle
import subprocess
import sys
from pathlib import Path
from typing import Any

import polars as pl
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────

UTILITY = "rie"
STATE = "ri"
YEAR = 2025
BATCH = "ri_20260507_r1-2_grid_cons_fix"
RUN_DIR = "20260507_213944_ri_rie_run1_up00_precalc__default"

S3_CAIRO_BASE = "s3://data.sb/switchbox/cairo/outputs/hp_rates"
PATH_BILLING_KWH_8760 = f"{S3_CAIRO_BASE}/{STATE}/{UTILITY}/{BATCH}/{RUN_DIR}/billing_kwh_8760.parquet"
PATH_MASTER_BAT_12 = f"{S3_CAIRO_BASE}/{STATE}/all_utilities/{BATCH}/run_1+2/cross_subsidization_BAT_values/"

S3_DIST_SUB_TX = (
    f"s3://data.sb/switchbox/marginal_costs/{STATE}/dist_and_sub_tx/utility={UTILITY}/year={YEAR}/data.parquet"
)
S3_BULK_TX = f"s3://data.sb/switchbox/marginal_costs/{STATE}/bulk_tx/utility={UTILITY}/year={YEAR}/data.parquet"

RDP_REF = "0b203bc"
RDP_GITHUB_BASE = "https://github.com/switchbox-data/rate-design-platform/blob"
REPORTS2_GITHUB_BASE = "https://github.com/switchbox-data/reports2/blob"

RDP_RR_YAML_PATH = "rate_design/hp_rates/ri/config/rev_requirement/rie_rate_case_test_year.yaml"
RDP_TARIFF_PATH = "rate_design/hp_rates/ri/config/tariffs/electric/rie_default_calibrated.json"

SUBCLASS_ORDER = [
    "heat_pump",
    "electrical_resistance",
    "natgas",
    "delivered_fuels",
    "other",
]
SUBCLASS_LABELS = {
    "heat_pump": "Heat pump",
    "electrical_resistance": "Electric resistance",
    "natgas": "Natural gas",
    "delivered_fuels": "Delivered fuels",
    "other": "Other",
}

REPORT_DIR = Path(__file__).resolve().parents[1]

DEFAULT_FOLDER_ID = "1uPcJbcOChD6zoFuPb-gsxSByPr7xwmCH"
DEFAULT_TITLE = "RIE 1-14a"

# ── Permalink helpers ─────────────────────────────────────────────────────────


def _rdp_permalink(rel_path: str) -> str:
    return f"{RDP_GITHUB_BASE}/{RDP_REF}/{rel_path}"


def _reports2_head_sha() -> str:
    if not hasattr(_reports2_head_sha, "_cached"):
        repo_root = Path(__file__).resolve().parents[3]
        sha = subprocess.check_output(
            ["git", "-C", str(repo_root), "rev-parse", "HEAD"],
            text=True,
        ).strip()
        _reports2_head_sha._cached = sha  # type: ignore[attr-defined]
    return _reports2_head_sha._cached  # type: ignore[attr-defined]


def _reports2_permalink(rel_path: str) -> str:
    return f"{REPORTS2_GITHUB_BASE}/{_reports2_head_sha()}/{rel_path}"


# ── Formatting helpers ────────────────────────────────────────────────────────

_README_BOLD_ROWS: list[int] = []


def _header_fill(ws, row: int, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="E8E8E8")
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = fill


def _autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Data loading ──────────────────────────────────────────────────────────────


def _load_mc_parquet(s3_path: str) -> pl.DataFrame:
    df = pl.read_parquet(s3_path)
    numeric_cols = [c for c in df.columns if c != "timestamp" and df[c].dtype.is_numeric() and c != "year"]
    return df.select("timestamp", numeric_cols[0]).rename({numeric_cols[0]: "mc_value"})


def load_delivery_mc() -> pl.DataFrame:
    """Load and combine the two delivery MC parquets into one 8760 DataFrame."""
    dist = _load_mc_parquet(S3_DIST_SUB_TX).rename({"mc_value": "mc_dist_sub_tx"})
    bulk = _load_mc_parquet(S3_BULK_TX).rename({"mc_value": "mc_bulk_tx"})
    combined = dist.join(bulk, on="timestamp", how="inner").sort("timestamp")
    return combined.with_columns(
        (pl.col("mc_dist_sub_tx") + pl.col("mc_bulk_tx")).alias("mc_delivery_total"),
    )


def load_billing_8760() -> pl.DataFrame:
    """Load per-building hourly grid consumption from CAIRO output."""
    return pl.read_parquet(PATH_BILLING_KWH_8760)


def load_master_bat() -> pl.DataFrame:
    """Load master BAT for run 1+2, filtered to RIE."""
    df = (
        pl.scan_parquet(PATH_MASTER_BAT_12, hive_partitioning=True)
        .filter(pl.col("sb.electric_utility") == UTILITY)
        .select(
            "bldg_id",
            "weight",
            "postprocess_group.heating_type_v2",
            "annual_bill_delivery",
            "economic_burden_delivery",
            "residual_share_epmc_delivery",
            "BAT_epmc_delivery",
        )
        .collect()
    )
    assert isinstance(df, pl.DataFrame)
    return df


def load_rr_yaml() -> dict:
    """Load the revenue requirement YAML from the RDP repo at RDP_REF."""
    from lib.rdp import fetch_rdp_file

    raw = fetch_rdp_file(RDP_RR_YAML_PATH, RDP_REF)
    return yaml.safe_load(raw)


def load_calibrated_tariff() -> dict:
    """Load the calibrated tariff JSON from the RDP repo at RDP_REF."""
    from lib.rdp import fetch_rdp_file

    raw = fetch_rdp_file(RDP_TARIFF_PATH, RDP_REF)
    return json.loads(raw)


# ── Build subclass 8760 loads ─────────────────────────────────────────────────


def build_subclass_8760(billing_8760: pl.DataFrame, bat: pl.DataFrame) -> pl.DataFrame:
    """Aggregate per-building 8760 kWh into weighted subclass totals.

    Returns an 8760-row DataFrame with columns:
        timestamp, kWh_heat_pump, kWh_electrical_resistance, kWh_natgas,
        kWh_delivered_fuels, kWh_other, kWh_total
    """
    bldg_info = bat.select("bldg_id", "weight", "postprocess_group.heating_type_v2")
    joined = billing_8760.join(bldg_info, on="bldg_id", how="inner")

    # weighted kWh per building-hour
    joined = joined.with_columns(
        (pl.col("grid_cons_kwh") * pl.col("weight")).alias("w_kwh"),
    )

    subclass_dfs = []
    for sc_key in SUBCLASS_ORDER:
        sc_agg = (
            joined.filter(pl.col("postprocess_group.heating_type_v2") == sc_key)
            .group_by("timestamp")
            .agg(pl.col("w_kwh").sum().alias(f"kWh_{sc_key}"))
        )
        subclass_dfs.append(sc_agg)

    result = subclass_dfs[0]
    for sc_df in subclass_dfs[1:]:
        result = result.join(sc_df, on="timestamp", how="full", coalesce=True)

    for sc_key in SUBCLASS_ORDER:
        col_name = f"kWh_{sc_key}"
        result = result.with_columns(pl.col(col_name).fill_null(0.0))

    kwh_cols = [f"kWh_{sc}" for sc in SUBCLASS_ORDER]
    result = result.with_columns(
        pl.sum_horizontal([pl.col(c) for c in kwh_cols]).alias("kWh_total"),
    ).sort("timestamp")

    return result


# ── Tab writers ───────────────────────────────────────────────────────────────


def _write_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README", 0)
    rows: list[list] = [
        [
            "RIE 1-14a: Subclass Cost-of-Service Workbook - RIE 2025 (delivery)",
            "",
            "",
        ],
        ["", "", ""],
        ["Item", "Source", "Notes"],
        [
            "RIE 1-10 (marginal costs derivation)",
            "https://docs.google.com/spreadsheets/d/1cYGVrLmDTKm9XXEw7jjLVrSqTr8xp7pVK1o0Kk8AOkw/edit",
            "Companion workbook with full 8760 MC derivation from raw ISO-NE/EIA data.",
        ],
        [
            "Revenue requirement YAML",
            _rdp_permalink(RDP_RR_YAML_PATH),
            "rie_rate_case_test_year.yaml: customer count, delivery RR, kWh scale factor.",
        ],
        [
            "Calibrated tariff JSON",
            _rdp_permalink(RDP_TARIFF_PATH),
            "rie_default_calibrated.json: volumetric rate and fixed charge after CAIRO precalc.",
        ],
        [
            "Billing kWh 8760",
            PATH_BILLING_KWH_8760,
            "CAIRO output: per-building hourly grid consumption (post-scaling, post-PV floor).",
        ],
        [
            "Master BAT (run 1+2)",
            PATH_MASTER_BAT_12,
            "Post-processed BAT: per-building delivery bill, economic burden, EPMC residual.",
        ],
        [
            "Cost-of-service notebook",
            _reports2_permalink("reports/ri_hp_rates/notebooks/cost_of_service_by_subclass.qmd"),
            "Quarto notebook that produces the subclass COSS tables in the expert testimony.",
        ],
        [
            "Expert testimony",
            _reports2_permalink("reports/ri_hp_rates/expert_testimony.qmd"),
            "Pre-Filed Direct Testimony of Juan-Pablo Velez. Key values cross-referenced in inputs_scalars notes and validation checks.",
        ],
        [
            "This workbook builder",
            _reports2_permalink("reports/ri_hp_rates/testimony_response/build_RIE_1_14a_workbook.py"),
            "Script that generated this workbook.",
        ],
        ["", "", ""],
        ["Sheet", "What it contains", ""],
        [
            "inputs_scalars",
            "Revenue requirement, customer count, tariff parameters, kWh scale factor.",
            "",
        ],
        [
            "mc_combined",
            "8760 rows: delivery MC per kWh (dist/sub-TX + bulk TX). Same data as RIE 1-10.",
            "",
        ],
        [
            "subclass_loads",
            "8760 rows: weighted grid kWh by heating subclass + total.",
            "",
        ],
        [
            "subclass_mc_alloc",
            "8760 rows: MC x load = hourly marginal cost allocation by subclass. All formulas.",
            "",
        ],
        [
            "subclass_coss",
            "Summary: customers, consumption, economic burden, EPMC residual, COS, revenue, cross-subsidy by subclass.",
            "",
        ],
        [
            "validation",
            "Formula-level checks against YAML targets, BAT totals, and expert testimony values.",
            "",
        ],
        [
            "validation_testimony",
            "Per-subclass comparison of workbook values against expert testimony cache (report_variables_cos_subclass.pkl).",
            "",
        ],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    bold_rows = []
    for row_idx, row_data in enumerate(rows, start=1):
        if row_data and row_data[0] in ("Item", "Sheet"):
            _header_fill(ws, row_idx, 3)
            bold_rows.append(row_idx)
    _README_BOLD_ROWS.clear()
    _README_BOLD_ROWS.extend(bold_rows)
    _autosize(ws, {"A": 40, "B": 80, "C": 80})
    ws.sheet_view.showGridLines = False


def _write_inputs_scalars(wb: Workbook, rr_yaml: dict, tariff: dict) -> None:
    ws = wb.create_sheet("inputs_scalars")
    vol_rate = float(tariff["items"][0]["energyratestructure"][0][0]["rate"])
    fixed_monthly = float(tariff["items"][0]["fixedchargefirstmeter"])
    fixed_annual = fixed_monthly * 12.0
    total_delivery_rr = float(rr_yaml["total_delivery_revenue_requirement"])
    ty_kwh = float(rr_yaml["test_year_residential_kwh"])
    ty_customers = float(rr_yaml["test_year_customer_count"])
    kwh_sf = float(rr_yaml["resstock_kwh_scale_factor"])

    yaml_url = _rdp_permalink(RDP_RR_YAML_PATH)
    tariff_url = _rdp_permalink(RDP_TARIFF_PATH)
    rows = [
        ["key", "value", "source", "notes"],
        [
            "total_delivery_revenue_requirement",
            total_delivery_rr,
            yaml_url,
            "Test Year total residential delivery revenue requirement ($). Expert testimony §III and §IX.",
        ],
        [
            "test_year_residential_kwh",
            ty_kwh,
            yaml_url,
            "Test Year total residential kWh. "
            "Expert testimony §IX: 'the Company's Test Year total of [2.82 billion] kWh'.",
        ],
        [
            "test_year_customer_count",
            ty_customers,
            yaml_url,
            "Test Year residential customer count. "
            "Expert testimony §IX: applied to RECS 2020 shares to derive subclass counts.",
        ],
        [
            "resstock_kwh_scale_factor",
            kwh_sf,
            yaml_url,
            "Scaling factor applied to ResStock kWh to match Test Year total. "
            "Expert testimony §IX: 'a small additional scaling factor'.",
        ],
        [
            "volumetric_rate_usd_per_kwh",
            vol_rate,
            tariff_url,
            "Calibrated volumetric delivery rate ($/kWh) from CAIRO precalc.",
        ],
        [
            "fixed_charge_usd_per_month",
            fixed_monthly,
            tariff_url,
            "Monthly fixed customer charge ($6.00/month). Expert testimony §III: '$6.00/month'.",
        ],
        [
            "fixed_charge_usd_per_year",
            fixed_annual,
            "Derived: monthly x 12",
            "Derived: monthly fixed charge x 12.",
        ],
    ]
    for r in rows:
        ws.append(r)
    # Overwrite B8 with live formula referencing B7 (monthly charge)
    ws.cell(row=8, column=2, value="=B7*12")
    _header_fill(ws, 1, 4)
    _autosize(ws, {"A": 40, "B": 24, "C": 80, "D": 80})
    ws.sheet_view.showGridLines = False


def _write_mc_combined(wb: Workbook, mc_df: pl.DataFrame) -> None:
    """mc_combined: 8760 rows of delivery MC (values, not formulas — same as RIE 1-10)."""
    ws = wb.create_sheet("mc_combined")
    headers = ["timestamp", "mc_dist_sub_tx", "mc_bulk_tx", "mc_delivery_total"]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    for i, row in enumerate(mc_df.iter_rows(named=True), start=2):
        ws.cell(row=i, column=1, value=row["timestamp"].strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=i, column=2, value=float(row["mc_dist_sub_tx"]))
        ws.cell(row=i, column=3, value=float(row["mc_bulk_tx"]))
        ws.cell(row=i, column=4, value=f"=B{i}+C{i}")

    _autosize(ws, {"A": 18, "B": 18, "C": 16, "D": 20})


def _write_subclass_loads(wb: Workbook, loads_8760: pl.DataFrame) -> None:
    """subclass_loads: 8760 rows of weighted kWh by subclass."""
    ws = wb.create_sheet("subclass_loads")

    sc_labels = [SUBCLASS_LABELS[sc] for sc in SUBCLASS_ORDER]
    headers = ["timestamp", *sc_labels, "Total"]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    n_sc = len(SUBCLASS_ORDER)
    for i, row in enumerate(loads_8760.iter_rows(named=True), start=2):
        ws.cell(row=i, column=1, value=row["timestamp"].strftime("%Y-%m-%d %H:%M"))
        for j, sc_key in enumerate(SUBCLASS_ORDER):
            ws.cell(row=i, column=2 + j, value=float(row[f"kWh_{sc_key}"]))
        # Total as formula: sum of subclass columns
        first_col = get_column_letter(2)
        last_col = get_column_letter(1 + n_sc)
        ws.cell(
            row=i,
            column=2 + n_sc,
            value=f"=SUM({first_col}{i}:{last_col}{i})",
        )

    widths = {"A": 18}
    for j in range(n_sc + 1):
        widths[get_column_letter(2 + j)] = 18
    _autosize(ws, widths)


def _write_subclass_mc_alloc(wb: Workbook, n_hours: int = 8760) -> None:
    """subclass_mc_alloc: MC x load for each subclass, all as cross-sheet formulas."""
    ws = wb.create_sheet("subclass_mc_alloc")

    sc_labels = [SUBCLASS_LABELS[sc] for sc in SUBCLASS_ORDER]
    headers = ["timestamp", "mc_delivery_per_kwh", *sc_labels, "Total"]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    n_sc = len(SUBCLASS_ORDER)
    for i in range(2, n_hours + 2):
        ws.cell(row=i, column=1, value=f"=mc_combined!A{i}")
        ws.cell(row=i, column=2, value=f"=mc_combined!D{i}")
        for j in range(n_sc):
            load_col = get_column_letter(2 + j)  # col in subclass_loads
            ws.cell(
                row=i,
                column=3 + j,
                value=f"=$B{i}*subclass_loads!{load_col}{i}",
            )
        # Total = sum of subclass MC allocations
        first = get_column_letter(3)
        last = get_column_letter(2 + n_sc)
        ws.cell(
            row=i,
            column=3 + n_sc,
            value=f"=SUM({first}{i}:{last}{i})",
        )

    widths = {"A": 18, "B": 20}
    for j in range(n_sc + 1):
        widths[get_column_letter(3 + j)] = 18
    _autosize(ws, widths)


# Scalar references in inputs_scalars (row indices match _write_inputs_scalars)
REF_DELIVERY_RR = "inputs_scalars!$B$2"
REF_TY_KWH = "inputs_scalars!$B$3"
REF_TY_CUSTOMERS = "inputs_scalars!$B$4"
REF_FIXED_ANNUAL = "inputs_scalars!$B$8"
REF_VOL_RATE = "inputs_scalars!$B$6"


def _write_subclass_coss(
    wb: Workbook,
    bat: pl.DataFrame,
    rr_yaml: dict,
    tariff: dict,
) -> None:
    """Summary table: customers, kWh, MC, residual, COS, revenue, cross-subsidy."""
    ws = wb.create_sheet("subclass_coss")

    headers = [
        "Subclass",
        "Customers",
        "% of customers",
        "Consumption (kWh)",
        "% of consumption",
        "Marginal cost (economic burden)",
        "% of MC",
        "EPMC residual",
        "Cost of service (MC + residual)",
        "% of COS",
        "Delivery revenue",
        "% of revenue",
        "Cross-subsidy (revenue - COS)",
        "Cross-subsidy / COS",
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    vol_rate = float(tariff["items"][0]["energyratestructure"][0][0]["rate"])
    fixed_monthly = float(tariff["items"][0]["fixedchargefirstmeter"])
    fixed_annual = fixed_monthly * 12.0

    n_sc = len(SUBCLASS_ORDER)
    all_rows_data: list[dict] = []

    for sc_key in SUBCLASS_ORDER:
        sc_bat = bat.filter(pl.col("postprocess_group.heating_type_v2") == sc_key)
        customers = float(sc_bat["weight"].sum())
        revenue = float((sc_bat["weight"] * sc_bat["annual_bill_delivery"]).sum())
        economic_burden = float((sc_bat["weight"] * sc_bat["economic_burden_delivery"]).sum())
        epmc_residual = float((sc_bat["weight"] * sc_bat["residual_share_epmc_delivery"]).sum())
        cos = economic_burden + epmc_residual
        cross_subsidy = revenue - cos
        # kWh derived from bills: (bill - fixed_annual) / vol_rate
        consumption = float((sc_bat["weight"] * ((sc_bat["annual_bill_delivery"] - fixed_annual) / vol_rate)).sum())
        all_rows_data.append(
            {
                "label": SUBCLASS_LABELS[sc_key],
                "customers": customers,
                "consumption": consumption,
                "mc": economic_burden,
                "residual": epmc_residual,
                "cos": cos,
                "revenue": revenue,
                "cross_subsidy": cross_subsidy,
            }
        )

    # Write subclass rows as values
    for r_idx, d in enumerate(all_rows_data, start=2):
        ws.cell(row=r_idx, column=1, value=d["label"])
        ws.cell(row=r_idx, column=2, value=d["customers"])
        ws.cell(row=r_idx, column=4, value=d["consumption"])
        ws.cell(row=r_idx, column=6, value=d["mc"])
        ws.cell(row=r_idx, column=8, value=d["residual"])
        ws.cell(row=r_idx, column=11, value=d["revenue"])

    total_row = 2 + n_sc
    first_data = 2
    last_data = first_data + n_sc - 1

    # Total row as SUM formulas
    ws.cell(row=total_row, column=1, value="All customers")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    for col_idx in (2, 4, 6, 8, 11):
        cl = get_column_letter(col_idx)
        ws.cell(
            row=total_row,
            column=col_idx,
            value=f"=SUM({cl}{first_data}:{cl}{last_data})",
        )

    # Formula columns for all rows (subclass + total)
    for r in range(first_data, total_row + 1):
        total_cust_ref = f"$B${total_row}"
        total_kwh_ref = f"$D${total_row}"
        total_mc_ref = f"$F${total_row}"
        total_cos_ref = f"$I${total_row}"
        total_rev_ref = f"$K${total_row}"

        # % of customers
        ws.cell(row=r, column=3, value=f"=B{r}/{total_cust_ref}")
        # % of consumption
        ws.cell(row=r, column=5, value=f"=D{r}/{total_kwh_ref}")
        # % of MC
        ws.cell(row=r, column=7, value=f"=F{r}/{total_mc_ref}")
        # COS = MC + residual
        ws.cell(row=r, column=9, value=f"=F{r}+H{r}")
        # % of COS
        ws.cell(row=r, column=10, value=f"=I{r}/{total_cos_ref}")
        # % of revenue
        ws.cell(row=r, column=12, value=f"=K{r}/{total_rev_ref}")
        # Cross-subsidy = revenue - COS
        ws.cell(row=r, column=13, value=f"=K{r}-I{r}")
        # Cross-subsidy / COS
        ws.cell(row=r, column=14, value=f"=IF(I{r}<>0,M{r}/I{r},0)")

    # Number formatting
    for r in range(first_data, total_row + 1):
        ws.cell(row=r, column=2).number_format = "#,##0.00"
        ws.cell(row=r, column=3).number_format = "0.0%"
        ws.cell(row=r, column=4).number_format = "#,##0"
        ws.cell(row=r, column=5).number_format = "0.0%"
        ws.cell(row=r, column=6).number_format = "$#,##0"
        ws.cell(row=r, column=7).number_format = "0.0%"
        ws.cell(row=r, column=8).number_format = "$#,##0"
        ws.cell(row=r, column=9).number_format = "$#,##0"
        ws.cell(row=r, column=10).number_format = "0.0%"
        ws.cell(row=r, column=11).number_format = "$#,##0"
        ws.cell(row=r, column=12).number_format = "0.0%"
        ws.cell(row=r, column=13).number_format = "$#,##0"
        ws.cell(row=r, column=14).number_format = "0.0%"

    _autosize(
        ws,
        {
            "A": 22,
            "B": 14,
            "C": 16,
            "D": 20,
            "E": 18,
            "F": 30,
            "G": 10,
            "H": 18,
            "I": 30,
            "J": 10,
            "K": 18,
            "L": 14,
            "M": 28,
            "N": 20,
        },
    )
    ws.sheet_view.showGridLines = False


def _write_validation(
    wb: Workbook,
    bat: pl.DataFrame,
    rr_yaml: dict,
    tariff: dict,
    loads_8760: pl.DataFrame,
) -> None:
    ws = wb.create_sheet("validation")
    headers = ["check", "actual", "expected", "abs_error", "tolerance", "ok"]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))

    total_delivery_rr = float(rr_yaml["total_delivery_revenue_requirement"])
    ty_kwh = float(rr_yaml["test_year_residential_kwh"])
    ty_customers = float(rr_yaml["test_year_customer_count"])

    bat_total_w = float(bat["weight"].sum())
    bat_total_rev = float((bat["weight"] * bat["annual_bill_delivery"]).sum())
    bat_total_eb = float((bat["weight"] * bat["economic_burden_delivery"]).sum())
    bat_total_res = float((bat["weight"] * bat["residual_share_epmc_delivery"]).sum())
    bat_total_cos = bat_total_eb + bat_total_res
    bat_total_xs = bat_total_rev - bat_total_cos
    loads_total_kwh = float(loads_8760["kWh_total"].sum())

    # MC from subclass_mc_alloc total column
    n_sc = len(SUBCLASS_ORDER)
    mc_total_col = get_column_letter(3 + n_sc)

    checks = [
        (
            "sum(weight) = test_year_customer_count [expert testimony §IX]",
            bat_total_w,
            ty_customers,
            1.0,
        ),
        (
            "sum(w * delivery_bill) = total_delivery_RR [expert testimony §III, §IX]",
            bat_total_rev,
            total_delivery_rr,
            100.0,
        ),
        (
            "sum(w * COS) ~ total_delivery_RR [expert testimony §III, §IX]",
            bat_total_cos,
            total_delivery_rr,
            5000.0,
        ),
        (
            "sum(cross-subsidy) nets to ~0 [expert testimony §IX Step 4: BAT]",
            abs(bat_total_xs),
            0.0,
            5000.0,
        ),
        (
            "8760 weighted grid kWh = test_year_residential_kwh [expert testimony §IX]",
            loads_total_kwh,
            ty_kwh,
            1000.0,
        ),
        (
            "subclass_mc_alloc sum = BAT economic_burden total [expert testimony §IX]",
            f"=SUM(subclass_mc_alloc!{mc_total_col}$2:{mc_total_col}$8761)",
            bat_total_eb,
            1000.0,
        ),
    ]

    for i, (name, actual, expected, tol) in enumerate(checks, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=actual)
        ws.cell(row=i, column=3, value=expected)
        ws.cell(row=i, column=4, value=f"=ABS(B{i}-C{i})")
        ws.cell(row=i, column=5, value=tol)
        ws.cell(row=i, column=6, value=f'=IF(D{i}<=E{i},"OK","FAIL")')

    _autosize(ws, {"A": 70, "B": 20, "C": 20, "D": 14, "E": 14, "F": 8})
    for r in range(2, 2 + len(checks)):
        ws[f"B{r}"].number_format = "#,##0.00"
        ws[f"C{r}"].number_format = "#,##0.00"
        ws[f"D{r}"].number_format = "#,##0.00"
    ws.sheet_view.showGridLines = False


def _write_validation_testimony(wb: Workbook) -> None:
    """Compare subclass_coss values against the expert testimony cache.

    Loads ``report_variables_cos_subclass.pkl`` (the same pickle that feeds
    the rendered expert testimony) and checks per-subclass customers, COS,
    revenue, and cross-subsidy, plus aggregate totals.  Mirrors the approach
    used in build_DIV_1_2_workbook.py § ``add_validation_sheet``.
    """
    ws = wb.create_sheet("validation_testimony")

    ws["A1"] = "Validation: Workbook vs. Expert Testimony"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:G1")

    hdr_row = 3
    headers = ["Metric", "Workbook", "Testimony", "Diff", "% Diff", "Tol", "Status"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8E8E8")

    # Load testimony pickle
    cos_pkl = REPORT_DIR / "cache" / "report_variables_cos_subclass.pkl"
    t_vars: dict[str, Any] = {}
    if cos_pkl.exists():
        t_vars = pickle.loads(cos_pkl.read_bytes())

    # subclass_coss layout - row 1 = header, rows 2-6 = subclasses, row 7 = total
    # Columns: B=Customers, D=Consumption, I=COS, K=Revenue, M=Cross-subsidy
    coss = "subclass_coss"
    n_sc = len(SUBCLASS_ORDER)
    total_row = 2 + n_sc  # row 7

    # Subclass row mapping: SUBCLASS_ORDER index → coss row
    prefix_map = {
        "heat_pump": "hp",
        "electrical_resistance": "er",
        "natgas": "ng",
        "delivered_fuels": "df",
    }

    checks: list[tuple[str, str | float, float | None, float]] = []

    # Aggregate parameter checks
    checks.append(
        (
            "Total Delivery RR [expert testimony §III]",
            f"={REF_DELIVERY_RR}",
            t_vars.get("rie_rev_req_total_delivery_rr"),
            0.01,
        )
    )
    checks.append(
        (
            "Test Year Customers [expert testimony §IX]",
            f"={REF_TY_CUSTOMERS}",
            t_vars.get("rie_rev_req_test_year_customer_count"),
            0.01,
        )
    )

    # Per-subclass checks (COS, revenue, cross-subsidy, customers)
    for sc_key in SUBCLASS_ORDER:
        prefix = prefix_map.get(sc_key)
        if prefix is None:
            continue
        sc_idx = SUBCLASS_ORDER.index(sc_key)
        sc_row = 2 + sc_idx
        label = SUBCLASS_LABELS[sc_key]

        checks.append(
            (
                f"{label} - Customers [expert testimony §V, §IX]",
                f"={coss}!B{sc_row}",
                t_vars.get(f"cos_default_{prefix}_group_customers"),
                1.0,
            )
        )
        checks.append(
            (
                f"{label} - COS [expert testimony §V]",
                f"={coss}!I{sc_row}",
                t_vars.get(f"cos_default_{prefix}_group_cos"),
                1.0,
            )
        )
        checks.append(
            (
                f"{label} - Revenue [expert testimony §V]",
                f"={coss}!K{sc_row}",
                t_vars.get(f"cos_default_{prefix}_group_rev"),
                1.0,
            )
        )
        checks.append(
            (
                f"{label} - Cross-subsidy [expert testimony §V]",
                f"={coss}!M{sc_row}",
                t_vars.get(f"cos_default_{prefix}_group_xs"),
                1.0,
            )
        )

    # Total row checks
    checks.append(
        (
            "Total customers [expert testimony §V]",
            f"={coss}!B{total_row}",
            t_vars.get("cos_subclass_total_customers"),
            1.0,
        )
    )
    checks.append(
        (
            "Total revenue [expert testimony §V]",
            f"={coss}!K{total_row}",
            t_vars.get("cos_subclass_total_delivery_rev"),
            100.0,
        )
    )
    checks.append(
        (
            "Total COS [expert testimony §V]",
            f"={coss}!I{total_row}",
            t_vars.get("cos_subclass_total_cos"),
            100.0,
        )
    )

    row = hdr_row + 1
    for metric, wb_val, t_val, tol in checks:
        ws.cell(row=row, column=1, value=metric)

        if (isinstance(wb_val, str) and wb_val.startswith("=")) or wb_val is not None:
            ws.cell(row=row, column=2, value=wb_val).number_format = "#,##0.00"
        else:
            ws.cell(row=row, column=2, value="N/A")

        if t_val is not None:
            ws.cell(row=row, column=3, value=t_val).number_format = "#,##0.00"
            ws.cell(row=row, column=4, value=f"=B{row}-C{row}").number_format = "#,##0.00"
            ws.cell(row=row, column=5, value=f"=IF(C{row}<>0,ABS(D{row}/C{row}),0)").number_format = "0.00%"
            ws.cell(row=row, column=6, value=tol).number_format = "#,##0.00"
            ws.cell(row=row, column=7, value=f'=IF(ABS(D{row})<={tol},"PASS","FAIL")')
        else:
            ws.cell(row=row, column=3, value="N/A")
            for c in (4, 5, 6):
                ws.cell(row=row, column=c, value="")
            ws.cell(row=row, column=7, value="N/A")

        row += 1

    _autosize(ws, {"A": 50, "B": 16, "C": 16, "D": 14, "E": 10, "F": 8, "G": 8})
    ws.freeze_panes = f"A{hdr_row + 1}"
    ws.sheet_view.showGridLines = False


# ── Main build ────────────────────────────────────────────────────────────────


def build_workbook(output_path: Path) -> Path:
    print("Building RIE 1-14a subclass COSS workbook ...", flush=True)

    print("Loading RR YAML and calibrated tariff ...", flush=True)
    rr_yaml = load_rr_yaml()
    tariff = load_calibrated_tariff()

    print("Loading delivery MCs from S3 ...", flush=True)
    mc_df = load_delivery_mc()
    print(f"  MC combined: {mc_df.height} hours", flush=True)

    print("Loading billing_kwh_8760 ...", flush=True)
    billing_8760 = load_billing_8760()
    print(f"  Billing 8760: {billing_8760.height} rows ({billing_8760.height // 8760} buildings)", flush=True)

    print("Loading master BAT (run 1+2) ...", flush=True)
    bat = load_master_bat()
    print(f"  BAT: {bat.height} buildings", flush=True)

    print("Building subclass 8760 loads ...", flush=True)
    loads_8760 = build_subclass_8760(billing_8760, bat)
    print(f"  Subclass loads: {loads_8760.height} hours", flush=True)

    print("Assembling workbook tabs ...", flush=True)
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    _write_readme(wb)
    _write_inputs_scalars(wb, rr_yaml, tariff)
    _write_mc_combined(wb, mc_df)
    _write_subclass_loads(wb, loads_8760)
    _write_subclass_mc_alloc(wb)
    _write_subclass_coss(wb, bat, rr_yaml, tariff)
    _write_validation(wb, bat, rr_yaml, tariff, loads_8760)
    _write_validation_testimony(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    size_kb = output_path.stat().st_size / 1024
    print(f"Wrote {output_path} ({size_kb:.1f} KB)", flush=True)
    return output_path


_TAB_FORMATTING: dict[str, dict] = {
    "README": {
        "wrap_columns": ["A:C"],
        "column_widths_px": {"A": 300, "B": 540, "C": 540},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "inputs_scalars": {
        "column_number_formats": {"B": "#,##0.000000"},
        "wrap_columns": ["C", "D"],
        "column_widths_px": {"A": 300, "B": 160, "C": 540, "D": 540},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "mc_combined": {
        "column_number_formats": {
            "B": "0.000000",
            "C": "0.000000",
            "D": "0.000000",
        },
        "column_widths_px": {"A": 130, "B": 130, "C": 115, "D": 144},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "subclass_loads": {
        "column_number_formats": {
            "B": "#,##0.000",
            "C": "#,##0.000",
            "D": "#,##0.000",
            "E": "#,##0.000",
            "F": "#,##0.000",
            "G": "#,##0.000",
        },
        "column_widths_px": {"A": 130, "B": 130, "C": 130, "D": 130, "E": 130, "F": 130, "G": 130},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "subclass_mc_alloc": {
        "column_number_formats": {
            "B": "0.000000",
            "C": "$#,##0.000000",
            "D": "$#,##0.000000",
            "E": "$#,##0.000000",
            "F": "$#,##0.000000",
            "G": "$#,##0.000000",
            "H": "$#,##0.000000",
        },
        "column_widths_px": {
            "A": 130,
            "B": 144,
            "C": 130,
            "D": 130,
            "E": 130,
            "F": 130,
            "G": 130,
            "H": 130,
        },
        "freeze_rows": 1,
        "bold_header": True,
    },
    "subclass_coss": {
        "column_widths_px": {
            "A": 158,
            "B": 100,
            "C": 115,
            "D": 144,
            "E": 130,
            "F": 216,
            "G": 72,
            "H": 130,
            "I": 216,
            "J": 72,
            "K": 130,
            "L": 100,
            "M": 200,
            "N": 144,
        },
        "freeze_rows": 1,
        "bold_header": True,
    },
    "validation": {
        "column_number_formats": {
            "B": "#,##0.00",
            "C": "#,##0.00",
            "D": "#,##0.00",
        },
        "column_widths_px": {"A": 504, "B": 144, "C": 144, "D": 100, "E": 100, "F": 58},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "validation_testimony": {
        "column_number_formats": {
            "B": "#,##0.00",
            "C": "#,##0.00",
            "D": "#,##0.00",
            "E": "0.00%",
        },
        "column_widths_px": {"A": 360, "B": 115, "C": 115, "D": 100, "E": 72, "F": 58, "G": 58},
        "freeze_rows": 3,
        "bold_header": True,
    },
}


def upload_to_folder(
    xlsx_path: Path,
    folder_id: str,
    title: str,
    formula_patches: dict[str, dict[str, str]] | None = None,
) -> None:
    """Create (or replace) a Google Sheet in the given Drive folder.

    Searches the folder for any existing non-trashed file with the same name and
    trashes it before creating a fresh spreadsheet. Mirrors the workbook contents
    with live formulas via ``xlsx_to_gsheet``, then applies tab formatting.

    ``formula_patches`` is an optional ``{sheet_name: {cell_addr: formula}}`` map of
    cross-sheet formulas to re-write *after* the full upload.  Because
    ``xlsx_to_gsheet`` writes sheets one at a time, cross-sheet references written
    early are flagged as broken until all tabs exist.  Writing them again here, after
    every tab is present, ensures they evaluate correctly.
    """
    from lib.data.gsheets import (
        apply_sheet_formatting,
        create_sheet_in_folder,
        write_values_with_formulas,
        xlsx_to_gsheet,
    )

    print(f"Uploading '{title}' to Drive folder {folder_id} ...", flush=True)
    spreadsheet = create_sheet_in_folder(title, folder_id)
    xlsx_to_gsheet(xlsx_path, spreadsheet.id, delete_other_tabs=True)

    if formula_patches:
        print("Patching cross-sheet formulas ...", flush=True)
        for sheet_name, patches in formula_patches.items():
            ws = spreadsheet.worksheet(sheet_name)
            for cell_addr, formula in patches.items():
                write_values_with_formulas(ws, [[formula]], start=cell_addr)

    print("Applying formatting ...", flush=True)
    for ws in spreadsheet.worksheets():
        spec = _TAB_FORMATTING.get(ws.title)
        if spec:
            fmt = dict(spec)
            if ws.title == "README" and _README_BOLD_ROWS:
                fmt["bold_rows"] = list(_README_BOLD_ROWS)
            apply_sheet_formatting(ws, **fmt)
    print(
        f"Done. View at https://docs.google.com/spreadsheets/d/{spreadsheet.id}/edit",
        flush=True,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__.splitlines()[0] if __doc__ else "",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("cache/rie_1_14a_subclass_coss.xlsx"),
        help="Output .xlsx path. Default: cache/rie_1_14a_subclass_coss.xlsx",
    )
    parser.add_argument(
        "--folder-id",
        default=DEFAULT_FOLDER_ID,
        help=f"Google Drive folder ID to upload into. Default: {DEFAULT_FOLDER_ID}",
    )
    parser.add_argument(
        "--title",
        default=DEFAULT_TITLE,
        help=f"Name for the Google Sheet. Default: '{DEFAULT_TITLE}'",
    )
    parser.add_argument(
        "--no-upload",
        action="store_true",
        help="Build the .xlsx locally without uploading to Google Drive.",
    )
    args = parser.parse_args(argv)

    out = build_workbook(args.output)
    if not args.no_upload:
        upload_to_folder(out, args.folder_id, args.title)
    return 0


if __name__ == "__main__":
    sys.exit(main())
