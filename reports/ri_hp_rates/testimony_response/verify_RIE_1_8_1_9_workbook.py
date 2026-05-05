"""Verify the RIE 1-8 and RIE 1-9 workbooks reproduce the testimony numbers.

Two checks per workbook:

1. **Structural** — open the xlsx and confirm expected sheets, formula columns.
2. **Numerical** — confirm the save/lose percentages match report_variables.pkl.
"""

from __future__ import annotations

import pickle
import sys
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

REPORT_DIR = Path("/ebs/home/alex_switch_box/reports2/reports/ri_hp_rates")
CACHE_DIR = REPORT_DIR / "cache"


def _load_report_vars() -> SimpleNamespace:
    pkl = CACHE_DIR / "report_variables.pkl"
    assert pkl.exists(), f"Missing {pkl}. Render analysis.qmd first."
    return SimpleNamespace(**pickle.loads(pkl.read_bytes()))


def _structural_check(xlsx_path: Path) -> None:
    print(f"=== Structural check: {xlsx_path.name} ===")
    wb = load_workbook(str(xlsx_path), data_only=False)

    expected_sheets = ["assumptions", "per_building", "result"]
    missing = [s for s in expected_sheets if s not in wb.sheetnames]
    assert not missing, f"missing sheets: {missing}"
    print(f"  sheets: {wb.sheetnames}")

    pb = wb["per_building"]
    headers = [pb.cell(row=1, column=c).value for c in range(1, 22)]
    assert headers[0] == "bldg_id", f"unexpected first header: {headers[0]}"
    assert headers[18] == "delta", f"expected 'delta' in col S, got: {headers[18]}"
    assert headers[19] == "saves", f"expected 'saves' in col T, got: {headers[19]}"
    assert headers[20] == "w_saves", f"expected 'w_saves' in col U, got: {headers[20]}"
    print("  per_building headers: OK")

    cell_s2 = pb["S2"].value
    assert isinstance(cell_s2, str) and cell_s2.startswith("="), f"S2 should be a formula, got: {cell_s2}"
    print("  formulas in per_building: OK")

    result_ws = wb["result"]
    b6 = result_ws["B6"].value
    assert isinstance(b6, str) and "B5" in b6, f"B6 should reference B5, got: {b6}"
    print("  result sheet formulas: OK")
    print()


def _numerical_check(xlsx_path: Path, expected_pct_save: float, label: str) -> None:
    """Check pct_save from raw data columns (formulas aren't evaluated in openpyxl)."""
    print(f"=== Numerical check: {label} ({xlsx_path.name}) ===")
    wb = load_workbook(str(xlsx_path), data_only=False)
    pb = wb["per_building"]

    # LMI bill columns (D=elec_lmi_before, F=gas_lmi_before, G=oil_before,
    # H=propane_before, J=elec_lmi_after, L=gas_lmi_after, M=oil_after,
    # N=propane_after); weight = col B (index 1)
    total_w = 0.0
    save_w = 0.0
    for row in pb.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        weight = float(row[1])
        total_w += weight
        # D(3) + F(5) + G(6) + H(7) = energy_bill_lmi_before
        before = float(row[3]) + float(row[5]) + float(row[6]) + float(row[7])
        # J(9) + L(11) + M(12) + N(13) = energy_bill_lmi_after
        after = float(row[9]) + float(row[11]) + float(row[12]) + float(row[13])
        if after - before < 0:
            save_w += weight

    actual_pct_save = save_w / total_w
    diff = abs(actual_pct_save - expected_pct_save)
    print(f"  expected pct_save: {expected_pct_save:.6f}")
    print(f"  actual pct_save:   {actual_pct_save:.6f}")
    print(f"  delta:             {diff:.6f}")
    assert diff < 0.001, f"pct_save mismatch for {label}: {diff:.6f}"
    print("  PASS")
    print()


def main() -> int:
    v = _load_report_vars()

    xlsx_1_8 = CACHE_DIR / "rie_1_8.xlsx"
    xlsx_1_9 = CACHE_DIR / "rie_1_9.xlsx"

    for p in (xlsx_1_8, xlsx_1_9):
        if not p.exists():
            print(f"ERROR: {p} not found. Run build_RIE_1_8_1_9_workbook.py first.")
            return 1

    _structural_check(xlsx_1_8)
    _structural_check(xlsx_1_9)

    _numerical_check(
        xlsx_1_8,
        expected_pct_save=v.pct_natgas_save_default_lmi40,
        label="RIE 1-8 (default rates, 79% lose)",
    )
    _numerical_check(
        xlsx_1_9,
        expected_pct_save=v.pct_natgas_save_hprate_lmi40,
        label="RIE 1-9 (HP rates, 87% save)",
    )

    print("All checks PASSED.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
