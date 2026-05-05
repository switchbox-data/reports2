"""Build workbooks for RIE 1-8 and RIE 1-9 discovery responses.

RIE 1-8: "Under current default rates, 79% of gas-heated households
in Rhode Island would lose money after switching to heat pumps."

RIE 1-9: "With such rates 87% of gas-heated households would save
money after switching to heat pumps."

Both percentages use LMI-adjusted bills (32% enrollment tier).

Each workbook has four sheets:
  - assumptions: tariff tables, data sources, column descriptions
  - monthly: 12 rows per building -- consumption x rate = bill derivation (no LMI)
  - annual: 1 row per building — SUM of monthly + LMI bills + delta/saves
  - result: headline percentage from annual sheet

Run from the report directory::

    uv run python -m testimony_response.build_RIE_1_8_1_9_workbook
    uv run python -m testimony_response.build_RIE_1_8_1_9_workbook --upload
"""

from __future__ import annotations

import argparse
import datetime
import subprocess
import sys
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from lib.rdp import fetch_rdp_file, parse_urdb_json

# ---------------------------------------------------------------------------
# Constants.
# ---------------------------------------------------------------------------
UTILITY = "rie"
BATCH = "ri_20260331_r1-20_rate_case_test_year"
S3_BASE = "s3://data.sb/switchbox/cairo/outputs/hp_rates"
_state = "ri"

PATH_BILLS_12 = f"{S3_BASE}/{_state}/all_utilities/{BATCH}/run_1+2/comb_bills_year_target/"
PATH_BILLS_34 = f"{S3_BASE}/{_state}/all_utilities/{BATCH}/run_3+4/comb_bills_year_target/"
PATH_BILLS_1920 = f"{S3_BASE}/{_state}/all_utilities/{BATCH}/run_19+20/comb_bills_year_target/"

KWH_BATCH = "ri_20260504_kwh_export_v2"
_KWH_BASE = f"{S3_BASE}/{_state}/rie/{KWH_BATCH}"
PATH_KWH_8760_U0 = f"{_KWH_BASE}/20260505_011359_ri_rie_run1_up00_precalc__default/billing_kwh_8760.parquet"
PATH_KWH_8760_U2 = f"{_KWH_BASE}/20260505_011437_ri_rie_run3_up02_default__default/billing_kwh_8760.parquet"

RESSTOCK_RELEASE = "s3://data.sb/nrel/resstock/res_2024_amy2018_2"
GAS_CONSUMPTION_COL = "out.natural_gas.total.energy_consumption"

RDP_REF = "e9e5088"
RDP_TARIFF_DIR = "rate_design/hp_rates/ri/config/tariffs/electric"

REPORTS2_GITHUB_BASE = "https://github.com/switchbox-data/reports2/blob"

SPREADSHEET_1_8 = "1TEURpTKhM3ddpPagNxq1bT0oB24RyoorD4y7EVfFeUY"
SPREADSHEET_1_9 = "1wocSf02gT7WS9G_vxs70d1SG1IOwhPLcUZGRuHXDRB8"

# ---------------------------------------------------------------------------
# Tariff rates.
# ---------------------------------------------------------------------------
ELEC_DELIVERY_RATE = 0.14058  # $/kWh, flat all months
ELEC_FIXED_PER_MONTH = 10.01  # $6.00 customer + $3.22 RE Growth + $0.79 LIHEAP

# Default supply rates (LRS seasonal, $/kWh)
_ELEC_SUPPLY_WINTER = 0.17254  # Jan-Mar
_ELEC_SUPPLY_SUMMER = 0.10674  # Apr-Sep
_ELEC_SUPPLY_FALL = 0.15376  # Oct-Dec

ELEC_SUPPLY_DEFAULT: dict[str, float] = {
    "Jan": _ELEC_SUPPLY_WINTER,
    "Feb": _ELEC_SUPPLY_WINTER,
    "Mar": _ELEC_SUPPLY_WINTER,
    "Apr": _ELEC_SUPPLY_SUMMER,
    "May": _ELEC_SUPPLY_SUMMER,
    "Jun": _ELEC_SUPPLY_SUMMER,
    "Jul": _ELEC_SUPPLY_SUMMER,
    "Aug": _ELEC_SUPPLY_SUMMER,
    "Sep": _ELEC_SUPPLY_SUMMER,
    "Oct": _ELEC_SUPPLY_FALL,
    "Nov": _ELEC_SUPPLY_FALL,
    "Dec": _ELEC_SUPPLY_FALL,
}
ELEC_COMBINED_DEFAULT: dict[str, float] = {m: ELEC_DELIVERY_RATE + s for m, s in ELEC_SUPPLY_DEFAULT.items()}

HP_FLAT_COMBINED_RATE = 0.23129  # $/kWh, delivery+supply bundled, all months

# Gas tariff rates in $/therm (URDB $/kWh x 29.3)
KWH_PER_THERM = 29.3
GAS_FIXED_PER_MONTH = 14.79

# Rate 12 — Residential Heating (before HP)
_GAS_HEAT_KWH = {
    "Jan": 0.063686,
    "Feb": 0.063686,
    "Mar": 0.063686,
    "Apr": 0.065945,
    "May": 0.063843,
    "Jun": 0.063843,
    "Jul": 0.063843,
    "Aug": 0.063843,
    "Sep": 0.063843,
    "Oct": 0.063843,
    "Nov": 0.058560,
    "Dec": 0.058560,
}
GAS_HEATING_PER_THERM: dict[str, float] = {m: r * KWH_PER_THERM for m, r in _GAS_HEAT_KWH.items()}

# Rate 10 — Residential Non-Heating (after HP)
_GAS_NONHEAT_KWH = {
    "Jan": 0.059563,
    "Feb": 0.059563,
    "Mar": 0.059563,
    "Apr": 0.061823,
    "May": 0.061823,
    "Jun": 0.061823,
    "Jul": 0.061823,
    "Aug": 0.061823,
    "Sep": 0.061823,
    "Oct": 0.061823,
    "Nov": 0.049116,
    "Dec": 0.049116,
}
GAS_NONHEATING_PER_THERM: dict[str, float] = {m: r * KWH_PER_THERM for m, r in _GAS_NONHEAT_KWH.items()}

MONTH_ORDER = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_INT_TO_STR: dict[int, str] = {i + 1: m for i, m in enumerate(MONTH_ORDER)}

# Build hour→month mapping for 2018 (AMY weather year).
_HOUR_TO_MONTH: list[int] = []
for _m in range(1, 13):
    _start = datetime.datetime(2018, _m, 1)
    _end = datetime.datetime(2018, _m + 1, 1) if _m < 12 else datetime.datetime(2019, 1, 1)
    _HOUR_TO_MONTH.extend([_m] * int((_end - _start).total_seconds() // 3600))
assert len(_HOUR_TO_MONTH) == 8760

# Bills columns used from master bills.
LMI_BILL_COLS = [
    "elec_total_bill_lmi_32",
    "gas_total_bill_lmi_32",
    "oil_total_bill",
    "propane_total_bill",
]

SCENARIOS: dict[str, dict[str, str]] = {
    "1-8": {
        "request": (
            "RIE 1-8: Under current default rates, 79% of gas-heated "
            "households in Rhode Island would lose money after switching "
            "to heat pumps."
        ),
        "before_desc": "Run 1+2: Current HVAC, default delivery + supply rates",
        "after_desc": "Run 3+4: Heat pump upgrade, same default rates",
        "before_path": PATH_BILLS_12,
        "after_path": PATH_BILLS_34,
        "headline": "Percentage that LOSE money (1 - pct_save)",
    },
    "1-9": {
        "request": (
            "RIE 1-9: With such rates 87% of gas-heated households would save money after switching to heat pumps."
        ),
        "before_desc": "Run 1+2: Current HVAC, default delivery + supply rates",
        "after_desc": "Run 19+20: Heat pump upgrade, proposed HP flat rate",
        "before_path": PATH_BILLS_12,
        "after_path": PATH_BILLS_1920,
        "headline": "Percentage that SAVE money (pct_save)",
    },
}


# ---------------------------------------------------------------------------
# Permalink helpers.
# ---------------------------------------------------------------------------
def _reports2_head_sha() -> str:
    if not hasattr(_reports2_head_sha, "_cached"):
        repo_root = Path(__file__).resolve().parents[3]
        sha = subprocess.check_output(
            ["git", "-C", str(repo_root), "rev-parse", "HEAD"],
            text=True,
        ).strip()
        _reports2_head_sha._cached = sha  # type: ignore[attr-defined]
    return _reports2_head_sha._cached  # type: ignore[attr-defined]


def _reports2_permalink(rel_path: str) -> str:
    return f"{REPORTS2_GITHUB_BASE}/{_reports2_head_sha()}/{rel_path}"


# ---------------------------------------------------------------------------
# Data loading.
# ---------------------------------------------------------------------------
def _load_monthly_elec_kwh(path_8760: str) -> pl.DataFrame:
    """Load 8760 billing kWh and roll to monthly."""
    hour_map = pl.DataFrame(
        {
            "hour": list(range(8760)),
            "month_int": _HOUR_TO_MONTH,
        }
    ).with_columns(pl.col("hour").cast(pl.Int16))

    return (
        pl.scan_parquet(path_8760)
        .join(hour_map.lazy(), on="hour")
        .group_by("bldg_id", "month_int")
        .agg(pl.col("grid_cons_kwh").sum().alias("elec_kwh"))
        .sort("bldg_id", "month_int")
        .with_columns(pl.col("month_int").replace_strict(MONTH_INT_TO_STR, return_dtype=pl.String).alias("month"))
        .select("bldg_id", "month", "elec_kwh")
        .collect()  # ty: ignore[invalid-return-type]
    )


def _load_monthly_gas_kwh(bldg_ids: list[int], upgrade: str) -> pl.DataFrame:
    """Load ResStock monthly gas consumption for specific buildings."""
    upgrade_int = str(int(upgrade))
    base = f"{RESSTOCK_RELEASE}/load_curve_monthly/state=RI/upgrade={upgrade}"
    paths = [f"{base}/{bid}-{upgrade_int}.parquet" for bid in bldg_ids]
    return (
        pl.scan_parquet(paths)
        .select(
            "bldg_id",
            pl.col("month").cast(pl.Int8).alias("month_int"),
            pl.col(GAS_CONSUMPTION_COL).fill_null(0.0).alias("gas_kwh"),
        )
        .sort("bldg_id", "month_int")
        .with_columns(pl.col("month_int").replace_strict(MONTH_INT_TO_STR, return_dtype=pl.String).alias("month"))
        .select("bldg_id", "month", "gas_kwh")
        .collect()  # ty: ignore[invalid-return-type]
    )


def _load_bills(path: str) -> pl.DataFrame:
    """Load master bills (all months including Annual) for RIE."""
    return (
        pl.scan_parquet(path, hive_partitioning=True)
        .filter(pl.col("sb.electric_utility") == UTILITY)
        .select(
            "bldg_id",
            "weight",
            "heats_with_natgas",
            "month",
            "elec_total_bill",
            "elec_total_bill_lmi_32",
            "gas_total_bill",
            "gas_total_bill_lmi_32",
            "oil_total_bill",
            "propane_total_bill",
        )
        .collect()  # ty: ignore[invalid-return-type]
    )


def _load_inputs() -> dict:
    """Pull tariff rates and report tariff-table values."""
    import pickle

    def _fetch_tariff(filename: str) -> dict:
        return parse_urdb_json(fetch_rdp_file(f"{RDP_TARIFF_DIR}/{filename}", RDP_REF))

    hp_flat_del = _fetch_tariff("rie_hp_flat_calibrated.json")
    hp_flat_sup = _fetch_tariff("rie_hp_flat_supply_calibrated.json")

    rv = pickle.loads(Path("cache/report_variables.pkl").read_bytes())

    return {
        "tariff_table": {
            "elec_customer_charge": rv["elec_customer_charge"],
            "elec_other_fixed_charges": rv["elec_other_fixed_charges"],
            "elec_delivery_avg_cents": rv["elec_delivery_avg_cents"],
            "elec_lrs_winter_cents": rv["elec_lrs_winter_cents"],
            "elec_lrs_summer_cents": rv["elec_lrs_summer_cents"],
            "elec_lrs_fall_cents": rv["elec_lrs_fall_cents"],
            "gas_fixed_charge": rv["gas_fixed_charge"],
            "gas_nonheating_fixed_charge": rv["gas_nonheating_fixed_charge"],
            "gas_avg_per_therm": rv["gas_avg_per_therm"],
            "gas_nonheating_avg_per_therm": rv["gas_nonheating_avg_per_therm"],
            "oil_price_range": rv.get("oil_price_range", "$3.48–$4.10 /gal (monthly prices)"),  # noqa: RUF001
            "propane_price_range": rv.get("propane_price_range", "$3.44–$3.67 /gal (monthly prices)"),  # noqa: RUF001
            "hp_flat_delivery_cents": float(hp_flat_del["items"][0]["energyratestructure"][0][0]["rate"]) * 100,
            "hp_flat_supply_cents": float(hp_flat_sup["items"][0]["energyratestructure"][0][0]["rate"]) * 100,
        },
    }


# ---------------------------------------------------------------------------
# Formatting helpers.
# ---------------------------------------------------------------------------
def _header_fill(ws, row: int, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="E8E8E8")
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = fill


def _autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Sheet writers.
# ---------------------------------------------------------------------------
def _write_assumptions(wb: Workbook, scenario_id: str, inputs: dict) -> None:
    ws = wb.create_sheet("assumptions", 0)
    scen = SCENARIOS[scenario_id]
    tt = inputs["tariff_table"]

    ws["A1"] = f"RIE {scenario_id} — Supporting calculation"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    row = 3

    kv_rows: list[tuple[str, str]] = [
        ("Discovery request", scen["request"]),
        ("", ""),
        ("Assumptions", ""),
        (
            "LMI discount tier",
            "32% of eligible households are enrolled and receive discount. The _lmi_32 bill columns "
            "apply existing utility LMI discounts to enrolled low-income "
            "households. Non-enrolled households see the undiscounted bill.",
        ),
        (
            "Bill columns used",
            "elec_total_bill_lmi_32 + gas_total_bill_lmi_32 + oil_total_bill + propane_total_bill",
        ),
        (
            "Gas-heated filter",
            "heats_with_natgas == true (ResStock heating fuel assignment)",
        ),
        (
            "Save definition",
            "delta < 0 (annual LMI-adjusted energy bill decreases after HP)",
        ),
        (
            "Lose definition",
            "delta > 0 (annual LMI-adjusted energy bill increases after HP)",
        ),
    ]
    bold_labels = {"Assumptions", "Data sources", "monthly columns", "annual columns"}
    for label, value in kv_rows:
        ws.cell(row=row, column=1, value=label)
        if label in bold_labels:
            ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        row += 1

    # --- Default tariff table ------------------------------------------------
    row += 1
    ws.cell(row=row, column=1, value="Default tariffs (before)")
    ws.cell(row=row, column=1).font = Font(bold=True)
    cite = "Rhode Island Energy, Docket No. 2545-GE (2025). Blazunas Schedules & Workpapers, Book 21, PRB-1-ELEC."
    ws.cell(row=row, column=2, value=cite)
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    row += 1

    _header_fill(ws, row, 3)
    for c, hdr in enumerate(["Tariff", "Fixed charge", "Volumetric rate"], 1):
        ws.cell(row=row, column=c, value=hdr)
    row += 1

    default_tariff_rows: list[tuple[str, str, str]] = [
        ("Electricity — customer charge (A-16)", f"${tt['elec_customer_charge']:.2f} /mo", "—"),
        ("Electricity — RE Growth + LIHEAP (A-16)", f"${tt['elec_other_fixed_charges']:.2f} /mo", "—"),
        ("Electricity — delivery (A-16)", "—", f"{tt['elec_delivery_avg_cents']:.2f} ¢/kWh"),
        ("Electricity — supply, winter (LRS, Jan–Mar)", "—", f"{tt['elec_lrs_winter_cents']:.2f} ¢/kWh"),  # noqa: RUF001
        ("Electricity — supply, summer (LRS, Apr–Sep)", "—", f"{tt['elec_lrs_summer_cents']:.2f} ¢/kWh"),  # noqa: RUF001
        ("Electricity — supply, fall (LRS, Oct–Dec)", "—", f"{tt['elec_lrs_fall_cents']:.2f} ¢/kWh"),  # noqa: RUF001
        (
            "Natural gas — Rate 12 (heating)",
            f"${tt['gas_fixed_charge']:.2f} /mo",
            f"${tt['gas_avg_per_therm']:.3f} /therm (avg)",
        ),
        (
            "Natural gas — Rate 10 (non-heating)",
            f"${tt['gas_nonheating_fixed_charge']:.2f} /mo",
            f"${tt['gas_nonheating_avg_per_therm']:.3f} /therm (avg)",
        ),
        ("Heating oil (EIA)", "—", "$3.48–$4.10 /gal (monthly prices)"),  # noqa: RUF001
        ("Propane (EIA)", "—", "$3.44–$3.67 /gal (monthly prices)"),  # noqa: RUF001
    ]
    for tariff, fixed, vol in default_tariff_rows:
        ws.cell(row=row, column=1, value=tariff)
        ws.cell(row=row, column=2, value=fixed)
        ws.cell(row=row, column=3, value=vol)
        row += 1

    # --- After tariffs (scenario-specific) ----------------------------------
    if scenario_id == "1-9":
        row += 1
        ws.cell(row=row, column=1, value="HP subclass tariffs (after)")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(
            row=row,
            column=2,
            value=(
                "Delivery: EPMC allocation. Supply: per-customer allocation. Gas, oil, and propane tariffs unchanged."
            ),
        )
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        row += 1
        _header_fill(ws, row, 3)
        for c, hdr in enumerate(["Tariff", "Fixed charge", "Volumetric rate"], 1):
            ws.cell(row=row, column=c, value=hdr)
        row += 1
        hp_rows: list[tuple[str, str, str]] = [
            ("Electricity — customer charge", f"${tt['elec_customer_charge']:.2f} /mo", "—"),
            ("Electricity — RE Growth + LIHEAP", f"${tt['elec_other_fixed_charges']:.2f} /mo", "—"),
            ("Electricity — HP delivery (EPMC)", "—", f"{tt['hp_flat_delivery_cents']:.2f} ¢/kWh"),
            ("Electricity — HP supply (per-customer)", "—", f"{tt['hp_flat_supply_cents']:.2f} ¢/kWh"),
        ]
        for tariff, fixed, vol in hp_rows:
            ws.cell(row=row, column=1, value=tariff)
            ws.cell(row=row, column=2, value=fixed)
            ws.cell(row=row, column=3, value=vol)
            row += 1
    else:
        row += 1
        ws.cell(row=row, column=1, value="After tariffs")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value="Same default tariffs as above (HP upgrade keeps current rates).")
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        row += 1

    # --- Gas seasonal rate detail -------------------------------------------
    row += 1
    ws.cell(row=row, column=1, value="Gas seasonal rates ($/therm)")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    _header_fill(ws, row, 4)
    for c, hdr in enumerate(["Month", "Heating (Rate 12)", "Non-heating (Rate 10)", "Notes"], 1):
        ws.cell(row=row, column=c, value=hdr)
    row += 1
    for m in MONTH_ORDER:
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=round(GAS_HEATING_PER_THERM[m], 4))
        ws.cell(row=row, column=3, value=round(GAS_NONHEATING_PER_THERM[m], 4))
        row += 1

    # --- Data sources -------------------------------------------------------
    row += 1
    ds_rows: list[tuple[str, str]] = [
        ("Data sources", ""),
        ("Before bills", f"{scen['before_desc']}\n{scen['before_path']}"),
        ("After bills", f"{scen['after_desc']}\n{scen['after_path']}"),
        ("Electric kWh (8760, upgrade 0)", PATH_KWH_8760_U0),
        ("Electric kWh (8760, upgrade 2)", PATH_KWH_8760_U2),
        ("Gas consumption (ResStock monthly)", f"{RESSTOCK_RELEASE}/load_curve_monthly/state=RI/upgrade={{00,02}}/"),
        (
            "Analysis notebook",
            _reports2_permalink("reports/ri_hp_rates/notebooks/analysis.qmd"),
        ),
    ]
    for label, value in ds_rows:
        ws.cell(row=row, column=1, value=label)
        if label in bold_labels:
            ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        row += 1

    # --- monthly columns description ----------------------------------------
    row += 1
    mcol_rows: list[tuple[str, str]] = [
        ("monthly columns", ""),
        ("bldg_id", "ResStock building identifier"),
        ("month", "Calendar month (Jan-Dec)"),
        (
            "elec_kwh_before",
            "Monthly electric kWh, upgrade 0. Rolled from CAIRO 8760 billing kWh (post PV-clip, scale factor, timeshift).",
        ),
        ("elec_kwh_after", "Monthly electric kWh, upgrade 2. Same source."),
        ("gas_therms_before", "Monthly gas therms, upgrade 0. From ResStock load_curve_monthly (/ 29.3 kWh/therm)."),
        ("gas_therms_after", "Monthly gas therms, upgrade 2. Same source."),
        ("elec_rate_before", "Combined electric rate (delivery + supply) for that month, $/kWh."),
        ("elec_rate_after", "Combined electric rate for after scenario, $/kWh."),
        ("elec_fixed", "$10.01/mo ($6.00 customer + $3.22 RE Growth + $0.79 LIHEAP)."),
        ("gas_rate_before", "Gas heating tariff (Rate 12), seasonal $/therm."),
        ("gas_rate_after", "Gas non-heating tariff (Rate 10), seasonal $/therm."),
        ("gas_fixed", "$14.79/mo."),
        ("elec_bill_before", "Formula: elec_kwh_before x elec_rate_before + elec_fixed."),
        ("elec_bill_after", "Formula: elec_kwh_after x elec_rate_after + elec_fixed."),
        ("gas_bill_before", "Formula: gas_therms_before x gas_rate_before + gas_fixed."),
        ("gas_bill_after", "Formula: gas_therms_after x gas_rate_after + gas_fixed."),
    ]
    for label, value in mcol_rows:
        ws.cell(row=row, column=1, value=label)
        if label in bold_labels:
            ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        row += 1

    # --- annual columns description -----------------------------------------
    row += 1
    acol_rows: list[tuple[str, str]] = [
        ("annual columns", ""),
        ("bldg_id", "ResStock building identifier"),
        ("weight", "CAIRO sample weight"),
        ("annual_kwh_before / after", "SUM of monthly electric kWh"),
        ("annual_therms_before / after", "SUM of monthly gas therms"),
        ("elec_bill_before / after", "SUM of monthly elec_bill formula"),
        (
            "elec_bill_lmi_before / after",
            "If building received LMI discount: discounted bill from master bills. Otherwise: =elec_bill (same as undiscounted).",
        ),
        ("gas_bill_before / after", "SUM of monthly gas_bill formula"),
        (
            "gas_bill_lmi_before / after",
            "If building received LMI discount: discounted bill from master bills. Otherwise: =gas_bill (same as undiscounted).",
        ),
        ("oil_bill_before / after", "Annual oil bill (data from master bills)"),
        ("propane_bill_before / after", "Annual propane bill (data from master bills)"),
        ("energy_bill_lmi_before / after", "Sum of LMI-adjusted fuel bills"),
        ("delta", "energy_bill_lmi_after - energy_bill_lmi_before"),
        ("saves", "1 if delta < 0, else 0"),
        ("w_saves", "weight x saves"),
    ]
    for label, value in acol_rows:
        ws.cell(row=row, column=1, value=label)
        if label in bold_labels:
            ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        row += 1

    _autosize(ws, {"A": 44, "B": 40, "C": 36, "D": 36})
    ws.sheet_view.showGridLines = False


def _write_monthly(
    wb: Workbook,
    monthly_data: pl.DataFrame,
    scenario_id: str,
) -> int:
    """Write monthly derivation sheet. Returns number of buildings."""
    ws = wb.create_sheet("monthly")

    headers = [
        "bldg_id",  # A
        "month",  # B
        "elec_kwh_before",  # C
        "elec_kwh_after",  # D
        "gas_therms_before",  # E  formula
        "gas_therms_after",  # F  formula
        "elec_rate_before",  # G  data
        "elec_rate_after",  # H  data
        "elec_fixed",  # I  data
        "gas_rate_before",  # J  data
        "gas_rate_after",  # K  data
        "gas_fixed",  # L  data
        "elec_bill_before",  # M  formula
        "elec_bill_after",  # N  formula
        "gas_bill_before",  # O  formula
        "gas_bill_after",  # P  formula
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    elec_rate_after_map: dict[str, float]
    if scenario_id == "1-8":
        elec_rate_after_map = ELEC_COMBINED_DEFAULT
    else:
        elec_rate_after_map = dict.fromkeys(MONTH_ORDER, HP_FLAT_COMBINED_RATE)

    bldg_ids = monthly_data["bldg_id"].unique().sort().to_list()
    n_buildings = len(bldg_ids)

    r = 2  # current row
    for bid in bldg_ids:
        bldg_rows = monthly_data.filter(pl.col("bldg_id") == bid)
        for m in MONTH_ORDER:
            mrow = bldg_rows.filter(pl.col("month") == m)
            elec_kwh_b = float(mrow["elec_kwh_before"][0])
            elec_kwh_a = float(mrow["elec_kwh_after"][0])
            gas_kwh_b = float(mrow["gas_kwh_before"][0])
            gas_kwh_a = float(mrow["gas_kwh_after"][0])

            elec_rate_b = ELEC_COMBINED_DEFAULT[m]
            elec_rate_a = elec_rate_after_map[m]
            gas_rate_b = GAS_HEATING_PER_THERM[m]
            gas_rate_a = GAS_NONHEATING_PER_THERM[m]

            ws.cell(row=r, column=1, value=bid)  # A: bldg_id
            ws.cell(row=r, column=2, value=m)  # B: month
            ws.cell(row=r, column=3, value=elec_kwh_b)  # C: elec_kwh_before
            ws.cell(row=r, column=4, value=elec_kwh_a)  # D: elec_kwh_after
            # E: gas_therms_before = gas_kwh / 29.3
            ws.cell(row=r, column=5, value=gas_kwh_b / KWH_PER_THERM)
            # F: gas_therms_after
            ws.cell(row=r, column=6, value=gas_kwh_a / KWH_PER_THERM)
            ws.cell(row=r, column=7, value=elec_rate_b)  # G
            ws.cell(row=r, column=8, value=elec_rate_a)  # H
            ws.cell(row=r, column=9, value=ELEC_FIXED_PER_MONTH)  # I
            ws.cell(row=r, column=10, value=gas_rate_b)  # J
            ws.cell(row=r, column=11, value=gas_rate_a)  # K
            ws.cell(row=r, column=12, value=GAS_FIXED_PER_MONTH)  # L
            # M: elec_bill_before = C*G + I
            ws.cell(row=r, column=13, value=f"=C{r}*G{r}+I{r}")
            # N: elec_bill_after = D*H + I
            ws.cell(row=r, column=14, value=f"=D{r}*H{r}+I{r}")
            # O: gas_bill_before = E*J + L
            ws.cell(row=r, column=15, value=f"=E{r}*J{r}+L{r}")
            # P: gas_bill_after = F*K + L
            ws.cell(row=r, column=16, value=f"=F{r}*K{r}+L{r}")
            r += 1

    # Number formatting
    last_row = r - 1
    for row_i in range(2, last_row + 1):
        for col in "CD":
            ws[f"{col}{row_i}"].number_format = "#,##0.0"
        for col in "EF":
            ws[f"{col}{row_i}"].number_format = "#,##0.00"
        for col in "GH":
            ws[f"{col}{row_i}"].number_format = "0.00000"
        ws[f"I{row_i}"].number_format = '"$"#,##0.00'
        for col in "JK":
            ws[f"{col}{row_i}"].number_format = '"$"#,##0.0000'
        ws[f"L{row_i}"].number_format = '"$"#,##0.00'
        for col in "MNOP":
            ws[f"{col}{row_i}"].number_format = '"$"#,##0.00'

    _autosize(
        ws,
        {
            "A": 10,
            "B": 8,
            "C": 16,
            "D": 16,
            "E": 18,
            "F": 18,
            "G": 16,
            "H": 16,
            "I": 10,
            "J": 16,
            "K": 16,
            "L": 10,
            "M": 16,
            "N": 16,
            "O": 16,
            "P": 16,
        },
    )
    return n_buildings


def _write_annual(
    wb: Workbook,
    annual_lmi: pl.DataFrame,
    scenario_id: str,
    n_buildings: int,
) -> int:
    """Write annual summary sheet. Returns the last data row number."""
    ws = wb.create_sheet("annual")

    headers = [
        "bldg_id",  # A
        "weight",  # B
        "annual_kwh_before",  # C  SUM
        "annual_kwh_after",  # D  SUM
        "annual_therms_before",  # E  SUM
        "annual_therms_after",  # F  SUM
        "elec_bill_before",  # G  SUM
        "elec_bill_lmi_before",  # H  data
        "gas_bill_before",  # I  SUM
        "gas_bill_lmi_before",  # J  data
        "oil_bill_before",  # K  data
        "propane_bill_before",  # L  data
        "elec_bill_after",  # M  SUM
        "elec_bill_lmi_after",  # N  data
        "gas_bill_after",  # O  SUM
        "gas_bill_lmi_after",  # P  data
        "oil_bill_after",  # Q  data
        "propane_bill_after",  # R  data
        "energy_bill_lmi_before",  # S  formula
        "energy_bill_lmi_after",  # T  formula
        "delta",  # U  formula
        "saves",  # V  formula
        "w_saves",  # W  formula
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    data_cols = [
        "bldg_id",
        "weight",
        "has_elec_discount",
        "has_gas_discount",
        "elec_total_bill_lmi_32_before",
        "gas_total_bill_lmi_32_before",
        "oil_total_bill_before",
        "propane_total_bill_before",
        "has_elec_discount_after",
        "has_gas_discount_after",
        "elec_total_bill_lmi_32_after",
        "gas_total_bill_lmi_32_after",
        "oil_total_bill_after",
        "propane_total_bill_after",
    ]

    rows_list = list(annual_lmi.select(data_cols).iter_rows())
    for i, row_data in enumerate(rows_list):
        ar = i + 2  # annual sheet row
        ms = 2 + i * 12  # monthly start row
        me = ms + 11  # monthly end row

        (
            bid,
            weight,
            has_elec_disc_b,
            has_gas_disc_b,
            elec_lmi_b,
            gas_lmi_b,
            oil_b,
            propane_b,
            has_elec_disc_a,
            has_gas_disc_a,
            elec_lmi_a,
            gas_lmi_a,
            oil_a,
            propane_a,
        ) = row_data

        ws.cell(row=ar, column=1, value=bid)  # A
        ws.cell(row=ar, column=2, value=float(weight))  # B
        ws.cell(row=ar, column=3, value=f"=SUM(monthly!C{ms}:C{me})")  # C: kwh before
        ws.cell(row=ar, column=4, value=f"=SUM(monthly!D{ms}:D{me})")  # D: kwh after
        ws.cell(row=ar, column=5, value=f"=SUM(monthly!E{ms}:E{me})")  # E: therms before
        ws.cell(row=ar, column=6, value=f"=SUM(monthly!F{ms}:F{me})")  # F: therms after
        ws.cell(row=ar, column=7, value=f"=SUM(monthly!M{ms}:M{me})")  # G: elec bill before (SUM)
        # H: elec LMI before — discounted value if LMI, else =G (matches undiscounted)
        if has_elec_disc_b:
            ws.cell(row=ar, column=8, value=float(elec_lmi_b))
        else:
            ws.cell(row=ar, column=8, value=f"=G{ar}")
        ws.cell(row=ar, column=9, value=f"=SUM(monthly!O{ms}:O{me})")  # I: gas bill before (SUM)
        # J: gas LMI before
        if has_gas_disc_b:
            ws.cell(row=ar, column=10, value=float(gas_lmi_b))
        else:
            ws.cell(row=ar, column=10, value=f"=I{ar}")
        ws.cell(row=ar, column=11, value=float(oil_b))  # K: oil before
        ws.cell(row=ar, column=12, value=float(propane_b))  # L: propane before
        ws.cell(row=ar, column=13, value=f"=SUM(monthly!N{ms}:N{me})")  # M: elec bill after (SUM)
        # N: elec LMI after
        if has_elec_disc_a:
            ws.cell(row=ar, column=14, value=float(elec_lmi_a))
        else:
            ws.cell(row=ar, column=14, value=f"=M{ar}")
        ws.cell(row=ar, column=15, value=f"=SUM(monthly!P{ms}:P{me})")  # O: gas bill after (SUM)
        # P: gas LMI after
        if has_gas_disc_a:
            ws.cell(row=ar, column=16, value=float(gas_lmi_a))
        else:
            ws.cell(row=ar, column=16, value=f"=O{ar}")
        ws.cell(row=ar, column=17, value=float(oil_a))  # Q: oil after
        ws.cell(row=ar, column=18, value=float(propane_a))  # R: propane after
        # S: energy_bill_lmi_before = H + J + K + L
        ws.cell(row=ar, column=19, value=f"=H{ar}+J{ar}+K{ar}+L{ar}")
        # T: energy_bill_lmi_after = N + P + Q + R
        ws.cell(row=ar, column=20, value=f"=N{ar}+P{ar}+Q{ar}+R{ar}")
        # U: delta
        ws.cell(row=ar, column=21, value=f"=T{ar}-S{ar}")
        # V: saves
        ws.cell(row=ar, column=22, value=f"=IF(U{ar}<0,1,0)")
        # W: w_saves
        ws.cell(row=ar, column=23, value=f"=B{ar}*V{ar}")

    last_row = 1 + len(rows_list)

    for r in range(2, last_row + 1):
        for col in "CD":
            ws[f"{col}{r}"].number_format = "#,##0"
        for col in "EF":
            ws[f"{col}{r}"].number_format = "#,##0.0"
        for col in "GHIJKLMNOPQRSTU":
            ws[f"{col}{r}"].number_format = '"$"#,##0.00'

    _autosize(
        ws,
        {
            "A": 10,
            "B": 10,
            "C": 18,
            "D": 16,
            "E": 18,
            "F": 18,
            "G": 16,
            "H": 18,
            "I": 16,
            "J": 18,
            "K": 14,
            "L": 16,
            "M": 16,
            "N": 18,
            "O": 16,
            "P": 18,
            "Q": 14,
            "R": 16,
            "S": 22,
            "T": 22,
            "U": 14,
            "V": 8,
            "W": 12,
        },
    )
    return last_row


def _write_result(wb: Workbook, last_row: int, scenario_id: str) -> None:
    """Summary sheet deriving the headline percentage."""
    ws = wb.create_sheet("result")
    scen = SCENARIOS[scenario_id]

    ws["A1"] = f"RIE {scenario_id} — Result"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    hdr_row = 3
    for ci, h in enumerate(["", "Value", "Formula / Notes"], start=1):
        ws.cell(row=hdr_row, column=ci, value=h)
    _header_fill(ws, hdr_row, 3)

    ws.cell(row=4, column=1, value="Gas-heated customers (weighted)")
    ws.cell(row=4, column=2, value=f"=SUM(annual!B2:B{last_row})")
    ws.cell(row=4, column=3, value="SUM of weights (all rows are gas-heated)")

    ws.cell(row=5, column=1, value="Gas-heated customers that save")
    ws.cell(row=5, column=2, value=f"=SUM(annual!W2:W{last_row})")
    ws.cell(row=5, column=3, value="SUM of weighted saves indicator")

    ws.cell(row=6, column=1, value="Percentage that save")
    ws.cell(row=6, column=2, value="=B5/B4")
    ws.cell(row=6, column=3, value="weighted_savers / total_weighted")
    ws.cell(row=6, column=2).number_format = "0.0%"

    ws.cell(row=7, column=1, value="Percentage that lose")
    ws.cell(row=7, column=2, value="=1-B6")
    ws.cell(row=7, column=3, value="1 - pct_save")
    ws.cell(row=7, column=2).number_format = "0.0%"

    ws.cell(row=9, column=1, value="Headline figure")
    ws.cell(row=9, column=1).font = Font(bold=True)
    ws.cell(row=9, column=2, value=scen["headline"])

    for r in (4, 5):
        ws[f"B{r}"].number_format = "#,##0.00"

    if scenario_id == "1-8":
        ws.cell(row=7, column=1).font = Font(bold=True)
        ws.cell(row=7, column=2).font = Font(bold=True)
    else:
        ws.cell(row=6, column=1).font = Font(bold=True)
        ws.cell(row=6, column=2).font = Font(bold=True)

    _autosize(ws, {"A": 32, "B": 20, "C": 40})
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Orchestration.
# ---------------------------------------------------------------------------
def build_workbook(
    output_path: Path,
    scenario_id: str,
    monthly_data: pl.DataFrame,
    annual_lmi: pl.DataFrame,
    inputs: dict,
) -> Path:
    """Build one workbook for the given scenario."""
    total_w = float(annual_lmi["weight"].sum())
    save_w = float(
        annual_lmi.with_columns(
            pl.sum_horizontal([f"{c}_before" for c in LMI_BILL_COLS]).alias("e_before"),
            pl.sum_horizontal([f"{c}_after" for c in LMI_BILL_COLS]).alias("e_after"),
        )
        .filter(pl.col("e_after") < pl.col("e_before"))["weight"]
        .sum()
    )
    pct_save = save_w / total_w

    print(f"\n  RIE {scenario_id}:", flush=True)
    print(f"    gas-heated buildings: {annual_lmi.height:,}", flush=True)
    print(f"    weighted customers: {total_w:,.1f}", flush=True)
    print(f"    pct save: {pct_save:.4f} ({pct_save * 100:.1f}%)", flush=True)
    print(f"    pct lose: {1 - pct_save:.4f} ({(1 - pct_save) * 100:.1f}%)", flush=True)

    if scenario_id == "1-8":
        expected_lose = 0.79
        actual_lose = 1 - pct_save
        assert abs(actual_lose - expected_lose) < 0.02, (
            f"RIE 1-8: expected ~{expected_lose:.0%} lose, got {actual_lose:.1%}"
        )
    else:
        expected_save = 0.87
        assert abs(pct_save - expected_save) < 0.02, f"RIE 1-9: expected ~{expected_save:.0%} save, got {pct_save:.1%}"

    wb = Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    _write_assumptions(wb, scenario_id, inputs)
    n_buildings = _write_monthly(wb, monthly_data, scenario_id)
    last_row = _write_annual(wb, annual_lmi, scenario_id, n_buildings)
    _write_result(wb, last_row, scenario_id)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"    Wrote {output_path} ({output_path.stat().st_size / 1024:.1f} KB)", flush=True)
    return output_path


# ---------------------------------------------------------------------------
# Google Sheets upload.
# ---------------------------------------------------------------------------
def _tab_formatting(scenario_id: str) -> dict[str, dict]:
    return {
        "assumptions": {
            "wrap_columns": ["B:B"],
            "column_widths_px": {"A": 180, "B": 640},
            "freeze_rows": 0,
            "bold_header": False,
        },
        "monthly": {
            "column_number_formats": {
                "C": "#,##0.0",
                "D": "#,##0.0",
                "E": "#,##0.00",
                "F": "#,##0.00",
                "G": "0.00000",
                "H": "0.00000",
                "I": '"$"#,##0.00',
                "J": '"$"#,##0.0000',
                "K": '"$"#,##0.0000',
                "L": '"$"#,##0.00',
                "M": '"$"#,##0.00',
                "N": '"$"#,##0.00',
                "O": '"$"#,##0.00',
                "P": '"$"#,##0.00',
            },
            "auto_resize_columns": ["A:P"],
            "freeze_rows": 1,
            "bold_header": True,
        },
        "annual": {
            "column_number_formats": {
                "C": "#,##0",
                "D": "#,##0",
                "E": "#,##0.0",
                "F": "#,##0.0",
                **dict.fromkeys("GHIJKLMNOPQRSTU", '"$"#,##0.00'),
            },
            "auto_resize_columns": ["A:W"],
            "freeze_rows": 1,
            "bold_header": True,
        },
        "result": {
            "auto_resize_columns": ["A:C"],
            "freeze_rows": 0,
            "bold_header": True,
        },
    }


def upload_to_sheet(xlsx_path: Path, spreadsheet_id: str, scenario_id: str) -> None:
    from lib.data.gsheets import apply_sheet_formatting, xlsx_to_gsheet

    print(f"  Uploading {xlsx_path} -> {spreadsheet_id} ...", flush=True)
    spreadsheet = xlsx_to_gsheet(xlsx_path, spreadsheet_id, delete_other_tabs=True)
    tab_fmt = _tab_formatting(scenario_id)
    for ws in spreadsheet.worksheets():
        spec = tab_fmt.get(ws.title)
        if spec:
            apply_sheet_formatting(ws, **spec)
    print(
        f"  Done. https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit",
        flush=True,
    )


# ---------------------------------------------------------------------------
# CLI.
# ---------------------------------------------------------------------------
def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0] if __doc__ else "")
    parser.add_argument("--upload", action="store_true", help="Upload to Google Sheets after building.")
    args = parser.parse_args(argv)

    print("Loading tariff inputs ...", flush=True)
    inputs = _load_inputs()

    # --- Load electric kWh (8760 → monthly) --------------------------------
    print("Loading billing kWh 8760 (upgrade 0) ...", flush=True)
    elec_monthly_u0 = _load_monthly_elec_kwh(PATH_KWH_8760_U0)
    print(f"  {elec_monthly_u0['bldg_id'].n_unique()} buildings, {elec_monthly_u0.height} rows", flush=True)

    print("Loading billing kWh 8760 (upgrade 2) ...", flush=True)
    elec_monthly_u2 = _load_monthly_elec_kwh(PATH_KWH_8760_U2)
    print(f"  {elec_monthly_u2['bldg_id'].n_unique()} buildings, {elec_monthly_u2.height} rows", flush=True)

    # --- Load master bills -------------------------------------------------
    print("Loading master bills (run 1+2) ...", flush=True)
    bills_12 = _load_bills(PATH_BILLS_12)
    print(f"  {bills_12.height:,} rows", flush=True)

    print("Loading master bills (run 3+4) ...", flush=True)
    bills_34 = _load_bills(PATH_BILLS_34)
    print(f"  {bills_34.height:,} rows", flush=True)

    print("Loading master bills (run 19+20) ...", flush=True)
    bills_1920 = _load_bills(PATH_BILLS_1920)
    print(f"  {bills_1920.height:,} rows", flush=True)

    # --- Get gas-heated building IDs from before bills ---------------------
    annual_before = bills_12.filter(pl.col("month") == "Annual")
    gas_bldg_ids: list[int] = annual_before.filter(pl.col("heats_with_natgas"))["bldg_id"].sort().to_list()
    print(f"Gas-heated buildings: {len(gas_bldg_ids)}", flush=True)

    # --- Load gas consumption from ResStock --------------------------------
    print("Loading ResStock monthly gas (upgrade 0) ...", flush=True)
    gas_monthly_u0 = _load_monthly_gas_kwh(gas_bldg_ids, "00")
    print(f"  {gas_monthly_u0.height} rows", flush=True)

    print("Loading ResStock monthly gas (upgrade 2) ...", flush=True)
    gas_monthly_u2 = _load_monthly_gas_kwh(gas_bldg_ids, "02")
    print(f"  {gas_monthly_u2.height} rows", flush=True)

    # --- Build monthly data for gas-heated buildings -----------------------
    def _build_monthly(scenario_id: str) -> pl.DataFrame:
        """Join electric + gas consumption for 12 months per gas-heated building."""
        elec_rate_after = (
            ELEC_COMBINED_DEFAULT if scenario_id == "1-8" else dict.fromkeys(MONTH_ORDER, HP_FLAT_COMBINED_RATE)
        )
        rate_df = pl.DataFrame(
            {
                "month": MONTH_ORDER,
                "elec_rate_before": [ELEC_COMBINED_DEFAULT[m] for m in MONTH_ORDER],
                "elec_rate_after": [elec_rate_after[m] for m in MONTH_ORDER],
                "gas_rate_before": [GAS_HEATING_PER_THERM[m] for m in MONTH_ORDER],
                "gas_rate_after": [GAS_NONHEATING_PER_THERM[m] for m in MONTH_ORDER],
            }
        )

        elec_before = elec_monthly_u0.filter(pl.col("bldg_id").is_in(gas_bldg_ids)).rename(
            {"elec_kwh": "elec_kwh_before"}
        )

        elec_after = elec_monthly_u2.filter(pl.col("bldg_id").is_in(gas_bldg_ids)).rename(
            {"elec_kwh": "elec_kwh_after"}
        )

        gas_before = gas_monthly_u0.rename({"gas_kwh": "gas_kwh_before"})
        gas_after = gas_monthly_u2.rename({"gas_kwh": "gas_kwh_after"})

        monthly = (
            elec_before.join(elec_after, on=["bldg_id", "month"])
            .join(gas_before, on=["bldg_id", "month"])
            .join(gas_after, on=["bldg_id", "month"])
            .join(rate_df, on="month")
        )

        # Sort by bldg_id, then month order
        month_enum = pl.Enum(MONTH_ORDER)
        return (
            monthly.with_columns(pl.col("month").cast(month_enum).alias("_month_sort"))
            .sort("bldg_id", "_month_sort")
            .drop("_month_sort")
        )

    def _build_annual_lmi(bills_before: pl.DataFrame, bills_after: pl.DataFrame) -> pl.DataFrame:
        """Annual bills (undiscounted + LMI + discount flags) for the annual sheet."""
        all_bill_cols = [*LMI_BILL_COLS]
        before_annual = (
            bills_before.filter((pl.col("month") == "Annual") & pl.col("heats_with_natgas"))
            .with_columns(
                (pl.col("elec_total_bill") != pl.col("elec_total_bill_lmi_32")).alias("has_elec_discount"),
                (pl.col("gas_total_bill") != pl.col("gas_total_bill_lmi_32")).alias("has_gas_discount"),
            )
            .select(
                "bldg_id",
                "weight",
                "has_elec_discount",
                "has_gas_discount",
                *[pl.col(c).alias(f"{c}_before") for c in all_bill_cols],
            )
        )
        after_annual = (
            bills_after.filter((pl.col("month") == "Annual") & pl.col("heats_with_natgas"))
            .with_columns(
                (pl.col("elec_total_bill") != pl.col("elec_total_bill_lmi_32")).alias("has_elec_discount_after"),
                (pl.col("gas_total_bill") != pl.col("gas_total_bill_lmi_32")).alias("has_gas_discount_after"),
            )
            .select(
                "bldg_id",
                "has_elec_discount_after",
                "has_gas_discount_after",
                *[pl.col(c).alias(f"{c}_after") for c in all_bill_cols],
            )
        )
        return before_annual.join(after_annual, on="bldg_id").sort("bldg_id")

    monthly_default = _build_monthly("1-8")
    monthly_hprate = _build_monthly("1-9")
    annual_lmi_default = _build_annual_lmi(bills_12, bills_34)
    annual_lmi_hprate = _build_annual_lmi(bills_12, bills_1920)

    out_1_8 = build_workbook(
        Path("cache/rie_1_8.xlsx"),
        "1-8",
        monthly_default,
        annual_lmi_default,
        inputs,
    )
    out_1_9 = build_workbook(
        Path("cache/rie_1_9.xlsx"),
        "1-9",
        monthly_hprate,
        annual_lmi_hprate,
        inputs,
    )

    if args.upload:
        print("\nUploading ...", flush=True)
        upload_to_sheet(out_1_8, SPREADSHEET_1_8, "1-8")
        upload_to_sheet(out_1_9, SPREADSHEET_1_9, "1-9")

    return 0


if __name__ == "__main__":
    sys.exit(main())
