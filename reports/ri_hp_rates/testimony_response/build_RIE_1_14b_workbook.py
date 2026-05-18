"""Build RIE 1-14b: External studies used in the subclass cost-of-service study.

Responds to RIE 1-14(b): "Please provide, in Excel with all formulas intact,
any external studies performed by Witness Velez whose results are used in the
'cost-of-service study at the subclass level' performed by Witness Velez."

Two external studies feed into the subclass COSS (RIE 1-14a):

1. **Hourly marginal delivery cost derivation** (delivery-only: dist/sub-TX
   and bulk TX). This is the same derivation documented in RIE 1-10, using:
     - EIA-930 utility demand for RIE (S3: eia/hourly_demand/utilities/)
     - ISO-NE 8-zone loads summed to NE system load (S3: isone/hourly_demand/zones/)
     - AESC 2024 capacity costs (Synapse Energy Economics)
   The two components are:
     a) Dist & sub-TX: probability-of-peak (PoP) allocation on top-100 RIE
        load hours. Annualized cost = $101.05/kW-yr (AESC 2024 $80.24/kW-yr
        in 2019$ CPI-adjusted to 2025$).
     b) Bulk TX: exceedance allocation on top-100 NE system load hours.
        Annualized cost = $69/kW-yr (AESC 2024 avoided PTF cost).

2. **RECS 2020 residential heating-system shares for Rhode Island.** EIA's
   Residential Energy Consumption Survey (RECS 2020) heating-fuel shares for
   RI are applied to the Test Year customer count to estimate per-subclass
   customer counts. In the workbook these shares are derived directly from the
   CAIRO simulation weights (which were themselves calibrated to RECS 2020 RI
   shares), using the same master BAT + ResStock metadata approach as DIV 1-2.

Sheets
------
README                  Overview, sources, and cross-references.
inputs_scalars          AESC 2024 costs, N_PEAK_HOURS, and REV_REQ parameters.
rie_load_8760           Raw EIA-930 RIE utility load (MW, 8760 rows).
ne_load_8760            Raw ISO-NE system load summed from 8 zones (MW, 8760 rows).
mc_dist_sub_tx          PoP allocation → dist/sub-TX MC ($/kWh). All formulas.
mc_bulk_tx              Exceedance allocation → bulk TX MC ($/kWh). All formulas.
mc_combined             Delivery MC = dist + bulk TX. Live cross-sheet formulas.
validation_mc           Checks: weight sums = 1.0, annual totals = AESC costs,
                        non-zero hour counts = 100.
validation_mc_hourly    8760-row comparison: workbook formula MC vs S3 parquet
                        MC (dist/sub-TX and bulk TX). All diff columns = 0.
bat_building_data       Per-building BAT: bldg_id, heating_type_v2, weight,
                        annual_bill_delivery, economic_burden_delivery.
recs_subclass_customers RECS 2020 RI heating-system shares (derived from CAIRO
                        simulation weights), Test Year customer count, and
                        per-subclass customer counts.
validation_customers    Cross-checks: shares sum to 1, customers sum to Test Year
                        total.

Usage::

    cd reports/ri_hp_rates

    # Build locally only:
    uv run python -m testimony_response.build_RIE_1_14b_workbook --no-upload

    # Build and upload to default Drive folder (default behavior):
    uv run python -m testimony_response.build_RIE_1_14b_workbook
"""

from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ── Constants (aligned with build_DIV_1_2_workbook.py @ 0b203bc) ─────────────

UTILITY = "rie"
STATE = "ri"
YEAR = 2025
BATCH = "ri_20260507_r1-2_grid_cons_fix"
RUN_DIR = "20260507_213944_ri_rie_run1_up00_precalc__default"
RUN_DELIVERY = "1"
RUN_SUPPLY = "2"

S3_BASE = "s3://data.sb/switchbox/cairo/outputs/hp_rates"
RESSTOCK_BASE = "s3://data.sb/nrel/resstock/res_2024_amy2018_2_sb"

S3_BILLING_KWH_8760 = f"{S3_BASE}/{STATE}/{UTILITY}/{BATCH}/{RUN_DIR}/billing_kwh_8760.parquet"
BILLING_KWH_COL = "grid_cons_kwh"  # matches DIV 1-2

S3_MASTER_BAT = (
    f"{S3_BASE}/{STATE}/all_utilities/{BATCH}/run_{RUN_DELIVERY}+{RUN_SUPPLY}/cross_subsidization_BAT_values/"
)

# Raw EIA / ISO-NE load data (mirrors build_marginal_costs_workbook.py)
S3_EIA_UTILITY_LOADS = "s3://data.sb/eia/hourly_demand/utilities/"
S3_ISONE_ZONE_LOADS = "s3://data.sb/isone/hourly_demand/zones/"

# Pre-computed MC parquets on S3 — used for hourly validation
S3_DIST_SUB_TX = f"s3://data.sb/switchbox/marginal_costs/ri/dist_and_sub_tx/utility={UTILITY}/year={YEAR}/data.parquet"
S3_BULK_TX = f"s3://data.sb/switchbox/marginal_costs/ri/bulk_tx/utility={UTILITY}/year={YEAR}/data.parquet"

# AESC 2024 scalar costs (mirrors build_marginal_costs_workbook.py)
AESC_PTF_KW_YEAR = 69.0
AESC_DIST_KW_YR_2019 = 80.24  # original AESC 2024 value in 2019$; stated in expert testimony §IX
SUB_TX_AND_DIST_MC_KW_YR_2025 = 101.05  # AESC_DIST_KW_YR_2019 CPI-adjusted to 2025$
TOTAL_ANNUALIZED_DELIVERY_MC_KW_YR = AESC_PTF_KW_YEAR + SUB_TX_AND_DIST_MC_KW_YR_2025  # $170.05
N_PEAK_HOURS = 100

# ISO-NE zones for NE system load
ISONE_ALL_ZONES = ["CT", "ME", "NEMA", "NH", "RI", "SEMA", "VT", "WCMA"]

# Revenue-requirement constants — hardcoded from RDP @ 0b203bc
# rate_design/hp_rates/ri/config/rev_requirement/rie_rate_case_test_year.yaml
REV_REQ: dict = {
    "total_delivery_revenue_requirement": 446463143.03,
    "test_year_customer_count": 419347.83,
    "test_year_residential_kwh": 2821237490.0,
    "resstock_kwh_scale_factor": 0.9594257590448669,
}

# Heating subclass definitions (matches cost_of_service_by_subclass.qmd and DIV 1-2)
HT_V2_ORDER = ("heat_pump", "electrical_resistance", "natgas", "delivered_fuels", "other")
HT_V2_LABELS: dict[str, str] = {
    "heat_pump": "Heat pump",
    "electrical_resistance": "Electric resistance",
    "natgas": "Natural gas",
    "delivered_fuels": "Delivered fuels",
    "other": "Other",
}

RDP_REF = "0b203bc"
RDP_GITHUB_BASE = "https://github.com/switchbox-data/rate-design-platform/blob"
REPORTS2_GITHUB_BASE = "https://github.com/switchbox-data/reports2/blob"

DEFAULT_FOLDER_ID = "1uPcJbcOChD6zoFuPb-gsxSByPr7xwmCH"
DEFAULT_TITLE = "RIE 1-14b"

# Styling (matches DIV 1-2)
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
NUMBER_FORMAT_MCPKWH = "0.000000"

# ── Permalink helpers (mirrors both reference scripts) ────────────────────────

_README_BOLD_ROWS: list[int] = []


def _rdp_permalink(rel_path: str) -> str:
    return f"{RDP_GITHUB_BASE}/{RDP_REF}/{rel_path}"


def _reports2_head_sha() -> str:
    if not hasattr(_reports2_head_sha, "_cached"):
        repo_root = Path(__file__).resolve().parents[3]
        sha = subprocess.check_output(
            ["git", "-C", str(repo_root), "rev-parse", "HEAD"],
            text=True,
        ).strip()
        _reports2_head_sha._cached = sha  # type: ignore[attr-defined]
    return _reports2_head_sha._cached  # type: ignore[attr-defined]


def _reports2_permalink(rel_path: str) -> str:
    return f"{REPORTS2_GITHUB_BASE}/{_reports2_head_sha()}/{rel_path}"


# ── Formatting helpers ────────────────────────────────────────────────────────


def _header_fill_light(ws, row: int, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="E8E8E8")
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = fill


def _header_fill_dark(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).font = HEADER_FONT
        ws.cell(row=row, column=c).fill = HEADER_FILL
        ws.cell(row=row, column=c).alignment = Alignment(horizontal="center", wrap_text=True)


def _autosize(ws, widths: dict[str, int]) -> None:
    """Set explicit fixed column widths (in character units). Never auto-fit to content."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Data loading (load curve approach from build_DIV_1_2_workbook.py) ─────────


def load_master_bat() -> pl.DataFrame:
    """Load master BAT data from S3 for run 1+2, filtered to RIE."""
    print(f"Loading master BAT from: {S3_MASTER_BAT}", flush=True)
    df = (
        pl.scan_parquet(S3_MASTER_BAT, hive_partitioning=True)
        .filter(pl.col("sb.electric_utility") == UTILITY)
        .collect()
    )
    assert isinstance(df, pl.DataFrame)
    print(f"  Loaded {df.height:,} buildings for utility={UTILITY}", flush=True)
    return df


def load_mc_parquets(rie_load: pl.DataFrame) -> pl.DataFrame:
    """Load the two pre-computed S3 MC parquets and align to rie_load timestamps.

    Reads the first numeric column from each parquet (same approach as the
    upstream generator scripts).  Results are left-joined to the same 8760
    timestamp sequence used by mc_dist_sub_tx and mc_bulk_tx so that row i in
    the returned DataFrame matches Excel row (i + 4) in those sheets.
    """

    def _load_single(s3_path: str, alias: str) -> pl.DataFrame:
        df = pl.read_parquet(s3_path)
        mc_col = next(c for c in df.columns if c != "timestamp" and df[c].dtype.is_numeric() and c not in ("year",))
        return _strip_tz(df.select("timestamp", pl.col(mc_col).alias(alias)), "timestamp")

    print(f"Loading dist/sub-TX MC parquet from S3: {S3_DIST_SUB_TX}", flush=True)
    dist_df = _load_single(S3_DIST_SUB_TX, "mc_dist_sub_tx_parquet")
    print(f"Loading bulk TX MC parquet from S3: {S3_BULK_TX}", flush=True)
    bulk_df = _load_single(S3_BULK_TX, "mc_bulk_tx_parquet")

    return (
        rie_load.select("timestamp")
        .join(dist_df, on="timestamp", how="left")
        .join(bulk_df, on="timestamp", how="left")
        .with_columns(
            (pl.col("mc_dist_sub_tx_parquet").fill_null(0.0) + pl.col("mc_bulk_tx_parquet").fill_null(0.0)).alias(
                "mc_delivery_total_parquet"
            )
        )
        .sort("timestamp")
    )


# ── Raw load data loaders (mirrors build_marginal_costs_workbook.py) ──────────


def _strip_tz(df: pl.DataFrame, col: str) -> pl.DataFrame:
    if df[col].dtype == pl.Datetime and df[col].dtype.time_zone is not None:  # type: ignore[union-attr]
        return df.with_columns(pl.col(col).dt.replace_time_zone(None))
    return df


def _dedup_to_8760(df: pl.DataFrame, ts_col: str = "timestamp") -> pl.DataFrame:
    numeric_cols = [c for c in df.columns if c != ts_col and df[c].dtype.is_numeric()]
    if df.select(pl.col(ts_col).n_unique()).item() < df.height:
        df = df.group_by(ts_col).agg([pl.col(c).mean() for c in numeric_cols])
    return df.sort(ts_col).head(8760)


def _read_isone_zone_year(base: str, zone: str, year: int) -> pl.DataFrame:
    path = f"{base.rstrip('/')}/zone={zone}/year={year}/"
    df = pl.read_parquet(path + "**/*.parquet")
    return _strip_tz(df, "interval_start_et")


def load_rie_hourly_load() -> pl.DataFrame:
    """Load EIA-930 hourly utility demand for RIE, year 2025.

    Mirrors build_marginal_costs_workbook.py load_rie_hourly_load() exactly.
    Source: s3://data.sb/eia/hourly_demand/utilities/region=isone/utility=rie/year=2025/
    """
    base = S3_EIA_UTILITY_LOADS.rstrip("/")
    path = f"{base}/region=isone/utility={UTILITY}/year={YEAR}/"
    df = pl.read_parquet(path + "**/*.parquet")
    df = _strip_tz(df, "timestamp")
    df = df.select("timestamp", "load_mw")
    return _dedup_to_8760(df)


def load_ne_system_load() -> pl.DataFrame:
    """Load ISO-NE 8-zone loads and sum to NE system load for year 2025.

    Mirrors build_marginal_costs_workbook.py load_ne_system_load() exactly.
    Reads each of the 8 zone partitions directly, then sums to system load.
    Source: s3://data.sb/isone/hourly_demand/zones/zone=*/year=2025/
    """
    base = S3_ISONE_ZONE_LOADS.rstrip("/") + "/"
    frames = []
    for zone in ISONE_ALL_ZONES:
        zdf = _read_isone_zone_year(base, zone, YEAR)
        frames.append(zdf.select("interval_start_et", "load_mw"))
    df = pl.concat(frames)
    ne_load = (
        df.group_by("interval_start_et")
        .agg(pl.col("load_mw").sum().alias("load_mw"))
        .rename({"interval_start_et": "timestamp"})
    )
    return _dedup_to_8760(ne_load)


# ── Peak-hour logic (mirrors build_marginal_costs_workbook.py) ────────────────


def _rank_load(load_df: pl.DataFrame) -> pl.DataFrame:
    """Rank hours by load descending; returns load_df with 'rank' column."""
    rank = (
        load_df.filter(pl.col("load_mw").is_not_null())
        .sort("load_mw", descending=True)
        .with_row_index("rank_0")
        .with_columns((pl.col("rank_0") + 1).cast(pl.Int32).alias("rank"))
        .select("timestamp", "rank")
    )
    return load_df.join(rank, on="timestamp", how="left").sort("timestamp")


# ── Formula cell references into inputs_scalars ───────────────────────────────
# Rows in inputs_scalars (1-indexed, header = row 1):
#   row 2: aesc_ptf_kw_year                     → REF_AESC_PTF
#   row 3: aesc_dist_kw_yr_2019                 (informational only, not used in formulas)
#   row 4: sub_tx_and_dist_mc_kw_yr             → REF_SUB_TX_DIST
#   row 5: total_annualized_delivery_mc_kw_yr   → REF_TOTAL_MC
#   row 6: n_peak_hours                         → REF_N_PEAK
#   row 7: (blank separator)
#   row 8: total_delivery_revenue_requirement
#   row 9: test_year_customer_count             → REF_TY_CUST
#   row 10: test_year_residential_kwh
#   row 11: resstock_kwh_scale_factor
REF_AESC_PTF = "inputs_scalars!$B$2"
REF_SUB_TX_DIST = "inputs_scalars!$B$4"
REF_TOTAL_MC = "inputs_scalars!$B$5"
REF_N_PEAK = "inputs_scalars!$B$6"
REF_TY_CUST = "inputs_scalars!$B$9"


# ── Tab writers ───────────────────────────────────────────────────────────────


def _write_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README", 0)
    rows: list[list] = [
        ["RIE 1-14b: External Studies Used in Subclass Cost-of-Service Study", "", ""],
        ["", "", ""],
        ["Item", "Source", "Notes"],
        [
            "RIE 1-14a (subclass COSS workbook)",
            "https://docs.google.com/spreadsheets/d/1cYGVrLmDTKm9XXEw7jjLVrSqTr8xp7pVK1o0Kk8AOkw/edit",
            "The cost-of-service study that these external studies feed into.",
        ],
        [
            "Generator: dist & sub-TX MC",
            _rdp_permalink("utils/pre/marginal_costs/generate_utility_tx_dx_mc.py"),
            "Produces dist_and_sub_tx parquet via PoP (probability of peak) on RIE utility load.",
        ],
        [
            "Generator: bulk TX MC",
            _rdp_permalink("utils/pre/marginal_costs/generate_bulk_tx_mc.py"),
            "AESC PTF exceedance on NE system load (top 100 hours).",
        ],
        [
            "Core: bulk_tx_isone.py",
            _rdp_permalink("utils/pre/marginal_costs/bulk_tx_isone.py"),
            "ISO-NE bulk TX logic (AESC PTF + NE system peak).",
        ],
        [
            "Scalars CSV",
            _rdp_permalink("rate_design/hp_rates/ri/config/marginal_costs/ri_marginal_costs_2025.csv"),
            "sub_tx_and_dist_mc_kw_yr = $80.24 (2019$), CPI-adjusted to $101.05 (2025$).",
        ],
        [
            "AESC 2024 (Synapse Energy Economics)",
            "https://www.synapse-energy.com/sites/default/files/Avoided-Energy-Supply-Components-2024-24-007.pdf",
            "Source of $69/kW-yr bulk TX (PTF) and $80.24/kW-yr dist capacity costs.",
        ],
        [
            "API: EIA-930 hourly utility loads",
            _rdp_permalink("data/eia/hourly_loads/fetch_zone_loads_parquet.py"),
            f"EIA Hourly Electric Grid Monitor (Form EIA-930). S3: {S3_EIA_UTILITY_LOADS}",
        ],
        [
            "API: ISO-NE zone loads",
            _rdp_permalink("data/isone/hourly_demand/fetch_isone_zone_loads.py"),
            f"ISO-NE CELT hourly demand, 8 zones. S3: {S3_ISONE_ZONE_LOADS}",
        ],
        [
            "Billing kWh 8760 (CAIRO output)",
            S3_BILLING_KWH_8760,
            f"Per-building hourly grid consumption. Column: {BILLING_KWH_COL}. Batch: {BATCH}.",
        ],
        [
            "ResStock metadata",
            f"{RESSTOCK_BASE}/metadata_utility/state=RI/utility_assignment.parquet",
            "bldg_id → weight and heating_type_v2; weights scaled to Test Year customer count.",
        ],
        [
            "Master BAT (run 1+2)",
            S3_MASTER_BAT,
            "Post-processed BAT: per-building delivery bill, economic burden, EPMC residual.",
        ],
        [
            "RECS 2020 (EIA)",
            "https://www.eia.gov/consumption/residential/data/2020/",
            "Residential Energy Consumption Survey. Heating-fuel shares used for subclass customer counts.",
        ],
        [
            "Revenue requirement YAML",
            _rdp_permalink("rate_design/hp_rates/ri/config/rev_requirement/rie_rate_case_test_year.yaml"),
            "rie_rate_case_test_year.yaml @ 0b203bc — hardcoded in this workbook.",
        ],
        [
            "This workbook builder",
            _reports2_permalink("reports/ri_hp_rates/testimony_response/build_RIE_1_14b_workbook.py"),
            "Script that generated this workbook.",
        ],
        ["", "", ""],
        ["Sheet", "What it contains", ""],
        ["inputs_scalars", "AESC 2024 cost parameters, N_PEAK_HOURS, and REV_REQ values.", ""],
        [
            "rie_load_8760",
            "8760 rows: EIA-930 RIE utility load (MW). Raw values.",
            "",
        ],
        [
            "ne_load_8760",
            "8760 rows: ISO-NE system load = sum of 8 zones (MW). Raw values.",
            "",
        ],
        [
            "mc_dist_sub_tx",
            "8760 rows: RIE load rank, PoP weight, dist/sub-TX MC per kWh. All formulas.",
            "",
        ],
        [
            "mc_bulk_tx",
            "8760 rows: NE load rank, exceedance weight, bulk TX MC per kWh. All formulas.",
            "",
        ],
        [
            "mc_combined",
            "8760 rows: delivery MC total = dist/sub-TX + bulk TX. Live cross-sheet formulas.",
            "",
        ],
        [
            "validation_mc",
            "Weight sums = 1.0, annual MC totals = AESC costs, non-zero hours = 100.",
            "",
        ],
        [
            "validation_mc_hourly",
            (
                "8760-row hour-by-hour comparison: workbook formula MC vs pre-computed S3 parquet MC. "
                f"Parquets: {S3_DIST_SUB_TX} and {S3_BULK_TX}. "
                "diff columns should all be 0."
            ),
            "",
        ],
        [
            "bat_building_data",
            "Per-building BAT: bldg_id, heating_type_v2, weight, bill, economic burden.",
            "",
        ],
        [
            "recs_subclass_customers",
            "RECS 2020 RI shares (from CAIRO weights) + Test Year customer counts per subclass.",
            "",
        ],
        ["validation_customers", "RECS shares sum to 1.0; customer counts sum to Test Year total.", ""],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    bold_rows: list[int] = []
    for row_idx, row_data in enumerate(rows, start=1):
        if row_data and row_data[0] in ("Item", "Sheet"):
            _header_fill_light(ws, row_idx, 3)
            bold_rows.append(row_idx)
    _README_BOLD_ROWS.clear()
    _README_BOLD_ROWS.extend(bold_rows)
    _autosize(ws, {"A": 40, "B": 80, "C": 80})
    ws.sheet_view.showGridLines = False


def _write_inputs_scalars(wb: Workbook) -> None:
    ws = wb.create_sheet("inputs_scalars")
    yaml_url = _rdp_permalink("rate_design/hp_rates/ri/config/rev_requirement/rie_rate_case_test_year.yaml")
    aesc_url = "https://www.synapse-energy.com/sites/default/files/Avoided-Energy-Supply-Components-2024-24-007.pdf"
    rows = [
        ["key", "value", "source", "notes"],
        [
            "aesc_ptf_kw_year",
            AESC_PTF_KW_YEAR,
            _rdp_permalink("utils/pre/marginal_costs/bulk_tx_isone.py"),
            "AESC 2024 avoided Pool Transmission Facility cost ($/kW-yr). "
            "Expert testimony §IX: '$69/kW-year' (table and text). "
            'Source: Synapse, "Avoided Energy Supply Components (AESC) in New England: 2024 Report."',
        ],
        [
            "aesc_dist_kw_yr_2019",
            AESC_DIST_KW_YR_2019,
            aesc_url,
            "AESC 2024 avoided distribution capacity cost in original 2019$ ($/kW-yr). "
            "Expert testimony §IX: '$80.24/kW-year' (table and text). "
            "CPI-adjusted to 2025$ below.",
        ],
        [
            "sub_tx_and_dist_mc_kw_yr",
            SUB_TX_AND_DIST_MC_KW_YR_2025,
            _rdp_permalink("rate_design/hp_rates/ri/config/marginal_costs/ri_marginal_costs_2025.csv"),
            "AESC 2024 distribution capacity cost $80.24/kW-yr (2019$) CPI-adjusted to 2025$. "
            "Expert testimony §IX states the original $80.24 value. "
            "Sources: Synapse AESC 2024; U.S. BLS CPIAUCSL.",
        ],
        [
            "total_annualized_delivery_mc_kw_yr",
            TOTAL_ANNUALIZED_DELIVERY_MC_KW_YR,
            aesc_url,
            "Sum of bulk TX ($69) + dist/sub-TX ($101.05) = $170.05/kW-yr. "
            "Both components stated in expert testimony §IX (table). "
            "This is the total annualized delivery marginal cost per kW-year.",
        ],
        [
            "n_peak_hours",
            N_PEAK_HOURS,
            _rdp_permalink("utils/pre/marginal_costs/supply_utils.py"),
            "Top-N hours for peak-driven allocations (PoP and exceedance). "
            "Expert testimony §III: 'top 100 hours of aggregate New England system load "
            "for bulk transmission and the top 100 hours of RIE's system load for "
            "distribution and sub-transmission.' Also §IX (table).",
        ],
        ["", "", "", ""],
        [
            "total_delivery_revenue_requirement",
            REV_REQ["total_delivery_revenue_requirement"],
            yaml_url,
            "Test Year total residential delivery revenue requirement ($). Expert testimony §III and §IX.",
        ],
        [
            "test_year_customer_count",
            REV_REQ["test_year_customer_count"],
            yaml_url,
            "Test Year residential customer count. "
            "Expert testimony §IX: 'the Company's Test Year residential customer count of [419,348]'.",
        ],
        [
            "test_year_residential_kwh",
            REV_REQ["test_year_residential_kwh"],
            yaml_url,
            "Test Year total residential kWh. "
            "Expert testimony §IX: 'the Company's Test Year total of [2.82 billion] kWh'.",
        ],
        [
            "resstock_kwh_scale_factor",
            REV_REQ["resstock_kwh_scale_factor"],
            yaml_url,
            "Scaling factor applied to ResStock kWh to match Test Year total. "
            "Expert testimony §IX: 'a small additional scaling factor'.",
        ],
    ]
    for r in rows:
        ws.append(r)
    _header_fill_light(ws, 1, 4)
    _autosize(ws, {"A": 36, "B": 18, "C": 80, "D": 80})
    ws.sheet_view.showGridLines = False


def _write_raw_load_8760(
    wb: Workbook, load_df: pl.DataFrame, sheet_name: str, load_label: str, source_note: str
) -> None:
    """Write a raw 8760 load table (timestamp + load_mw), values only."""
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"{sheet_name}: {load_label} — 8760 Hours"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:B1")
    ws["A2"] = source_note
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:B2")
    ws.row_dimensions[2].height = 40

    for col_idx, h in enumerate(["timestamp", "load_mw"], start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(load_df.iter_rows(named=True), start=5):
        ws.cell(row=i, column=1, value=row["timestamp"].strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=i, column=2, value=float(row["load_mw"])).number_format = "#,##0.000"

    ws.freeze_panes = "A5"
    _autosize(ws, {"A": 18, "B": 14})
    ws.sheet_view.showGridLines = False


def _write_mc_dist_sub_tx(wb: Workbook, rie_load: pl.DataFrame) -> None:
    """Dist & sub-TX: probability-of-peak (PoP) top-100 on RIE utility load.

    Mirrors build_marginal_costs_workbook.py _write_mc_dist_sub_tx() exactly.
    """
    ws = wb.create_sheet("mc_dist_sub_tx")
    ws["A1"] = "Distribution & Sub-Transmission MC — Probability-of-Peak (PoP) Allocation"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        "PoP method: annualized capacity cost is spread across the top-100 RIE load hours, "
        "weighted in proportion to each hour's load. pop_weight = load_mw / SUM(top-100 loads). "
        f"Annualized cost = ${SUB_TX_AND_DIST_MC_KW_YR_2025}/kW-yr ({REF_SUB_TX_DIST}). "
        f"mc_dist_sub_tx_per_kwh = pop_weight * {REF_SUB_TX_DIST}."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 50

    headers = ["timestamp", "rie_load_mw", "rank", "is_peak_top100", "pop_weight", "mc_dist_sub_tx_per_kwh"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A5"

    joined = _rank_load(rie_load)

    for i, row in enumerate(joined.iter_rows(named=True), start=5):
        r = i  # Excel row = data row
        ws.cell(row=r, column=1, value=row["timestamp"].strftime("%Y-%m-%d %H:%M"))
        if row["load_mw"] is not None:
            ws.cell(row=r, column=2, value=float(row["load_mw"])).number_format = "#,##0.0"
        if row["rank"] is not None:
            ws.cell(row=r, column=3, value=int(row["rank"]))
        ws.cell(row=r, column=4, value=f"=IF(ISNUMBER(C{r}),C{r}<={REF_N_PEAK},FALSE)")
        ws.cell(
            row=r,
            column=5,
            value=f"=IFERROR(IF(D{r},B{r}/SUMPRODUCT(($C$5:$C$8764<={REF_N_PEAK})*IFERROR($B$5:$B$8764,0)),0),0)",
        ).number_format = NUMBER_FORMAT_MCPKWH
        ws.cell(row=r, column=6, value=f"=E{r}*{REF_SUB_TX_DIST}").number_format = NUMBER_FORMAT_MCPKWH

    _autosize(ws, {"A": 18, "B": 14, "C": 8, "D": 14, "E": 14, "F": 24})
    ws.sheet_view.showGridLines = False


def _write_mc_bulk_tx(wb: Workbook, ne_load: pl.DataFrame) -> None:
    """Bulk TX: exceedance allocation on top-100 NE system load hours.

    Mirrors build_marginal_costs_workbook.py _write_exceedance_tab() exactly.
    """
    ws = wb.create_sheet("mc_bulk_tx")
    ws["A1"] = "Bulk Transmission MC — Exceedance Allocation on NE System Load"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        "Exceedance method: the threshold = 101st-highest load. Each of the top-100 hours receives "
        "cost proportional to how far its load exceeds the threshold. "
        "exceedance_weight = max(0, load - threshold) / SUM(top-100 exceedances). "
        f"Annualized cost = ${AESC_PTF_KW_YEAR}/kW-yr ({REF_AESC_PTF}). "
        f"mc_bulk_tx_per_kwh = exceedance_weight * {REF_AESC_PTF}."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 60

    headers = [
        "timestamp",
        "ne_system_load_mw",
        "rank",
        "is_peak_top100",
        "exceedance_threshold",
        "exceedance",
        "exceedance_weight",
        "mc_bulk_tx_per_kwh",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A5"

    joined = _rank_load(ne_load)

    for i, row in enumerate(joined.iter_rows(named=True), start=5):
        r = i
        ws.cell(row=r, column=1, value=row["timestamp"].strftime("%Y-%m-%d %H:%M"))
        if row["load_mw"] is not None:
            ws.cell(row=r, column=2, value=float(row["load_mw"])).number_format = "#,##0.0"
        if row["rank"] is not None:
            ws.cell(row=r, column=3, value=int(row["rank"]))
        ws.cell(row=r, column=4, value=f"=IF(ISNUMBER(C{r}),C{r}<={REF_N_PEAK},FALSE)")
        ws.cell(row=r, column=5, value=f"=LARGE($B$5:$B$8764,{REF_N_PEAK}+1)").number_format = "#,##0.0"
        ws.cell(row=r, column=6, value=f"=IFERROR(IF(D{r},MAX(0,B{r}-E{r}),0),0)").number_format = "#,##0.00"
        ws.cell(
            row=r,
            column=7,
            value=f"=IFERROR(IF(D{r},F{r}/SUMPRODUCT(($C$5:$C$8764<={REF_N_PEAK})*IFERROR(IF($B$5:$B$8764>E{r},$B$5:$B$8764-E{r},0),0)),0),0)",
        ).number_format = NUMBER_FORMAT_MCPKWH
        ws.cell(row=r, column=8, value=f"=G{r}*{REF_AESC_PTF}").number_format = NUMBER_FORMAT_MCPKWH

    _autosize(ws, {"A": 18, "B": 18, "C": 8, "D": 14, "E": 20, "F": 14, "G": 16, "H": 24})
    ws.sheet_view.showGridLines = False


def _write_mc_combined(wb: Workbook, n_hours: int = 8760) -> None:
    """mc_combined: delivery MC = dist/sub-TX + bulk TX (live cross-sheet formulas).

    Data rows start at row 5 (header at row 4) to match mc_dist_sub_tx and mc_bulk_tx.
    """
    ws = wb.create_sheet("mc_combined")
    ws["A1"] = "Delivery Marginal Cost — Combined (dist/sub-TX + bulk TX)"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:D1")
    ws["A2"] = "All values are cross-sheet formulas referencing mc_dist_sub_tx (col F) and mc_bulk_tx (col H)."
    ws.merge_cells("A2:D2")

    headers = ["timestamp", "mc_dist_sub_tx", "mc_bulk_tx", "mc_delivery_total"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A5"

    for i in range(5, n_hours + 5):
        ws.cell(row=i, column=1, value=f"=mc_dist_sub_tx!A{i}")
        ws.cell(row=i, column=2, value=f"=mc_dist_sub_tx!F{i}").number_format = NUMBER_FORMAT_MCPKWH
        ws.cell(row=i, column=3, value=f"=mc_bulk_tx!H{i}").number_format = NUMBER_FORMAT_MCPKWH
        ws.cell(row=i, column=4, value=f"=B{i}+C{i}").number_format = NUMBER_FORMAT_MCPKWH

    _autosize(ws, {"A": 18, "B": 16, "C": 14, "D": 18})
    ws.sheet_view.showGridLines = False


def _write_validation_mc(wb: Workbook) -> None:
    """Validation for MC derivation — mirrors build_marginal_costs_workbook.py."""
    ws = wb.create_sheet("validation_mc")
    headers = ["check", "actual", "expected", "abs_error", "tolerance", "ok"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Data rows in mc_dist_sub_tx and mc_bulk_tx start at row 5
    checks = [
        ("sum(pop_weight) for dist/sub_tx = 1.0", "=SUM(mc_dist_sub_tx!E$5:E$8764)", "=1", 1e-6),
        (
            "sum(mc_dist_sub_tx) = $101.05/kW-yr [expert testimony §IX: $80.24 CPI-adj]",
            "=SUM(mc_dist_sub_tx!F$5:F$8764)",
            f"={REF_SUB_TX_DIST}",
            0.01,
        ),
        ("sum(exceedance_weight) for bulk_tx = 1.0", "=SUM(mc_bulk_tx!G$5:G$8764)", "=1", 1e-6),
        (
            "sum(mc_bulk_tx) = $69/kW-yr [expert testimony §IX: '$69/kW-year']",
            "=SUM(mc_bulk_tx!H$5:H$8764)",
            f"={REF_AESC_PTF}",
            0.01,
        ),
        (
            "sum(mc_delivery_total) = $170.05/kW-yr [expert testimony §IX: sum of both AESC components]",
            "=SUM(mc_combined!D$5:D$8764)",
            f"={REF_TOTAL_MC}",
            0.02,
        ),
        (
            "non-zero hours dist_sub_tx = 100 [expert testimony §III and §IX: 'top 100 hours']",
            '=COUNTIF(mc_dist_sub_tx!F$5:F$8764,">0")',
            f"={REF_N_PEAK}",
            0,
        ),
        (
            "non-zero hours bulk_tx = 100 [expert testimony §III and §IX: 'top 100 hours']",
            '=COUNTIF(mc_bulk_tx!H$5:H$8764,">0")',
            f"={REF_N_PEAK}",
            0,
        ),
    ]
    for i, (name, actual, expected, tol) in enumerate(checks, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=actual).number_format = "#,##0.000000"
        ws.cell(row=i, column=3, value=expected).number_format = "#,##0.000000"
        ws.cell(row=i, column=4, value=f"=ABS(B{i}-C{i})").number_format = "#,##0.000000"
        ws.cell(row=i, column=5, value=tol)
        ws.cell(row=i, column=6, value=f'=IF(D{i}<=E{i},"OK","FAIL")')
    _autosize(ws, {"A": 72, "B": 18, "C": 18, "D": 14, "E": 14, "F": 8})
    ws.sheet_view.showGridLines = False


def _write_validation_mc_hourly(wb: Workbook, mc_parquet_df: pl.DataFrame) -> None:
    """Hour-by-hour comparison: workbook formula MC vs S3 pre-computed MC parquets.

    Columns A-J for each of the 8760 hours:
      A  timestamp            (formula -> mc_dist_sub_tx col A)
      B  mc_dist_parquet      (hardcoded from S3 parquet)
      C  mc_dist_workbook     (formula -> mc_dist_sub_tx col F)
      D  diff_dist            (= C - B)
      E  mc_bulk_parquet      (hardcoded from S3 parquet)
      F  mc_bulk_workbook     (formula -> mc_bulk_tx col H)
      G  diff_bulk            (= F - E)
      H  mc_total_parquet     (hardcoded = B + E)
      I  mc_total_workbook    (formula -> mc_combined col D)
      J  diff_total           (= I - H)

    All diff columns should be ≈ 0.  Any non-zero value indicates a discrepancy
    between the formula derivation in this workbook and the parquet stored on S3.
    """
    ws = wb.create_sheet("validation_mc_hourly")
    ws["A1"] = "MC Hourly Validation — Workbook Formulas vs S3 Parquet (8760 rows)"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:J1")
    ws["A2"] = (
        f"Parquet sources: {S3_DIST_SUB_TX}  |  {S3_BULK_TX}. "
        "diff = workbook formula - parquet value.  All diffs should be 0 (<= 1e-9 floating-point rounding)."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 40

    headers = [
        "timestamp",
        "mc_dist_parquet ($/kWh)",
        "mc_dist_workbook ($/kWh)",
        "diff_dist",
        "mc_bulk_parquet ($/kWh)",
        "mc_bulk_workbook ($/kWh)",
        "diff_bulk",
        "mc_total_parquet ($/kWh)",
        "mc_total_workbook ($/kWh)",
        "diff_total",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A5"

    fmt = NUMBER_FORMAT_MCPKWH
    for i, row in enumerate(mc_parquet_df.iter_rows(named=True), start=5):
        r = i
        # A: timestamp linked from mc_dist_sub_tx (same row offset)
        ws.cell(row=r, column=1, value=f"=mc_dist_sub_tx!A{r}")
        # B: dist/sub-TX parquet value (Python-computed, written as a literal)
        v_dist = row["mc_dist_sub_tx_parquet"]
        ws.cell(row=r, column=2, value=float(v_dist) if v_dist is not None else 0.0).number_format = fmt
        # C: dist/sub-TX workbook formula
        ws.cell(row=r, column=3, value=f"=mc_dist_sub_tx!F{r}").number_format = fmt
        # D: diff dist
        ws.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = fmt
        # E: bulk TX parquet value
        v_bulk = row["mc_bulk_tx_parquet"]
        ws.cell(row=r, column=5, value=float(v_bulk) if v_bulk is not None else 0.0).number_format = fmt
        # F: bulk TX workbook formula
        ws.cell(row=r, column=6, value=f"=mc_bulk_tx!H{r}").number_format = fmt
        # G: diff bulk
        ws.cell(row=r, column=7, value=f"=F{r}-E{r}").number_format = fmt
        # H: total parquet (dist + bulk, Python-computed)
        v_total = row["mc_delivery_total_parquet"]
        ws.cell(row=r, column=8, value=float(v_total) if v_total is not None else 0.0).number_format = fmt
        # I: total workbook formula
        ws.cell(row=r, column=9, value=f"=mc_combined!D{r}").number_format = fmt
        # J: diff total
        ws.cell(row=r, column=10, value=f"=I{r}-H{r}").number_format = fmt

    _autosize(ws, {"A": 18, "B": 18, "C": 18, "D": 14, "E": 18, "F": 18, "G": 14, "H": 18, "I": 18, "J": 14})
    ws.sheet_view.showGridLines = False


def _write_bat_building_data(wb: Workbook, bat_df: pl.DataFrame) -> None:
    """Per-building BAT data — source for RECS share calculations in next sheet.

    Follows build_DIV_1_2_workbook.py add_bat_building_data_sheet() exactly.
    Columns: bldg_id, heating_type_v2, weight, annual_bill_delivery,
    economic_burden_delivery, weight*bill (formula), weight*EB (formula).
    """
    ws = wb.create_sheet("bat_building_data")
    ws["A1"] = "CAIRO Master BAT — Per-Building Data (RIE Residential, Test Year)"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:G1")
    ws["A2"] = f"S3 source: {S3_MASTER_BAT}"
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:G2")

    col_headers = [
        "bldg_id",
        "heating_type_v2",
        "weight",
        "annual_bill_delivery ($)",
        "economic_burden_delivery ($)",
        "weight * bill",
        "weight * EB",
    ]
    HDR_ROW = 3
    for ci, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    order_map = {k: i for i, k in enumerate(HT_V2_ORDER)}
    sorted_df = (
        bat_df.select(
            "bldg_id",
            "postprocess_group.heating_type_v2",
            "weight",
            "annual_bill_delivery",
            "economic_burden_delivery",
        )
        .with_columns(pl.col("postprocess_group.heating_type_v2").replace(order_map).alias("_sort_order"))
        .sort(["_sort_order", "bldg_id"])
        .drop("_sort_order")
    )

    DATA_START = HDR_ROW + 1
    n_bldg = sorted_df.height
    for i, row_data in enumerate(sorted_df.iter_rows(named=True)):
        row = DATA_START + i
        label = HT_V2_LABELS[row_data["postprocess_group.heating_type_v2"]]
        ws.cell(row=row, column=1, value=row_data["bldg_id"])
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=float(row_data["weight"])).number_format = "#,##0.000"
        ws.cell(row=row, column=4, value=float(row_data["annual_bill_delivery"])).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=5, value=float(row_data["economic_burden_delivery"])).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=6, value=f"=C{row}*D{row}").number_format = '"$"#,##0.00'
        ws.cell(row=row, column=7, value=f"=C{row}*E{row}").number_format = '"$"#,##0.00'

    _autosize(ws, {"A": 10, "B": 18, "C": 10, "D": 14, "E": 14, "F": 14, "G": 14})
    ws.freeze_panes = f"A{DATA_START}"
    ws.sheet_view.showGridLines = False

    # Store for cross-sheet SUMIF formulas in recs_subclass_customers
    wb._1_14b_bat_sheet = "bat_building_data"  # type: ignore[attr-defined]
    wb._1_14b_bat_data_start = DATA_START  # type: ignore[attr-defined]
    wb._1_14b_bat_last_row = DATA_START + n_bldg - 1  # type: ignore[attr-defined]


def _write_recs_subclass_customers(wb: Workbook, bat_df: pl.DataFrame) -> None:
    """RECS 2020 RI heating-system shares and per-subclass customer counts.

    The RECS shares are derived from the CAIRO simulation weights, which were
    calibrated to RECS 2020 RI state-level heating-fuel shares from EIA Table
    CE1.1.  Each subclass's share = SUMIF(weights, subclass) / Test Year total.
    Per-subclass customer count = share * Test Year customer count.

    Customers column uses SUMIF formulas over bat_building_data.
    Share column = Customers / Test Year total (formula).
    """
    ws = wb.create_sheet("recs_subclass_customers")
    ws["A1"] = "RECS 2020 RI Heating-System Shares and Test Year Customer Counts by Subclass"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:E1")
    ws["A2"] = (
        "Heating-system shares are derived from CAIRO simulation weights, which are calibrated to "
        "EIA Residential Energy Consumption Survey (RECS) 2020 Rhode Island heating-fuel shares "
        "(Table CE1.1).  SUMIF formulas aggregate weighted customer counts by subclass from "
        "the bat_building_data sheet.  Per-subclass count = share * Test Year total customer count."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 60

    # Parameters block
    ws["A4"] = "Parameters"
    ws["A4"].font = SECTION_FONT
    ws["A4"].fill = SECTION_FILL
    ws.merge_cells("A4:E4")

    ws.cell(row=5, column=1, value="test_year_customer_count").font = Font(bold=True)
    ws.cell(row=5, column=2, value=f"={REF_TY_CUST}").number_format = "#,##0.00"
    ws.cell(row=5, column=3, value=f"From {REF_TY_CUST} (formula link). Expert testimony §IX.")
    ws.cell(row=5, column=3).font = Font(italic=True)
    TY_CUST_REF = "$B$5"  # absolute ref within this sheet

    # Table header
    HDR_ROW = 7
    ws["A6"] = "Subclass Table"
    ws["A6"].font = SECTION_FONT
    ws["A6"].fill = SECTION_FILL
    ws.merge_cells("A6:E6")

    col_headers = [
        "Heating subclass",
        "Customers (= SUMIF weights)",
        "RECS share (% of total)",
        "Source",
        "Notes",
    ]
    for ci, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # BAT data sheet references
    bat_sheet = f"'{wb._1_14b_bat_sheet}'"  # type: ignore[attr-defined]
    bat_ds = wb._1_14b_bat_data_start  # type: ignore[attr-defined]
    bat_lr = wb._1_14b_bat_last_row  # type: ignore[attr-defined]
    bat_label_range = f"{bat_sheet}!$B${bat_ds}:$B${bat_lr}"
    bat_weight_range = f"{bat_sheet}!$C${bat_ds}:$C${bat_lr}"

    DATA_START = HDR_ROW + 1
    n_sub = len(HT_V2_ORDER)
    TOTAL_ROW = DATA_START + n_sub

    recs_notes = {
        "heat_pump": "Includes air-source and ground-source heat pumps.",
        "electrical_resistance": "Electric baseboard, furnace, or boiler.",
        "natgas": "Natural gas furnace, boiler, or wall heater.",
        "delivered_fuels": "Fuel oil, kerosene, propane, or wood (delivered fuels).",
        "other": "District steam, solar, or no heating equipment.",
    }
    recs_source = "EIA RECS 2020, Table CE1.1 (Space Heating Fuel, by State — Rhode Island)"

    for idx, code in enumerate(HT_V2_ORDER):
        row = DATA_START + idx
        ws.cell(row=row, column=1, value=HT_V2_LABELS[code])
        # Customers = SUMIF over BAT weight column matching label
        ws.cell(
            row=row,
            column=2,
            value=f"=SUMIF({bat_label_range},A{row},{bat_weight_range})",
        ).number_format = "#,##0.00"
        # RECS share = customers / total
        ws.cell(row=row, column=3, value=f"=B{row}/{TY_CUST_REF}").number_format = "0.00%"
        ws.cell(row=row, column=4, value=recs_source)
        ws.cell(row=row, column=5, value=recs_notes.get(code, ""))

    # Totals row
    ws.cell(row=TOTAL_ROW, column=1, value="All customers").font = Font(bold=True)
    ws.cell(row=TOTAL_ROW, column=2, value=f"=SUM(B{DATA_START}:B{TOTAL_ROW - 1})").number_format = "#,##0.00"
    ws.cell(row=TOTAL_ROW, column=3, value=f"=SUM(C{DATA_START}:C{TOTAL_ROW - 1})").number_format = "0.00%"
    for c in range(1, 4):
        ws.cell(row=TOTAL_ROW, column=c).font = Font(bold=True)

    _autosize(ws, {"A": 22, "B": 26, "C": 18, "D": 60, "E": 44})
    ws.freeze_panes = f"A{HDR_ROW + 1}"
    ws.sheet_view.showGridLines = False

    # Store for validation
    wb._1_14b_recs_total_row = TOTAL_ROW  # type: ignore[attr-defined]
    wb._1_14b_recs_data_start = DATA_START  # type: ignore[attr-defined]


def _write_validation_customers(wb: Workbook) -> None:
    """Cross-checks for RECS shares and customer counts."""
    ws = wb.create_sheet("validation_customers")
    headers = ["check", "actual", "expected", "abs_error", "tolerance", "ok"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    recs_sheet = "'recs_subclass_customers'"
    total_row = wb._1_14b_recs_total_row  # type: ignore[attr-defined]
    data_start = wb._1_14b_recs_data_start  # type: ignore[attr-defined]
    n_sub = len(HT_V2_ORDER)

    checks = [
        (
            "RECS shares sum to 1.0 [expert testimony §IX: RECS 2020 heating-fuel shares for RI]",
            f"={recs_sheet}!C{total_row}",
            "=1",
            0.0001,
        ),
        (
            "Total customers = Test Year customer count [expert testimony §IX: applied to RECS shares]",
            f"={recs_sheet}!B{total_row}",
            f"={REF_TY_CUST}",
            1.0,
        ),
        (
            "No subclass has negative customers",
            f'=COUNTIF({recs_sheet}!B{data_start}:B{data_start + n_sub - 1},"<0")',
            "=0",
            0,
        ),
    ]
    for i, (name, actual, expected, tol) in enumerate(checks, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=actual).number_format = "#,##0.000000"
        ws.cell(row=i, column=3, value=expected).number_format = "#,##0.000000"
        ws.cell(row=i, column=4, value=f"=ABS(B{i}-C{i})").number_format = "#,##0.000000"
        ws.cell(row=i, column=5, value=tol)
        ws.cell(row=i, column=6, value=f'=IF(D{i}<=E{i},"OK","FAIL")')

    _autosize(ws, {"A": 72, "B": 18, "C": 18, "D": 14, "E": 14, "F": 8})
    ws.sheet_view.showGridLines = False


# ── Tab formatting spec for Google Sheets upload ──────────────────────────────

_TAB_FORMATTING: dict[str, dict] = {
    "README": {
        "wrap_columns": ["A:C"],
        "column_widths_px": {"A": 300, "B": 540, "C": 540},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "inputs_scalars": {
        "column_number_formats": {"B": "#,##0.000000"},
        "wrap_columns": ["C:D"],
        "column_widths_px": {"A": 260, "B": 130, "C": 540, "D": 540},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "rie_load_8760": {
        "column_number_formats": {"B": "#,##0.000"},
        "column_widths_px": {"A": 130, "B": 100},
        "freeze_rows": 4,
        "bold_header": True,
    },
    "ne_load_8760": {
        "column_number_formats": {"B": "#,##0.000"},
        "column_widths_px": {"A": 130, "B": 100},
        "freeze_rows": 4,
        "bold_header": True,
    },
    "mc_dist_sub_tx": {
        "column_number_formats": {
            "B": "#,##0.0",
            "E": "0.000000",
            "F": "0.000000",
        },
        "column_widths_px": {"A": 130, "B": 100, "C": 60, "D": 100, "E": 100, "F": 175},
        "freeze_rows": 4,
        "bold_header": True,
    },
    "mc_bulk_tx": {
        "column_number_formats": {
            "B": "#,##0.0",
            "E": "#,##0.0",
            "F": "#,##0.00",
            "G": "0.000000",
            "H": "0.000000",
        },
        "column_widths_px": {"A": 130, "B": 130, "C": 60, "D": 100, "E": 145, "F": 100, "G": 115, "H": 175},
        "freeze_rows": 4,
        "bold_header": True,
    },
    "mc_combined": {
        "column_number_formats": {
            "B": "0.000000",
            "C": "0.000000",
            "D": "0.000000",
        },
        "column_widths_px": {"A": 130, "B": 115, "C": 100, "D": 130},
        "freeze_rows": 4,
        "bold_header": True,
    },
    "validation_mc": {
        "column_number_formats": {"B": "#,##0.000000", "C": "#,##0.000000", "D": "#,##0.000000"},
        "column_widths_px": {"A": 520, "B": 130, "C": 130, "D": 100, "E": 100, "F": 60},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "validation_mc_hourly": {
        "column_number_formats": {
            "B": "0.000000",
            "C": "0.000000",
            "D": "0.000000",
            "E": "0.000000",
            "F": "0.000000",
            "G": "0.000000",
            "H": "0.000000",
            "I": "0.000000",
            "J": "0.000000",
        },
        "column_widths_px": {
            "A": 130,
            "B": 130,
            "C": 130,
            "D": 100,
            "E": 130,
            "F": 130,
            "G": 100,
            "H": 130,
            "I": 130,
            "J": 100,
        },
        "freeze_rows": 4,
        "bold_header": True,
    },
    "bat_building_data": {
        "freeze_rows": 3,
        "bold_header": True,
        "column_widths_px": {"A": 75, "B": 140, "C": 80, "D": 110, "E": 110, "F": 100, "G": 100},
    },
    "recs_subclass_customers": {
        "freeze_rows": 7,
        "bold_header": True,
        "column_widths_px": {"A": 160, "B": 200, "C": 130, "D": 420, "E": 320},
    },
    "validation_customers": {
        "column_number_formats": {"B": "#,##0.000000", "C": "#,##0.000000", "D": "#,##0.000000"},
        "column_widths_px": {"A": 520, "B": 130, "C": 130, "D": 100, "E": 100, "F": 60},
        "freeze_rows": 1,
        "bold_header": True,
    },
}


# ── Main build ────────────────────────────────────────────────────────────────


def build_workbook(output_path: Path) -> Path:
    """Load all inputs, build every sheet, write the .xlsx."""
    print("Building RIE 1-14b external-studies workbook ...", flush=True)

    print("Loading master BAT from S3 ...", flush=True)
    bat_df = load_master_bat()

    print("Loading RIE EIA-930 utility load from S3 ...", flush=True)
    rie_load = load_rie_hourly_load()
    print(f"  RIE load: {rie_load.height} hours", flush=True)

    print("Loading NE system load (8 ISO-NE zones) from S3 ...", flush=True)
    ne_load = load_ne_system_load()
    print(f"  NE system load: {ne_load.height} hours", flush=True)

    print("Loading pre-computed MC parquets for hourly validation ...", flush=True)
    mc_parquet_df = load_mc_parquets(rie_load)

    print("Assembling workbook ...", flush=True)
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    _write_readme(wb)
    _write_inputs_scalars(wb)
    _write_raw_load_8760(
        wb,
        rie_load,
        "rie_load_8760",
        "EIA-930 RIE Utility Load",
        f"EIA Form EIA-930 hourly utility demand for RIE, year {YEAR}. "
        f"S3: {S3_EIA_UTILITY_LOADS}region=isone/utility={UTILITY}/year={YEAR}/. "
        "DST fall-back duplicates averaged and series truncated to 8760 rows.",
    )
    _write_raw_load_8760(
        wb,
        ne_load,
        "ne_load_8760",
        "ISO-NE NE System Load (8 zones summed)",
        f"ISO-NE CELT hourly demand, 8 zones (CT, ME, NEMA, NH, RI, SEMA, VT, WCMA), "
        f"summed to NE system total, year {YEAR}. "
        f"S3: {S3_ISONE_ZONE_LOADS}zone=*/year={YEAR}/. "
        "DST fall-back duplicates averaged and series truncated to 8760 rows.",
    )
    _write_mc_dist_sub_tx(wb, rie_load)
    _write_mc_bulk_tx(wb, ne_load)
    _write_mc_combined(wb)
    _write_validation_mc(wb)
    _write_validation_mc_hourly(wb, mc_parquet_df)
    _write_bat_building_data(wb, bat_df)
    _write_recs_subclass_customers(wb, bat_df)
    _write_validation_customers(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    size_kb = output_path.stat().st_size / 1024
    print(f"Wrote {output_path} ({size_kb:.1f} KB)", flush=True)
    return output_path


# ── Google Sheets upload ──────────────────────────────────────────────────────


def upload_to_folder(xlsx_path: Path, folder_id: str, title: str) -> None:
    """Create (or replace) a Google Sheet in the given Drive folder.

    Searches the folder for any existing non-trashed file with the same name
    and trashes it before creating a fresh spreadsheet — identical pattern to
    build_DIV_1_2_workbook.py and build_RIE_1_14a_workbook.py.
    """
    from lib.data.gsheets import (
        apply_sheet_formatting,
        create_sheet_in_folder,
        xlsx_to_gsheet,
    )

    print(f"Uploading '{title}' to Drive folder {folder_id} ...", flush=True)
    spreadsheet = create_sheet_in_folder(title, folder_id)
    xlsx_to_gsheet(xlsx_path, spreadsheet.id, delete_other_tabs=True)

    print("Applying formatting ...", flush=True)
    for ws in spreadsheet.worksheets():
        spec = _TAB_FORMATTING.get(ws.title)
        if spec:
            fmt = dict(spec)
            if ws.title == "README" and _README_BOLD_ROWS:
                fmt["bold_rows"] = list(_README_BOLD_ROWS)
            apply_sheet_formatting(ws, **fmt)

    print(
        f"Done. View at https://docs.google.com/spreadsheets/d/{spreadsheet.id}/edit",
        flush=True,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__.splitlines()[0] if __doc__ else "",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("cache/rie_1_14b_external_studies.xlsx"),
        help="Output .xlsx path. Default: cache/rie_1_14b_external_studies.xlsx",
    )
    parser.add_argument(
        "--folder-id",
        default=DEFAULT_FOLDER_ID,
        help=f"Google Drive folder ID to upload into. Default: {DEFAULT_FOLDER_ID}",
    )
    parser.add_argument(
        "--title",
        default=DEFAULT_TITLE,
        help=f"Name for the Google Sheet. Default: '{DEFAULT_TITLE}'",
    )
    parser.add_argument(
        "--no-upload",
        action="store_true",
        help="Build the .xlsx locally without uploading to Google Drive.",
    )
    args = parser.parse_args(argv)

    out = build_workbook(args.output)
    if not args.no_upload:
        upload_to_folder(out, args.folder_id, args.title)
    return 0


if __name__ == "__main__":
    sys.exit(main())
