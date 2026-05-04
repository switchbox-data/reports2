"""Verify the revenue-neutrality workbook reproduces the testimony numbers.

Two checks:

1. **Structural** — open the xlsx and confirm expected sheets and sum
   formulas exist.
2. **Numerical** — confirm HP + non-HP subclass RRs sum to the total
   delivery RR for each allocation method.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

REPORT_DIR = Path("/ebs/home/alex_switch_box/reports2/reports/ri_hp_rates")
sys.path.insert(0, str(REPORT_DIR))

from testimony_response.build_RIE_1_12_workbook import (  # noqa: E402
    DELIVERY_METHOD,
    load_inputs,
)


def _structural_check(xlsx_path: Path) -> None:
    print(f"=== Structural check: {xlsx_path} ===")
    wb = load_workbook(str(xlsx_path), data_only=False)

    expected_sheets = ["README", "subclass_revenue_proof"]
    missing = [s for s in expected_sheets if s not in wb.sheetnames]
    assert not missing, f"missing sheets: {missing}"
    print(f"  sheets: {wb.sheetnames}")

    proof_ws = wb["subclass_revenue_proof"]
    assert "proof" in str(proof_ws["A1"].value).lower()
    print("  subclass_revenue_proof title: OK")

    found_sum = False
    for row in proof_ws.iter_rows(min_col=2, max_col=2, values_only=False):
        cell = row[0]
        if isinstance(cell.value, str) and cell.value.startswith("=") and "+" in cell.value:
            found_sum = True
            break
    assert found_sum, "no sum formula found in column B"
    print("  sum formula in column B: OK")
    print()


def _numerical_check() -> None:
    print("=== Numerical check (expected values from YAMLs) ===")
    inputs = load_inputs()

    total_rr = inputs["total_delivery_revenue_requirement"]
    subs = inputs["subclass_revenue_requirements"]
    print(f"  total_delivery_RR: ${total_rr:,.2f}")

    hp_rr = float(subs["delivery"][DELIVERY_METHOD]["hp"])
    nonhp_rr = float(subs["delivery"][DELIVERY_METHOD]["non-hp"])
    total = hp_rr + nonhp_rr
    delta = abs(total - total_rr)
    print(f"  EPMC: hp=${hp_rr:,.2f} + nonhp=${nonhp_rr:,.2f} = ${total:,.2f} (delta=${delta:.2f})")
    assert delta < 1.0, f"EPMC delivery sum mismatch: {delta}"

    epmc_hp = hp_rr
    print(f"\n  EPMC delivery HP RR: ${epmc_hp:,.2f}")
    print("  (testimony cos_default_hp_group_cos should = $11,940,870.79)")
    assert abs(epmc_hp - 11_940_870.79) < 0.01, f"EPMC HP mismatch: {epmc_hp}"
    print("  PASS: matches testimony value")

    print("\n  All numerical checks: PASS")


def main() -> int:
    xlsx_path = REPORT_DIR / "cache" / "revenue_neutrality.xlsx"
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found. Run build_RIE_1_12_workbook.py first.")
        return 1
    _structural_check(xlsx_path)
    _numerical_check()
    return 0


if __name__ == "__main__":
    sys.exit(main())
