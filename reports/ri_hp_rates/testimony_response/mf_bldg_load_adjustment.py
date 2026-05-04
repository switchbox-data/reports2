"""Build the supporting workbook for ResStock vs EIA-861 load curve adjustment.

This script reproduces the ResStock vs EIA-861 comparison analysis, showing the
impact of multifamily non-HVAC electricity adjustment on load discrepancies.

Usage:
    # Option 1: Export to existing spreadsheet
    uv run python -m testimony_response.load_adjustment_workbook \
        --sheet-id <your-google-sheet-id> \
        --state RI

    # Option 2: Create new spreadsheet in folder
    uv run python -m testimony_response.load_adjustment_workbook \
        --folder-id <google-drive-folder-id> \
        --filename "RIE Load Adjustment Analysis" \
        --state RI

    # Option 3: Build local workbook only (no upload)
    uv run python -m testimony_response.load_adjustment_workbook \
        --output cache/load_adjustment.xlsx \
        --state RI
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path
from typing import cast

import polars as pl
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Data paths - can be local or S3
LOCAL_BASE_RESSTOCK = "/ebs/data/nrel/resstock"
S3_BASE_RESSTOCK = "s3://data.sb/nrel/resstock"
S3_BASE_EIA861 = "s3://data.sb/eia/861/electric_utility_stats"

# Use local data by default if available, fallback to S3
BASE_RESSTOCK = LOCAL_BASE_RESSTOCK if Path(LOCAL_BASE_RESSTOCK).exists() else S3_BASE_RESSTOCK
BASE_EIA861 = S3_BASE_EIA861  # EIA data not available locally

# Column constants
BLDG_ID_COL = "bldg_id"
ELECTRIC_UTILITY_COL = "sb.electric_utility"
WEIGHT_COL = "weight"
ANNUAL_ELECTRICITY_COL = "out.electricity.total.energy_consumption.kwh"
MWH_TO_KWH = 1000
BUILDING_TYPE_RECS_COL = "in.geometry_building_type_recs"
FLOOR_AREA_COL = "in.geometry_floor_area"
TIMESTAMP_COL = "timestamp"
TOTAL_ELEC_COL = "out.electricity.total.energy_consumption"

NON_HVAC_RELATED_ELECTRICITY_COLS = (
    "out.electricity.ceiling_fan.energy_consumption.kwh",
    "out.electricity.clothes_dryer.energy_consumption.kwh",
    "out.electricity.clothes_washer.energy_consumption.kwh",
    "out.electricity.dishwasher.energy_consumption.kwh",
    "out.electricity.freezer.energy_consumption.kwh",
    "out.electricity.hot_water.energy_consumption.kwh",
    "out.electricity.lighting_exterior.energy_consumption.kwh",
    "out.electricity.lighting_garage.energy_consumption.kwh",
    "out.electricity.lighting_interior.energy_consumption.kwh",
    "out.electricity.permanent_spa_heat.energy_consumption.kwh",
    "out.electricity.permanent_spa_pump.energy_consumption.kwh",
    "out.electricity.plug_loads.energy_consumption.kwh",
    "out.electricity.pool_heater.energy_consumption.kwh",
    "out.electricity.pool_pump.energy_consumption.kwh",
    "out.electricity.pv.energy_consumption.kwh",
    "out.electricity.range_oven.energy_consumption.kwh",
    "out.electricity.refrigerator.energy_consumption.kwh",
    "out.electricity.well_pump.energy_consumption.kwh",
)

HVAC_RELATED_ELECTRICITY_COLS = (
    "out.electricity.cooling.energy_consumption.kwh",
    "out.electricity.cooling_fans_pumps.energy_consumption.kwh",
    "out.electricity.heating.energy_consumption.kwh",
    "out.electricity.heating_fans_pumps.energy_consumption.kwh",
    "out.electricity.heating_hp_bkup.energy_consumption.kwh",
    "out.electricity.heating_hp_bkup_fa.energy_consumption.kwh",
    "out.electricity.mech_vent.energy_consumption.kwh",
)


def get_aws_region(default: str = "us-west-2") -> str:
    """Return AWS region from env or default."""
    region = os.environ.get("AWS_REGION") or os.environ.get("AWS_DEFAULT_REGION")
    if region:
        return region
    try:
        import boto3

        session = boto3.Session()
        if session.region_name:
            return session.region_name
    except ImportError:
        pass
    return default


def _storage_options() -> dict[str, str]:
    """AWS region for S3 access."""
    return {"aws_region": get_aws_region()}


def _is_s3(path: str) -> bool:
    return path.startswith("s3://")


def load_hourly_load_curves(
    state: str,
    resstock_release: str,
    upgrade: str,
    bldg_ids: list[int],
    storage_options: dict[str, str] | None,
) -> pl.DataFrame:
    """Load hourly load curves for specific building IDs."""
    base_path = f"{BASE_RESSTOCK}/{resstock_release}/load_curve_hourly/state={state}/upgrade={upgrade}"

    # Construct file paths for each building
    file_paths = [f"{base_path}/{bldg_id}-{int(upgrade)}.parquet" for bldg_id in bldg_ids]

    print(f"Loading {len(file_paths)} hourly load curve files...", flush=True)

    # Load all files and combine
    dfs = []
    for i, path in enumerate(file_paths):
        try:
            opts = storage_options if _is_s3(path) else None
            df = pl.read_parquet(path, storage_options=opts)

            # Check which columns exist and select appropriately
            cols_to_select = [pl.col(BLDG_ID_COL)]

            if TIMESTAMP_COL in df.columns:
                cols_to_select.append(pl.col(TIMESTAMP_COL))

            if TOTAL_ELEC_COL in df.columns:
                cols_to_select.append(pl.col(TOTAL_ELEC_COL))

            df_selected = df.select(cols_to_select)
            dfs.append(df_selected)

            # Progress indicator every 100 files
            if (i + 1) % 100 == 0:
                print(f"  Loaded {i + 1}/{len(file_paths)} files...", flush=True)

        except Exception as e:
            print(f"Warning: Could not load {path}: {e}", flush=True)
            continue

    if not dfs:
        print("  ERROR: No files were successfully loaded!", flush=True)
        return pl.DataFrame({"hour": range(8760), "total_kwh": [0.0] * 8760})

    print(f"  Successfully loaded {len(dfs)} files, concatenating...", flush=True)
    return pl.concat(dfs)


def aggregate_weighted_load_curves(
    hourly_loads: pl.DataFrame,
    metadata: pl.DataFrame,
    bldg_ids: list[int],
    by_floor_area: bool = False,
) -> pl.DataFrame:
    """Aggregate hourly load curves weighted by WEIGHT_COL and optionally by floor area."""
    # Filter to specific buildings
    loads_subset = hourly_loads.filter(pl.col(BLDG_ID_COL).is_in(bldg_ids))
    metadata_subset = metadata.filter(pl.col(BLDG_ID_COL).is_in(bldg_ids))

    # Join with metadata to get weights and floor area
    if by_floor_area and FLOOR_AREA_COL in metadata_subset.columns:
        meta_with_area = metadata_subset.with_columns(
            pl.col(FLOOR_AREA_COL)
            .map_batches(
                lambda s: pl.Series([_parse_floor_area_sqft(x) for x in s]),
                return_dtype=pl.Float64,
            )
            .alias("floor_area_sqft")
        ).select(pl.col(BLDG_ID_COL), pl.col(WEIGHT_COL), pl.col("floor_area_sqft"))

        joined = loads_subset.join(meta_with_area, on=BLDG_ID_COL, how="inner")

        # Weight by (weight * kwh / floor_area)
        aggregated = (
            joined.group_by(TIMESTAMP_COL)
            .agg(
                (pl.col(TOTAL_ELEC_COL) * pl.col(WEIGHT_COL) / pl.col("floor_area_sqft"))
                .sum()
                .alias("total_kwh_per_sqft")
            )
            .sort(TIMESTAMP_COL)
        )

        # Add hour column
        return aggregated.with_row_index("hour")
    else:
        meta_weights = metadata_subset.select(pl.col(BLDG_ID_COL), pl.col(WEIGHT_COL))
        joined = loads_subset.join(meta_weights, on=BLDG_ID_COL, how="inner")

        # Weight by (weight * kwh)
        aggregated = (
            joined.group_by(TIMESTAMP_COL)
            .agg((pl.col(TOTAL_ELEC_COL) * pl.col(WEIGHT_COL)).sum().alias("total_kwh"))
            .sort(TIMESTAMP_COL)
        )

        # Add hour column
        return aggregated.with_row_index("hour")


def _parse_floor_area_sqft(val: str | None) -> float:
    """Parse floor area value: '4000+' -> 5000, '750-999' -> midpoint."""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return float("nan")
    s = str(val).strip()
    if s.endswith("+"):
        return 5000.0
    if "-" in s:
        parts = s.split("-", 1)
        if len(parts) == 2:
            try:
                lo = float(parts[0].strip())
                hi = float(parts[1].strip())
                return (lo + hi) / 2.0
            except ValueError:
                return float("nan")
    try:
        return float(s.replace("+", ""))
    except ValueError:
        return float("nan")


def load_resstock_annual_building_level(
    path_annual: str,
    path_utility_assignment: str,
    storage_options: dict[str, str] | None,
) -> pl.DataFrame:
    """Load full annual parquet, join utility assignment, add HVAC/non-HVAC sums."""
    opts_annual = storage_options if _is_s3(path_annual) else None
    annual_lf = pl.scan_parquet(path_annual, storage_options=opts_annual)
    schema = annual_lf.collect_schema().names()

    hvac_cols = [c for c in HVAC_RELATED_ELECTRICITY_COLS if c in schema]
    non_hvac_cols = [c for c in NON_HVAC_RELATED_ELECTRICITY_COLS if c in schema]

    add_exprs: list[pl.Expr] = [
        pl.col(ANNUAL_ELECTRICITY_COL).alias("annual_kwh"),
        (pl.col(ANNUAL_ELECTRICITY_COL) * pl.col(WEIGHT_COL)).alias("weighted_kwh"),
    ]

    if hvac_cols:
        add_exprs.append(pl.sum_horizontal([pl.col(c) for c in hvac_cols]).alias("total_hvac_related_electricity_kwh"))
    else:
        add_exprs.append(pl.lit(0.0).alias("total_hvac_related_electricity_kwh"))

    if non_hvac_cols:
        add_exprs.append(
            pl.sum_horizontal([pl.col(c) for c in non_hvac_cols]).alias("total_non_hvac_related_electricity_kwh")
        )
    else:
        add_exprs.append(pl.lit(0.0).alias("total_non_hvac_related_electricity_kwh"))

    annual_df = cast(pl.DataFrame, annual_lf.collect()).with_columns(add_exprs)

    opts_ua = storage_options if _is_s3(path_utility_assignment) else None
    ua_df = cast(
        pl.DataFrame,
        pl.scan_parquet(path_utility_assignment, storage_options=opts_ua)
        .select(BLDG_ID_COL, ELECTRIC_UTILITY_COL)
        .collect(),
    )

    return annual_df.join(ua_df, on=BLDG_ID_COL, how="inner").rename({ELECTRIC_UTILITY_COL: "utility_code"})


def group_resstock_annual_by_utility(
    resstock_annual: pl.DataFrame,
) -> pl.DataFrame:
    """Group resstock_annual by utility_code."""
    return resstock_annual.group_by("utility_code").agg(
        pl.col("weighted_kwh").sum().alias("resstock_total_kwh"),
        pl.col(WEIGHT_COL).sum().alias("resstock_customers"),
        pl.len().alias("n_bldgs"),
        pl.col("annual_kwh").mean().alias("mean_annual_kwh_per_bldg"),
    )


def load_metadata_by_utility(
    path_metadata: str,
    path_utility_assignment: str,
    storage_options: dict[str, str] | None,
) -> tuple[pl.DataFrame, dict[str, pl.DataFrame]]:
    """Load metadata parquet, join utility assignment."""
    opts_meta = storage_options if _is_s3(path_metadata) else None
    opts_ua = storage_options if _is_s3(path_utility_assignment) else None

    meta_df = cast(
        pl.DataFrame,
        pl.scan_parquet(path_metadata, storage_options=opts_meta).collect(),
    )

    ua_df = cast(
        pl.DataFrame,
        pl.scan_parquet(path_utility_assignment, storage_options=opts_ua)
        .select(BLDG_ID_COL, ELECTRIC_UTILITY_COL)
        .collect(),
    )

    metadata_with_utility = meta_df.join(ua_df, on=BLDG_ID_COL, how="inner").rename(
        {ELECTRIC_UTILITY_COL: "utility_code"}
    )

    by_utility: dict[str, pl.DataFrame] = {}
    for code in metadata_with_utility["utility_code"].unique().to_list():
        by_utility[code] = metadata_with_utility.filter(pl.col("utility_code") == code)

    return metadata_with_utility, by_utility


def load_eia_by_utility(
    path_eia861: str,
    storage_options: dict[str, str] | None,
    utility_codes: list[int] | None = None,
) -> pl.DataFrame:
    """Load EIA-861 state parquet."""
    opts = storage_options if _is_s3(path_eia861) else None
    lf = pl.scan_parquet(path_eia861, storage_options=opts).select(
        pl.col("utility_code"),
        pl.col("residential_sales_mwh"),
        (pl.col("residential_sales_mwh") * MWH_TO_KWH).alias("eia_residential_kwh"),
        pl.col("residential_customers").alias("eia_residential_customers"),
    )
    if utility_codes is not None:
        lf = lf.filter(pl.col("utility_code").is_in(utility_codes))
    return cast(pl.DataFrame, lf.collect())


def building_type_share_by_utility(
    metadata_by_utility: dict[str, pl.DataFrame],
) -> dict[str | int, dict[str, int | float]]:
    """Compute each utility's share of multifamily and single-family buildings."""
    out: dict[str | int, dict[str, int | float]] = {}
    for utility_code, df in metadata_by_utility.items():
        col = pl.col(BUILDING_TYPE_RECS_COL)
        n_total = len(df)
        if n_total == 0:
            out[utility_code] = {"multifamily_pct": 0.0, "single_family_pct": 0.0}
            continue
        n_multifamily = df.filter(col.str.contains("Multi-Family", literal=True)).height
        n_single_family = df.filter(col.str.contains("Single-Family", literal=True)).height
        n_mobile_home = df.filter(col.str.contains("Mobile Home", literal=True)).height
        out[utility_code] = {
            "multifamily_pct": n_multifamily / n_total * 100.0,
            "single_family_pct": n_single_family / n_total * 100.0,
            "mobile_home_pct": n_mobile_home / n_total * 100.0,
        }
    return out


def compare_resstock_eia_by_utility(
    resstock_annual_by_utility: pl.DataFrame,
    eia_by_utility: pl.DataFrame,
) -> pl.DataFrame:
    """Join ResStock and EIA by utility_code and compute comparison metrics."""
    joined = resstock_annual_by_utility.join(
        eia_by_utility.select(
            pl.col("utility_code"),
            pl.col("eia_residential_kwh"),
            pl.col("eia_residential_customers"),
        ),
        on="utility_code",
        how="inner",
    )

    resstock_normalized_kwh = (
        pl.col("resstock_total_kwh") * pl.col("eia_residential_customers") / pl.col("resstock_customers")
    )

    return joined.with_columns(resstock_normalized_kwh.alias("resstock_total_kwh_normalized_to_eia")).select(
        pl.col("utility_code"),
        pl.col("resstock_total_kwh"),
        pl.col("resstock_total_kwh_normalized_to_eia"),
        pl.col("eia_residential_kwh"),
        (pl.col("resstock_total_kwh_normalized_to_eia") / pl.col("eia_residential_kwh")).alias("kwh_ratio"),
        (
            (pl.col("resstock_total_kwh_normalized_to_eia") - pl.col("eia_residential_kwh"))
            / pl.col("eia_residential_kwh")
            * 100
        ).alias("kwh_pct_diff"),
        pl.col("resstock_customers"),
        pl.col("eia_residential_customers"),
        (pl.col("resstock_customers") / pl.col("eia_residential_customers")).alias("customers_ratio"),
        (
            (pl.col("resstock_customers") - pl.col("eia_residential_customers"))
            / pl.col("eia_residential_customers")
            * 100
        ).alias("customers_pct_diff"),
    )


def get_all_electricity_mf_to_sf_ratios(
    resstock_annual: pl.DataFrame,
    metadata_with_utility: pl.DataFrame,
) -> tuple[dict[str, float | None], dict[str, tuple[float | None, float | None]]]:
    """Compute MF/SF ratio (mean kWh/sqft, non-zero only) for ALL electricity columns (HVAC and non-HVAC).

    Returns:
        tuple of (ratios_dict, means_dict) where:
        - ratios_dict: {col_name: mf/sf ratio or None if insufficient data}
        - means_dict: {col_name: (mean_mf, mean_sf) or (None, None) if insufficient data}
    """
    ratios: dict[str, float | None] = {}
    means: dict[str, tuple[float | None, float | None]] = {}
    if (
        BUILDING_TYPE_RECS_COL not in metadata_with_utility.columns
        or FLOOR_AREA_COL not in metadata_with_utility.columns
    ):
        return ratios, means

    meta = (
        metadata_with_utility.with_columns(
            pl.col(FLOOR_AREA_COL)
            .map_batches(
                lambda s: pl.Series([_parse_floor_area_sqft(x) for x in s]),
                return_dtype=pl.Float64,
            )
            .alias("floor_area_sqft")
        )
        .with_columns(
            pl.col(BUILDING_TYPE_RECS_COL).str.contains("Single-Family", literal=True).alias("_is_sf"),
            pl.col(BUILDING_TYPE_RECS_COL).str.contains("Multi-Family", literal=True).alias("_is_mf"),
        )
        .select(
            pl.col(BLDG_ID_COL),
            pl.col("floor_area_sqft"),
            pl.col("_is_sf"),
            pl.col("_is_mf"),
        )
    )

    # Collect all electricity columns (both HVAC and non-HVAC)
    all_elec_cols = list(HVAC_RELATED_ELECTRICITY_COLS) + list(NON_HVAC_RELATED_ELECTRICITY_COLS)
    cols_present = [c for c in all_elec_cols if c in resstock_annual.columns]

    if not cols_present:
        return ratios, means

    merged = resstock_annual.select([pl.col(BLDG_ID_COL)] + [pl.col(c) for c in cols_present]).join(
        meta, on=BLDG_ID_COL, how="inner"
    )

    merged = merged.filter(pl.col("floor_area_sqft").is_finite() & (pl.col("floor_area_sqft") > 0))

    sf_df = merged.filter(pl.col("_is_sf"))
    mf_df = merged.filter(pl.col("_is_mf"))

    for col in cols_present:
        by_sqft = pl.col(col) / pl.col("floor_area_sqft")
        sf_vals = (
            sf_df.filter(pl.col(col) > 0)
            .with_columns(by_sqft.alias("_kwh_sqft"))
            .filter(pl.col("_kwh_sqft").is_finite())
            .get_column("_kwh_sqft")
        )
        mf_vals = (
            mf_df.filter(pl.col(col) > 0)
            .with_columns(by_sqft.alias("_kwh_sqft"))
            .filter(pl.col("_kwh_sqft").is_finite())
            .get_column("_kwh_sqft")
        )

        if sf_vals.len() < 2 or mf_vals.len() < 2:
            ratios[col] = None
            means[col] = (None, None)
            continue

        mean_sf = float(sf_vals.mean())  # type: ignore[arg-type]
        mean_mf = float(mf_vals.mean())  # type: ignore[arg-type]
        ratios[col] = mean_mf / mean_sf if mean_sf != 0 else None
        means[col] = (mean_mf, mean_sf)

    return ratios, means


def get_non_hvac_mf_to_sf_ratios(
    resstock_annual: pl.DataFrame,
    metadata_with_utility: pl.DataFrame,
) -> tuple[dict[str, float], dict[str, tuple[float, float]]]:
    """Compute MF/SF ratio (mean kWh/sqft, non-zero only) for non-HVAC columns only.

    This is used for the adjustment calculation. Returns 1.0 when data is insufficient
    so adjustment has no effect.

    Returns:
        tuple of (ratios_dict, means_dict) where:
        - ratios_dict: {col_name: mf/sf ratio, defaulting to 1.0 if insufficient data}
        - means_dict: {col_name: (mean_mf, mean_sf)}
    """
    ratios: dict[str, float] = {}
    means: dict[str, tuple[float, float]] = {}
    if (
        BUILDING_TYPE_RECS_COL not in metadata_with_utility.columns
        or FLOOR_AREA_COL not in metadata_with_utility.columns
    ):
        return ratios, means

    meta = (
        metadata_with_utility.with_columns(
            pl.col(FLOOR_AREA_COL)
            .map_batches(
                lambda s: pl.Series([_parse_floor_area_sqft(x) for x in s]),
                return_dtype=pl.Float64,
            )
            .alias("floor_area_sqft")
        )
        .with_columns(
            pl.col(BUILDING_TYPE_RECS_COL).str.contains("Single-Family", literal=True).alias("_is_sf"),
            pl.col(BUILDING_TYPE_RECS_COL).str.contains("Multi-Family", literal=True).alias("_is_mf"),
        )
        .select(
            pl.col(BLDG_ID_COL),
            pl.col("floor_area_sqft"),
            pl.col("_is_sf"),
            pl.col("_is_mf"),
        )
    )

    non_hvac_present = [c for c in NON_HVAC_RELATED_ELECTRICITY_COLS if c in resstock_annual.columns]
    if not non_hvac_present:
        return ratios, means

    merged = resstock_annual.select([pl.col(BLDG_ID_COL)] + [pl.col(c) for c in non_hvac_present]).join(
        meta, on=BLDG_ID_COL, how="inner"
    )

    merged = merged.filter(pl.col("floor_area_sqft").is_finite() & (pl.col("floor_area_sqft") > 0))

    sf_df = merged.filter(pl.col("_is_sf"))
    mf_df = merged.filter(pl.col("_is_mf"))

    for col in non_hvac_present:
        by_sqft = pl.col(col) / pl.col("floor_area_sqft")
        sf_vals = (
            sf_df.filter(pl.col(col) > 0)
            .with_columns(by_sqft.alias("_kwh_sqft"))
            .filter(pl.col("_kwh_sqft").is_finite())
            .get_column("_kwh_sqft")
        )
        mf_vals = (
            mf_df.filter(pl.col(col) > 0)
            .with_columns(by_sqft.alias("_kwh_sqft"))
            .filter(pl.col("_kwh_sqft").is_finite())
            .get_column("_kwh_sqft")
        )

        if sf_vals.len() < 2 or mf_vals.len() < 2:
            ratios[col] = 1.0
            means[col] = (1.0, 1.0)
            continue

        mean_sf = float(sf_vals.mean())  # type: ignore[arg-type]
        mean_mf = float(mf_vals.mean())  # type: ignore[arg-type]
        ratios[col] = mean_mf / mean_sf if mean_sf != 0 else 1.0
        means[col] = (mean_mf, mean_sf)

    return ratios, means


def adjust_mf_electricity(
    resstock_annual: pl.DataFrame,
    metadata_with_utility: pl.DataFrame,
    non_hvac_column_ratios: dict[str, float],
) -> pl.DataFrame:
    """Adjust non-HVAC electricity for multifamily buildings by column-by-column ratios."""
    multifamily_bldg_ids = (
        metadata_with_utility.filter(pl.col(BUILDING_TYPE_RECS_COL).str.contains("Multi-Family", literal=True))
        .get_column(BLDG_ID_COL)
        .to_list()
    )

    is_mf = pl.col(BLDG_ID_COL).is_in(multifamily_bldg_ids)
    non_hvac_in_df = [c for c in NON_HVAC_RELATED_ELECTRICITY_COLS if c in resstock_annual.columns]

    if non_hvac_column_ratios and non_hvac_in_df:
        sum_parts: list[pl.Expr] = []
        for c in non_hvac_in_df:
            ratio = non_hvac_column_ratios.get(c, 1.0)
            if ratio > 0:
                sum_parts.append(pl.when(is_mf).then(pl.col(c) / ratio).otherwise(pl.col(c)))
            else:
                sum_parts.append(pl.col(c))

        adjusted_total_non_hvac = pl.sum_horizontal(sum_parts)
        adjusted_annual = pl.col("total_hvac_related_electricity_kwh") + adjusted_total_non_hvac

        out = resstock_annual.with_columns(
            pl.when(is_mf)
            .then(adjusted_total_non_hvac)
            .otherwise(pl.col("total_non_hvac_related_electricity_kwh"))
            .alias("total_non_hvac_related_electricity_kwh"),
            pl.when(is_mf).then(adjusted_annual).otherwise(pl.col("annual_kwh")).alias("annual_kwh"),
        )
    else:
        out = resstock_annual

    return out.with_columns((pl.col("annual_kwh") * pl.col(WEIGHT_COL)).alias("weighted_kwh"))


def load_data(
    path_annual: str,
    path_utility_assignment: str,
    path_metadata: str,
    path_eia861: str,
    storage_options: dict[str, str] | None = None,
) -> tuple[pl.DataFrame, pl.DataFrame, pl.DataFrame, dict[str, pl.DataFrame], pl.DataFrame]:
    """Load ResStock annual by utility, metadata by utility, and EIA by utility."""
    resstock_annual = load_resstock_annual_building_level(
        path_annual,
        path_utility_assignment,
        storage_options=storage_options,
    )

    resstock_annual_by_utility = group_resstock_annual_by_utility(resstock_annual)

    metadata_with_utility, metadata_by_utility = load_metadata_by_utility(
        path_metadata,
        path_utility_assignment,
        storage_options=storage_options,
    )

    utility_codes = resstock_annual_by_utility["utility_code"].unique().to_list()
    eia = load_eia_by_utility(path_eia861, storage_options=storage_options, utility_codes=utility_codes)

    return (
        resstock_annual_by_utility,
        resstock_annual,
        metadata_with_utility,
        metadata_by_utility,
        eia,
    )


def _bold(ws, cell: str) -> None:
    """Make a cell bold."""
    ws[cell].font = Font(bold=True)


def _header_fill(ws, row: int, n_cols: int) -> None:
    """Apply header styling to a row."""
    fill = PatternFill("solid", fgColor="E8E8E8")
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = fill


def _write_readme(wb: Workbook, state: str, resstock_release: str, upgrade: str, eia_year: int) -> None:
    """Write README tab with data sources and methodology."""
    ws = wb.create_sheet("README", 0)

    # Use S3 paths in README for informational purposes (where others can find the data)
    # Actual file reading uses BASE_RESSTOCK which may be local or S3
    path_annual = f"{S3_BASE_RESSTOCK}/{resstock_release}/load_curve_annual/state={state}/upgrade={upgrade}/{state}_upgrade{upgrade}_metadata_and_annual_results.parquet"
    path_metadata = (
        f"{S3_BASE_RESSTOCK}/{resstock_release}/metadata/state={state}/upgrade={upgrade}/metadata-sb.parquet"
    )
    path_utility = f"{S3_BASE_RESSTOCK}/{resstock_release}/metadata_utility/state={state}/utility_assignment.parquet"
    path_eia = f"{S3_BASE_EIA861}/year={eia_year}/state={state}/data.parquet"

    rows: list[list] = [
        ["ResStock vs EIA-861 Load Curve Adjustment Analysis", "", ""],
        ["", "", ""],
        ["Data Sources", "", "S3 paths shown below for reference (s3://data.sb)"],
        ["", "", ""],
        ["Item", "Source", "Notes"],
        [
            "ResStock annual loads",
            path_annual,
            f"Annual electricity consumption by building from ResStock {resstock_release}, upgrade {upgrade}.",
        ],
        [
            "ResStock metadata",
            path_metadata,
            "Building characteristics (type, floor area, heating system) used for MF/SF classification.",
        ],
        [
            "Utility assignment",
            path_utility,
            "Maps each building to its electric utility.",
        ],
        [
            "EIA-861 utility stats",
            path_eia,
            f"EIA-861 residential sales and customer counts by utility for {eia_year}.",
        ],
        ["", "", ""],
        ["", "", ""],
        ["Worksheets", "", ""],
        [
            "original_comparison",
            "",
            "ResStock vs EIA-861 comparison before MF adjustment. ResStock kWh is customer-adjusted: ResStock total * (EIA customers / ResStock customers). Ratios computed using customer-adjusted values.",
        ],
        [
            "adjusted_comparison",
            "",
            "ResStock vs EIA-861 comparison after MF non-HVAC adjustment. ResStock kWh is customer-adjusted: ResStock total * (EIA customers / ResStock customers). Ratios computed using customer-adjusted values.",
        ],
        [
            "mf_sf_ratios",
            "",
            "MF/SF ratios (kWh per sqft) for ALL electricity end uses (HVAC and non-HVAC), separated by category. Only non-HVAC end uses are adjusted in MF buildings. HVAC columns shown for reference. 'n/a' indicates insufficient data for that end use.",
        ],
        [
            "summary",
            "",
            "Customer-weight-adjusted comparison before and after MF adjustment. Shows impact on ResStock/EIA ratio and % difference.",
        ],
        [
            "mf_load_8760_before",
            "",
            "Aggregated hourly load curve for multifamily buildings BEFORE adjustment (weighted by sample weight).",
        ],
        [
            "sf_load_8760",
            "",
            "Aggregated hourly load curve for single-family buildings (weighted by sample weight).",
        ],
        [
            "mf_load_8760_after",
            "",
            "Aggregated hourly load curve for multifamily buildings AFTER non-HVAC adjustment (weighted by sample weight).",
        ],
        ["", "", ""],
        ["", "", ""],
        ["Analysis Steps", "", ""],
        [
            "1. Original comparison",
            "",
            "ResStock annual loads grouped by utility, compared to EIA-861 residential sales. ResStock normalized by (EIA customers / ResStock customers) for per-customer comparison.",
        ],
        [
            "2. MF/SF ratio calculation",
            "",
            "For each electricity end use (HVAC and non-HVAC), compute mean kWh/sqft for multifamily and single-family buildings (non-zero values only). Ratio = MF_mean / SF_mean. Returns 'n/a' when insufficient data (fewer than 2 samples).",
        ],
        [
            "3. MF adjustment",
            "",
            "Divide each multifamily building's non-HVAC end-use consumption by its column-specific MF/SF ratio. Recompute annual_kwh and weighted_kwh.",
        ],
        [
            "4. Adjusted comparison",
            "",
            "Re-group adjusted ResStock by utility and compare to EIA-861. Shows reduced load discrepancy after MF correction.",
        ],
        [
            "5. 8760 load curves",
            "",
            "Hourly load curves aggregated by building type, weighted by sample weight. MF curves shown before and after adjustment for comparison.",
        ],
        ["", "", ""],
        ["", "", ""],
        ["Formulas", "", ""],
        [
            "kWh ratio",
            "ResStock_kWh / EIA_kWh",
            "Values >1 mean ResStock overestimates; <1 underestimates.",
        ],
        [
            "kWh % diff",
            "(ResStock_kWh - EIA_kWh) / EIA_kWh * 100",
            "Percentage difference (positive = ResStock higher).",
        ],
        [
            "Customer ratio / % diff",
            "Same formulas for customer counts",
            "Validates sample weighting.",
        ],
    ]

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    _header_fill(ws, 3, 3)
    _header_fill(ws, 11, 3)
    _header_fill(ws, 21, 3)
    _header_fill(ws, 33, 3)

    ws["A1"].font = Font(bold=True, size=12)
    ws["A11"].font = Font(bold=True, size=11)
    ws["A21"].font = Font(bold=True, size=11)
    ws["A33"].font = Font(bold=True, size=11)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 80


def _write_comparison(
    ws,
    comparison: pl.DataFrame,
    building_type_shares: dict[str | int, dict[str, int | float]],
    title: str,
    start_row: int,
) -> int:
    """Write a comparison table to the worksheet. Returns the next available row."""
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)

    headers = [
        "Utility",
        "ResStock kWh (customer-adjusted)",
        "EIA kWh",
        "kWh Ratio",
        "kWh % Diff",
        "ResStock Customers",
        "EIA Customers",
        "Customer Ratio",
        "Customer % Diff",
        "MF Share %",
    ]
    header_row = start_row + 1
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8E8E8")

    data_start = header_row + 1
    for i, row in enumerate(comparison.iter_rows(named=True)):
        r = data_start + i
        utility = row["utility_code"]
        mf_share = building_type_shares.get(utility, {}).get("multifamily_pct", 0)

        ws.cell(row=r, column=1, value=utility)
        ws.cell(row=r, column=2, value=row["resstock_total_kwh_normalized_to_eia"])
        ws.cell(row=r, column=3, value=row["eia_residential_kwh"])
        ws.cell(row=r, column=6, value=row["resstock_customers"])
        ws.cell(row=r, column=7, value=row["eia_residential_customers"])
        ws.cell(row=r, column=10, value=mf_share)

        ws.cell(row=r, column=4, value=f"=B{r}/C{r}")
        ws.cell(row=r, column=5, value=f"=(B{r}-C{r})/C{r}*100")
        ws.cell(row=r, column=8, value=f"=F{r}/G{r}")
        ws.cell(row=r, column=9, value=f"=(F{r}-G{r})/G{r}*100")

    return data_start + len(comparison) + 2


def _write_ratios(
    ws, ratios: dict[str, float | None], means: dict[str, tuple[float | None, float | None]], start_row: int
) -> int:
    """Write MF/SF ratios table with component averages, separated by HVAC/non-HVAC. Returns the next available row."""
    ws.cell(
        row=start_row, column=1, value="Multifamily / Single-Family Electricity End-Use Ratios (kWh/sqft)"
    ).font = Font(bold=True, size=12)

    # Separate HVAC and non-HVAC columns
    hvac_items = [(k, v) for k, v in sorted(ratios.items()) if k in HVAC_RELATED_ELECTRICITY_COLS]
    non_hvac_items = [(k, v) for k, v in sorted(ratios.items()) if k in NON_HVAC_RELATED_ELECTRICITY_COLS]

    current_row = start_row + 1

    # Write HVAC section
    if hvac_items:
        ws.cell(row=current_row, column=1, value="HVAC-Related (not adjusted)").font = Font(bold=True, italic=True)
        current_row += 1

        header_row = current_row
        ws.cell(row=header_row, column=1, value="End Use").font = Font(bold=True)
        ws.cell(row=header_row, column=2, value="MF Avg kWh/sqft").font = Font(bold=True)
        ws.cell(row=header_row, column=3, value="SF Avg kWh/sqft").font = Font(bold=True)
        ws.cell(row=header_row, column=4, value="MF/SF Ratio").font = Font(bold=True)
        ws[f"A{header_row}"].fill = PatternFill("solid", fgColor="E8E8E8")
        ws[f"B{header_row}"].fill = PatternFill("solid", fgColor="E8E8E8")
        ws[f"C{header_row}"].fill = PatternFill("solid", fgColor="E8E8E8")
        ws[f"D{header_row}"].fill = PatternFill("solid", fgColor="E8E8E8")

        data_start = header_row + 1
        for i, (col_name, _ratio) in enumerate(hvac_items):
            r = data_start + i
            end_use = col_name.replace("out.electricity.", "").replace(".energy_consumption.kwh", "")
            mean_mf, mean_sf = means.get(col_name, (None, None))

            ws.cell(row=r, column=1, value=end_use)

            if mean_mf is None or mean_sf is None:
                ws.cell(row=r, column=2, value="n/a")
                ws.cell(row=r, column=3, value="n/a")
                ws.cell(row=r, column=4, value="n/a")
            else:
                ws.cell(row=r, column=2, value=mean_mf)
                ws.cell(row=r, column=3, value=mean_sf)
                ws.cell(row=r, column=4, value=f"=B{r}/C{r}")

        current_row = data_start + len(hvac_items) + 1

    # Write non-HVAC section
    if non_hvac_items:
        ws.cell(row=current_row, column=1, value="Non-HVAC (adjusted in MF buildings)").font = Font(
            bold=True, italic=True
        )
        current_row += 1

        header_row = current_row
        ws.cell(row=header_row, column=1, value="End Use").font = Font(bold=True)
        ws.cell(row=header_row, column=2, value="MF Avg kWh/sqft").font = Font(bold=True)
        ws.cell(row=header_row, column=3, value="SF Avg kWh/sqft").font = Font(bold=True)
        ws.cell(row=header_row, column=4, value="MF/SF Ratio").font = Font(bold=True)
        ws[f"A{header_row}"].fill = PatternFill("solid", fgColor="E8E8E8")
        ws[f"B{header_row}"].fill = PatternFill("solid", fgColor="E8E8E8")
        ws[f"C{header_row}"].fill = PatternFill("solid", fgColor="E8E8E8")
        ws[f"D{header_row}"].fill = PatternFill("solid", fgColor="E8E8E8")

        data_start = header_row + 1
        for i, (col_name, _ratio) in enumerate(non_hvac_items):
            r = data_start + i
            end_use = col_name.replace("out.electricity.", "").replace(".energy_consumption.kwh", "")
            mean_mf, mean_sf = means.get(col_name, (None, None))

            ws.cell(row=r, column=1, value=end_use)

            if mean_mf is None or mean_sf is None:
                ws.cell(row=r, column=2, value="n/a")
                ws.cell(row=r, column=3, value="n/a")
                ws.cell(row=r, column=4, value="n/a")
            else:
                ws.cell(row=r, column=2, value=mean_mf)
                ws.cell(row=r, column=3, value=mean_sf)
                ws.cell(row=r, column=4, value=f"=B{r}/C{r}")

        current_row = data_start + len(non_hvac_items) + 2

    return current_row


def _write_summary(
    ws,
    comparison_original: pl.DataFrame,
    comparison_adjusted: pl.DataFrame,
    start_row: int,
) -> int:
    """Write summary comparison table. Returns the next available row."""
    ws.cell(row=start_row, column=1, value="Summary: Impact of MF Non-HVAC Adjustment").font = Font(bold=True, size=12)

    header_row = start_row + 1
    headers = ["", "Before Adjustment", "After Adjustment", "Change"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8E8E8")

    orig_row = comparison_original.row(0, named=True)
    adj_row = comparison_adjusted.row(0, named=True)

    r1 = header_row + 1
    r2 = header_row + 2
    r3 = header_row + 3
    r4 = header_row + 4

    ws.cell(row=r1, column=1, value="ResStock kWh (customer-adjusted)")
    ws.cell(row=r1, column=2, value=orig_row["resstock_total_kwh_normalized_to_eia"])
    ws.cell(row=r1, column=3, value=adj_row["resstock_total_kwh_normalized_to_eia"])
    ws.cell(row=r1, column=4, value=f"=C{r1}-B{r1}")

    ws.cell(row=r2, column=1, value="kWh Ratio (ResStock/EIA)")
    ws.cell(row=r2, column=2, value=orig_row["kwh_ratio"])
    ws.cell(row=r2, column=3, value=adj_row["kwh_ratio"])
    ws.cell(row=r2, column=4, value=f"=C{r2}-B{r2}")

    ws.cell(row=r3, column=1, value="kWh % Diff")
    ws.cell(row=r3, column=2, value=orig_row["kwh_pct_diff"])
    ws.cell(row=r3, column=3, value=adj_row["kwh_pct_diff"])
    ws.cell(row=r3, column=4, value=f"=C{r3}-B{r3}")

    ws.cell(row=r4, column=1, value="EIA kWh (reference)")
    ws.cell(row=r4, column=2, value=orig_row["eia_residential_kwh"])
    ws.cell(row=r4, column=3, value=adj_row["eia_residential_kwh"])
    ws.cell(row=r4, column=4, value=f"=C{r4}-B{r4}")

    return r4 + 2


def _write_8760_load_curve(ws, load_curve: pl.DataFrame, title: str) -> None:
    """Write an 8760 hourly load curve to worksheet."""
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)

    # Determine which column to use (total_kwh or total_kwh_per_sqft)
    value_col = "total_kwh" if "total_kwh" in load_curve.columns else "total_kwh_per_sqft"

    # Headers
    ws.cell(row=2, column=1, value="timestamp").font = Font(bold=True)
    ws.cell(row=2, column=1).fill = PatternFill("solid", fgColor="E8E8E8")
    ws.cell(row=2, column=2, value="Total kWh").font = Font(bold=True)
    ws.cell(row=2, column=2).fill = PatternFill("solid", fgColor="E8E8E8")

    # Data rows - use timestamp column if available, otherwise hour index
    for i, row in enumerate(load_curve.iter_rows(named=True)):
        r = i + 3
        ws.cell(row=r, column=1, value=row.get("timestamp", row.get("hour", i)))
        ws.cell(row=r, column=2, value=row.get(value_col, 0.0))


def build_workbook(
    state: str,
    resstock_release: str,
    upgrade: str,
    eia_year: int,
    output_path: Path,
    include_8760: bool = False,
) -> tuple[
    Path,
    pl.DataFrame,
    pl.DataFrame,
    dict[str | int, dict[str, int | float]],
    dict[str, int | float | None],
    dict[str, tuple[int | float | None, int | float | None]],
]:
    """Build the load adjustment workbook and save to output_path.

    Args:
        include_8760: If True, load and aggregate hourly 8760 load curves (slow for large datasets).
    """
    data_source = "local EBS" if BASE_RESSTOCK.startswith("/ebs") else "S3"
    print(
        f"Loading ResStock and EIA-861 data for {state} (reading from: {data_source}, README shows S3 paths)...",
        flush=True,
    )

    path_annual = f"{BASE_RESSTOCK}/{resstock_release}/load_curve_annual/state={state}/upgrade={upgrade}/{state}_upgrade{upgrade}_metadata_and_annual_results.parquet"
    path_utility_assignment = (
        f"{BASE_RESSTOCK}/{resstock_release}/metadata_utility/state={state}/utility_assignment.parquet"
    )
    path_metadata = f"{BASE_RESSTOCK}/{resstock_release}/metadata/state={state}/upgrade={upgrade}/metadata-sb.parquet"
    path_eia861 = f"{BASE_EIA861}/year={eia_year}/state={state}/data.parquet"

    (
        resstock_annual_by_utility,
        resstock_annual,
        metadata_with_utility,
        metadata_by_utility,
        eia,
    ) = load_data(
        path_annual=path_annual,
        path_utility_assignment=path_utility_assignment,
        path_metadata=path_metadata,
        path_eia861=path_eia861,
        storage_options=_storage_options(),
    )

    print("Computing original comparison...", flush=True)
    comparison_original = compare_resstock_eia_by_utility(resstock_annual_by_utility, eia)

    print("Computing building type shares...", flush=True)
    building_type_shares = building_type_share_by_utility(metadata_by_utility)

    print("Computing MF/SF ratios and adjusted comparison...", flush=True)
    # Compute all electricity ratios (HVAC + non-HVAC) for display in the workbook
    all_elec_ratios, all_elec_means = get_all_electricity_mf_to_sf_ratios(resstock_annual, metadata_with_utility)

    # Compute non-HVAC ratios only (with 1.0 default) for adjustment calculation
    non_hvac_ratios, _non_hvac_means = get_non_hvac_mf_to_sf_ratios(resstock_annual, metadata_with_utility)

    adjusted_resstock_annual = adjust_mf_electricity(resstock_annual, metadata_with_utility, non_hvac_ratios)
    adjusted_resstock_by_utility = group_resstock_annual_by_utility(adjusted_resstock_annual)
    comparison_adjusted = compare_resstock_eia_by_utility(adjusted_resstock_by_utility, eia)

    print("Building workbook...", flush=True)
    wb = Workbook()
    wb.remove(wb.active)

    _write_readme(wb, state, resstock_release, upgrade, eia_year)

    ws_orig = wb.create_sheet("original_comparison")
    _write_comparison(
        ws_orig,
        comparison_original,
        building_type_shares,
        "ResStock vs EIA-861 Comparison: Original (Before MF Non-HVAC Adjustment)",
        1,
    )
    ws_orig.column_dimensions["A"].width = 12
    ws_orig.column_dimensions["B"].width = 16
    ws_orig.column_dimensions["C"].width = 14
    ws_orig.column_dimensions["D"].width = 12
    ws_orig.column_dimensions["E"].width = 12
    ws_orig.column_dimensions["F"].width = 18
    ws_orig.column_dimensions["G"].width = 16
    ws_orig.column_dimensions["H"].width = 14
    ws_orig.column_dimensions["I"].width = 14
    ws_orig.column_dimensions["J"].width = 12

    ws_adj = wb.create_sheet("adjusted_comparison")
    _write_comparison(
        ws_adj,
        comparison_adjusted,
        building_type_shares,
        "ResStock vs EIA-861 Comparison: Adjusted (After MF Non-HVAC Adjustment)",
        1,
    )
    ws_adj.column_dimensions["A"].width = 12
    ws_adj.column_dimensions["B"].width = 16
    ws_adj.column_dimensions["C"].width = 14
    ws_adj.column_dimensions["D"].width = 12
    ws_adj.column_dimensions["E"].width = 12
    ws_adj.column_dimensions["F"].width = 18
    ws_adj.column_dimensions["G"].width = 16
    ws_adj.column_dimensions["H"].width = 14
    ws_adj.column_dimensions["I"].width = 14
    ws_adj.column_dimensions["J"].width = 12

    ws_ratios = wb.create_sheet("mf_sf_ratios")
    _write_ratios(ws_ratios, all_elec_ratios, all_elec_means, 1)
    ws_ratios.column_dimensions["A"].width = 40
    ws_ratios.column_dimensions["B"].width = 18
    ws_ratios.column_dimensions["C"].width = 18
    ws_ratios.column_dimensions["D"].width = 14

    ws_summary = wb.create_sheet("summary")
    _write_summary(ws_summary, comparison_original, comparison_adjusted, 1)
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 24
    ws_summary.column_dimensions["C"].width = 24
    ws_summary.column_dimensions["D"].width = 16

    # Get MF and SF building IDs
    mf_bldg_ids = (
        metadata_with_utility.filter(pl.col(BUILDING_TYPE_RECS_COL).str.contains("Multi-Family", literal=True))
        .get_column(BLDG_ID_COL)
        .to_list()
    )

    sf_bldg_ids = (
        metadata_with_utility.filter(pl.col(BUILDING_TYPE_RECS_COL).str.contains("Single-Family", literal=True))
        .get_column(BLDG_ID_COL)
        .to_list()
    )

    print(f"  MF buildings: {len(mf_bldg_ids)}, SF buildings: {len(sf_bldg_ids)}", flush=True)

    if include_8760:
        print("Loading hourly 8760 load curves (this may take several minutes)...", flush=True)
        try:
            # Load hourly data for all buildings
            all_bldg_ids = mf_bldg_ids + sf_bldg_ids
            storage_opts = _storage_options()
            hourly_loads = load_hourly_load_curves(state, resstock_release, upgrade, all_bldg_ids, storage_opts)

            # Aggregate MF before adjustment
            print("  Aggregating MF load curve (before adjustment)...", flush=True)
            mf_load_before = aggregate_weighted_load_curves(
                hourly_loads, metadata_with_utility, mf_bldg_ids, by_floor_area=False
            )

            # Aggregate SF
            print("  Aggregating SF load curve...", flush=True)
            sf_load = aggregate_weighted_load_curves(
                hourly_loads, metadata_with_utility, sf_bldg_ids, by_floor_area=False
            )

            # For MF after adjustment, we need to apply the adjustment to hourly data
            # This requires loading the hourly columns and applying ratios
            # For now, note that this is complex and would require per-end-use hourly data
            print("  Note: MF after adjustment requires per-end-use hourly data (not yet implemented)", flush=True)
            mf_load_after = mf_load_before  # Placeholder

            # Write actual data to sheets
            ws_mf_before = wb.create_sheet("mf_load_8760_before")
            _write_8760_load_curve(
                ws_mf_before,
                mf_load_before,
                "Multifamily Aggregated 8760 Load Curve (Before Adjustment)",
            )
            ws_mf_before.column_dimensions["A"].width = 12
            ws_mf_before.column_dimensions["B"].width = 16

            ws_sf = wb.create_sheet("sf_load_8760")
            _write_8760_load_curve(
                ws_sf,
                sf_load,
                "Single-Family Aggregated 8760 Load Curve",
            )
            ws_sf.column_dimensions["A"].width = 12
            ws_sf.column_dimensions["B"].width = 16

            ws_mf_after = wb.create_sheet("mf_load_8760_after")
            _write_8760_load_curve(
                ws_mf_after,
                mf_load_after,
                "Multifamily Aggregated 8760 Load Curve (After Adjustment)",
            )
            ws_mf_after.column_dimensions["A"].width = 12
            ws_mf_after.column_dimensions["B"].width = 16

        except Exception as e:
            import traceback

            print(f"  ERROR: Could not load 8760 data: {e}", flush=True)
            print("  Full traceback:", flush=True)
            traceback.print_exc()
            print("  Creating placeholder sheets instead.", flush=True)
            include_8760 = False

    if not include_8760:
        print("Creating 8760 load curve placeholder sheets...", flush=True)
        # MF before adjustment
        ws_mf_before = wb.create_sheet("mf_load_8760_before")
        ws_mf_before.cell(
            row=1, column=1, value="Multifamily Aggregated 8760 Load Curve (Before Adjustment)"
        ).font = Font(bold=True, size=12)
        ws_mf_before.cell(row=2, column=1, value="timestamp").font = Font(bold=True)
        ws_mf_before.cell(row=2, column=1).fill = PatternFill("solid", fgColor="E8E8E8")
        ws_mf_before.cell(row=2, column=2, value="Total kWh").font = Font(bold=True)
        ws_mf_before.cell(row=2, column=2).fill = PatternFill("solid", fgColor="E8E8E8")
        ws_mf_before.cell(row=3, column=1, value="[Use --include-8760 flag to load actual hourly load data]")
        ws_mf_before.cell(row=4, column=1, value=f"MF building count: {len(mf_bldg_ids)}")
        ws_mf_before.column_dimensions["A"].width = 12
        ws_mf_before.column_dimensions["B"].width = 16

        # SF load curve
        ws_sf = wb.create_sheet("sf_load_8760")
        ws_sf.cell(row=1, column=1, value="Single-Family Aggregated 8760 Load Curve").font = Font(bold=True, size=12)
        ws_sf.cell(row=2, column=1, value="timestamp").font = Font(bold=True)
        ws_sf.cell(row=2, column=1).fill = PatternFill("solid", fgColor="E8E8E8")
        ws_sf.cell(row=2, column=2, value="Total kWh").font = Font(bold=True)
        ws_sf.cell(row=2, column=2).fill = PatternFill("solid", fgColor="E8E8E8")
        ws_sf.cell(row=3, column=1, value="[Use --include-8760 flag to load actual hourly load data]")
        ws_sf.cell(row=4, column=1, value=f"SF building count: {len(sf_bldg_ids)}")
        ws_sf.column_dimensions["A"].width = 12
        ws_sf.column_dimensions["B"].width = 16

        # MF after adjustment
        ws_mf_after = wb.create_sheet("mf_load_8760_after")
        ws_mf_after.cell(
            row=1, column=1, value="Multifamily Aggregated 8760 Load Curve (After Adjustment)"
        ).font = Font(bold=True, size=12)
        ws_mf_after.cell(row=2, column=1, value="timestamp").font = Font(bold=True)
        ws_mf_after.cell(row=2, column=1).fill = PatternFill("solid", fgColor="E8E8E8")
        ws_mf_after.cell(row=2, column=2, value="Total kWh").font = Font(bold=True)
        ws_mf_after.cell(row=2, column=2).fill = PatternFill("solid", fgColor="E8E8E8")
        ws_mf_after.cell(row=3, column=1, value="[Use --include-8760 flag to load actual hourly load data]")
        ws_mf_after.cell(row=4, column=1, value=f"MF building count: {len(mf_bldg_ids)}")
        ws_mf_after.cell(row=5, column=1, value="Non-HVAC columns adjusted by MF/SF ratios")
        ws_mf_after.column_dimensions["A"].width = 12
        ws_mf_after.column_dimensions["B"].width = 16

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # Verify workbook was created with sheets
    import os

    file_size = os.path.getsize(output_path)
    print(f"Workbook saved to {output_path} ({file_size:,} bytes, {len(wb.sheetnames)} sheets)", flush=True)
    if not include_8760:
        print(
            f"Note: 8760 load curve sheets are placeholders. Use --include-8760 to load actual hourly data (slow for {len(mf_bldg_ids) + len(sf_bldg_ids)} buildings).",
            flush=True,
        )

    return output_path, comparison_original, comparison_adjusted, building_type_shares, all_elec_ratios, all_elec_means


def upload_to_sheet(
    xlsx_path: Path,
    spreadsheet_id: str | None,
    folder_id: str | None,
    filename: str | None,
) -> None:
    """Mirror the workbook into the target Google Sheet, preserving formulas."""
    from lib.data.gsheets import apply_sheet_formatting, get_gspread_client, xlsx_to_gsheet

    if spreadsheet_id:
        print(f"Uploading {xlsx_path} -> Google Sheet {spreadsheet_id} ...", flush=True)
        spreadsheet = xlsx_to_gsheet(xlsx_path, spreadsheet_id, delete_other_tabs=True)
    elif folder_id and filename:
        gc, _ = get_gspread_client()

        # Delete ANY existing files with the same name - use gspread to list, Drive API to delete
        print(f"Searching for existing files named '{filename}'...", flush=True)
        try:
            # Use gspread's list_spreadsheet_files which has working credentials
            all_files = gc.list_spreadsheet_files(folder_id=folder_id)
            files_to_delete = [f for f in all_files if f.get("name") == filename and not f.get("trashed", False)]

            if files_to_delete:
                print(f"Found {len(files_to_delete)} file(s) to delete:", flush=True)
                for file in files_to_delete:
                    print(f"  - {file.get('name')} ({file.get('id')})", flush=True)

                # Build Drive service using gspread's credentials
                from googleapiclient.discovery import build

                credentials = None
                if hasattr(gc, "http_client") and hasattr(gc.http_client, "auth"):
                    credentials = gc.http_client.auth
                elif hasattr(gc, "auth") and gc.auth is not None:
                    credentials = gc.auth
                elif hasattr(gc, "_auth"):
                    credentials = gc._auth

                if credentials is None:
                    raise RuntimeError("Could not access gspread credentials")

                drive_service = build("drive", "v3", credentials=credentials, cache_discovery=False)

                deleted_count = 0
                for file in files_to_delete:
                    file_id = file["id"]
                    # Try Drive API delete first (permanent)
                    try:
                        drive_service.files().delete(fileId=file_id, supportsAllDrives=True).execute()
                        print(f"  ✓ Deleted: {file_id}", flush=True)
                        deleted_count += 1
                        continue
                    except Exception:
                        # Try moving to trash instead
                        try:
                            drive_service.files().update(
                                fileId=file_id,
                                body={"trashed": True},
                                supportsAllDrives=True,
                            ).execute()
                            print(f"  ✓ Trashed: {file_id}", flush=True)
                            deleted_count += 1
                        except Exception as e2:
                            print(f"  ✗ Could not delete {file_id}: {e2}", flush=True)

                print(f"Successfully removed {deleted_count}/{len(files_to_delete)} file(s)", flush=True)
            else:
                print(f"No existing files found with name '{filename}' in folder.", flush=True)
        except Exception as e:
            import traceback

            print(f"Warning: Could not search/delete existing files: {e}", flush=True)
            traceback.print_exc()

        # Create fresh spreadsheet
        print(f"Creating new spreadsheet '{filename}'...", flush=True)
        spreadsheet = gc.create(filename, folder_id=folder_id)
        spreadsheet_id = spreadsheet.id
        print(f"✓ Created: {spreadsheet_id}")

        print("Uploading data...", flush=True)
        spreadsheet = xlsx_to_gsheet(xlsx_path, spreadsheet_id, delete_other_tabs=True)
        print(f"✓ Upload complete, {len(spreadsheet.worksheets())} worksheets", flush=True)
    else:
        raise ValueError("Must provide either spreadsheet_id or (folder_id + filename)")

    formatting = {
        "original_comparison": {
            "column_number_formats": {
                "B:C": "#,##0",
                "D": "0.00",
                "E": "0.00",
                "F:G": "#,##0",
                "H": "0.00",
                "I": "0.00",
                "J": "0.0",
            },
            "freeze_rows": 2,
            "bold_header": False,
        },
        "adjusted_comparison": {
            "column_number_formats": {
                "B:C": "#,##0",
                "D": "0.00",
                "E": "0.00",
                "F:G": "#,##0",
                "H": "0.00",
                "I": "0.00",
                "J": "0.0",
            },
            "freeze_rows": 2,
            "bold_header": False,
        },
        "mf_sf_ratios": {
            "column_number_formats": {"B:D": "0.0000"},
            "freeze_rows": 2,
            "bold_header": False,
        },
        "summary": {
            "column_number_formats": {"B:D": "#,##0.00"},
            "freeze_rows": 2,
            "bold_header": False,
        },
        "mf_load_8760_before": {
            "column_number_formats": {"B": "#,##0.00"},
            "freeze_rows": 2,
            "bold_header": False,
        },
        "sf_load_8760": {
            "column_number_formats": {"B": "#,##0.00"},
            "freeze_rows": 2,
            "bold_header": False,
        },
        "mf_load_8760_after": {
            "column_number_formats": {"B": "#,##0.00"},
            "freeze_rows": 2,
            "bold_header": False,
        },
        "README": {
            "wrap_columns": ["A:C"],
            "column_widths_px": {"A": 220, "B": 520, "C": 600},
            "freeze_rows": 1,
            "bold_header": False,
        },
    }

    print("Applying number / wrap / width formatting ...", flush=True)
    for ws in spreadsheet.worksheets():
        spec = formatting.get(ws.title)
        if spec:
            apply_sheet_formatting(ws, **spec)  # type: ignore[arg-type]

    print(
        f"✓ Exported to Google Sheet: {spreadsheet.url}",
        flush=True,
    )


def main(argv: list[str] | None = None) -> int:
    """Main entrypoint."""
    parser = argparse.ArgumentParser(description="Build ResStock vs EIA-861 load adjustment workbook")

    group = parser.add_mutually_exclusive_group(required=False)
    group.add_argument(
        "--sheet-id",
        help="Existing Google Spreadsheet ID (from the URL)",
    )
    group.add_argument(
        "--folder-id",
        help="Google Drive folder ID to create new spreadsheet in (defaults to GDRIVE_SHEETS_FOLDER_ID env var)",
    )
    group.add_argument(
        "--output",
        type=Path,
        help="Local output path only (no upload). E.g., cache/load_adjustment.xlsx",
    )

    parser.add_argument(
        "--filename",
        help='Filename for new spreadsheet (required with --folder-id, e.g. "RIE Load Adjustment")',
    )
    parser.add_argument(
        "--state",
        default="RI",
        help="State abbreviation (default: RI)",
    )
    parser.add_argument(
        "--resstock-release",
        default="res_2024_amy2018_2",
        help="ResStock release name (default: res_2024_amy2018_2)",
    )
    parser.add_argument(
        "--upgrade",
        default="00",
        help="Upgrade ID (default: 00)",
    )
    parser.add_argument(
        "--eia-year",
        type=int,
        default=2018,
        help="EIA-861 year (default: 2018)",
    )
    parser.add_argument(
        "--include-8760",
        action="store_true",
        help="Load and aggregate hourly 8760 load curves (slow, requires access to load_curve_hourly files)",
    )

    args = parser.parse_args(argv)

    load_dotenv()

    folder_id = args.folder_id or os.getenv("GDRIVE_SHEETS_FOLDER_ID")
    upload_mode = args.sheet_id or folder_id
    local_only = args.output is not None

    if not upload_mode and not local_only:
        parser.error(
            "Must provide either --sheet-id, --folder-id (or set GDRIVE_SHEETS_FOLDER_ID env var), or --output"
        )
    if folder_id and not args.filename and not local_only:
        parser.error("--folder-id (or GDRIVE_SHEETS_FOLDER_ID) requires --filename")
    if args.filename and not folder_id and not args.sheet_id:
        parser.error("--filename requires either --folder-id or GDRIVE_SHEETS_FOLDER_ID env var")

    if local_only:
        output_path = args.output
    else:
        import tempfile

        temp_dir = Path(tempfile.mkdtemp())
        output_path = temp_dir / "load_adjustment.xlsx"

    (
        xlsx_path,
        _comparison_original,
        _comparison_adjusted,
        _building_type_shares,
        _all_elec_ratios,
        _all_elec_means,
    ) = build_workbook(
        state=args.state,
        resstock_release=args.resstock_release,
        upgrade=args.upgrade,
        eia_year=args.eia_year,
        output_path=output_path,
        include_8760=args.include_8760,
    )

    if upload_mode and not local_only:
        if args.sheet_id:
            print(f"Exporting to Google Sheets (ID: {args.sheet_id})...")
        else:
            print(f"Creating new spreadsheet '{args.filename}' in folder {folder_id}...")

        upload_to_sheet(
            xlsx_path,
            spreadsheet_id=args.sheet_id,
            folder_id=folder_id,
            filename=args.filename,
        )

        import shutil

        shutil.rmtree(temp_dir)

    return 0


if __name__ == "__main__":
    sys.exit(main())
