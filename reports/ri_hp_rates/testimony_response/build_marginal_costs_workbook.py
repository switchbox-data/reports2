"""Build the supporting workbook for the RIE marginal cost 8760 derivation.

This script documents the start-to-finish derivation of the five RIE / 2025
marginal cost (MC) 8760 parquets used by CAIRO for cost-of-service allocation.
Each tab shows the raw ISO-NE or EIA data, the peak-hour identification method,
and the resulting hourly marginal cost — all as **live formulas** that can be
audited in Excel or Google Sheets.

Components:
  1. Distribution & sub-transmission (PoP top-100 on RIE load)
  2. Bulk transmission (exceedance top-100 on NE system load)
  3. Supply energy (RI zone LMP / 1000)
  4. Supply capacity (FCA exceedance top-100 on SENE load)
  5. Supply ancillary (regulation sum / 1000)

Run from the report directory::

    uv run python -m testimony_response.build_marginal_costs_workbook \\
        --output cache/marginal_costs_rie_2025.xlsx
    uv run python -m testimony_response.build_marginal_costs_workbook --upload
"""

from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ── Constants ─────────────────────────────────────────────────────────────────

UTILITY = "rie"
YEAR = 2025
N_PEAK_HOURS = 100

# S3 paths for the five MC output parquets
S3_DIST_SUB_TX = f"s3://data.sb/switchbox/marginal_costs/ri/dist_and_sub_tx/utility={UTILITY}/year={YEAR}/data.parquet"
S3_BULK_TX = f"s3://data.sb/switchbox/marginal_costs/ri/bulk_tx/utility={UTILITY}/year={YEAR}/data.parquet"
S3_SUPPLY_ENERGY = f"s3://data.sb/switchbox/marginal_costs/ri/supply/energy/utility={UTILITY}/year={YEAR}/data.parquet"
S3_SUPPLY_CAPACITY = (
    f"s3://data.sb/switchbox/marginal_costs/ri/supply/capacity/utility={UTILITY}/year={YEAR}/data.parquet"
)
S3_SUPPLY_ANCILLARY = (
    f"s3://data.sb/switchbox/marginal_costs/ri/supply/ancillary/utility={UTILITY}/year={YEAR}/data.parquet"
)

# S3 raw data sources
S3_EIA_UTILITY_LOADS = "s3://data.sb/eia/hourly_demand/utilities/"
S3_ISONE_ZONE_LOADS = "s3://data.sb/isone/hourly_demand/zones/"
S3_ISONE_LMP = "s3://data.sb/isone/lmp/real_time/zones/"
S3_ISONE_ANCILLARY = "s3://data.sb/isone/ancillary/"
S3_ISONE_FCA = "s3://data.sb/isone/capacity/fca/data.parquet"

# Scalar inputs
AESC_PTF_KW_YEAR = 69.0
SUB_TX_AND_DIST_MC_KW_YR_2019 = 80.24
SUB_TX_AND_DIST_MC_KW_YR_2025 = 101.05  # CPI-adjusted 2019$ -> 2025$
FCA_SENE_BLENDED_KW_YR = 38.373  # 3.980*5 + 2.639*7

# RDP permalink infrastructure
RDP_REF = "eed3f0d"
RDP_GITHUB_BASE = "https://github.com/switchbox-data/rate-design-platform/blob"
REPORTS2_GITHUB_BASE = "https://github.com/switchbox-data/reports2/blob"

DEFAULT_SPREADSHEET_ID = "1GmLOgM90orMbFnhFti169idl9Li6QM8_IQKRWGhEct8"


# ── Permalink helpers ─────────────────────────────────────────────────────────


def _rdp_permalink(rel_path: str) -> str:
    """SHA-pinned GitHub permalink for a rate-design-platform file."""
    return f"{RDP_GITHUB_BASE}/{RDP_REF}/{rel_path}"


def _reports2_head_sha() -> str:
    """Current HEAD sha of the reports2 repo. Cached."""
    if not hasattr(_reports2_head_sha, "_cached"):
        repo_root = Path(__file__).resolve().parents[3]
        sha = subprocess.check_output(
            ["git", "-C", str(repo_root), "rev-parse", "HEAD"],
            text=True,
        ).strip()
        _reports2_head_sha._cached = sha  # type: ignore[attr-defined]
    return _reports2_head_sha._cached  # type: ignore[attr-defined]


def _reports2_permalink(rel_path: str) -> str:
    """SHA-pinned GitHub permalink for a file in the reports2 repo."""
    return f"{REPORTS2_GITHUB_BASE}/{_reports2_head_sha()}/{rel_path}"


# ── Formatting helpers ────────────────────────────────────────────────────────


def _header_fill(ws, row: int, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="E8E8E8")
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = fill


def _autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Data loading ──────────────────────────────────────────────────────────────


def _load_mc_parquet(s3_path: str) -> pl.DataFrame:
    """Load an MC parquet and return only (timestamp, mc_value) columns."""
    df = pl.read_parquet(s3_path)
    numeric_cols = [c for c in df.columns if c != "timestamp" and df[c].dtype.is_numeric() and c not in ("year",)]
    if not numeric_cols:
        raise ValueError(f"No numeric MC column found in {s3_path}: cols={df.columns}")
    return df.select("timestamp", numeric_cols[0]).rename({numeric_cols[0]: "mc_value"})


def load_dist_sub_tx_parquet() -> pl.DataFrame:
    """Load the dist_and_sub_tx MC parquet (8760 rows)."""
    return _load_mc_parquet(S3_DIST_SUB_TX)


def load_bulk_tx_parquet() -> pl.DataFrame:
    """Load the bulk_tx MC parquet (8760 rows)."""
    return _load_mc_parquet(S3_BULK_TX)


def load_supply_energy_parquet() -> pl.DataFrame:
    """Load the supply energy MC parquet (8760 rows)."""
    return _load_mc_parquet(S3_SUPPLY_ENERGY)


def load_supply_capacity_parquet() -> pl.DataFrame:
    """Load the supply capacity MC parquet (8760 rows)."""
    return _load_mc_parquet(S3_SUPPLY_CAPACITY)


def load_supply_ancillary_parquet() -> pl.DataFrame:
    """Load the supply ancillary MC parquet (8760 rows)."""
    return _load_mc_parquet(S3_SUPPLY_ANCILLARY)


def _strip_tz(df: pl.DataFrame, col: str) -> pl.DataFrame:
    """Cast a timezone-aware datetime column to naive (drop tz info)."""
    if df[col].dtype == pl.Datetime and df[col].dtype.time_zone is not None:  # type: ignore[union-attr]
        return df.with_columns(pl.col(col).dt.replace_time_zone(None))
    return df


def _dedup_to_8760(df: pl.DataFrame, ts_col: str = "timestamp") -> pl.DataFrame:
    """Average duplicate timestamps (DST fall-back) and truncate to 8760 rows."""
    numeric_cols = [c for c in df.columns if c != ts_col and df[c].dtype.is_numeric()]
    if df.select(pl.col(ts_col).n_unique()).item() < df.height:
        df = df.group_by(ts_col).agg([pl.col(c).mean() for c in numeric_cols])
    return df.sort(ts_col).head(8760)


def _read_isone_zone_year(base: str, zone: str, year: int) -> pl.DataFrame:
    """Read a single ISO-NE zone/year from Hive-partitioned S3 by direct path."""
    path = f"{base.rstrip('/')}/zone={zone}/year={year}/"
    df = pl.read_parquet(path + "**/*.parquet")
    return _strip_tz(df, "interval_start_et")


def load_rie_hourly_load() -> pl.DataFrame:
    """Load EIA-930 hourly utility demand for RIE, year 2025.

    Reads month-partitioned parquets directly to avoid full Hive scan.
    """
    base = S3_EIA_UTILITY_LOADS.rstrip("/")
    path = f"{base}/region=isone/utility={UTILITY}/year={YEAR}/"
    df = pl.read_parquet(path + "**/*.parquet")
    df = _strip_tz(df, "timestamp")
    df = df.select("timestamp", "load_mw")
    return _dedup_to_8760(df)


ISONE_ALL_ZONES = ["CT", "ME", "NEMA", "NH", "RI", "SEMA", "VT", "WCMA"]


def load_ne_system_load() -> pl.DataFrame:
    """Load ISO-NE 8-zone loads and sum to NE system load for year 2025.

    Reads each zone partition directly (8 reads) rather than scanning the
    entire Hive tree.
    """
    base = S3_ISONE_ZONE_LOADS.rstrip("/") + "/"
    frames = []
    for zone in ISONE_ALL_ZONES:
        zdf = _read_isone_zone_year(base, zone, YEAR)
        zdf = zdf.select("interval_start_et", "load_mw")
        frames.append(zdf)
    df = pl.concat(frames)
    ne_load = (
        df.group_by("interval_start_et")
        .agg(pl.col("load_mw").sum().alias("load_mw"))
        .rename({"interval_start_et": "timestamp"})
    )
    return _dedup_to_8760(ne_load)


def load_sene_load() -> pl.DataFrame:
    """Load ISO-NE RI + SEMA zone loads summed (SENE aggregate) for year 2025."""
    base = S3_ISONE_ZONE_LOADS.rstrip("/") + "/"
    frames = []
    for zone in ["RI", "SEMA"]:
        zdf = _read_isone_zone_year(base, zone, YEAR)
        zdf = zdf.select("interval_start_et", "load_mw")
        frames.append(zdf)
    df = pl.concat(frames)
    sene_load = (
        df.group_by("interval_start_et")
        .agg(pl.col("load_mw").sum().alias("load_mw"))
        .rename({"interval_start_et": "timestamp"})
    )
    return _dedup_to_8760(sene_load)


def load_ri_lmp() -> pl.DataFrame:
    """Load ISO-NE real-time LMP for RI zone, year 2025."""
    base = S3_ISONE_LMP.rstrip("/")
    path = f"{base}/zone=RI/year={YEAR}/"
    df = pl.read_parquet(path + "**/*.parquet")
    df = _strip_tz(df, "interval_start_et")
    df = df.select("interval_start_et", "lmp_usd_per_mwh").rename({"interval_start_et": "timestamp"})
    return _dedup_to_8760(df)


def load_ancillary_prices() -> pl.DataFrame:
    """Load ISO-NE ancillary regulation prices for year 2025."""
    base = S3_ISONE_ANCILLARY.rstrip("/")
    path = f"{base}/year={YEAR}/"
    df = pl.read_parquet(path + "**/*.parquet")
    df = _strip_tz(df, "interval_start_et")
    df = df.select(
        "interval_start_et",
        "reg_service_price_usd_per_mwh",
        "reg_capacity_price_usd_per_mwh",
    ).rename({"interval_start_et": "timestamp"})
    return _dedup_to_8760(df)


# ── Cross-sheet formula references ───────────────────────────────────────────

REF_AESC_PTF = "inputs_scalars!$B$2"
REF_SUB_TX_DIST = "inputs_scalars!$B$3"
REF_FCA_BLENDED = "inputs_scalars!$B$4"
REF_N_PEAK = "inputs_scalars!$B$5"


# ── Tab writers ───────────────────────────────────────────────────────────────


def _write_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README", 0)
    rows: list[list] = [
        ["Marginal Costs Workbook — RIE 2025 (5-component 8760 derivation)", "", ""],
        ["", "", ""],
        # --- Item / Source / Notes section ---
        ["Item", "Source", "Notes"],
        [
            "Generator: dist & sub-TX MC",
            _rdp_permalink("utils/pre/marginal_costs/generate_utility_tx_dx_mc.py"),
            "Produces dist_and_sub_tx parquet via PoP (probability of peak) on utility load.",
        ],
        [
            "Generator: supply energy MC",
            _rdp_permalink("utils/pre/marginal_costs/generate_supply_energy_mc.py"),
            "LMP/1000 direct conversion.",
        ],
        [
            "Generator: supply capacity MC",
            _rdp_permalink("utils/pre/marginal_costs/generate_supply_capacity_mc.py"),
            "FCA exceedance on SENE aggregate load.",
        ],
        [
            "Generator: supply ancillary MC",
            _rdp_permalink("utils/pre/marginal_costs/generate_supply_ancillary_mc.py"),
            "Regulation (service + capacity) / 1000.",
        ],
        [
            "Generator: bulk TX MC",
            _rdp_permalink("utils/pre/marginal_costs/generate_bulk_tx_mc.py"),
            "AESC PTF exceedance on NE system load.",
        ],
        [
            "Core: supply_energy.py",
            _rdp_permalink("utils/pre/marginal_costs/supply_energy.py"),
            "ISO-NE LMP loading and energy MC computation.",
        ],
        [
            "Core: supply_capacity_isone.py",
            _rdp_permalink("utils/pre/marginal_costs/supply_capacity_isone.py"),
            "FCA price resolution and capacity exceedance allocation.",
        ],
        [
            "Core: supply_ancillary.py",
            _rdp_permalink("utils/pre/marginal_costs/supply_ancillary.py"),
            "Ancillary regulation price loading.",
        ],
        [
            "Core: bulk_tx_isone.py",
            _rdp_permalink("utils/pre/marginal_costs/bulk_tx_isone.py"),
            "ISO-NE bulk TX logic (AESC PTF + NE system peak).",
        ],
        [
            "Core: supply_utils.py",
            _rdp_permalink("utils/pre/marginal_costs/supply_utils.py"),
            "Shared utilities: exceedance allocation, prepare_component_output, 8760 alignment.",
        ],
        [
            "Scalars CSV",
            _rdp_permalink("rate_design/hp_rates/ri/config/marginal_costs/ri_marginal_costs_2025.csv"),
            "sub_tx_and_dist_mc_kw_yr = 80.24 (2019$), CPI-adjusted to $101.05 (2025$).",
        ],
        [
            "API: EIA hourly utility loads",
            _rdp_permalink("data/eia/hourly_loads/fetch_zone_loads_parquet.py"),
            'Fetches EIA-930 hourly demand by utility. Source: EIA, "Hourly Electric Grid Monitor (Form EIA-930)," 2025, https://www.eia.gov/electricity/gridmonitor/.',
        ],
        [
            "API: EIA utility load aggregation",
            _rdp_permalink("data/eia/hourly_loads/aggregate_eia_utility_loads.py"),
            "Aggregates raw EIA zone loads into utility-level profiles.",
        ],
        [
            "API: ISO-NE LMP",
            _rdp_permalink("data/isone/lmp/fetch_isone_lmp_parquet.py"),
            'Fetches ISO-NE real-time LMP. Source: ISO NE, "ISO-NE Web Services API: Hourly Locational Marginal Prices," 2025, https://webservices.iso-ne.com/api/v1.1.',
        ],
        [
            "API: ISO-NE zone loads",
            _rdp_permalink("data/isone/hourly_demand/fetch_isone_zone_loads.py"),
            'Fetches ISO-NE 8-zone hourly demand (CELT). Source: ISO NE, "Capacity, Energy, Loads, and Transmission (CELT) Report," 2025, https://www.iso-ne.com/system-planning/system-plans-studies/celt.',
        ],
        [
            "API: ISO-NE ancillary prices",
            _rdp_permalink("data/isone/ancillary/fetch_isone_ancillary_parquet.py"),
            'Fetches ISO-NE regulation clearing prices. Source: ISO NE, "Five-Minute Regulation Clearing Prices (Final)," 2025, https://webservices.iso-ne.com/api/v1.1/fiveminutercp/final/day/.',
        ],
        [
            "FCA clearing prices CSV",
            _rdp_permalink("data/isone/capacity/fca/fca_clearing_prices.csv"),
            'Historical FCA results. Source: ISO NE, "Forward Capacity Auction Results Report," 2024, https://www.iso-ne.com/static-assets/documents/2018/05/fca-results-report.pdf.',
        ],
        [
            "Methodology: RI bulk TX",
            _rdp_permalink("context/methods/marginal_costs/ri_bulk_transmission_marginal_cost.md"),
            "Documents the NE system peak exceedance method for RNS/PTF allocation.",
        ],
        [
            "Methodology: RI supply cost recovery",
            _rdp_permalink("context/domain/marginal_costs/ri_supply_cost_recovery.md"),
            "Documents supply component methodology for RI.",
        ],
        [
            "This workbook builder",
            _reports2_permalink("reports/ri_hp_rates/testimony_response/build_marginal_costs_workbook.py"),
            "Script that generated this workbook.",
        ],
        ["", "", ""],
        # --- Sheet directory ---
        ["Sheet", "What it contains", ""],
        ["inputs_scalars", "Key scalar cost parameters with sources and Zotero citations.", ""],
        [
            "mc_dist_sub_tx",
            "8760 rows: RIE utility load, rank, PoP weight, and dist+sub-TX MC per kWh. Top 100 hours are non-zero.",
            "",
        ],
        [
            "mc_bulk_tx",
            "8760 rows: NE system load, rank, exceedance weight, and bulk TX MC per kWh. Top 100 hours are non-zero.",
            "",
        ],
        [
            "mc_supply_energy",
            "8760 rows: RI zone LMP ($/MWh) and supply energy MC (= LMP / 1000).",
            "",
        ],
        [
            "mc_supply_capacity",
            "8760 rows: SENE aggregate load, rank, exceedance weight, and supply capacity MC. Top 100 hours non-zero.",
            "",
        ],
        [
            "mc_supply_ancillary",
            "8760 rows: regulation service + capacity prices, and supply ancillary MC (= sum / 1000).",
            "",
        ],
        [
            "mc_combined",
            "8760 rows joining all 5 components: delivery total, supply total, grand total MC per kWh.",
            "",
        ],
        ["validation", "Formula-level checks: weight sums, annual cost totals, non-zero hour counts.", ""],
        ["", "", ""],
        # --- Key scalar inputs ---
        ["Key scalar inputs (also live in inputs_scalars)", "Value", "Source"],
        [
            "aesc_ptf_kw_year ($/kW-yr)",
            AESC_PTF_KW_YEAR,
            'AESC 2024 avoided PTF. Source: Synapse, "Avoided Energy Supply Components (AESC) in New England: 2024 Report," 2024.',
        ],
        [
            "sub_tx_and_dist_mc_kw_yr ($/kW-yr, 2025$)",
            SUB_TX_AND_DIST_MC_KW_YR_2025,
            'AESC 2024 dist $80.24/kW-yr (2019$) x CPI 2025/2019. Sources: Synapse, "AESC in New England: 2024 Report," 2024; U.S. BLS, "Consumer Price Index (CUUR0000SA0)," 2025, https://fred.stlouisfed.org/series/CUUR0000SA0.',
        ],
        [
            "fca_sene_blended_kw_yr ($/kW-yr)",
            FCA_SENE_BLENDED_KW_YR,
            'FCA15 SENE $3.980 x 5mo + FCA16 SENE $2.639 x 7mo. Source: ISO NE, "Forward Capacity Auction Results Report," 2024.',
        ],
        ["n_peak_hours", N_PEAK_HOURS, "Convention: top 100 hours for all peak-driven allocations."],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    # Bold section headers
    for header_row in (3, 27, 37):
        _header_fill(ws, header_row, 3)
    _autosize(ws, {"A": 44, "B": 80, "C": 80})
    ws.sheet_view.showGridLines = False


def _write_inputs_scalars(wb: Workbook) -> None:
    ws = wb.create_sheet("inputs_scalars")
    rows = [
        ["key", "value", "source", "notes"],
        [
            "aesc_ptf_kw_year",
            AESC_PTF_KW_YEAR,
            _rdp_permalink("utils/pre/marginal_costs/bulk_tx_isone.py"),
            'AESC 2024 avoided PTF ($/kW-yr). Source: Synapse, "Avoided Energy Supply Components (AESC) in New England: 2024 Report," 2024.',
        ],
        [
            "sub_tx_and_dist_mc_kw_yr",
            SUB_TX_AND_DIST_MC_KW_YR_2025,
            _rdp_permalink("rate_design/hp_rates/ri/config/marginal_costs/ri_marginal_costs_2025.csv"),
            'AESC 2024 dist $80.24 (2019$) adjusted to 2025$ via CPIAUCSL. Sources: Synapse, "AESC in New England: 2024 Report," 2024; U.S. BLS, "Consumer Price Index (CUUR0000SA0)," 2025, https://fred.stlouisfed.org/series/CUUR0000SA0.',
        ],
        [
            "fca_sene_blended_kw_yr",
            FCA_SENE_BLENDED_KW_YR,
            _rdp_permalink("data/isone/capacity/fca/fca_clearing_prices.csv"),
            'Calendar-year 2025 blended: FCA15 SENE $3.980/kW-mo x 5 + FCA16 SENE $2.639/kW-mo x 7 = $38.373/kW-yr. Source: ISO NE, "Forward Capacity Auction Results Report," 2024.',
        ],
        [
            "n_peak_hours",
            N_PEAK_HOURS,
            _rdp_permalink("utils/pre/marginal_costs/supply_utils.py"),
            "Top-N hours for exceedance and PoP allocation (consistent across all peak-driven components).",
        ],
    ]
    for r in rows:
        ws.append(r)
    _header_fill(ws, 1, 4)
    _autosize(ws, {"A": 30, "B": 14, "C": 80, "D": 80})
    ws.sheet_view.showGridLines = False


def _rank_load(load_df: pl.DataFrame) -> pl.DataFrame:
    """Rank load hours descending. Missing hours left as null."""
    rank = (
        load_df.filter(pl.col("load_mw").is_not_null())
        .sort("load_mw", descending=True)
        .with_row_index("rank_0")
        .with_columns((pl.col("rank_0") + 1).cast(pl.Int32).alias("rank"))
        .select("timestamp", "rank")
    )
    return load_df.join(rank, on="timestamp", how="left").sort("timestamp")


def _write_mc_dist_sub_tx(wb: Workbook, rie_load: pl.DataFrame) -> None:
    """Dist & sub-TX: PoP top-100 on RIE utility load."""
    ws = wb.create_sheet("mc_dist_sub_tx")
    headers = [
        "timestamp",
        "rie_load_mw",
        "rank",
        "is_peak_top100",
        "pop_weight",
        "mc_dist_sub_tx_per_kwh",
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    joined = _rank_load(rie_load)

    for i, row in enumerate(joined.iter_rows(named=True), start=2):
        ws.cell(row=i, column=1, value=row["timestamp"].strftime("%Y-%m-%d %H:%M"))
        if row["load_mw"] is not None:
            ws.cell(row=i, column=2, value=float(row["load_mw"]))
        if row["rank"] is not None:
            ws.cell(row=i, column=3, value=int(row["rank"]))
        ws.cell(row=i, column=4, value=f"=IF(ISNUMBER(C{i}),C{i}<={REF_N_PEAK},FALSE)")
        ws.cell(
            row=i,
            column=5,
            value=f"=IFERROR(IF(D{i},B{i}/SUMPRODUCT(($C$2:$C$8761<={REF_N_PEAK})*IFERROR($B$2:$B$8761,0)),0),0)",
        )
        ws.cell(row=i, column=6, value=f"=E{i}*{REF_SUB_TX_DIST}")

    _autosize(ws, {"A": 18, "B": 14, "C": 8, "D": 14, "E": 14, "F": 22})


def _write_exceedance_tab(
    wb: Workbook,
    sheet_name: str,
    load_label: str,
    load_df: pl.DataFrame,
    cost_ref: str,
    mc_label: str,
) -> None:
    """Generic exceedance allocation tab (used for bulk_tx and supply capacity)."""
    ws = wb.create_sheet(sheet_name)
    headers = [
        "timestamp",
        load_label,
        "rank",
        "is_peak_top100",
        "exceedance_threshold",
        "exceedance",
        "exceedance_weight",
        mc_label,
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    joined = _rank_load(load_df)

    for i, row in enumerate(joined.iter_rows(named=True), start=2):
        ws.cell(row=i, column=1, value=row["timestamp"].strftime("%Y-%m-%d %H:%M"))
        if row["load_mw"] is not None:
            ws.cell(row=i, column=2, value=float(row["load_mw"]))
        if row["rank"] is not None:
            ws.cell(row=i, column=3, value=int(row["rank"]))
        ws.cell(row=i, column=4, value=f"=IF(ISNUMBER(C{i}),C{i}<={REF_N_PEAK},FALSE)")
        ws.cell(row=i, column=5, value=f"=LARGE($B$2:$B$8761,{REF_N_PEAK}+1)")
        ws.cell(row=i, column=6, value=f"=IFERROR(IF(D{i},MAX(0,B{i}-E{i}),0),0)")
        ws.cell(
            row=i,
            column=7,
            value=f"=IFERROR(IF(D{i},F{i}/SUMPRODUCT(($C$2:$C$8761<={REF_N_PEAK})*IFERROR(IF($B$2:$B$8761>E{i},$B$2:$B$8761-E{i},0),0)),0),0)",
        )
        ws.cell(row=i, column=8, value=f"=G{i}*{cost_ref}")

    _autosize(
        ws,
        {"A": 18, "B": 18, "C": 8, "D": 14, "E": 20, "F": 14, "G": 16, "H": 24},
    )


def _write_mc_bulk_tx(wb: Workbook, ne_load: pl.DataFrame) -> None:
    """Bulk TX: exceedance top-100 on NE system load."""
    _write_exceedance_tab(
        wb,
        "mc_bulk_tx",
        "ne_system_load_mw",
        ne_load,
        cost_ref=REF_AESC_PTF,
        mc_label="mc_bulk_tx_per_kwh",
    )


def _write_mc_supply_energy(wb: Workbook, lmp_df: pl.DataFrame) -> None:
    """Supply energy: LMP / 1000."""
    ws = wb.create_sheet("mc_supply_energy")
    headers = ["timestamp", "ri_zone_lmp_mwh", "mc_supply_energy_per_kwh"]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    lmp_sorted = lmp_df.sort("timestamp")

    for i, row in enumerate(lmp_sorted.iter_rows(named=True), start=2):
        ws.cell(row=i, column=1, value=row["timestamp"].strftime("%Y-%m-%d %H:%M"))
        if row["lmp_usd_per_mwh"] is not None:
            ws.cell(row=i, column=2, value=float(row["lmp_usd_per_mwh"]))
        ws.cell(row=i, column=3, value=f"=IFERROR(B{i}/1000,0)")

    _autosize(ws, {"A": 18, "B": 18, "C": 24})


def _write_mc_supply_capacity(wb: Workbook, sene_load: pl.DataFrame) -> None:
    """Supply capacity: FCA exceedance top-100 on SENE load."""
    _write_exceedance_tab(
        wb,
        "mc_supply_capacity",
        "sene_load_mw",
        sene_load,
        cost_ref=REF_FCA_BLENDED,
        mc_label="mc_supply_capacity_per_kwh",
    )


def _write_mc_supply_ancillary(wb: Workbook, ancillary_df: pl.DataFrame) -> None:
    """Supply ancillary: (reg_service + reg_capacity) / 1000."""
    ws = wb.create_sheet("mc_supply_ancillary")
    headers = [
        "timestamp",
        "reg_service_mwh",
        "reg_capacity_mwh",
        "mc_supply_ancillary_per_kwh",
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    anc_sorted = ancillary_df.sort("timestamp")

    for i, row in enumerate(anc_sorted.iter_rows(named=True), start=2):
        ws.cell(row=i, column=1, value=row["timestamp"].strftime("%Y-%m-%d %H:%M"))
        if row["reg_service_price_usd_per_mwh"] is not None:
            ws.cell(row=i, column=2, value=float(row["reg_service_price_usd_per_mwh"]))
        if row["reg_capacity_price_usd_per_mwh"] is not None:
            ws.cell(row=i, column=3, value=float(row["reg_capacity_price_usd_per_mwh"]))
        ws.cell(row=i, column=4, value=f"=IFERROR((B{i}+C{i})/1000,0)")

    _autosize(ws, {"A": 18, "B": 16, "C": 16, "D": 26})


def _write_mc_combined(wb: Workbook) -> None:
    """Combined tab: cross-sheet references summing the 5 components."""
    ws = wb.create_sheet("mc_combined")
    headers = [
        "timestamp",
        "mc_dist_sub_tx",
        "mc_bulk_tx",
        "mc_supply_energy",
        "mc_supply_capacity",
        "mc_supply_ancillary",
        "mc_delivery_total",
        "mc_supply_total",
        "mc_total",
    ]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    for i in range(2, 8762):
        # Timestamp from dist_sub_tx tab
        ws.cell(row=i, column=1, value=f"=mc_dist_sub_tx!A{i}")
        # Individual components
        ws.cell(row=i, column=2, value=f"=mc_dist_sub_tx!F{i}")
        ws.cell(row=i, column=3, value=f"=mc_bulk_tx!H{i}")
        ws.cell(row=i, column=4, value=f"=mc_supply_energy!C{i}")
        ws.cell(row=i, column=5, value=f"=mc_supply_capacity!H{i}")
        ws.cell(row=i, column=6, value=f"=mc_supply_ancillary!D{i}")
        # Aggregates
        ws.cell(row=i, column=7, value=f"=B{i}+C{i}")  # delivery total
        ws.cell(row=i, column=8, value=f"=D{i}+E{i}+F{i}")  # supply total
        ws.cell(row=i, column=9, value=f"=G{i}+H{i}")  # grand total

    _autosize(
        ws,
        {
            "A": 18,
            "B": 16,
            "C": 14,
            "D": 18,
            "E": 18,
            "F": 20,
            "G": 18,
            "H": 16,
            "I": 12,
        },
    )


def _write_validation(wb: Workbook) -> None:
    """Validation tab: formula-level checks on weight sums and annual totals."""
    ws = wb.create_sheet("validation")
    headers = ["check", "actual", "expected", "abs_error", "tolerance", "ok"]
    ws.append(headers)
    _header_fill(ws, 1, len(headers))

    checks = [
        (
            "sum(pop_weight) for dist/sub_tx = 1.0",
            "=SUM(mc_dist_sub_tx!E$2:E$8761)",
            "=1",
            1e-6,
        ),
        (
            "sum(mc_dist_sub_tx_per_kwh) = sub_tx_and_dist_mc_kw_yr",
            "=SUM(mc_dist_sub_tx!F$2:F$8761)",
            f"={REF_SUB_TX_DIST}",
            0.01,
        ),
        (
            "sum(exceedance_weight) for bulk_tx = 1.0",
            "=SUM(mc_bulk_tx!G$2:G$8761)",
            "=1",
            1e-6,
        ),
        (
            "sum(mc_bulk_tx_per_kwh) = aesc_ptf_kw_year",
            "=SUM(mc_bulk_tx!H$2:H$8761)",
            f"={REF_AESC_PTF}",
            0.01,
        ),
        (
            "sum(exceedance_weight) for capacity = 1.0",
            "=SUM(mc_supply_capacity!G$2:G$8761)",
            "=1",
            1e-6,
        ),
        (
            "sum(mc_supply_capacity_per_kwh) = fca_blended_kw_yr",
            "=SUM(mc_supply_capacity!H$2:H$8761)",
            f"={REF_FCA_BLENDED}",
            0.01,
        ),
        (
            "non-zero hours dist_sub_tx = 100",
            '=COUNTIF(mc_dist_sub_tx!F$2:F$8761,">0")',
            f"={REF_N_PEAK}",
            0,
        ),
        (
            "non-zero hours bulk_tx = 100",
            '=COUNTIF(mc_bulk_tx!H$2:H$8761,">0")',
            f"={REF_N_PEAK}",
            0,
        ),
        (
            "non-zero hours capacity = 100",
            '=COUNTIF(mc_supply_capacity!H$2:H$8761,">0")',
            f"={REF_N_PEAK}",
            0,
        ),
    ]

    for i, (name, actual, expected, tol) in enumerate(checks, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=actual)
        ws.cell(row=i, column=3, value=expected)
        ws.cell(row=i, column=4, value=f"=ABS(B{i}-C{i})")
        ws.cell(row=i, column=5, value=tol)
        ws.cell(row=i, column=6, value=f'=IF(D{i}<=E{i},"OK","FAIL")')

    _autosize(ws, {"A": 52, "B": 18, "C": 18, "D": 14, "E": 14, "F": 8})
    for r in range(2, 2 + len(checks)):
        ws[f"B{r}"].number_format = "#,##0.000000"
        ws[f"C{r}"].number_format = "#,##0.000000"
        ws[f"D{r}"].number_format = "#,##0.000000"
    ws.sheet_view.showGridLines = False


# ── Main build ────────────────────────────────────────────────────────────────


def _derive_pop_mc(load_df: pl.DataFrame, annual_cost: float, n_peak: int) -> pl.DataFrame:
    """Reproduce PoP allocation in Python (mirrors the spreadsheet formulas)."""
    ranked = _rank_load(load_df)
    peak = ranked.filter(pl.col("rank").is_not_null(), pl.col("rank") <= n_peak)
    peak_load_sum = float(peak["load_mw"].sum())
    return (
        ranked.with_columns(
            pl.when(pl.col("rank").is_not_null() & (pl.col("rank") <= n_peak))
            .then(pl.col("load_mw") / peak_load_sum * annual_cost)
            .otherwise(0.0)
            .alias("mc_derived")
        )
        .select("timestamp", "mc_derived")
        .sort("timestamp")
    )


def _derive_exceedance_mc(load_df: pl.DataFrame, annual_cost: float, n_peak: int) -> pl.DataFrame:
    """Reproduce exceedance allocation in Python (mirrors the spreadsheet formulas)."""
    ranked = _rank_load(load_df)
    threshold = float(ranked.filter(pl.col("rank") == n_peak + 1)["load_mw"].item())
    return (
        ranked.with_columns(
            pl.when(pl.col("rank").is_not_null() & (pl.col("rank") <= n_peak))
            .then(
                (pl.col("load_mw") - threshold).clip(lower_bound=0)
                / (
                    ranked.filter(pl.col("rank").is_not_null(), pl.col("rank") <= n_peak)
                    .select((pl.col("load_mw") - threshold).clip(lower_bound=0).sum())
                    .item()
                )
                * annual_cost
            )
            .otherwise(0.0)
            .alias("mc_derived")
        )
        .select("timestamp", "mc_derived")
        .sort("timestamp")
    )


def _validate_mc_against_parquets(
    rie_load: pl.DataFrame,
    ne_load: pl.DataFrame,
    sene_load: pl.DataFrame,
    lmp_df: pl.DataFrame,
    ancillary_df: pl.DataFrame,
) -> None:
    """Assert that our derivation reproduces the MC parquets passed to CAIRO."""
    tol = 1e-4

    def _check(name: str, derived: pl.DataFrame, parquet: pl.DataFrame) -> None:
        joined = derived.join(parquet, on="timestamp", how="inner").sort("timestamp")
        max_err = float((joined["mc_derived"] - joined["mc_value"]).abs().max())  # type: ignore[arg-type]
        assert max_err < tol, f"{name}: max hourly error = {max_err:.2e} (tol = {tol:.0e})"
        print(f"  {name}: PASS (max error = {max_err:.2e})", flush=True)

    _check(
        "dist_sub_tx",
        _derive_pop_mc(rie_load, SUB_TX_AND_DIST_MC_KW_YR_2025, N_PEAK_HOURS),
        load_dist_sub_tx_parquet(),
    )
    _check(
        "bulk_tx",
        _derive_exceedance_mc(ne_load, AESC_PTF_KW_YEAR, N_PEAK_HOURS),
        load_bulk_tx_parquet(),
    )
    # Supply parquets are stored in $/MWh (energy, ancillary) or scaled by
    # 1000 (capacity), so we derive in those same native units for comparison.
    cap_derived = _derive_exceedance_mc(sene_load, FCA_SENE_BLENDED_KW_YR, N_PEAK_HOURS)
    cap_derived = cap_derived.with_columns(pl.col("mc_derived") * 1000.0)
    _check("supply_capacity", cap_derived, load_supply_capacity_parquet())

    energy_derived = (
        lmp_df.sort("timestamp")
        .with_columns(pl.col("lmp_usd_per_mwh").alias("mc_derived"))
        .select("timestamp", "mc_derived")
    )
    _check("supply_energy", energy_derived, load_supply_energy_parquet())

    anc_derived = (
        ancillary_df.sort("timestamp")
        .with_columns(
            (pl.col("reg_service_price_usd_per_mwh") + pl.col("reg_capacity_price_usd_per_mwh")).alias("mc_derived")
        )
        .select("timestamp", "mc_derived")
    )
    _check("supply_ancillary", anc_derived, load_supply_ancillary_parquet())


def build_workbook(output_path: Path) -> Path:
    """Build and save the .xlsx workbook. Returns the output path."""
    print("Loading raw ISO-NE / EIA data from S3 ...", flush=True)
    rie_load = load_rie_hourly_load()
    print(f"  RIE utility load: {rie_load.height} hours", flush=True)
    ne_load = load_ne_system_load()
    print(f"  NE system load: {ne_load.height} hours", flush=True)
    sene_load = load_sene_load()
    print(f"  SENE aggregate load: {sene_load.height} hours", flush=True)
    lmp_df = load_ri_lmp()
    print(f"  RI zone LMP: {lmp_df.height} hours", flush=True)
    ancillary_df = load_ancillary_prices()
    print(f"  Ancillary prices: {ancillary_df.height} hours", flush=True)

    print("Validating derivation against MC parquets passed to CAIRO ...", flush=True)
    _validate_mc_against_parquets(rie_load, ne_load, sene_load, lmp_df, ancillary_df)

    print("Building workbook ...", flush=True)
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    _write_readme(wb)
    _write_inputs_scalars(wb)
    _write_mc_dist_sub_tx(wb, rie_load)
    _write_mc_bulk_tx(wb, ne_load)
    _write_mc_supply_energy(wb, lmp_df)
    _write_mc_supply_capacity(wb, sene_load)
    _write_mc_supply_ancillary(wb, ancillary_df)
    _write_mc_combined(wb)
    _write_validation(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Wrote {output_path} ({output_path.stat().st_size / 1024:.1f} KB)", flush=True)
    return output_path


# ── Formatting & upload ───────────────────────────────────────────────────────

_TAB_FORMATTING: dict[str, dict] = {
    "README": {
        "wrap_columns": ["A:C"],
        "column_widths_px": {"A": 300, "B": 540, "C": 540},
        "freeze_rows": 1,
        "bold_header": True,
        "bold_rows": [3, 27, 37],
    },
    "inputs_scalars": {
        "column_number_formats": {"B": "#,##0.000"},
        "wrap_columns": ["C:D"],
        "column_widths_px": {"A": 240, "B": 120, "C": 540, "D": 540},
        "freeze_rows": 1,
        "bold_header": True,
    },
    "mc_dist_sub_tx": {
        "column_number_formats": {
            "B": "#,##0.0",
            "E": "0.000000",
            "F": "0.000000",
        },
        "auto_resize_columns": ["A:F"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "mc_bulk_tx": {
        "column_number_formats": {
            "B": "#,##0.0",
            "E": "#,##0.0",
            "F": "#,##0.00",
            "G": "0.000000",
            "H": "0.000000",
        },
        "auto_resize_columns": ["A:H"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "mc_supply_energy": {
        "column_number_formats": {
            "B": "#,##0.00",
            "C": "0.000000",
        },
        "auto_resize_columns": ["A:C"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "mc_supply_capacity": {
        "column_number_formats": {
            "B": "#,##0.0",
            "E": "#,##0.0",
            "F": "#,##0.00",
            "G": "0.000000",
            "H": "0.000000",
        },
        "auto_resize_columns": ["A:H"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "mc_supply_ancillary": {
        "column_number_formats": {
            "B": "#,##0.00",
            "C": "#,##0.00",
            "D": "0.000000",
        },
        "auto_resize_columns": ["A:D"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "mc_combined": {
        "column_number_formats": {
            "B": "0.000000",
            "C": "0.000000",
            "D": "0.000000",
            "E": "0.000000",
            "F": "0.000000",
            "G": "0.000000",
            "H": "0.000000",
            "I": "0.000000",
        },
        "auto_resize_columns": ["A:I"],
        "freeze_rows": 1,
        "bold_header": True,
    },
    "validation": {
        "column_number_formats": {"B": "#,##0.000000", "C": "#,##0.000000", "D": "#,##0.000000"},
        "auto_resize_columns": ["A:F"],
        "freeze_rows": 1,
        "bold_header": True,
    },
}


def upload_to_sheet(xlsx_path: Path, spreadsheet_id: str) -> None:
    """Mirror the workbook into the target Google Sheet, preserving formulas."""
    from lib.data.gsheets import apply_sheet_formatting, xlsx_to_gsheet

    print(f"Uploading {xlsx_path} -> Google Sheet {spreadsheet_id} ...", flush=True)
    spreadsheet = xlsx_to_gsheet(xlsx_path, spreadsheet_id, delete_other_tabs=True)
    print("Applying number / wrap / width formatting ...", flush=True)
    for ws in spreadsheet.worksheets():
        spec = _TAB_FORMATTING.get(ws.title)
        if spec:
            apply_sheet_formatting(ws, **spec)
    print(
        f"Done. View at https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit",
        flush=True,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build the RIE marginal costs 8760 workbook.")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("cache/marginal_costs_rie_2025.xlsx"),
        help="Output .xlsx path. Default: cache/marginal_costs_rie_2025.xlsx",
    )
    parser.add_argument(
        "--upload",
        action="store_true",
        help="Upload to the default Google Sheet after building.",
    )
    parser.add_argument(
        "--spreadsheet-id",
        default=DEFAULT_SPREADSHEET_ID,
        help=f"Override the upload target Sheet id. Default: {DEFAULT_SPREADSHEET_ID}",
    )
    args = parser.parse_args(argv)

    out = build_workbook(args.output)
    if args.upload:
        upload_to_sheet(out, args.spreadsheet_id)
    return 0


if __name__ == "__main__":
    sys.exit(main())
